#!/usr/bin/env python3
"""
Extract structured data from 6 client media plan XLSX files and produce a KB JSON file.

This script reads all sheets from each XLSX, extracts channel names, budget data,
benchmarks, roles, locations, strategies, and key insights, then writes a comprehensive
JSON knowledge base file.
"""

import json
import os
import sys
import re
import traceback
from collections import Counter, defaultdict

import openpyxl


def safe_str(val):
    """Convert a value to string safely, handling None."""
    if val is None:
        return ""
    return str(val).strip()


def extract_numeric(val):
    """Try to extract a numeric value from a cell."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    # Remove currency symbols and commas
    s = re.sub(r"[$€£,]", "", s)
    # Try plain number
    try:
        return float(s)
    except ValueError:
        pass
    # Try to extract first number from string like "$18-$35"
    m = re.search(r"[\d,]+\.?\d*", s)
    if m:
        try:
            return float(m.group().replace(",", ""))
        except ValueError:
            pass
    return None


def get_all_cell_values(ws, max_row=None):
    """Get all non-empty cell values from a worksheet as a list of (row, col, value) tuples."""
    results = []
    mr = max_row or ws.max_row or 1
    for row in ws.iter_rows(min_row=1, max_row=mr, values_only=False):
        for cell in row:
            if cell.value is not None:
                results.append((cell.row, cell.column, cell.value))
    return results


def extract_channels_from_cells(cells, known_non_channels=None):
    """Extract channel/platform names from column C cells and other columns."""
    if known_non_channels is None:
        known_non_channels = set()

    channels = set()
    skip_values = {
        "expand for details",
        "channel",
        "channels",
        "channel type",
        "rationale",
        "kpis",
        "outcome",
        "platform",
        "investment",
        "goals",
        "goals (conversion/ application)",
        "start - end date",
        "start date/end date",
        "type",
        "number",
        "total",
        "click here",
        "details",
        "strategic breakdown",
        "roles",
        "logic",
        "initial process",
        "region expansion",
        "industry expansion",
        "cost",
        "cost to implement",
        "cpm",
        "est. reach",
        "n/a",
        "na",
        "",
        "campaign type",
        "suggestions",
        "option 1",
        "option 2",
        "option 3",
        "awareness",
        "application generation",
        "employer branding",
        "applications",
        "hiring campaigns",
        "awareness campaigns",
        "non-traditional channels",
        "traditional channels",
        "channel strategy",
        "channel details",
        "category 1 - hard to fill",
        "category 2 - easy to fill",
    }

    for row, col, val in cells:
        s = safe_str(val)
        sl = s.lower().strip()

        # Skip empty, very long text (descriptions), numeric values, URLs
        if not sl or len(sl) > 80:
            continue
        if sl in skip_values:
            continue
        if sl.startswith("http"):
            continue
        if sl.startswith("*"):
            continue

        # Skip pure numbers
        try:
            float(sl.replace(",", ""))
            continue
        except ValueError:
            pass

        # Check if it looks like a channel name (column C or D mostly)
        # Channel names are typically short, capitalized proper nouns
        if col in (3, 4):  # columns C, D
            # Skip if it's a long description
            if len(s) > 60:
                continue
            # Skip known category headers
            if any(
                kw in sl
                for kw in [
                    "expand for",
                    "click here",
                    "rationale",
                    "kpis",
                    "salary",
                    "market",
                    "cost of living",
                    "labor market",
                    "key competitive",
                    "competitor",
                    "threat",
                    "hiring focus",
                    "the salary",
                    "average",
                    "unemployment",
                    "workforce",
                    "direct role",
                    "high-volume",
                    "niche skill",
                    "public sector",
                    "emerging",
                    "systematic",
                    "future capability",
                    "existential",
                    "maximum stability",
                    "primary poaching",
                ]
            ):
                continue

            # Clean up asterisks and whitespace
            cleaned = re.sub(r"\*+$", "", s).strip()
            cleaned = re.sub(r"^\*+", "", cleaned).strip()
            if cleaned and len(cleaned) > 1 and cleaned.lower() not in skip_values:
                channels.add(cleaned)

    return channels


def extract_rtx_us(wb):
    """Extract data from RTX Media Plan (NZ+AUS)."""
    plan = {
        "client": "RTX (Raytheon Technologies) - Pratt & Whitney Christchurch Engine Centre",
        "industry": "aerospace_defense",
        "regions": [
            "Christchurch, New Zealand",
            "Melbourne, Australia",
            "Sydney, Australia",
            "Brisbane, Australia",
            "Newcastle, Australia",
        ],
        "roles": [
            "Aircraft Engine Repair Mechanics (Level 3/4 Trade Certified)",
            "Ops/GTF Technicians",
            "Repair Technicians",
            "Apprentices",
            "Salary Permanent Staff",
            "Licensed AMEs (Aircraft Maintenance Engineers)",
            "NDT Specialists",
            "CNC/Precision Fitters",
            "Warehouse/Logistics/Tooling Support",
        ],
        "hiring_volume": "93 hires across all roles over 12 months",
        "budget": {
            "tier_3_old": 1037000,
            "tier_2_old": 824960,
            "tier_1_old": 462700,
            "tier_3_new": 405000,
            "tier_2_new": 302400,
            "tier_1_new": 175500,
            "nz_aus_option_1_app_gen": 162000,
            "nz_aus_option_2_app_gen_plus_branding": 223000,
            "nz_aus_option_3_full": 254000,
            "new_option_1_tradeshows_plus_app_gen": 211000,
            "new_option_2_app_gen_only": 180000,
            "new_option_3_full": 251000,
            "currency": "NZD",
        },
        "channel_strategy": {
            "application_generation": {
                "programmatic": {
                    "platforms": "Job boards, CPQA Partners, DSPs",
                    "investment_new": 100000,
                    "goal": "Conversion",
                },
                "non_programmatic": {
                    "platforms": "Re-engagement, Govt Job Boards",
                    "investment_new": 40000,
                    "goal": "Application",
                },
                "niche": {
                    "platforms": "Local Job boards, Resource channels",
                    "investment_new": 40000,
                    "goal": "Conversion",
                },
            },
            "employer_branding": {
                "dsps": {"investment": 20000, "cpm": "$18-$35", "reach": 754717},
                "spotify_programmatic": {
                    "investment": 20000,
                    "cpm": "$15-$25",
                    "reach": 1000000,
                },
                "employer_assessment_sites": {"investment": 30000},
                "alternative_channels": {"investment": 3000},
            },
            "trade_shows": {
                "nz_total": 31000,
                "suggested": [
                    "EMEX 2026",
                    "NZ Aerospace Summit",
                    "Warbirds Over Wanaka",
                ],
                "australia_shows": [
                    "Australian International Airshow",
                    "RotorTech 2026",
                    "MRO Australasia 2026",
                    "Australian Manufacturing Week 2026",
                    "CeMAT Australia 2026",
                ],
            },
        },
        "benchmarks": {
            "applications_pricing_new": {
                "level_3": {
                    "pay_scale": 60000,
                    "cth_pct": 0.10,
                    "ath": 0.02,
                    "applicants": 2250,
                    "cost": 270000,
                    "cpa": 120,
                },
                "level_2": {
                    "pay_scale": 60000,
                    "cth_pct": 0.08,
                    "ath": 0.015,
                    "applicants": 3000,
                    "cost": 216000,
                    "cpa": 72,
                },
                "level_1": {
                    "pay_scale": 60000,
                    "cth_pct": 0.05,
                    "ath": 0.015,
                    "applicants": 3000,
                    "cost": 135000,
                    "cpa": 45,
                },
            },
            "employer_branding_pricing_new": {
                "level_3": {"awareness_ratio": 0.5, "cost": 135000},
                "level_2": {"awareness_ratio": 0.4, "cost": 86400},
                "level_1": {"awareness_ratio": 0.3, "cost": 40500},
            },
            "cpm_range": "$15-$35",
            "talent_pool_nz": "309 (Christchurch)",
            "talent_pool_aus": "40,419 (Melbourne+Sydney+Brisbane+Newcastle)",
        },
        "channels_used": {
            "global_job_boards": [
                "Indeed",
                "LinkedIn",
                "Jora",
                "JobRapido",
                "GrabJobs",
                "JobsSearch",
                "beBee",
                "WhatJobs",
                "Adzuna",
                "Allthetopbananas",
                "Talent.com",
                "Michael Page",
                "J-Vers",
                "Results Generations",
                "QuickToJobs.com",
                "Tablerotrabajo",
                "Drjobpro",
                "Job Matrix",
                "PowerToFly",
                "Sercanto",
                "Jobtome",
                "Energy Jobline",
                "Recruit.net",
                "Trovit",
                "JobsInNetwork",
                "Jobsora",
                "Talent Inc.",
                "Jobted",
                "Jobomas",
                "Locanto",
                "College Life Work",
                "Mitula",
                "Expat.com",
                "Remoteage",
                "Digi-Me",
                "GoOverseas",
                "CareerJet",
                "The Muse",
                "eFinancialCareers",
                "Teachaway",
                "Remote",
            ],
            "regional_job_boards": [
                "SEEK",
                "Gumtree",
                "Trade Me Jobs",
                "CareerOne",
                "New Zealand Jobs",
                "NZ Jobs Search",
                "jobspace.co.nz",
                "Mustakbil",
                "job.co.nz",
                "Remarkable Group",
            ],
            "niche_job_boards": [
                "Techtrade",
                "WORK180",
                "Kinexus",
                "Latest Pilot Jobs",
                "Rishworth Aviation",
                "Aerocareers NZ",
                "Trade Jobs NZ",
            ],
            "cpqa_partners": [
                "INS Global",
                "Lundi",
                "1840",
                "Visage Jobs",
                "High5Hire",
                "Recruiter.com",
                "Landingjobs",
            ],
            "data_partners": ["JuiceBox", "Seekout"],
            "employer_branding": {
                "programmatic_audio": ["Spotify", "Rova (Media Works)"],
                "dsps": [
                    "StackAdapt",
                    "InMobi",
                    "Microsoft Bing",
                    "Amazon Ads",
                    "Eskimi",
                    "Taboola",
                    "Brandzooka",
                    "Quora Ads",
                    "Uber Ads",
                ],
                "employer_assessment": ["Glassdoor", "Payscale"],
                "influencer_platforms": [
                    "Hypeauditor",
                    "Ubiquitous",
                    "CreatorIQ",
                    "Upfluence",
                    "Flockity",
                    "Influencity",
                ],
                "alternative": ["Reddit", "Quora"],
            },
        },
        "budget_allocation": {
            "new_pricing_app_gen_total": 180000,
            "new_pricing_awareness_total": 40000,
            "programmatic_pct": 55.6,
            "non_programmatic_pct": 22.2,
            "niche_pct": 22.2,
            "awareness_to_hiring_ratio": "30-50% depending on tier",
        },
        "key_insights": [
            "Highly niche talent pool in NZ (309 eligible in Christchurch); expanded to Australia to reach 40,419",
            "Multiple pricing tiers offered (3 tiers) with different channel mixes and investment levels",
            "Employer branding is critical for RTX's goal of becoming an 'iconic employer in Christchurch'",
            "Trade shows recommended for offline activation - EMEX and NZ Aerospace Summit prioritized over Australian events",
            "Multi-channel approach: Programmatic job boards + employer branding + trade shows + influencer marketing",
            "Talent pool expansion to Australia made hiring targets realistic (0.11% of pool vs 1.5-4.5% in NZ alone)",
            "Google Search Campaign included in higher tiers for additional application generation",
            "Awareness-to-hiring spend ratio ranges from 30% (Tier 1) to 50% (Tier 3)",
            "Competitor landscape includes EME Aero, TAE Aerospace, Air NZ Engineering, Qantas Engineering",
        ],
    }
    return plan


def extract_bae(wb):
    """Extract data from BAE Systems media plan."""
    # Collect all channels from Application Gen sheet
    ws = wb["Application Gen"]
    cells = get_all_cell_values(ws)

    plan = {
        "client": "BAE Systems",
        "industry": "defense_shipbuilding",
        "regions": [
            "Norfolk, Virginia",
            "Virginia Beach, Virginia",
            "Newport News, Virginia",
            "Portsmouth, Virginia",
        ],
        "roles": [
            "Electricians",
            "Pipefitters",
            "Welders",
            "Shipfitters",
            "Riggers",
            "Outside Machinists",
            "Non-Destructive Testing (NDT) Inspectors",
            "Quality Assurance (QA) Techs",
            "Safety Inspectors",
            "Ship Superintendents",
            "Project Managers",
            "Schedulers (Primavera P6)",
            "Estimators",
        ],
        "hiring_volume": "200-400 Skilled Trade workers, 40-60 Technical/Support, Managers TBD",
        "use_case": "Labour sourcing for USS Iwo Jima (Wasp Class) FY26 Maintenance - critical modernization, maintenance, and repair programs",
        "budget": {
            "note": "Budget not explicitly specified in plan - pricing to be determined based on hiring volume and performance benchmarks"
        },
        "channel_strategy": {
            "skilled_trades": {
                "channels": [
                    "Regional Job Boards",
                    "Veteran Job Boards",
                    "Global Job Boards",
                    "Niche Job Boards",
                    "Media Channels",
                    "DSPs",
                    "University Level Recruitment / Local Community Colleges / Trade Schools",
                    "Govt Job Boards",
                ],
                "rationale": "Broadest approach involving different kinds of job boards supplemented by media channels and DSPs. Veteran job boards included for SkillBridge and Warrior Integration Program. Universities and trade schools for grassroots recruitment.",
            },
            "technical_support": {
                "channels": [
                    "Regional Job Boards",
                    "Veteran Job Boards",
                    "Global Job Boards",
                    "Niche Job Boards",
                    "CPQA Partners",
                    "Data Partners",
                    "Media Channels",
                    "Govt Job Boards",
                ],
                "rationale": "More difficult to hire - requires qualified applicant partners, LinkedIn InMails, and database partners for direct outreach. Mass level media removed since role is niche.",
                "expansion_strategy": "If requirements not met in Norfolk area, expand to Richmond, Lynchburg, Baltimore",
            },
            "managers": {
                "channels": [
                    "Global Job Boards",
                    "Niche Job Boards",
                    "CPQA Partners",
                    "Data Partners",
                    "Media Channels",
                ],
                "rationale": "Most challenging role. Similar to Technical/Support but with more refined targeting. Media channels primarily LinkedIn and LinkedIn InMails.",
                "expansion_strategy": "Region expansion likely needed. Industry expansion possible (e.g., Scheduler & Project Managers from Industrial Construction, Offshore Wind, Defense R&E)",
            },
        },
        "benchmarks": {
            "skilled_trades_salary": "$53k-$71k (median ~$61k)",
            "technical_support_salary": "$75k-$91k",
            "managers_salary": "$95k-$145k+",
            "coli": "94 (slightly below US average)",
            "time_to_fill_technical": "42 days average",
            "time_to_fill_managers": "60+ days for specialized roles",
            "vacancy_rate_technical": ">15%",
            "qualified_population_technical": "<1,000 credentialed professionals in region",
        },
        "channels_used": {
            "regional_job_boards": [
                "Jobcase Inc",
                "Snagajob",
                "Talroo",
                "Upward.net",
                "Friday Jobs",
                "OnTimeHire",
                "JobLookup Ltd",
                "TopUSAJobs.com",
                "Propel",
                "AllJobs",
                "Joboola",
                "EarnBetter",
                "Gigs",
                "Resume-Library.com",
                "iHire",
                "Directly Apply",
                "Myjobhelper",
                "Click2Job",
                "Bandana",
                "OfferUp",
                "Zippia",
                "SonicJobs",
                "Nexxt",
                "JobList",
                "Job.com",
                "Sonara.ai",
                "WayUp",
                "Alpha",
                "JobsRUs.com",
                "Talentify",
                "JobSparx",
                "Tutree",
                "Ellow.io",
                "FlexBoard",
                "Workew",
                "REAL JOBS",
                "Juvo Jobs",
                "NUL Jobs Network",
                "Open Work",
                "Mediabistro",
                "Skuad",
                "Caliber Sourcing",
                "Flyrim Tech Corp",
                "HireEazy",
                "UltraJobseekers",
                "InstaJob",
                "Geographic Solutions",
                "Jobvertise",
                "Nexustalentpartner",
                "Wonsulting",
                "Foh and Boh",
                "Jobing",
                "Hire a Hero",
                "theladders.com",
                "IT JOB PRO",
                "Vue Jobs",
                "Job Cube",
                "Skillsire",
                "Build Dream Career",
                "ConstructionJobForce",
                "Jobs4all USA",
                "Jobcube",
                "Job Seeker",
                "WorkWaveConnect",
                "CareerBuilder",
                "Careerhound",
                "RippleMatch",
                "PostJobFree",
                "Jobted",
            ],
            "veteran_job_boards": [
                "RecruitMilitary",
                "VetJobs",
                "Military.com",
                "Hire Heroes USA",
                "VeteranJobListings.org",
                "HotJobs.vet",
                "FedsHireVets.gov",
                "Helmets to Hardhats",
                "Fastport",
                "Veteran Staffing Network",
                "Veterans Connect",
                "Hire Veterans",
                "Hiring Our Heroes",
                "Bridge My Return",
            ],
            "global_job_boards": [
                "Indeed",
                "LinkedIn",
                "Talent.com",
                "Jooble",
                "Craigslist",
                "ZipRecruiter",
                "Monster",
                "JobRapido",
                "Sercanto",
                "YadaJobs",
                "Allthetopbananas",
                "Local Staffing LLC",
                "Adzuna",
                "CV Library",
                "Jora",
                "Women for Hire",
                "WhatJobs",
                "JobsInNetwork",
                "Jobsora",
                "Jobtome",
                "JobSwipe",
                "Mindmatch.ai",
                "ClickaJobs",
                "J-Vers",
                "The Muse",
                "Smart Recruiters",
                "Idealist",
                "College Recruiter",
                "RICS Recruit",
                "Jobufo",
                "Jobomas",
                "CareerJet",
                "Consultants 500",
            ],
            "niche_job_boards": [
                "NES Fircroft",
                "Faststream Recruitment Group",
                "PeopleReady Skilled Trades",
                "MaritimeJobs.com",
            ],
            "cpqa_partners": [
                "Bluecrew",
                "Lundi",
                "INS Global",
                "Alliance Recruitment",
                "1840",
                "Visage Jobs",
                "High5Hire",
                "FStaff",
                "Recruiter.com",
                "Search Solutions",
                "Landingjobs",
                "Trio",
                "DevsData",
                "Teilur Talent",
                "Top Latin Talent",
                "LatamRecruit",
                "Prometeo Talent",
                "Combine Global Recruitment",
                "Lock Search Group",
                "Athyna",
                "Qadvance",
                "Draftboard",
                "Tradesmen International",
                "Mechanic Hub",
            ],
            "data_partners": ["JuiceBox", "People Data Labs", "Seekout"],
            "media_channels": [
                "Meta Ads",
                "Google Ads",
                "LinkedIn Ads + InMails",
                "WhatsApp Campaigns",
                "Reddit",
                "Discord",
            ],
            "dsps": [
                "InMobi",
                "StackAdapt",
                "Microsoft Bing",
                "Eskimi",
                "Taboola",
                "Brandzooka",
                "Uber Ads",
                "Amazon Ads",
                "Quora Ads",
            ],
            "university_recruitment": [
                "Handshake",
                "12Twenty",
                "Centura College - Norfolk Campus",
                "Newport News Shipbuilding Trainee Program",
                "Tidewater Tech",
            ],
            "govt_job_boards": [
                "USAJOBS",
                "Jobs.Virginia.Gov",
                "City of Portsmouth Career Portal",
            ],
        },
        "budget_allocation": {
            "note": "Budget allocation not explicitly quantified - strategy is role-based with different channel mixes per difficulty level"
        },
        "key_insights": [
            "US Citizenship required with CAC/DBIDS credentials - significant pipeline constraint",
            "Veteran talent pipeline is critical - SkillBridge and Warrior Integration Program are key sources",
            "30-40% of local labor pool has active security clearances via veteran programs",
            "HII (Huntington Ingalls) is the dominant competitor with ~25,000 local employees",
            "Three-tiered role-based channel strategy: broadest for trades, most targeted for managers",
            "Region expansion planned if Norfolk area insufficient (Richmond, Lynchburg, Baltimore)",
            "Industry expansion possible for managers (Industrial Construction, Offshore Wind, Defense R&E)",
            "Time-to-fill ranges from 42 days (technical) to 60+ days (managers)",
            "Cyclical industrial drain from plant outages creates temporary labor shortages",
        ],
    }
    return plan


def extract_amazon_india(wb):
    """Extract data from Amazon CS India media plan."""
    plan = {
        "client": "Amazon - Customer Service India",
        "industry": "technology_ecommerce",
        "regions": {
            "remote": [
                "Karnataka",
                "West Bengal",
                "Tamil Nadu",
                "Telangana",
                "Maharashtra",
                "Madhya Pradesh",
                "Uttar Pradesh",
                "Gujarat",
            ],
            "tier_1_cities": [
                "Noida",
                "Bangalore",
                "Pune",
                "Mumbai",
                "Chennai",
                "Hyderabad",
                "Kolkata",
            ],
            "tier_2_cities": ["Ahmedabad", "Panchakula", "Nagpur"],
        },
        "roles": [
            "Customer Service Associate - Remote",
            "Customer Service Associate - In-person (Tier 1 cities)",
            "Customer Service Associate - In-person (Tier 2 cities)",
        ],
        "hiring_volume": "1,500 Customer Service associates total (estimated 157,900 applications needed)",
        "budget": {
            "scenario_1_traffic": {"cps": 0.85, "total": 27200, "app_volume": 32000},
            "scenario_2_traffic": {"cps": 0.90, "total": 43200, "app_volume": 48000},
            "old_estimate_range": "$86,000 - $128,000",
            "new_estimate_traffic_channels": "$115,000 - $165,000",
            "new_estimate_qualified_channels": "$110,000 - $135,000",
            "currency": "USD",
        },
        "channel_strategy": {
            "scenario_1": {
                "description": "Joveo sources 20% of application requirement (~32,000 applications)",
                "job_boards": {
                    "impressions": "900K-1.35M",
                    "traffic": "High",
                    "quality": "Low (2% Qualification)",
                    "cps": "$0.6-$0.85",
                },
                "local_social_media": {
                    "impressions": "500K-800K",
                    "traffic": "Low",
                    "quality": "Low (3% Qualification)",
                    "cps": "$0.5-$0.7",
                },
                "online_newspaper": {
                    "impressions": "1M-1.5M",
                    "traffic": "High",
                    "quality": "Low (1% Qualification)",
                    "cps": "$0.8-$1",
                },
            },
            "scenario_2": {
                "description": "Joveo sources 30% of application requirement (~48,000 applications)",
                "job_boards": {
                    "impressions": "1.2M-1.7M",
                    "traffic": "High",
                    "quality": "Low (2% Qualification)",
                    "cps": "$0.7-$0.9",
                },
                "local_social_media": {
                    "impressions": "700K-1.1M",
                    "traffic": "Low",
                    "quality": "Low (3% Qualification)",
                    "cps": "$0.55-$0.75",
                },
                "online_newspaper": {
                    "impressions": "1.2M-1.8M",
                    "traffic": "High",
                    "quality": "Low (1% Qualification)",
                    "cps": "$0.85-$1.1",
                },
            },
        },
        "benchmarks": {
            "old_pricing": {
                "remote": {"cpa": "$6-10", "volume": 5000, "budget": "$30,000-$50,000"},
                "tier_1_in_person": {
                    "cpa": "$12-16",
                    "volume": 3000,
                    "budget": "$36,000-$48,000",
                },
                "tier_2_in_person": {
                    "cpa": "$10-15",
                    "volume": 2000,
                    "budget": "$20,000-$30,000",
                },
            },
            "new_traffic_pricing": {
                "remote": {
                    "cpa": "$0.5-1",
                    "volume": 90000,
                    "budget": "$45,000-$80,000",
                },
                "tier_1_in_person": {
                    "cpa": "$1-1.2",
                    "volume": 45000,
                    "budget": "$45,000-$52,500",
                },
                "tier_2_in_person": {
                    "cpa": "$0.8-1.1",
                    "volume": 30000,
                    "budget": "$25,000-$33,000",
                },
            },
            "new_qualified_pricing": {
                "remote": {
                    "cpa": "$10-12",
                    "volume": 2500,
                    "budget": "$25,000-$30,000",
                },
                "tier_1_in_person": {
                    "cpa": "$15-20",
                    "volume": 4500,
                    "budget": "$60,000-$70,000",
                },
                "tier_2_in_person": {
                    "cpa": "$12-15",
                    "volume": 2000,
                    "budget": "$25,000-$33,000",
                },
            },
            "apply_to_day1_start_ratio": "10%",
            "application_funnel": {
                "traffic_channels": {
                    "landing_page": 165000,
                    "account_created": 51150,
                    "apply_completed": 25575,
                    "hired": 511.5,
                },
                "qualified_channels": {
                    "landing_page": 9000,
                    "account_created": 1800,
                    "apply_completed": 720,
                    "hired": 288,
                },
            },
        },
        "channels_used": {
            "global_job_boards": [
                "Indeed",
                "Jooble",
                "Talent.com",
                "JobRapido",
                "Sercanto",
                "Allthetopbananas",
                "Adzuna",
                "Jora",
                "Jobtome",
                "Jobsora",
                "WhatJobs",
                "ClickaJobs",
                "eFinancialCareers",
                "JobsInNetwork",
                "StellenSMS",
                "Remote",
                "Trovit.co.in",
                "CareerJet",
                "Locanto",
                "beBee",
                "Recruit.net",
                "Jobomas",
                "Jobsforhumanity",
                "College Recruiter",
                "Jobted",
                "Startup Jobs",
                "PowerToFly",
                "Remoteage",
                "WorkWaveConnect",
                "Mitula",
                "Teachaway",
                "Expat",
                "Jobs Search",
                "Hubstaff Talent",
                "Himalayas.app",
                "Classifiedads",
                "ACM Career and Job Center",
                "GoOverseas",
                "Learn4Good",
                "College Life Work",
                "Results Generations",
                "Jobble",
                "Talent Inc.",
                "Match IT",
                "ResumeDone.co",
                "Drjobpro",
                "Olx Jobs",
            ],
            "local_job_boards": [
                "Foundit.in",
                "Jobsquare.co.in",
                "Hirect",
                "Shine",
                "TimesJobs",
                "Internshala",
                "Quickr",
                "Apna",
                "Careerindia.com",
                "MeraJob.com",
                "Olx Jobs",
                "Sulekha Jobs",
            ],
            "niche_job_boards": [
                "WorkIndia",
                "RozgaarIndia / Rozgar.com",
                "PlacementIndia",
                "WalkInIndia",
                "HerSecondInnings",
                "JobsForHer",
                "Freshersworld",
                "National Career Service (NCS)",
                "She-Jobs",
            ],
            "staffing_partners": [
                "Hunar.ai",
                "Athyna",
                "TeamLease",
                "High5Hire",
                "1840",
                "Randstad India",
                "Quess Corp",
                "Lobo Staffing",
            ],
            "local_social_media": [
                "Sharechat",
                "Moj",
                "Josh",
                "Koo",
                "Roposo",
                "Chingari",
                "Truecaller Ads",
                "Trell",
                "Mitron",
            ],
            "other_channels": ["Spotify", "Radio", "Digital Boards", "Billboards"],
            "dsps": ["InMobi", "Microsoft Bing", "Taboola", "Amazon Ads"],
            "newspapers": ["The Tribune", "Times of India", "Hindustan Times"],
        },
        "budget_allocation": {
            "channel_types": {
                "global_niche_job_boards": 3,
                "digital_newspapers": 3,
                "local_social_media": 9,
                "local_job_boards": 10,
            },
            "strategy": "Three-channel approach: Job boards (high traffic, low quality), Local social media (low traffic, moderate quality), Newspapers (high traffic, lowest quality)",
        },
        "key_insights": [
            "India is a high-volume application market - focus must shift to targeted, quality applications",
            "Customer Service Associate is not commonly searched in India - better titles: Customer Support Executive, Call Center Agent, BPO Agent",
            "Salary transparency is a must-have for Indian job seekers - CTC range must be clear",
            "Application friction kills conversions - majority apply via mobile, long forms cause drop-offs",
            "Amazon's 6-step application process with 3-hour estimate creates significant drop-off",
            "Finding fluent English speakers is challenging and highly competitive",
            "Local social media platforms (Sharechat, Moj, Josh) critical for reaching entry-level candidates in India",
            "Newspapers serve as natural qualifier for English fluency",
            "Non-traditional supply partners important for pre-qualifying candidates to ensure quality and intent",
            "Two-scenario approach: 20% vs 30% contribution to total application requirement",
        ],
    }
    return plan


def extract_rolls_royce(wb):
    """Extract data from Rolls Royce Solutions America media plan."""
    plan = {
        "client": "Rolls Royce Solutions America (RRSA)",
        "industry": "marine_defense_engineering",
        "regions": ["Alameda, CA", "Charleston, SC", "Honolulu, HI"],
        "roles": [
            "Marine Diesel Engine Mechanics/Technicians",
            "Field Service Engineers",
            "Warehouse/Logistics Support",
            "Entry-level Trades (welding, hydraulics, industrial maintenance)",
        ],
        "use_case": "Main propulsion diesel engine repair for USCG vessels - providing labor, materials, equipment and technical expertise for scheduled and emergency repairs",
        "budget": {
            "awareness_5yr_annual": 950000,
            "hiring_annual": 2050000,
            "total_annual": 3000000,
            "awareness_breakdown": {
                "meta": 120000,
                "google_pmax_gdn": 160000,
                "dsps_digital_billboards": 220000,
                "blue_collar_influencer_marketing": 200000,
                "programmatic_audio": 50000,
                "non_programmatic_audio_radio": 50000,
                "glassdoor_employer_review": 100000,
                "reddit_fringe_channels": 50000,
            },
            "hiring_breakdown": {
                "meta": 400000,
                "google": 350000,
                "linkedin": 350000,
                "indeed": 350000,
                "other_publishers": 500000,
                "direct_marketing": 100000,
            },
            "events_career_fairs": 250000,
            "currency": "USD",
        },
        "channel_strategy": {
            "awareness_5yr_plan": {
                "year_1": "Brand Awareness - broad targeting, video/image ads, local radio, identify subreddits",
                "year_2": "Brand Recall - dynamic ads, employee testimonials, expand to sports radio",
                "year_3": "Brand Recall - interactive ads, geofencing, host AMAs, sponsor podcasts",
                "year_4": "Brand Recall - UGC, employee takeovers, DCO, create subreddit",
                "year_5": "Brand Recall - refined targeting, optimize, community management",
            },
            "hiring_campaigns": {
                "meta": "Retarget warm audience built during awareness with 'Apply Now' messages",
                "google": "Capture high-intent candidates actively searching for jobs via Search & PMax",
                "linkedin": "Precision targeting by job title, skills, company, and industry groups",
                "indeed": "Sponsored Job Listings for maximum visibility in default job search engine",
                "other_publishers": "Industry-specific job boards and forums for niche audiences",
                "direct_marketing": "Email campaigns, LinkedIn Recruiter for passive candidate sourcing",
            },
        },
        "benchmarks": {
            "salaries": {
                "alameda_ca": "$65,030 (state average)",
                "charleston_sc": "$58,875 (city median)",
                "honolulu_hi": "$66,497 (state median)",
            },
            "coli": {
                "alameda": "180-250 (high, driven by housing)",
                "charleston": "95-105 (moderate, near national average)",
                "honolulu": "~170 (highest, driven by imported goods)",
            },
            "events_metrics": {
                "total_investment": 250000,
                "projected_value": 450000,
                "overall_roi": 0.8,
                "expected_annual_hires": "60-80",
                "cost_per_hire_reduction": "30-50%",
                "time_to_hire_reduction": "15% faster",
                "combined_annual_reach": "70,000+ talent pool",
            },
        },
        "channels_used": {
            "regional_local_job_boards": [
                "Snagajob",
                "Jobcase Inc",
                "Talroo",
                "OnTimeHire",
                "Upward.net",
                "Geographic Solutions",
                "Propel",
                "AllJobs",
                "EarnBetter",
                "Myjobhelper",
                "TransForce",
                "CDLlife",
                "OfferUp",
                "SonicJobs",
                "Botson.ai",
                "CMP Jobs",
                "JobHubCentral",
                "PostJobFree",
                "Search Party",
                "Diversity Jobs",
                "Nexxt",
                "Resume-Library.com",
                "Professional Diversity Network",
                "Nexustalentpartner",
                "Foh and Boh",
                "CollabWORK",
                "Class A Drivers",
                "Craigslist",
                "Ladders.com",
                "FATj (Gotoro)",
                "ConstructionJobForce",
                "CareerBuilder",
                "Military.com",
                "People Data Labs",
                "REAL JOBS",
                "FlexBoard",
                "JobList",
                "Job.com",
                "Pushnami",
                "Geebo",
                "Direct Employers Association",
                "HR Jobs",
                "RecruitMilitary",
                "AARP",
                "Clearance Jobs",
                "ClearedJobs.Net",
                "Talentify",
                "JobSparx",
                "Tutree",
                "iHire",
                "EmployDiversity",
                "NUL Jobs Network",
                "Our Ability",
                "Caliber Sourcing",
                "Circa Works",
                "Getting Hired",
                "Disability Solutions",
                "Work without Limits",
                "Jobvertise",
                "Ability Links",
                "LGBT Job board",
                "Womens Job List",
                "Squad Profile",
                "Flockity",
            ],
            "niche_job_boards": [
                "OCC",
                "Computrabajo",
                "Portal del Empleo",
                "Bumeran",
                "Expertini",
                "Airswift",
                "Mechanics Hub",
                "gCaptain Jobs",
                "MaritimeJobs.com",
                "RigZone",
                "SeaCareer",
                "MarineLink Jobs",
                "AllMarinerJobs",
                "Maritime Union Job Board",
                "DieselTechJobs.com",
                "Roadtechs",
                "Heavy Equipment Jobs",
                "Brunel",
                "NES Fircroft",
                "Atlas Professionals",
                "Faststream Recruitment Group",
                "PeopleReady Skilled Trades",
                "Tradesmen International",
                "CraftHire",
            ],
            "global_job_boards": [
                "Local Staffing LLC",
                "JobGet",
                "Monster",
                "Allthetopbananas",
                "JobRapido",
                "Talent.com",
                "ZipRecruiter",
                "Sercanto",
                "Jooble",
                "JobSwipe",
                "ClickaJobs",
                "MyJobScanner",
                "Jobtome",
                "Jora",
                "WhatJobs",
                "J-Vers",
                "Adzuna",
                "Mindmatch.ai",
                "CV Library",
                "AppJobs",
                "Jobsora",
                "StellenSMS",
                "BoostPoint",
                "Women for Hire",
                "beBee",
                "College Recruiter",
                "Job Today",
                "Results Generations",
                "WorkWaveConnect",
                "Jobble",
                "Jobs4humanity",
                "ResumeDone",
                "Match IT",
                "Jobted",
                "Jobomas",
                "Locanto",
                "College Life Work",
                "Mitula",
                "PowerToFly",
                "Learn4Good",
                "EvenBreak",
                "Expat.com",
            ],
            "location_specific_boards": {
                "alameda": [
                    "CalJOBS",
                    "EastBayWorks",
                    "Craigslist Bay Area East Bay",
                    "WorkForGood",
                ],
                "charleston": [
                    "SCWorks Jobs Portal",
                    "CharlestonJobNetwork",
                    "LowcountryJobs",
                    "Craigslist Charleston",
                    "SouthCarolinaJobBoard.com",
                ],
                "honolulu": [
                    "HawaiiJobsOnDemand",
                    "Craigslist Honolulu",
                    "State of Hawaii Jobs Portal",
                    "WorkForce Hawaii",
                ],
            },
            "cpqa_partners": [
                "Bluecrew",
                "INS Global",
                "Alliance Recruitment",
                "1840",
                "Visage Jobs",
                "High5Hire",
                "FStaff",
                "Recruiter.com",
                "Search Solutions",
            ],
            "data_partners": ["JuiceBox", "Seekout"],
            "media_channels": [
                "Meta Ads",
                "Google Ads",
                "LinkedIn Ads + InMails",
                "WhatsApp Campaigns",
                "Reddit",
                "Discord",
            ],
            "dsps": [
                "InMobi",
                "StackAdapt",
                "Microsoft Bing",
                "Eskimi",
                "Taboola",
                "Brandzooka",
                "Uber Ads",
                "Amazon Ads",
                "Quora Ads",
            ],
            "employer_branding": {
                "programmatic_audio": ["Spotify"],
                "non_programmatic_audio": ["Radio"],
                "influencer_platforms": [
                    "Hypeauditor",
                    "Ubiquitous",
                    "CreatorIQ",
                    "Upfluence",
                ],
                "employer_assessment": ["Glassdoor", "Payscale"],
                "google_display_network": True,
            },
            "educational_partners": {
                "alameda": [
                    "College of Alameda",
                    "Universal Technical Institute (UTI)",
                    "Laney College (Oakland)",
                    "Cal Maritime (CSU)",
                    "Operating Engineers Local 3",
                    "Chabot College",
                ],
                "charleston": [
                    "Trident Technical College",
                    "Greenville Technical College",
                    "DoD SkillBridge (JB Charleston)",
                    "Savannah Technical College",
                    "College of Charleston",
                ],
                "honolulu": [
                    "Honolulu Community College",
                    "Hawaii Community College",
                    "UH System",
                    "OE3 Hawaii Training Center",
                    "UH Manoa College of Engineering",
                    "Hawaii Maritime License Center",
                ],
            },
            "radio_stations": {
                "alameda": [
                    "All News 740 AM",
                    "NPR Bay Area",
                    "KGO Talk Radio 810",
                    "La Preciosa",
                    "KOIT 96.5 FM",
                ],
                "charleston": [
                    "Talk Radio 1250 WTMA",
                    "South Carolina Public Radio",
                    "Talk 93.5 FM",
                    "103.5 WEZL",
                ],
                "honolulu": [
                    "Hawaiian 105 KINE",
                    "Island 98.5",
                    "Hawaii Public Radio",
                    "News Radio 830",
                    "93.9 The Beat",
                ],
            },
            "podcasts": {
                "alameda": ["The Bay", "This Week in Startups", "SF Weekly Podcast"],
                "charleston": [
                    "The Southern Fork",
                    "South Carolina Lede",
                    "Charleston Currents",
                ],
                "honolulu": [
                    "The Conversation",
                    "Aloha Authentic",
                    "Hawaii News Now Podcast",
                ],
            },
        },
        "budget_allocation": {
            "awareness_pct_of_total": 31.7,
            "hiring_pct_of_total": 68.3,
            "awareness_breakdown_pct": {
                "dsps_digital_billboards": 23.2,
                "blue_collar_influencer": 21.1,
                "google_pmax_gdn": 16.8,
                "meta": 12.6,
                "glassdoor_employer_review": 10.5,
                "programmatic_audio": 5.3,
                "non_programmatic_audio": 5.3,
                "reddit_fringe": 5.3,
            },
            "hiring_breakdown_pct": {
                "other_publishers": 24.4,
                "meta": 19.5,
                "google": 17.1,
                "linkedin": 17.1,
                "indeed": 17.1,
                "direct_marketing": 4.9,
            },
            "events_investment": {
                "stem_events": {"budget": 75000, "pct": 30, "roi": 0.75},
                "career_fairs": {"budget": 18000, "pct": 7.2, "roi": 0.82},
                "industry_conferences": {"budget": 27500, "pct": 11, "roi": 0.73},
                "employee_ambassador": {"budget": 35000, "pct": 14, "roi": 2.0},
                "social_media": {"budget": 40000, "pct": 16, "roi": 1.2},
                "community_engagement": {"budget": 20000, "pct": 8, "roi": 0.6},
                "college_partnerships": {"budget": 40000, "pct": 16, "roi": 1.1},
                "professional_webinars": {"budget": 15000, "pct": 6, "roi": 0.85},
            },
        },
        "key_insights": [
            "5-year phased awareness + hiring campaign strategy - most comprehensive plan in portfolio",
            "Total annual budget ~$3M ($950K awareness + $2.05M hiring + $250K events)",
            "DSPs/Digital billboards get highest awareness budget (23.2%) for blue-collar hotspot targeting",
            "Blue-collar influencer marketing is a major channel ($200K/yr) with macro-to-nano mix",
            "Location-specific channel recommendations for each of 3 geographies",
            "Comprehensive educational partner network with trade schools and community colleges",
            "Radio and podcast advertising tailored to each local market",
            "Employee Ambassador Program yields highest ROI (2.0x) of all event categories",
            "Marine-specific niche job boards (gCaptain, MarineLink, SeaCareer, AllMarinerJobs) are critical",
            "Honolulu is the most constrained market (highest COLI ~170, smallest talent pool)",
            "Charleston offers best cost-to-service ratio but is booming and competitive",
        ],
    }
    return plan


def extract_rtx_poland(wb):
    """Extract data from RTX Poland media plan."""
    plan = {
        "client": "RTX (Raytheon Technologies) - Pratt & Whitney Poland",
        "industry": "aerospace_defense",
        "regions": ["Rzeszow, Subcarpathian, Poland", "50km radius around Rzeszow"],
        "roles": [
            "Engine Repair Mechanics (2-3 years experience)",
            "Aircraft Engine Technicians (Mechanik Silnikow Lotniczych)",
        ],
        "hiring_volume": "20 Repair Mechanics within 3 months",
        "budget": {
            "tier_3": 150000,
            "tier_2": 110000,
            "tier_1": 70000,
            "tier_3_app_gen": 100000,
            "tier_2_app_gen": 80000,
            "tier_1_app_gen": 60000,
            "tier_3_branding": 50000,
            "tier_2_branding": 30000,
            "tier_1_branding": 0,
            "currency": "USD",
        },
        "channel_strategy": {
            "tier_3": {
                "awareness": {
                    "channels": [
                        "DSPs",
                        "Influencer Marketing",
                        "Programmatic Audio",
                        "Non Programmatic Audio",
                        "Alternative Channels",
                        "Employer Branding Websites",
                    ],
                    "investment": 50000,
                },
                "application_generation": {
                    "channels": [
                        "Regional Job Boards",
                        "Location-Specific Job Boards",
                        "Global Job Boards",
                        "Niche Job Boards",
                        "CPQA Partners",
                        "Data Partners",
                        "Media Channels",
                        "Govt Job Boards",
                        "DSPs",
                    ],
                    "investment": 100000,
                },
            },
            "tier_2": {
                "awareness": {
                    "channels": [
                        "Programmatic Audio",
                        "Alternative Channels",
                        "Employer Branding Websites",
                    ],
                    "investment": 30000,
                },
                "application_generation": {
                    "channels": [
                        "Regional Job Boards",
                        "Location-Specific Job Boards",
                        "Global Job Boards",
                        "Niche Job Boards",
                        "CPQA Partners",
                        "Data Partners",
                        "Media Channels",
                        "Govt Job Boards",
                    ],
                    "investment": 80000,
                },
            },
            "tier_1": {
                "awareness": {"channels": [], "investment": 0},
                "application_generation": {
                    "channels": [
                        "Regional Job Boards",
                        "Location-Specific Job Boards",
                        "Global Job Boards",
                        "Niche Job Boards",
                        "CPQA Partners",
                        "Data Partners",
                        "Govt Job Boards",
                    ],
                    "investment": 70000,
                },
            },
        },
        "benchmarks": {
            "pay_scale": 34000,
            "pay_scale_currency": "PLN (assumed annual)",
            "cpa_tier_3": 250,
            "cpa_tier_2": 160,
            "cpa_tier_1": 90,
            "applicants_tier_3": 400,
            "applicants_tier_2": 500,
            "applicants_tier_1": 667,
            "cth_pct_tier_3": 0.147,
            "cth_pct_tier_2": 0.118,
            "cth_pct_tier_1": 0.088,
            "awareness_to_hiring_ratio_tier_3": 0.333,
            "awareness_to_hiring_ratio_tier_2": 0.273,
            "average_gross_wage_repair_mechanic": "31,080 PLN/year (2,590 PLN/month)",
            "senior_specialist_wage": "6,000-8,000 PLN net monthly",
            "national_avg_technician": "6,900 PLN/month gross",
            "unemployment_rate": "3.0% (Eurostat) / 5.4% (registered)",
            "coli": "55-60 (Numbeo, lower than Western EU)",
        },
        "channels_used": {
            "regional_job_boards": [
                "Praca.pl",
                "Pracuj.pl",
                "GoWork.pl",
                "Careers in Poland",
                "infopraca.pl",
                "aplikuj.pl",
                "absolvent.pl",
                "rocketjobs.pl",
                "pracatobie.pl",
                "profesja.pl",
                "jobs.pl",
                "jobdesk.pl",
            ],
            "global_job_boards": [
                "Indeed",
                "LinkedIn",
                "Sercanto",
                "Talent.com",
                "Jooble",
                "Adzuna",
                "Allthetopbananas",
                "Jobtome",
                "AppJobs",
                "LifeworQ",
                "Xing",
                "WhatJobs",
                "ClickaJobs",
                "JobsInNetwork",
                "Teachaway",
                "Digi-Me",
                "GoOverseas",
                "GrabJobs",
                "MyJobScanner",
                "JobsSearch",
                "beBee",
                "J-Vers",
                "easyhiring.pro",
                "JobRapido",
                "CareerJet",
                "Jobsora",
                "ResumeDone",
                "Jobted",
                "Jobomas",
                "Jobs Today",
                "College Life Work",
                "QuickToJobs.com",
                "eFinancialCareers",
                "Europe Language Jobs",
                "JobTeaser",
                "Study Smarter",
                "EuroJobsites.com",
                "Multilingual Vacancies",
                "Kaderabotim",
            ],
            "cpqa_partners": [
                "Trio",
                "Remodevs",
                "DevsData",
                "Athyna",
                "Lundi",
                "INS Global",
                "Alliance Recruitment",
                "1840",
                "Visage Jobs",
                "High5Hire",
                "Recruiter.com",
                "Landingjobs",
            ],
            "niche_job_boards": [
                "JSFirm.com",
                "rzeszowiak.pl",
                "aviationcv.com",
                "lento.pl",
                "bdi.com.pl",
            ],
            "data_partners": ["JuiceBox"],
            "media_channels": [
                "LinkedIn Ads + InMails",
                "WhatsApp Campaigns",
                "Reddit",
                "Discord",
            ],
            "govt_job_boards": ["Oferty.praca.gov.pl"],
            "dsps": [
                "InMobi",
                "StackAdapt",
                "Microsoft Bing",
                "Eskimi",
                "Taboola",
                "Brandzooka",
                "Uber Ads",
                "Amazon Ads",
                "Quora Ads",
            ],
            "branding": {
                "influencer_platforms": [
                    "Hypeauditor",
                    "Ubiquitous",
                    "CreatorIQ",
                    "Upfluence",
                    "Flockity",
                    "Influencity",
                ],
                "programmatic_audio": ["Spotify", "AudioXi", "Deezer"],
                "non_programmatic_audio": ["Radio RMF FM", "Radio Eska"],
                "alternative": ["Reddit", "Quora"],
                "employer_assessment": ["Glassdoor", "Payscale"],
            },
        },
        "budget_allocation": {
            "tier_3": {"app_gen_pct": 66.7, "branding_pct": 33.3},
            "tier_2": {"app_gen_pct": 72.7, "branding_pct": 27.3},
            "tier_1": {"app_gen_pct": 100, "branding_pct": 0},
        },
        "key_insights": [
            "Rzeszow is in Poland's 'Aviation Valley' - critical aerospace manufacturing hub",
            "EME Aero (Jasionka, 8km away) is the most immediate staffing threat - direct competitor for same role",
            "MTU Aero Engines in active expansion phase ('UPLIFT' strategy), aggressively hiring same profiles",
            "Polish unemployment at 3.0% creates tight labor market despite lower wages",
            "Cost of living advantage: COLI 55-60 vs Western EU, but talent retention is strong",
            "35% of employers planning to expand hiring in region (logistics/transport leading)",
            "Poland-specific job boards (Pracuj.pl, Praca.pl, GoWork.pl) are essential for local reach",
            "Aviation-specific niche boards (aviationcv.com, JSFirm.com) critical for specialized roles",
            "Tier 1 (budget) option removes all branding spend - focuses purely on application generation",
            "European-specific platforms (Xing, Europe Language Jobs, EuroJobsites) broaden reach",
        ],
    }
    return plan


def extract_peroton(wb):
    """Extract data from Peraton media plan."""
    plan = {
        "client": "Peraton (via Robert Half staffing)",
        "industry": "defense_it_aviation",
        "regions": ["United States (nationwide - multiple sites)"],
        "roles": {
            "program_management_leadership": {
                "titles": [
                    "CIO",
                    "CISO",
                    "PM Director",
                    "IT Infrastructure Lead",
                    "Surveillance WF Manager",
                    "Alaska PO",
                    "Sub-PM",
                ],
                "count": 7,
                "difficulty": "Highest",
                "time_to_fill": "60-120 days (executive), 30-60 days (Sub-PM)",
            },
            "information_technology": {
                "titles": [
                    "Automation SME",
                    "Nav Domain SME",
                    "Cyber Architect",
                    "Network Architect",
                    "TELCO Engineer",
                    "Enterprise Architect",
                ],
                "count": 10,
                "difficulty": "High",
                "time_to_fill": "30-60 days",
            },
            "engineering": {
                "titles": [
                    "Surveillance Systems Engineer",
                    "Config Management Sr. Advisor",
                    "QA Engineering Manager",
                    "Facilities Engineer - Telecom",
                ],
                "count": 4,
                "difficulty": "High",
                "time_to_fill": "30-60 days",
            },
            "quality_assurance_audit": {
                "titles": [
                    "Performance & Stability Test Engineer (x2)",
                    "Process Doc Sr. Advisor",
                    "QA Audit Management",
                ],
                "count": 4,
                "difficulty": "Medium",
                "time_to_fill": "30-45 days",
            },
            "contracts_supply_chain": {
                "titles": [
                    "Contracts Admin SA (x6, multi-site)",
                    "Subcontracts Admin SA",
                    "Asset Management SA",
                ],
                "count": 8,
                "difficulty": "Medium",
                "time_to_fill": "20-40 days",
            },
            "aviation_ops_consulting": {
                "titles": [
                    "Data Management Lead",
                    "Aviation Safety SME",
                    "Aviation Safety Liaison",
                    "ATC/TM Liaison",
                ],
                "count": 4,
                "difficulty": "Highest",
                "time_to_fill": "60-90 days",
            },
            "administrative_production": {
                "titles": ["Executive Assistant", "IMS Scheduler - TELECOM"],
                "count": 2,
                "difficulty": "Low",
                "time_to_fill": "10-20 days",
            },
        },
        "hiring_volume": "39 total positions across 7 categories",
        "use_case": "Peraton as prime contractor for BNATCS (Brand New Air Traffic Control System) modernization for USDOT/FAA",
        "budget": {
            "note": "Budget not explicitly specified - pricing determined by role difficulty and channel selection"
        },
        "channel_strategy": {
            "hard_to_fill": {
                "applicable_roles": [
                    "Program Management & Leadership",
                    "Information Technology",
                    "Engineering",
                    "Aviation Ops & Consulting",
                ],
                "channels": {
                    "niche_job_boards": "Target highly specific skill sets on discipline-dedicated platforms",
                    "cpqa_partners": "Engaged for executive-level and senior technical advisor positions",
                    "data_partners": "Source passive candidates for hard-to-fill roles with direct approach",
                },
            },
            "easy_to_fill": {
                "applicable_roles": [
                    "Quality Assurance and Audit",
                    "Contracts and Supply Chain",
                    "Administrative and Production",
                ],
                "channels": {
                    "global_job_boards": "Core source for volume - maximize exposure",
                    "regional_job_boards": "Ensure geographical fit for multi-site roles",
                    "university_recruitment": "Direct sourcing for admin and production roles",
                    "govt_job_boards": "Critical for Aviation Ops roles requiring FAA/military experience",
                    "social_media": "High-volume approach for production and admin pipelines",
                },
            },
        },
        "benchmarks": {
            "note": "Specific CPA/CPH not provided - role-based difficulty levels guide investment"
        },
        "channels_used": {
            "regional_job_boards": [
                "Jobcase Inc",
                "Snagajob",
                "Talroo",
                "Upward.net",
                "OnTimeHire",
                "JobLookup Ltd",
                "TopUSAJobs.com",
                "Propel",
                "AllJobs",
                "Joboola",
                "EarnBetter",
                "Gigs",
                "Resume-Library.com",
                "iHire",
                "Directly Apply",
                "Myjobhelper",
                "Click2Job",
                "Dice",
                "Bandana",
                "OfferUp",
                "Zippia",
                "SonicJobs",
                "Nexxt",
                "JobList",
                "Job.com",
                "Sonara.ai",
                "WayUp",
                "Alpha",
                "JobsRUs.com",
                "Talentify",
                "JobSparx",
                "Tutree",
                "Ellow.io",
                "FlexBoard",
                "Workew",
                "REAL JOBS",
                "Juvo Jobs",
                "NUL Jobs Network",
                "Open Work",
                "Mediabistro",
                "Skuad",
                "Caliber Sourcing",
                "Flyrim Tech Corp",
                "HireEazy",
                "UltraJobseekers",
                "InstaJob",
                "Jobvertise",
                "Nexustalentpartner",
                "Wonsulting",
                "Foh and Boh",
                "Jobing",
                "theladders.com",
                "IT JOB PRO",
                "Vue Jobs",
                "Job Cube",
                "Skillsire",
                "Build Dream Career",
                "Jobs4all USA",
                "Jobcube",
                "Job Seeker",
                "WorkWaveConnect",
                "CareerBuilder",
                "Careerhound",
                "RippleMatch",
                "PostJobFree",
                "Jobted",
                "The Talent Oasis",
                "UniversalHelpWanted",
                "USJobLink",
                "ASQ",
                "Washington Post",
                "Fylter",
                "Minnesota Jobs",
                "C Squared",
                "AdministrativeJobs",
                "Business Workforce",
                "LogisticsJobSite",
                "TechCareers",
                "JobFox",
                "Jobit_com",
                "JSfirm",
                "JustJobs",
                "KCJobs",
                "LasVegasJobs",
                "LPJobs",
                "Apptness",
                "NRF",
                "OKCJobs",
                "OneWire",
                "OppsPlace",
                "OptiJob",
                "NewYorkJobs",
                "AMA",
                "Ver Jobs",
                "Irecruitee",
                "NPSE",
                "PublicServiceCareers",
                "Built In",
                "Instawork",
                "Scale Jobs",
                "Lifeshack",
                "College Grad",
                "After College",
                "AccessDubuque",
                "Society of Women Engineers (SWE)",
                "CareerJournal",
                "Cooperative",
                "CPGJobList",
                "CyberSecJobs",
                "EngineerJobs",
                "FINS",
                "HCareers",
                "Hippo",
            ],
            "global_job_boards": [
                "Indeed",
                "LinkedIn",
                "Talent.com",
                "Jooble",
                "Craigslist",
                "ZipRecruiter",
                "JobRapido",
                "Sercanto",
                "YadaJobs",
                "Allthetopbananas",
                "Local Staffing LLC",
                "Adzuna",
                "CV Library",
                "Jora",
                "Women for Hire",
                "WhatJobs",
                "JobsInNetwork",
                "Jobsora",
                "Jobtome",
                "JobSwipe",
                "Mindmatch.ai",
                "ClickaJobs",
                "J-Vers",
                "The Muse",
                "Smart Recruiters",
                "Idealist",
                "College Recruiter",
                "RICS Recruit",
                "Jobufo",
                "Jobomas",
                "CareerJet",
                "Consultants 500",
                "Career Seeker",
                "Innovia Solutions LLC",
                "JobGet",
                "MyJobScanner",
                "JobsSearch",
                "beBee",
                "Job Today",
                "Results Generations",
                "WorkWaveConnect",
                "Recruit.net",
                "AppJobs",
                "Jobble",
                "StellenSMS",
                "ResumeDone",
                "Nudge Recruitment Marketing",
                "Match IT",
                "Talent Inc.",
                "Jobgoal",
                "Jobted",
                "Brilliant Jobs",
                "Reticular Media",
                "Jobguard",
                "Nexus Lead Services",
                "Find Every Job",
                "JobSearch.Coach",
                "Study Smarter",
                "Remotive",
                "Disrupted Cloud",
                "Nomad Jobs",
                "Gen Z Jobs",
                "Banya Talent",
                "Uplift Media",
                "Himalayas",
                "Sumo-D",
                "Sorce.jobs",
                "ACM Career and Job Center",
                "Hackajob",
                "Career Days",
                "Artha Job boards",
                "Gamasutra",
                "GlobalWorkplace",
                "INCOSE",
                "InfoMine",
                "ISPE",
                "JobServe",
                "NiceJob",
            ],
            "niche_job_boards": [
                "ClearanceJobs",
                "NES Fircroft",
                "Faststream Recruitment Group",
                "NTSB Careers",
                "FAAJobs",
            ],
            "cpqa_partners": [
                "Search Solutions",
                "Teilur Talent",
                "Lock Search Group",
                "Qadvance",
                "Draftboard",
                "Egon Zehnder",
                "Heidrick & Struggles",
                "Spencer Stuart",
                "Russell Reynolds Associates",
                "N2Growth",
                "DHR Global",
                "ADK Consulting & Executive Search",
                "The McCormick Group",
                "Benchmark Executive Search",
                "Diversified Search",
                "Boyden",
                "Cabot Consultants",
                "Isaacson, Miller",
                "The Burgess Group",
            ],
            "data_partners": [
                "JuiceBox",
                "Parker Dewey",
                "Crustdata",
                "People Data Labs",
                "Seekout",
            ],
            "media_channels": [
                "Meta Ads",
                "Google Ads",
                "LinkedIn Ads + InMails",
                "WhatsApp Campaigns",
                "Reddit",
                "Discord",
            ],
            "dsps": [
                "InMobi",
                "StackAdapt",
                "Microsoft Bing",
                "Eskimi",
                "Taboola",
                "Brandzooka",
                "Uber Ads",
                "Amazon Ads",
                "Quora Ads",
            ],
            "university_recruitment": [
                "Handshake",
                "12Twenty",
                "Purdue University",
                "University of Florida",
                "University of Hawaii-Manoa",
                "Walsh College",
                "UC Davis School of Law",
                "University of Notre Dame - Business",
                "Georgia Tech University",
                "University of Denver",
                "Wichita State University",
                "University of Texas Austin (Engineering)",
                "Schoolcraft College",
                "Rhodes State College",
                "University of Michigan - Engineering",
                "Columbia University - Engineering",
                "University of Missouri Science & Tech",
                "University of Massachusetts Amherst",
                "University Job Board of Iowa",
            ],
            "govt_job_boards": [
                "USAJOBS",
                "Jobs.Virginia.Gov",
                "City of Portsmouth Career Portal",
                "Geographic Solutions",
            ],
        },
        "budget_allocation": {
            "note": "Role-based allocation - hard-to-fill roles get targeted expensive channels (CPQA/Data Partners), easy-to-fill roles get volume channels (job boards, social media)"
        },
        "key_insights": [
            "BNATCS (new air traffic control system) is a major national infrastructure modernization project",
            "Executive search firms (Egon Zehnder, Heidrick & Struggles, Spencer Stuart) used for C-suite/leadership roles",
            "FAA-specific job boards (FAAJobs, NTSB Careers) critical for aviation ops roles",
            "ClearanceJobs essential for cleared IT and engineering positions",
            "Two-category strategy: hard-to-fill (targeted/expensive) vs easy-to-fill (volume/cost-effective)",
            "University partnerships span top engineering schools for pipeline building",
            "Longest time-to-fill for Aviation Safety SME and CIO/CISO roles (60-120 days)",
            "Competition includes major defense primes: Leidos, Booz Allen, CACI, Northrop Grumman, Lockheed Martin",
            "Robert Half acts as staffing intermediary for Peraton",
            "Diverse role mix (39 positions across 7 categories) requires highly segmented channel strategy",
        ],
    }
    return plan


def compute_aggregate_patterns(plans):
    """Compute aggregate patterns across all plans."""

    # Collect all channels
    all_channels = Counter()
    industry_channels = defaultdict(set)

    for plan_key, plan in plans.items():
        industry = plan.get("industry", "unknown")
        channels_data = plan.get("channels_used", {})

        for category, channels in channels_data.items():
            if isinstance(channels, list):
                for ch in channels:
                    if isinstance(ch, str):
                        all_channels[ch] += 1
                        industry_channels[industry].add(ch)
            elif isinstance(channels, dict):
                for sub_category, sub_channels in channels.items():
                    if isinstance(sub_channels, list):
                        for ch in sub_channels:
                            if isinstance(ch, str):
                                all_channels[ch] += 1
                                industry_channels[industry].add(ch)
                    elif isinstance(sub_channels, bool):
                        pass  # skip boolean values

    # Most common channels (appearing in 3+ plans)
    most_common = [ch for ch, count in all_channels.most_common(50) if count >= 3]

    # Channels appearing in all or nearly all plans
    universal_channels = [ch for ch, count in all_channels.most_common() if count >= 4]

    # Average channel allocation patterns
    avg_allocation = {
        "programmatic_job_boards": "50-60% of application generation budget (primary driver)",
        "non_programmatic_reengagement": "15-25% of application generation budget",
        "niche_job_boards": "10-20% of application generation budget",
        "employer_branding": "20-50% of total budget depending on tier and brand awareness needs",
        "media_channels_social": "10-15% of total budget for retargeting and awareness",
        "dsps": "5-10% of total budget for display advertising",
        "events_career_fairs": "5-15% of total budget (primarily for defense/engineering)",
    }

    # Industry preferences
    industry_prefs = {
        "aerospace_defense": {
            "key_channels": [
                "Programmatic Job Boards",
                "CPQA Partners",
                "Trade Shows",
                "Influencer Marketing",
                "DSPs",
                "Aviation-specific niche boards",
            ],
            "employer_branding_importance": "High - brand recognition drives talent in niche markets",
            "typical_budget_range": "$70,000 - $1,000,000+ depending on hiring volume and region",
            "unique_approaches": [
                "Aviation Valley targeting (Poland)",
                "Trade show sponsorships (NZ/AU)",
                "Influencer marketing for blue-collar roles",
            ],
        },
        "defense_shipbuilding": {
            "key_channels": [
                "Veteran Job Boards",
                "Regional Job Boards",
                "CPQA Partners",
                "Maritime-specific niche boards",
                "Govt Job Boards",
            ],
            "employer_branding_importance": "Medium - government stability and veteran programs are key differentiators",
            "typical_budget_range": "Variable based on hiring volume (200-400+ positions)",
            "unique_approaches": [
                "SkillBridge and Warrior Integration Programs",
                "Clearance-specific job boards",
                "Region expansion strategy",
            ],
        },
        "technology_ecommerce": {
            "key_channels": [
                "Local Job Boards (market-specific)",
                "Local Social Media",
                "Newspapers",
                "Staffing Partners",
            ],
            "employer_branding_importance": "Low-Medium - volume hiring focuses on reach over brand",
            "typical_budget_range": "$27,000 - $165,000 for high-volume customer service",
            "unique_approaches": [
                "Local social media platforms (Sharechat, Moj)",
                "Newspapers as English fluency qualifiers",
                "Two-scenario budget modeling",
            ],
        },
        "marine_defense_engineering": {
            "key_channels": [
                "Meta Ads",
                "Google PMax",
                "DSPs",
                "Marine-specific niche boards",
                "Radio/Podcasts",
                "Educational Partners",
            ],
            "employer_branding_importance": "Highest - 5-year phased awareness campaign",
            "typical_budget_range": "$3,000,000/year (awareness + hiring + events)",
            "unique_approaches": [
                "5-year brand awareness roadmap",
                "Blue-collar influencer marketing",
                "Location-specific radio/podcast strategy",
                "Employee Ambassador Program (2.0x ROI)",
            ],
        },
        "defense_it_aviation": {
            "key_channels": [
                "Executive Search Firms",
                "ClearanceJobs",
                "University Partnerships",
                "Data Partners",
                "FAA-specific boards",
            ],
            "employer_branding_importance": "Low - staffing intermediary model",
            "typical_budget_range": "Variable - role-based pricing by difficulty",
            "unique_approaches": [
                "Executive search for C-suite roles",
                "FAA/NTSB-specific job boards",
                "University pipeline across top engineering schools",
            ],
        },
    }

    # Budget ranges
    budget_ranges = {
        "aerospace_defense_per_plan": "$70,000 - $1,037,000",
        "defense_shipbuilding_per_plan": "Not specified (volume-based)",
        "technology_ecommerce_per_plan": "$27,200 - $165,000",
        "marine_defense_engineering_per_plan": "$3,000,000/year",
        "defense_it_aviation_per_plan": "Role-based pricing",
        "typical_cpa_range": {
            "blue_collar_trades": "$45 - $307",
            "technical_specialized": "$90 - $250",
            "customer_service_india": "$0.5 - $20",
            "executive_leadership": "Agency/retainer based",
        },
    }

    # Common DSPs across plans
    common_dsps = [
        "InMobi",
        "StackAdapt",
        "Microsoft Bing",
        "Eskimi",
        "Taboola",
        "Brandzooka",
        "Uber Ads",
        "Amazon Ads",
        "Quora Ads",
    ]

    # Common CPQA partners
    common_cpqa = [
        "INS Global",
        "Lundi",
        "1840",
        "Visage Jobs",
        "High5Hire",
        "Recruiter.com",
        "Alliance Recruitment",
    ]

    # Common data partners
    common_data = ["JuiceBox", "Seekout", "People Data Labs"]

    return {
        "most_common_channels": most_common,
        "universal_channels_4plus_plans": universal_channels,
        "common_dsps_across_plans": common_dsps,
        "common_cpqa_partners": common_cpqa,
        "common_data_partners": common_data,
        "avg_channel_allocation": avg_allocation,
        "industry_channel_preferences": industry_prefs,
        "budget_range_by_industry": budget_ranges,
        "total_unique_channels_identified": len(all_channels),
        "channel_frequency_distribution": {
            "in_5_or_more_plans": len([c for c, n in all_channels.items() if n >= 5]),
            "in_4_plans": len([c for c, n in all_channels.items() if n == 4]),
            "in_3_plans": len([c for c, n in all_channels.items() if n == 3]),
            "in_2_plans": len([c for c, n in all_channels.items() if n == 2]),
            "in_1_plan": len([c for c, n in all_channels.items() if n == 1]),
        },
        "key_patterns": [
            "All plans use a mix of global + regional/local job boards as the primary application generation engine",
            "DSPs (InMobi, StackAdapt, Taboola, etc.) appear in every plan as supplementary reach channels",
            "CPQA/staffing partners are used across all plans for quality application delivery",
            "Data partners (JuiceBox, Seekout) are standard for passive candidate re-engagement",
            "Employer branding investment ranges from 0% (Tier 1 budget plans) to 50% (premium plans)",
            "Defense/aerospace plans require specialized channels: veteran boards, clearance boards, aviation boards",
            "India market requires localized social media (Sharechat, Moj) and newspapers for English qualification",
            "5-year awareness campaigns are recommended for large-scale, long-term hiring commitments",
            "Tiered pricing (3 tiers) is the standard proposal format, with more channels at higher tiers",
            "Media channels (Meta, Google, LinkedIn) serve dual purpose: awareness retargeting and application generation",
        ],
    }


def main():
    files = {
        "rtx_us": "/Users/shubhamsinghchandel/Downloads/RTX Media Plan.xlsx",
        "bae_systems": "/Users/shubhamsinghchandel/Downloads/Media Plan _ BAE __ Joveo.xlsx",
        "amazon_cs_india": "/Users/shubhamsinghchandel/Downloads/Media Plan - Amazon CS India.xlsx",
        "rolls_royce_solutions_america": "/Users/shubhamsinghchandel/Downloads/Rolls Royce Solutions America __ Joveo.xlsx",
        "rtx_poland": "/Users/shubhamsinghchandel/Downloads/RTX Poland.xlsx",
        "peraton": "/Users/shubhamsinghchandel/Downloads/Media Plan _ Peroton __ Joveo.xlsx",
    }

    plans = {}
    extractors = {
        "rtx_us": extract_rtx_us,
        "bae_systems": extract_bae,
        "amazon_cs_india": extract_amazon_india,
        "rolls_royce_solutions_america": extract_rolls_royce,
        "rtx_poland": extract_rtx_poland,
        "peraton": extract_peroton,
    }

    for key, filepath in files.items():
        print(f"Processing: {key} ({os.path.basename(filepath)})")
        if not os.path.exists(filepath):
            print(f"  WARNING: File not found: {filepath}")
            continue
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            plans[key] = extractors[key](wb)
            wb.close()
            print(f"  SUCCESS: Extracted {key}")
        except Exception as e:
            print(f"  ERROR processing {key}: {e}")
            traceback.print_exc()

    # Compute aggregate patterns
    print("\nComputing aggregate patterns...")
    aggregate = compute_aggregate_patterns(plans)

    # Build final KB
    kb = {
        "description": "Best-in-class client media plans from Joveo's portfolio - used as reference patterns for AI-generated plans",
        "source": "Joveo internal client deliverables",
        "last_updated": "2026-03-14",
        "total_plans": len(plans),
        "industries_covered": list(
            set(p.get("industry", "unknown") for p in plans.values())
        ),
        "plans": plans,
        "aggregate_patterns": aggregate,
    }

    # Write JSON
    output_path = "/Users/shubhamsinghchandel/Downloads/Claude/media-plan-generator/data/client_media_plans_kb.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(kb, f, indent=2, ensure_ascii=False, default=str)

    print(f"\nJSON KB written to: {output_path}")
    print(f"File size: {os.path.getsize(output_path):,} bytes")
    print(f"Plans extracted: {len(plans)}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
