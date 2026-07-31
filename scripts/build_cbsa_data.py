#!/usr/bin/env python3
"""Build the optional CBSA (Core Based Statistical Area) county-join layer
from the US Census Bureau / OMB metro/micropolitan delineation file.

A CBSA is OMB's free, public-domain grouping of counties into metro-level
markets ("Atlanta-Sandy Springs-Roswell, GA", "Houston-Pasadena-The
Woodlands, TX", etc.) -- see
https://www.census.gov/geographies/reference-files/time-series/demo/metro-micro/delineation-files.html.
This is a genuine substitute for metro-level market grouping built entirely
from public-domain government data (17 U.S.C. 105), independent of any
licensed commercial market-area product.

Downloads the "List 1" delineation file (CBSA + Metropolitan Division + CSA
codes joined to their constituent counties), joins it onto the county_fips
values already present in the committed ``data/geo/us_counties.tsv`` (built
separately by ``scripts/build_geo_data.py`` -- this script does NOT
regenerate that file), and emits a compact ``data/geo/cbsa_by_county.tsv``
with columns: county_fips, cbsa_code, cbsa_title, area_type
("Metropolitan Statistical Area" | "Micropolitan Statistical Area").

Re-run any time to refresh the data:

    python3 scripts/build_cbsa_data.py

Idempotent: the download is cached under ``data/.geo_build_cache/``
(gitignored, shared with build_geo_data.py's cache), so a re-run with a warm
cache does no network I/O. Pass --no-cache to force a fresh download of
just this script's source file (does not touch the other cached Gazetteer
files).

Source (verified HTTP 200 as of 2026-07-31 -- see data/geo/README.md for
the full citation):
    - Census/OMB "List 1" CBSA delineation file, July 2023 vintage (OMB
      Bulletin 23-01) -- the newest delineation published as of this build;
      no newer file is listed on the Census delineation-files index page.

Format note (schema drift from the original brief): the brief that
requested this script assumed a plain CSV. The actual, current file at the
documented URL is an .xlsx workbook ("List 1"), not a CSV -- confirmed by
fetching it directly. There is no CSV variant of this particular file
published by Census, so this script parses the xlsx with ``openpyxl``
(already a project dependency -- see requirements.txt) instead of the
delimited-text reader used elsewhere in scripts/build_geo_data.py. Two
title rows precede the header ("List 1. CORE BASED STATISTICAL AREAS...")
and two footer rows follow the data ("Note: ...", "Source: ..."); both are
Census-authored formatting, not data, and are recognized and skipped
explicitly below rather than silently mis-parsed as rows.

This script produces DATA ONLY. It does not import or modify app.py,
plan_location.py, or plan_geo.py, and does not implement any runtime
resolver.
"""

from __future__ import annotations

import argparse
import csv
import io
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path
from typing import Dict, List

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
CACHE_DIR = PROJECT_ROOT / "data" / ".geo_build_cache"
OUTPUT_DIR = PROJECT_ROOT / "data" / "geo"
COUNTIES_TSV = OUTPUT_DIR / "us_counties.tsv"

USER_AGENT = "media-plan-generator-geo-build/1.0 (+internal offline data pipeline)"
TIMEOUT_SECONDS = 60

# Verified by fetching https://www.census.gov/geographies/reference-files/
# time-series/demo/metro-micro/delineation-files.html with curl on
# 2026-07-31 and reading the "Jul. 2023" link off the page -- the URL was
# not guessed. This is the current file; the index page lists no newer
# vintage as of this build.
CBSA_DELINEATION_URL = (
    "https://www2.census.gov/programs-surveys/metro-micro/geographies/"
    "reference-files/2023/delineation-files/list1_2023.xlsx"
)

# Header actually present in row 3 of the "List 1" sheet (rows 1-2 are
# Census title/caption rows, not part of the tabular header). Confirmed by
# loading the real downloaded workbook, not assumed from the brief.
EXPECTED_HEADER = [
    "CBSA Code",
    "Metropolitan Division Code",
    "CSA Code",
    "CBSA Title",
    "Metropolitan/Micropolitan Statistical Area",
    "Metropolitan Division Title",
    "CSA Title",
    "County/County Equivalent",
    "State Name",
    "FIPS State Code",
    "FIPS County Code",
    "Central/Outlying County",
]

ALLOWED_AREA_TYPES = frozenset({"Metropolitan Statistical Area", "Micropolitan Statistical Area"})

_CBSA_CODE_RE = re.compile(r"^\d{5}$")


class BuildError(RuntimeError):
    """Raised on any condition that would otherwise silently corrupt output."""


# --------------------------------------------------------------------------
# Download / cache (same convention as scripts/build_geo_data.py's fetch())
# --------------------------------------------------------------------------


def _cache_path_for(url: str) -> Path:
    name = url.split("/")[-1]
    return CACHE_DIR / name


def fetch(url: str) -> bytes:
    """Download url, caching to CACHE_DIR. Fails loudly on any HTTP error."""
    dest = _cache_path_for(url)
    if dest.exists() and dest.stat().st_size > 0:
        return dest.read_bytes()

    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=TIMEOUT_SECONDS) as resp:
            status = getattr(resp, "status", 200)
            if status != 200:
                raise BuildError(f"HTTP {status} fetching {url}")
            data = resp.read()
    except urllib.error.HTTPError as exc:
        raise BuildError(f"HTTP {exc.code} fetching {url}: {exc.reason}") from exc
    except urllib.error.URLError as exc:
        raise BuildError(f"Network error fetching {url}: {exc.reason}") from exc

    if not data:
        raise BuildError(f"Empty response body fetching {url}")

    tmp = dest.with_suffix(dest.suffix + ".part")
    tmp.write_bytes(data)
    tmp.replace(dest)
    return data


def _require_header(actual: List, expected: List[str], source: str) -> None:
    normalized = list(actual)
    if normalized != expected:
        raise BuildError(
            f"{source}: unexpected header.\n  expected: {expected}\n  actual:   {normalized}"
        )


# --------------------------------------------------------------------------
# Parsing
# --------------------------------------------------------------------------


def load_county_fips_set() -> set:
    """Read the ALREADY-COMMITTED data/geo/us_counties.tsv (built separately
    by build_geo_data.py) so the CBSA join can be validated against real,
    known county_fips values. This script never regenerates us_counties.tsv."""
    if not COUNTIES_TSV.exists():
        raise BuildError(f"{COUNTIES_TSV} not found -- run scripts/build_geo_data.py first.")
    fips: set = set()
    with COUNTIES_TSV.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f, delimiter="\t"):
            v = (row.get("county_fips") or "").strip()
            if v:
                fips.add(v)
    if not fips:
        raise BuildError(f"{COUNTIES_TSV} contained no county_fips values")
    return fips


def parse_cbsa_rows(raw: bytes) -> List[Dict[str, str]]:
    """Parse the List-1 xlsx workbook into CBSA-county rows (not yet
    filtered against us_counties.tsv -- see main())."""
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, row in enumerate(all_rows):
        if row and row[0] == "CBSA Code":
            header_idx = i
            break
    if header_idx is None:
        raise BuildError("list1 CBSA delineation file: could not find the 'CBSA Code' header row")
    _require_header(all_rows[header_idx], EXPECTED_HEADER, "list1_2023.xlsx")

    out: List[Dict[str, str]] = []
    for row in all_rows[header_idx + 1 :]:
        if row is None or all(c is None for c in row):
            continue  # blank separator row between data and footer

        code = row[0]
        if not (isinstance(code, str) and _CBSA_CODE_RE.match(code)):
            # Known trailing footer rows ("Note: ...", "Source: ..."). Any
            # other non-data row is an unrecognized format change and must
            # fail loudly rather than be silently dropped.
            text = str(code or "")
            if text.startswith("Note:") or text.startswith("Source:"):
                continue
            raise BuildError(f"list1_2023.xlsx: unexpected non-data row: {row!r}")

        area_type = (row[4] or "").strip()
        if area_type not in ALLOWED_AREA_TYPES:
            raise BuildError(f"list1_2023.xlsx: unexpected area type {area_type!r} for CBSA {code}")

        fips_state = (row[9] or "").strip()
        fips_county = (row[10] or "").strip()
        county_fips = fips_state + fips_county
        if len(county_fips) != 5 or not county_fips.isdigit():
            raise BuildError(
                f"list1_2023.xlsx: bad county FIPS state={fips_state!r} county={fips_county!r} for CBSA {code}"
            )

        out.append(
            {
                "county_fips": county_fips,
                "cbsa_code": code,
                "cbsa_title": (row[3] or "").strip(),
                "area_type": area_type,
            }
        )
    return out


# --------------------------------------------------------------------------
# Writers
# --------------------------------------------------------------------------


def write_tsv(path: Path, fieldnames: List[str], rows) -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    count = 0
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
            count += 1
    return count


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--no-cache",
        action="store_true",
        help="Delete this script's cached download before building (forces a fresh download).",
    )
    args = parser.parse_args()

    if args.no_cache:
        cached = _cache_path_for(CBSA_DELINEATION_URL)
        if cached.exists():
            cached.unlink()

    print("Building CBSA county-join layer into", OUTPUT_DIR)

    print(f"Reading known county_fips values from {COUNTIES_TSV}...")
    county_fips_set = load_county_fips_set()
    print(f"  {len(county_fips_set)} known county_fips values")

    print("Fetching CBSA delineation file (list1_2023.xlsx)...")
    raw = fetch(CBSA_DELINEATION_URL)
    rows = parse_cbsa_rows(raw)
    print(f"  parsed {len(rows)} CBSA-county rows from the source file")

    skipped_no_county_join = 0
    joined_rows: List[Dict[str, str]] = []
    seen_fips: set = set()
    for r in rows:
        if r["county_fips"] not in county_fips_set:
            # Vintage mismatch between the 2023 CBSA delineation and
            # us_counties.tsv's vintage (e.g. a county FIPS that changed --
            # Connecticut's 2022 switch from counties to planning regions is
            # the known historical example of this class of drift). Reported
            # in the build summary, never silently dropped -- same
            # convention as build_geo_data.py's places_skipped/zips_skipped.
            skipped_no_county_join += 1
            continue
        if r["county_fips"] in seen_fips:
            raise BuildError(
                f"duplicate county_fips {r['county_fips']} across CBSA rows -- "
                "expected at most one CBSA per county"
            )
        seen_fips.add(r["county_fips"])
        joined_rows.append(r)

    print("Writing cbsa_by_county.tsv...")
    n = write_tsv(
        OUTPUT_DIR / "cbsa_by_county.tsv",
        ["county_fips", "cbsa_code", "cbsa_title", "area_type"],
        sorted(joined_rows, key=lambda r: r["county_fips"]),
    )

    metro_count = sum(1 for r in joined_rows if r["area_type"] == "Metropolitan Statistical Area")
    micro_count = sum(1 for r in joined_rows if r["area_type"] == "Micropolitan Statistical Area")
    counties_without_cbsa = len(county_fips_set) - n

    print()
    print("=== Build summary ===")
    print(
        f"cbsa_by_county.tsv : {n} rows "
        f"({skipped_no_county_join} skipped -- county_fips not found in us_counties.tsv, see README.md)"
    )
    print(f"  Metropolitan Statistical Area rows : {metro_count}")
    print(f"  Micropolitan Statistical Area rows : {micro_count}")
    print(f"  counties with no CBSA (outside any metro/micro area): {counties_without_cbsa}")
    print()
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
