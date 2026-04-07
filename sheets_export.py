"""Google Sheets export for media plan data.

Uses Google Sheets API v4 with service account credentials (REST only,
no google-api-python-client dependency).  Falls back to CSV download if
Google Sheets is not configured.

Environment variable:
    GOOGLE_SHEETS_CREDENTIALS  -- path to the service account JSON file.

Fallback chain:
    1. Google Sheets API  (if credentials configured)
    2. XLSX via openpyxl   (if available)
    3. CSV                 (stdlib, always works)
"""

from __future__ import annotations

import csv
import io
import json
import logging
import os
import ssl
import time
import urllib.parse
import urllib.request
import urllib.error
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]
_TOKEN_URI = "https://oauth2.googleapis.com/token"
_SHEETS_BASE = "https://sheets.googleapis.com/v4/spreadsheets"
_DRIVE_BASE = "https://www.googleapis.com/drive/v3/files"

# Cache the access token (valid for ~1 hour)
_token_cache: Dict[str, Any] = {"token": None, "expires_at": 0.0}


# ---------------------------------------------------------------------------
# Service-account credential helpers (stdlib only)
# ---------------------------------------------------------------------------


def _load_credentials() -> Optional[Dict[str, str]]:
    """Load service-account JSON from env vars.

    Tries two sources in order:
        1. GOOGLE_SHEETS_CREDENTIALS -- path to a service account JSON file.
        2. GOOGLE_SLIDES_CREDENTIALS_B64 -- base64-encoded service account JSON
           string (shared with Google Slides, used on Render).

    Returns:
        Parsed JSON dict or None if not configured / file missing.
    """
    import base64

    required = ("client_email", "private_key", "token_uri")

    # --- Try 1: file path ---
    cred_path = os.environ.get("GOOGLE_SHEETS_CREDENTIALS") or ""
    if cred_path:
        try:
            path = Path(cred_path)
            if not path.is_file():
                logger.warning(
                    "Google Sheets credentials file not found: %s", cred_path
                )
            else:
                with open(path, "r", encoding="utf-8") as fh:
                    creds = json.load(fh)
                for field in required:
                    if field not in creds:
                        logger.error(
                            "Service account JSON missing required field: %s", field
                        )
                        return None
                return creds
        except (json.JSONDecodeError, OSError) as exc:
            logger.error(
                "Failed to load Google Sheets credentials from file: %s",
                exc,
                exc_info=True,
            )

    # --- Try 2: base64-encoded JSON (shared with Google Slides on Render) ---
    b64_creds = os.environ.get("GOOGLE_SLIDES_CREDENTIALS_B64") or ""
    if b64_creds:
        try:
            decoded = base64.b64decode(b64_creds)
            creds = json.loads(decoded)
            for field in required:
                if field not in creds:
                    logger.error(
                        "B64 service account JSON missing required field: %s", field
                    )
                    return None
            logger.info(
                "Loaded Google Sheets credentials from GOOGLE_SLIDES_CREDENTIALS_B64 (service account: %s)",
                creds.get("client_email", "unknown"),
            )
            return creds
        except Exception as exc:
            logger.error("Failed to load B64 credentials: %s", exc, exc_info=True)

    return None


def _build_jwt(creds: Dict[str, str]) -> str:
    """Build a signed JWT for the service-account OAuth2 token exchange.

    Uses stdlib only (hmac is not sufficient -- RS256 requires real RSA).
    We use the PyJWT-free approach: base64-encode header+claims, sign with
    the private key via the ``cryptography`` or ``rsa`` package if available,
    otherwise fall back to shelling out to ``openssl``.

    Returns:
        Signed JWT string.

    Raises:
        RuntimeError: If no RSA signing method is available.
    """
    import base64
    import struct
    import subprocess
    import tempfile

    now = int(time.time())
    header = {"alg": "RS256", "typ": "JWT"}
    claims = {
        "iss": creds["client_email"],
        "scope": " ".join(_SCOPES),
        "aud": creds.get("token_uri") or _TOKEN_URI,
        "iat": now,
        "exp": now + 3600,
    }

    def _b64url(data: bytes) -> str:
        return base64.urlsafe_b64encode(data).rstrip(b"=").decode("ascii")

    header_b64 = _b64url(json.dumps(header, separators=(",", ":")).encode())
    claims_b64 = _b64url(json.dumps(claims, separators=(",", ":")).encode())
    signing_input = f"{header_b64}.{claims_b64}".encode("ascii")

    private_key_pem = creds["private_key"]

    # Try 1: cryptography library (most common in production)
    try:
        from cryptography.hazmat.primitives import hashes, serialization
        from cryptography.hazmat.primitives.asymmetric import padding

        key = serialization.load_pem_private_key(
            private_key_pem.encode("utf-8"), password=None
        )
        signature = key.sign(signing_input, padding.PKCS1v15(), hashes.SHA256())  # type: ignore[union-attr]
        return f"{header_b64}.{claims_b64}.{_b64url(signature)}"
    except ImportError:
        pass

    # Try 2: openssl CLI (available on most Unix systems)
    try:
        with tempfile.NamedTemporaryFile(mode="w", suffix=".pem", delete=False) as kf:
            kf.write(private_key_pem)
            key_path = kf.name
        try:
            proc = subprocess.run(
                ["openssl", "dgst", "-sha256", "-sign", key_path],
                input=signing_input,
                capture_output=True,
                timeout=10,
            )
            if proc.returncode == 0 and proc.stdout:
                return f"{header_b64}.{claims_b64}.{_b64url(proc.stdout)}"
        finally:
            try:
                os.unlink(key_path)
            except OSError:
                pass
    except FileNotFoundError:
        pass

    raise RuntimeError(
        "Cannot sign JWT: install 'cryptography' package or ensure 'openssl' is on PATH"
    )


def _get_access_token() -> Optional[str]:
    """Obtain a Google OAuth2 access token using the service account.

    Caches the token until 5 minutes before expiry.

    Returns:
        Bearer token string, or None if credentials not configured.
    """
    now = time.time()
    if _token_cache["token"] and _token_cache["expires_at"] > now + 300:
        return _token_cache["token"]

    creds = _load_credentials()
    if not creds:
        return None

    try:
        jwt_token = _build_jwt(creds)
    except RuntimeError as exc:
        logger.error("JWT signing failed: %s", exc, exc_info=True)
        return None

    payload = urllib.parse.urlencode(
        {
            "grant_type": "urn:ietf:params:oauth:2.0-jwt-bearer",
            "assertion": jwt_token,
        }
    ).encode("utf-8")

    token_uri = creds.get("token_uri") or _TOKEN_URI
    req = urllib.request.Request(
        token_uri,
        data=payload,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
    )

    try:
        ctx = ssl.create_default_context()
        with urllib.request.urlopen(req, context=ctx, timeout=15) as resp:
            token_data = json.loads(resp.read().decode("utf-8"))
        _token_cache["token"] = token_data["access_token"]
        _token_cache["expires_at"] = now + token_data.get("expires_in", 3600)
        return _token_cache["token"]
    except (urllib.error.URLError, json.JSONDecodeError, KeyError) as exc:
        logger.error("Google OAuth2 token exchange failed: %s", exc, exc_info=True)
        return None


# ---------------------------------------------------------------------------
# Google Sheets API helpers (REST, stdlib only)
# ---------------------------------------------------------------------------


def _sheets_request(
    method: str,
    url: str,
    body: Optional[dict] = None,
    token: Optional[str] = None,
) -> Optional[dict]:
    """Make an authenticated request to the Google Sheets/Drive API.

    Args:
        method: HTTP method (GET, POST, PUT, etc.).
        url: Full API URL.
        body: Optional JSON body.
        token: Bearer token. If None, will attempt to obtain one.

    Returns:
        Parsed JSON response, or None on failure.
    """
    if token is None:
        token = _get_access_token()
    if not token:
        return None

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    data = json.dumps(body).encode("utf-8") if body else None

    req = urllib.request.Request(url, data=data, headers=headers, method=method)

    try:
        ctx = ssl.create_default_context()
        with urllib.request.urlopen(req, context=ctx, timeout=30) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        error_body = ""
        try:
            error_body = exc.read().decode("utf-8")
        except Exception:
            pass
        logger.error(
            "Google API %s %s returned %d: %s",
            method,
            url,
            exc.code,
            error_body,
            exc_info=True,
        )
        return None
    except (urllib.error.URLError, json.JSONDecodeError) as exc:
        logger.error("Google API request failed: %s", exc, exc_info=True)
        return None


# ---------------------------------------------------------------------------
# Public API -- Google Sheets creation
# ---------------------------------------------------------------------------


def create_spreadsheet(title: str, data: Dict[str, List[List[str]]]) -> Optional[str]:
    """Create a new Google Sheet with multiple named tabs populated with data.

    Args:
        title: Spreadsheet title.
        data: Mapping of sheet_name -> list of rows (each row is a list of
              cell values as strings).

    Returns:
        URL of the created spreadsheet, or None on failure.
    """
    token = _get_access_token()
    if not token:
        logger.warning("Google Sheets not configured -- cannot create spreadsheet")
        return None

    # Step 1: Create the spreadsheet with named sheets
    sheets_spec = []
    for idx, sheet_name in enumerate(data.keys()):
        sheets_spec.append(
            {
                "properties": {
                    "sheetId": idx,
                    "title": sheet_name,
                    "index": idx,
                }
            }
        )

    create_body = {
        "properties": {"title": title},
        "sheets": sheets_spec,
    }

    result = _sheets_request("POST", _SHEETS_BASE, body=create_body, token=token)
    if not result:
        return None

    spreadsheet_id = result.get("spreadsheetId") or ""
    if not spreadsheet_id:
        logger.error("Google Sheets API returned no spreadsheetId")
        return None

    # Step 2: Batch-update all sheet data
    value_ranges = []
    for sheet_name, rows in data.items():
        if rows:
            value_ranges.append(
                {
                    "range": f"'{sheet_name}'!A1",
                    "majorDimension": "ROWS",
                    "values": rows,
                }
            )

    if value_ranges:
        batch_url = f"{_SHEETS_BASE}/{spreadsheet_id}/values:batchUpdate"
        batch_body = {
            "valueInputOption": "USER_ENTERED",
            "data": value_ranges,
        }
        batch_result = _sheets_request("POST", batch_url, body=batch_body, token=token)
        if not batch_result:
            logger.warning("Spreadsheet created but data population failed")

    # Step 3: Format header rows (bold, frozen)
    format_requests = []
    for idx, (sheet_name, rows) in enumerate(data.items()):
        if rows:
            # Freeze first row
            format_requests.append(
                {
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": idx,
                            "gridProperties": {"frozenRowCount": 1},
                        },
                        "fields": "gridProperties.frozenRowCount",
                    }
                }
            )
            # Bold header row
            format_requests.append(
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": idx,
                            "startRowIndex": 0,
                            "endRowIndex": 1,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "textFormat": {"bold": True},
                                "backgroundColor": {
                                    "red": 0.125,
                                    "green": 0.125,
                                    "blue": 0.345,
                                },
                                "horizontalAlignment": "CENTER",
                            }
                        },
                        "fields": "userEnteredFormat(textFormat,backgroundColor,horizontalAlignment)",
                    }
                }
            )
            # Auto-resize columns
            format_requests.append(
                {
                    "autoResizeDimensions": {
                        "dimensions": {
                            "sheetId": idx,
                            "dimension": "COLUMNS",
                            "startIndex": 0,
                            "endIndex": len(rows[0]) if rows else 10,
                        }
                    }
                }
            )

    if format_requests:
        fmt_url = f"{_SHEETS_BASE}/{spreadsheet_id}:batchUpdate"
        _sheets_request(
            "POST", fmt_url, body={"requests": format_requests}, token=token
        )

    # Step 4: Make spreadsheet readable by anyone with the link
    share_url = f"{_DRIVE_BASE}/{spreadsheet_id}/permissions"
    share_body = {
        "role": "reader",
        "type": "domain",
        "domain": "joveo.com",
    }
    _sheets_request("POST", share_url, body=share_body, token=token)

    sheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
    logger.info("Created Google Sheet: %s", sheet_url)
    return sheet_url


def _safe_str(value: Any) -> str:
    """Convert a value to a safe string for spreadsheet cells.

    Escapes Excel formula indicators (=, +, @, -) by prefixing with apostrophe
    to force text interpretation and prevent formula injection attacks.
    """
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value)

    # Convert to string
    str_val = str(value)

    # Escape Excel formula indicators by prefixing with apostrophe
    # This prevents formula injection attacks where malicious formulas execute code
    if str_val and str_val[0] in ("=", "+", "@", "-"):
        return f"'{str_val}"

    return str_val


def _format_currency(value: Any) -> str:
    """Format a numeric value as currency string."""
    if value is None:
        return "$0"
    try:
        num = float(value)
        if num >= 1_000_000:
            return f"${num / 1_000_000:,.1f}M"
        if num >= 1_000:
            return f"${num:,.0f}"
        return f"${num:,.2f}"
    except (ValueError, TypeError):
        return _safe_str(value)


def _format_number(value: Any) -> str:
    """Format a numeric value with commas."""
    if value is None:
        return "0"
    try:
        num = float(value)
        if num == int(num):
            return f"{int(num):,}"
        return f"{num:,.2f}"
    except (ValueError, TypeError):
        return _safe_str(value)


# ---------------------------------------------------------------------------
# Media-plan-specific sheet builders
# ---------------------------------------------------------------------------


def _build_summary_sheet(plan_data: Dict[str, Any]) -> List[List[str]]:
    """Build the Summary sheet rows from plan data.

    Args:
        plan_data: The media plan generation result dict.

    Returns:
        List of rows (each row a list of cell-value strings).
    """
    client = plan_data.get("client_name") or "Client"
    industry = plan_data.get("industry") or "General"
    budget = plan_data.get("budget") or plan_data.get("monthly_budget") or ""
    locations = plan_data.get("locations") or plan_data.get("location") or ""
    if isinstance(locations, list):
        locations = ", ".join(str(loc) for loc in locations)
    job_title = plan_data.get("job_title") or plan_data.get("role") or ""
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    total_budget = _format_currency(budget)
    channels = (
        plan_data.get("channels") or plan_data.get("channel_recommendations") or []
    )
    total_channels = str(len(channels))

    rows: List[List[str]] = [
        ["Field", "Value"],
        ["Client Name", _safe_str(client)],
        ["Industry", _safe_str(industry)],
        ["Job Title / Role", _safe_str(job_title)],
        ["Total Budget", total_budget],
        ["Location(s)", _safe_str(locations)],
        ["Number of Channels", total_channels],
        ["Generated Date", generated],
        [""],
        ["PLAN OVERVIEW"],
    ]

    # Add summary stats if available
    summary = plan_data.get("summary") or plan_data.get("executive_summary") or {}
    if isinstance(summary, dict):
        for key, val in summary.items():
            label = key.replace("_", " ").title()
            rows.append([label, _safe_str(val)])
    elif isinstance(summary, str):
        rows.append(["Summary", summary])

    return rows


def _build_channels_sheet(plan_data: Dict[str, Any]) -> List[List[str]]:
    """Build the Channel Recommendations sheet.

    Args:
        plan_data: The media plan generation result dict.

    Returns:
        List of rows for the Channels sheet.
    """
    channels = (
        plan_data.get("channels") or plan_data.get("channel_recommendations") or []
    )

    header = [
        "Channel",
        "Category",
        "CPC",
        "CPA",
        "Budget Allocation",
        "Est. Clicks",
        "Est. Applies",
        "Confidence",
        "Notes",
    ]
    rows: List[List[str]] = [header]

    for ch in channels:
        if isinstance(ch, dict):
            rows.append(
                [
                    _safe_str(ch.get("name") or ch.get("channel") or ""),
                    _safe_str(ch.get("category") or ch.get("type") or ""),
                    _format_currency(ch.get("cpc") or ch.get("cost_per_click")),
                    _format_currency(ch.get("cpa") or ch.get("cost_per_apply")),
                    _format_currency(
                        ch.get("budget")
                        or ch.get("allocation")
                        or ch.get("monthly_budget")
                    ),
                    _format_number(ch.get("estimated_clicks") or ch.get("clicks")),
                    _format_number(ch.get("estimated_applies") or ch.get("applies")),
                    _safe_str(ch.get("confidence") or ch.get("confidence_score") or ""),
                    _safe_str(ch.get("notes") or ch.get("rationale") or ""),
                ]
            )
        elif isinstance(ch, str):
            rows.append([ch, "", "", "", "", "", "", "", ""])

    return rows


def _build_budget_sheet(plan_data: Dict[str, Any]) -> List[List[str]]:
    """Build the Budget Breakdown sheet.

    Args:
        plan_data: The media plan generation result dict.

    Returns:
        List of rows for the Budget sheet.
    """
    channels = (
        plan_data.get("channels") or plan_data.get("channel_recommendations") or []
    )
    budget_raw = plan_data.get("budget") or plan_data.get("monthly_budget") or 0

    try:
        total_budget = float(str(budget_raw).replace("$", "").replace(",", ""))
    except (ValueError, TypeError):
        total_budget = 0.0

    header = [
        "Channel",
        "Monthly Budget",
        "% of Total",
        "Quarterly Projection",
        "Annual Projection",
    ]
    rows: List[List[str]] = [header]

    for ch in channels:
        if isinstance(ch, dict):
            ch_budget_raw = (
                ch.get("budget")
                or ch.get("allocation")
                or ch.get("monthly_budget")
                or 0
            )
            try:
                ch_budget = float(str(ch_budget_raw).replace("$", "").replace(",", ""))
            except (ValueError, TypeError):
                ch_budget = 0.0

            pct = (
                f"{(ch_budget / total_budget * 100):.1f}%" if total_budget > 0 else "0%"
            )
            rows.append(
                [
                    _safe_str(ch.get("name") or ch.get("channel") or ""),
                    _format_currency(ch_budget),
                    pct,
                    _format_currency(ch_budget * 3),
                    _format_currency(ch_budget * 12),
                ]
            )

    # Totals row
    if channels:
        rows.append([])
        rows.append(
            [
                "TOTAL",
                _format_currency(total_budget),
                "100%",
                _format_currency(total_budget * 3),
                _format_currency(total_budget * 12),
            ]
        )

    return rows


def _build_benchmarks_sheet(plan_data: Dict[str, Any]) -> List[List[str]]:
    """Build the Industry Benchmarks sheet.

    Args:
        plan_data: The media plan generation result dict.

    Returns:
        List of rows for the Benchmarks sheet.
    """
    header = ["Metric", "Industry Average", "Your Plan", "Difference", "Source"]
    rows: List[List[str]] = [header]

    benchmarks = (
        plan_data.get("benchmarks") or plan_data.get("industry_benchmarks") or {}
    )
    if isinstance(benchmarks, dict):
        for metric, data in benchmarks.items():
            label = metric.replace("_", " ").title()
            if isinstance(data, dict):
                rows.append(
                    [
                        label,
                        _safe_str(
                            data.get("industry_avg") or data.get("average") or ""
                        ),
                        _safe_str(
                            data.get("plan_value") or data.get("your_value") or ""
                        ),
                        _safe_str(data.get("difference") or data.get("delta") or ""),
                        _safe_str(data.get("source") or ""),
                    ]
                )
            else:
                rows.append([label, _safe_str(data), "", "", ""])
    elif isinstance(benchmarks, list):
        for item in benchmarks:
            if isinstance(item, dict):
                rows.append(
                    [
                        _safe_str(item.get("metric") or item.get("name") or ""),
                        _safe_str(
                            item.get("industry_avg") or item.get("average") or ""
                        ),
                        _safe_str(
                            item.get("plan_value") or item.get("your_value") or ""
                        ),
                        _safe_str(item.get("difference") or item.get("delta") or ""),
                        _safe_str(item.get("source") or ""),
                    ]
                )

    # Add channel-level CPC/CPA benchmarks
    channels = (
        plan_data.get("channels") or plan_data.get("channel_recommendations") or []
    )
    if channels:
        rows.append([])
        rows.append(["CHANNEL BENCHMARKS", "", "", "", ""])
        rows.append(["Channel", "CPC Range", "CPA Range", "Avg. Apply Rate", "Source"])
        for ch in channels:
            if isinstance(ch, dict):
                rows.append(
                    [
                        _safe_str(ch.get("name") or ch.get("channel") or ""),
                        _safe_str(ch.get("cpc_range") or ch.get("benchmark_cpc") or ""),
                        _safe_str(ch.get("cpa_range") or ch.get("benchmark_cpa") or ""),
                        _safe_str(
                            ch.get("apply_rate") or ch.get("conversion_rate") or ""
                        ),
                        _safe_str(ch.get("benchmark_source") or "Nova AI Suite"),
                    ]
                )

    return rows


def _build_timeline_sheet(plan_data: Dict[str, Any]) -> List[List[str]]:
    """Build the Timeline / Campaign Plan sheet.

    Args:
        plan_data: The media plan generation result dict.

    Returns:
        List of rows for the Timeline sheet.
    """
    header = ["Week", "Phase", "Channels Active", "Budget", "Key Actions", "KPIs"]
    rows: List[List[str]] = [header]

    timeline = plan_data.get("timeline") or plan_data.get("campaign_timeline") or []
    if isinstance(timeline, list):
        for entry in timeline:
            if isinstance(entry, dict):
                rows.append(
                    [
                        _safe_str(entry.get("week") or entry.get("period") or ""),
                        _safe_str(entry.get("phase") or ""),
                        _safe_str(
                            entry.get("channels") or entry.get("channels_active") or ""
                        ),
                        _format_currency(entry.get("budget") or entry.get("spend")),
                        _safe_str(
                            entry.get("actions") or entry.get("key_actions") or ""
                        ),
                        _safe_str(entry.get("kpis") or entry.get("targets") or ""),
                    ]
                )

    # If no explicit timeline, generate a default 4-week ramp plan
    if len(rows) <= 1:
        channels = (
            plan_data.get("channels") or plan_data.get("channel_recommendations") or []
        )
        budget_raw = plan_data.get("budget") or plan_data.get("monthly_budget") or 0
        try:
            total_budget = float(str(budget_raw).replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            total_budget = 0.0

        ch_names = (
            ", ".join(
                _safe_str(ch.get("name") or ch.get("channel") or "")
                for ch in channels[:5]
                if isinstance(ch, dict)
            )
            or "All channels"
        )

        phases = [
            (
                "Week 1",
                "Launch",
                "20%",
                "Setup tracking, launch initial ads, A/B test creatives",
            ),
            (
                "Week 2",
                "Optimize",
                "25%",
                "Review CTR data, pause underperformers, scale winners",
            ),
            (
                "Week 3",
                "Scale",
                "30%",
                "Increase budget on top channels, expand targeting",
            ),
            (
                "Week 4",
                "Refine",
                "25%",
                "Final optimization, compile performance report",
            ),
        ]
        for week, phase, pct_str, actions in phases:
            pct = float(pct_str.rstrip("%")) / 100
            rows.append(
                [
                    week,
                    phase,
                    ch_names,
                    _format_currency(total_budget * pct),
                    actions,
                    "CTR, CPC, CPA, Apply Rate",
                ]
            )

    return rows


# ---------------------------------------------------------------------------
# Main export functions
# ---------------------------------------------------------------------------


def export_media_plan(plan_data: Dict[str, Any]) -> Optional[str]:
    """Export a media plan to Google Sheets with 5 structured tabs.

    Args:
        plan_data: The media plan data dict (as returned by /api/generate).

    Returns:
        URL of the created Google Sheet, or None if Google Sheets is not
        configured (caller should fall back to CSV/XLSX).
    """
    if not _load_credentials():
        logger.info("Google Sheets not configured -- skipping Sheets export")
        return None

    client = plan_data.get("client_name") or "Client"
    title = f"{client} - Media Plan ({datetime.now(timezone.utc).strftime('%Y-%m-%d')})"

    sheet_data = {
        "Summary": _build_summary_sheet(plan_data),
        "Channel Recommendations": _build_channels_sheet(plan_data),
        "Budget Breakdown": _build_budget_sheet(plan_data),
        "Benchmarks": _build_benchmarks_sheet(plan_data),
        "Timeline": _build_timeline_sheet(plan_data),
    }

    try:
        url = create_spreadsheet(title, sheet_data)
        return url
    except Exception as exc:
        logger.error("Google Sheets export failed: %s", exc, exc_info=True)
        return None


def export_to_csv(data: Dict[str, Any]) -> bytes:
    """Generate a CSV file from media plan data (stdlib fallback).

    Produces a single CSV combining the most important data: summary
    fields followed by channel recommendations.

    Args:
        data: The media plan data dict.

    Returns:
        CSV file content as bytes (UTF-8 encoded with BOM for Excel compat).
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Summary section
    client = data.get("client_name") or "Client"
    industry = data.get("industry") or "General"
    budget = data.get("budget") or data.get("monthly_budget") or ""
    locations = data.get("locations") or data.get("location") or ""
    if isinstance(locations, list):
        locations = ", ".join(str(loc) for loc in locations)

    writer.writerow(["Media Plan Export"])
    writer.writerow(
        ["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")]
    )
    writer.writerow(["Client", _safe_str(client)])
    writer.writerow(["Industry", _safe_str(industry)])
    writer.writerow(["Budget", _format_currency(budget)])
    writer.writerow(["Location(s)", _safe_str(locations)])
    writer.writerow([])

    # Channel recommendations
    writer.writerow(
        [
            "Channel",
            "Category",
            "CPC",
            "CPA",
            "Budget Allocation",
            "Est. Clicks",
            "Est. Applies",
        ]
    )

    channels = data.get("channels") or data.get("channel_recommendations") or []
    for ch in channels:
        if isinstance(ch, dict):
            writer.writerow(
                [
                    _safe_str(ch.get("name") or ch.get("channel") or ""),
                    _safe_str(ch.get("category") or ch.get("type") or ""),
                    _format_currency(ch.get("cpc") or ch.get("cost_per_click")),
                    _format_currency(ch.get("cpa") or ch.get("cost_per_apply")),
                    _format_currency(
                        ch.get("budget")
                        or ch.get("allocation")
                        or ch.get("monthly_budget")
                    ),
                    _format_number(ch.get("estimated_clicks") or ch.get("clicks")),
                    _format_number(ch.get("estimated_applies") or ch.get("applies")),
                ]
            )

    # BOM + content for Excel compatibility
    csv_bytes = b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")
    return csv_bytes


def export_to_xlsx(data: Dict[str, Any]) -> Optional[bytes]:
    """Generate an XLSX workbook from media plan data using openpyxl.

    Falls back to None if openpyxl is not available (caller should use CSV).

    Args:
        data: The media plan data dict.

    Returns:
        XLSX file bytes, or None if openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.info("openpyxl not available -- XLSX export disabled")
        return None

    wb = Workbook()

    # Design tokens
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="202058", end_color="202058", fill_type="solid"
    )
    body_font = Font(name="Calibri", size=10, color="1C1917")
    border = Border(
        bottom=Side(style="thin", color="E7E5E4"),
    )

    def _add_sheet(
        ws: Any,
        title: str,
        rows: List[List[str]],
    ) -> None:
        """Populate a worksheet with rows and styling."""
        ws.title = title
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.font = body_font
                    cell.border = border
        # Auto-width columns
        for col_idx in range(1, (len(rows[0]) if rows else 0) + 1):
            max_len = 0
            for row in rows:
                if col_idx <= len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = (
                min(max_len + 4, 50)
            )
        ws.freeze_panes = "A2"

    # Build sheets
    sheets_data = {
        "Summary": _build_summary_sheet(data),
        "Channels": _build_channels_sheet(data),
        "Budget": _build_budget_sheet(data),
        "Benchmarks": _build_benchmarks_sheet(data),
        "Timeline": _build_timeline_sheet(data),
    }

    first = True
    for sheet_name, rows in sheets_data.items():
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        _add_sheet(ws, sheet_name, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Status check
# ---------------------------------------------------------------------------


def get_status() -> Dict[str, Any]:
    """Check whether Google Sheets export is configured and operational.

    Returns:
        Status dict with 'configured', 'service_account', and 'fallback' keys.
    """
    creds = _load_credentials()
    has_openpyxl = False
    try:
        import openpyxl  # noqa: F401

        has_openpyxl = True
    except ImportError:
        pass

    if creds:
        return {
            "configured": True,
            "service_account": creds.get("client_email") or "unknown",
            "fallback": "xlsx" if has_openpyxl else "csv",
            "formats": ["sheets", "xlsx", "csv"] if has_openpyxl else ["sheets", "csv"],
        }
    return {
        "configured": False,
        "service_account": None,
        "fallback": "xlsx" if has_openpyxl else "csv",
        "formats": ["xlsx", "csv"] if has_openpyxl else ["csv"],
    }
