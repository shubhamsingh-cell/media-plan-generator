#!/usr/bin/env python3
"""
Consolidated 5-Sheet Excel Generator (v2) for AI Media Plan Generator.

Replaces the 26+ sheet original with up to 9 focused sheets:
    1. Executive Summary     — overview, budget, benchmarks, recommendations
    2. Channels & Strategy   — vetted channels, ad platform analysis, niche boards
    3. Market Intelligence   — labour market, locations, competition, salary, demand
    4. Sources & Confidence  — data quality, API status, methodology
    5. ROI Projections       — per-channel hire forecasts, cost-per-hire, time-to-fill
    6. Quality Intelligence  — gold standard gates (conditional)
    7. 90-Day Forecast       — rolling monthly spend, apps, hires, CPA trend
    8. Confidence Intervals  — low/expected/high ranges for CPA, CPH, apps, hires
    9. Niche Board Matching  — role-level specialty job board recommendations

Design: Joveo 2026 deck palette (Indigo/Purple/Teal/Magenta on lavender surfaces),
Poppins headings + Inter body, clean professional layout.
All content starts at column B (col A = left margin).

Function signature mirrors generate_excel() — receives the same enriched data dict
and returns bytes (BytesIO.getvalue()).
"""

from __future__ import annotations

import io
import logging
import re
import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

from shared_utils import (
    parse_budget,
    INDUSTRY_LABEL_MAP,
    internal_qc_mode as _internal_qc_mode,
)

from joveo_brand_2026 import (
    INDIGO,
    PURPLE,
    PURPLE_LIGHT,
    TEAL,
    MAGENTA,
    LAVENDER_100,
    LAVENDER_50,
    INK,
    MUTED as _BRAND_MUTED,
    BORDER,
    CANVAS,
)
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# S48: Channel Recommender (optional)
try:
    from channel_recommender import recommend_channels as _recommend_channels_fn

    _HAS_CHANNEL_RECOMMENDER = True
except ImportError:
    _HAS_CHANNEL_RECOMMENDER = False

try:
    import plan_currency as _plan_currency
except ImportError:  # pragma: no cover - plan_currency ships with the repo
    _plan_currency = None

import plan_geo
import display_format
import insight_composer

# NOTE: aliased -- several functions in this module already use a local
# variable/parameter literally named `gold_standard` (the enriched
# ``data["_gold_standard"]`` dict from apply_all_quality_gates), which would
# shadow a bare `import gold_standard` inside those functions' scopes.
import gold_standard as gs_lib

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# S3: Plan currency (mirrors the thread-local pattern in ppt_generator.py).
# Every money figure that is the PLAN'S OWN budget/spend/CPA/CPH/allocation
# number must render in the plan's local currency, consistently between the
# deck and this workbook. Only fixed US-benchmark constants (FRED macro data,
# Intl Benchmarks sheet, Joveo Campaign Warehouse) may stay USD, and those
# call sites pass currency="USD" explicitly plus carry an inline "(USD)"
# marker directly on the figure/header -- never a bare "$".
#
# THREAD-LOCAL: generate_excel_v2 can run on concurrent per-request threads
# under the threading HTTP server, so (as in ppt_generator.py) the active
# currency is stored per-thread, not as a shared module global.
# ---------------------------------------------------------------------------
import threading as _threading  # noqa: E402

_currency_tls = _threading.local()


def _get_active_currency() -> str:
    """Active plan currency for THIS thread (defaults to USD)."""
    return getattr(_currency_tls, "code", "USD") or "USD"


def _plan_currency_code(data: Optional[dict]) -> str:
    """Resolve the ISO currency code for a plan from its data. Defaults to USD."""
    if not isinstance(data, dict):
        return "USD"
    explicit = data.get("currency_code") or data.get("currency")
    if isinstance(explicit, str) and explicit.strip():
        return explicit.strip().upper()
    if _plan_currency is None:
        return "USD"
    candidates: List[str] = []
    for key in ("country", "primary_location"):
        val = data.get(key)
        if isinstance(val, str) and val.strip():
            candidates.append(val)
    locs = data.get("locations") or []
    if isinstance(locs, (list, tuple)):
        for loc in locs:
            if isinstance(loc, str) and loc.strip():
                candidates.append(loc)
            elif isinstance(loc, dict):
                country = loc.get("country") or loc.get("location") or ""
                if isinstance(country, str) and country.strip():
                    candidates.append(country)
    for cand in candidates:
        try:
            code = _plan_currency.currency_for_country(cand)
        except Exception:  # noqa: BLE001 - resolution is best-effort
            code = None
        if code:
            return code
    return "USD"


def _set_active_currency(data: Optional[dict]) -> str:
    """Resolve and remember the plan currency for the duration of generation."""
    code = _plan_currency_code(data)
    _currency_tls.code = code
    if isinstance(data, dict):
        data["_plan_currency_code"] = code
    return code


def _cur_symbol(currency: Optional[str] = None) -> str:
    """Return the display symbol for ``currency`` (or the active plan currency)."""
    code = (currency or _get_active_currency() or "USD").strip().upper()
    if _plan_currency is not None:
        return _plan_currency.symbol_for_code(code)
    return "$" if code == "USD" else code + " "


def _usd_number_format(base_pattern: str) -> str:
    """Build a number_format string with the ACTIVE currency's symbol baked in.

    ``base_pattern`` is one of the ``#,##0`` / ``#,##0.00`` digit patterns
    (no symbol). The symbol is quoted so openpyxl/Excel treat it as a literal
    prefix rather than a format directive.
    """
    sym = _cur_symbol()
    return f'"{sym}"{base_pattern}'


def _usd0_fmt() -> str:
    """Active-currency whole-dollar number format (e.g. NZ$#,##0)."""
    return _usd_number_format("#,##0")


def _usd2_fmt() -> str:
    """Active-currency per-unit number format (e.g. NZ$#,##0.00)."""
    return _usd_number_format("#,##0.00")

# ---------------------------------------------------------------------------
# Seasonal Hiring Trends -- loaded once from data/seasonal_hiring_trends.json
# Used to adjust 90-day forecast phasing based on industry seasonality.
# ---------------------------------------------------------------------------
_SEASONAL_PATTERNS: dict = {}


def _load_seasonal_patterns() -> dict:
    """Load seasonal hiring trends from JSON. Cached after first call."""
    global _SEASONAL_PATTERNS
    if _SEASONAL_PATTERNS:
        return _SEASONAL_PATTERNS
    import json
    from pathlib import Path

    _path = Path(__file__).parent / "data" / "seasonal_hiring_trends.json"
    try:
        with open(_path, encoding="utf-8") as f:
            raw = json.load(f)
        _SEASONAL_PATTERNS = raw.get("seasonal_patterns", {})
        logger.info(
            "Seasonal hiring patterns loaded: %d industries", len(_SEASONAL_PATTERNS)
        )
    except FileNotFoundError:
        logger.warning("Seasonal hiring data not found: %s", _path)
    except (json.JSONDecodeError, OSError) as exc:
        logger.error("Failed to load seasonal hiring data: %s", exc, exc_info=True)
    return _SEASONAL_PATTERNS


def _seasonal_monthly_phasing(industry: str, campaign_start_month: int) -> list[float]:
    """Compute 3-month budget phasing adjusted for seasonal hiring patterns.

    Falls back to the standard ramp-up curve [0.25, 0.35, 0.40] when no
    seasonal data is available for the given industry.

    Args:
        industry: Raw industry string from form input.
        campaign_start_month: 1-12, the month the campaign begins.

    Returns:
        List of 3 floats summing to 1.0 representing monthly budget shares.
    """
    default_phasing = [0.25, 0.35, 0.40]
    patterns = _load_seasonal_patterns()
    if not patterns or not industry:
        return default_phasing

    # Normalize industry to match seasonal_hiring_trends.json keys
    ind_lower = industry.lower().strip()
    # Direct and substring matching
    matched_key = ""
    for key in patterns:
        if key in ind_lower or ind_lower in key:
            matched_key = key
            break
    # Broader keyword mapping for common industry names
    if not matched_key:
        _industry_map = {
            "tech": "technology",
            "software": "technology",
            "it ": "technology",
            "information technology": "technology",
            "saas": "technology",
            "health": "healthcare",
            "medical": "healthcare",
            "pharma": "healthcare",
            "hospital": "healthcare",
            "nursing": "healthcare",
            "retail": "retail",
            "ecommerce": "retail",
            "e-commerce": "retail",
            "hospitality": "hospitality",
            "hotel": "hospitality",
            "restaurant": "hospitality",
            "food service": "hospitality",
            "construction": "construction",
            "building": "construction",
            "education": "education",
            "university": "education",
            "school": "education",
            "finance": "finance",
            "banking": "finance",
            "insurance": "finance",
            "financial": "finance",
            "manufactur": "manufacturing",
            "industrial": "manufacturing",
            "logistics": "logistics",
            "warehouse": "logistics",
            "supply chain": "logistics",
            "shipping": "logistics",
            "freight": "logistics",
            "staffing": "staffing",
            "recruiting": "staffing",
            "temp agency": "staffing",
            "transport": "transportation",
            "trucking": "transportation",
            "driving": "transportation",
            "cdl": "transportation",
            "government": "government",
            "federal": "government",
            "public sector": "government",
        }
        for keyword, seasonal_key in _industry_map.items():
            if keyword in ind_lower:
                matched_key = seasonal_key
                break

    if not matched_key or matched_key not in patterns:
        return default_phasing

    pattern = patterns[matched_key]
    peak_months = set(pattern.get("peak_months", []))
    low_months = set(pattern.get("low_months", []))
    peak_mult = pattern.get("peak_multiplier", 1.15)
    low_mult = pattern.get("low_multiplier", 0.85)

    # Build raw weights for the 3 campaign months
    raw_weights = []
    for i in range(3):
        m = ((campaign_start_month - 1 + i) % 12) + 1
        if m in peak_months:
            raw_weights.append(peak_mult)
        elif m in low_months:
            raw_weights.append(low_mult)
        else:
            raw_weights.append(1.0)

    # Apply standard ramp-up curve as a base, then modulate by seasonal weights.
    # This preserves the ramp-up shape (month 1 < month 2 < month 3) while
    # shifting budget toward peak hiring months.
    base = [0.25, 0.35, 0.40]
    adjusted = [b * w for b, w in zip(base, raw_weights)]

    # Normalize to sum to 1.0
    total = sum(adjusted)
    if total <= 0:
        return default_phasing
    shares = [round(a / total, 4) for a in adjusted]
    # Rounding each share to 4 decimals independently can leave the list
    # summing to e.g. 0.9999 instead of 1.0 (observed: seasonal weights on
    # a 24-week logistics campaign summed to 0.9999, silently under-spending
    # the printed 90-Day Total by ~$8 relative to the sum of the three
    # printed monthly Spend cells). Push the residual onto the largest
    # share so the three values always sum to EXACTLY 1.0 -- monthly_spend
    # (== _budget_90d * share) then foots exactly to _budget_90d.
    residual = round(1.0 - sum(shares), 4)
    if residual:
        max_idx = max(range(len(shares)), key=lambda i: shares[i])
        shares[max_idx] = round(shares[max_idx] + residual, 4)
    return shares


# ---------------------------------------------------------------------------
# Design Tokens -- Joveo 2026 deck palette (canonical: joveo_brand_2026.py)
# openpyxl needs bare hex (no '#'), so each constant is derived from the
# canonical hex via .lstrip("#").upper(). Names kept for minimal diff; the
# resulting bare-hex strings are byte-identical to the previous literals
# (202058, 5A54BE, ECEAF7, F4F4FF, 1F2937, 6E6E8C, E3E1F1, FFFCF9).
# ---------------------------------------------------------------------------
NAVY = INDIGO.lstrip("#").upper()  # 202058 INDIGO — deep brand navy (headers, bars)
SAPPHIRE = PURPLE.lstrip("#").upper()  # 5A54BE PURPLE — primary accent
BLUE_LIGHT = LAVENDER_100.lstrip("#").upper()  # ECEAF7 — light tint fill (badges)
BLUE_PALE = LAVENDER_50.lstrip("#").upper()  # F4F4FF — pale alt-row / background fill
STONE = INK.lstrip("#").upper()  # 1F2937 INK — body text
MUTED = _BRAND_MUTED.lstrip("#").upper()  # 6E6E8C — footnotes / secondary labels
WARM_GRAY = BORDER.lstrip("#").upper()  # E3E1F1 BORDER — grid / borders
OFF_WHITE = CANVAS.lstrip("#").upper()  # FFFCF9 CANVAS — warm off-white surface
TEAL_HEX = TEAL.lstrip("#").upper()  # 6BB5CE — secondary accent / tab
MAGENTA_HEX = MAGENTA.lstrip("#").upper()  # B7669E — pop accent / tab
PURPLE_LIGHT_HEX = PURPLE_LIGHT.lstrip("#").upper()  # 8680D6 — tertiary accent / tab

# ---------------------------------------------------------------------------
# Excel number-format strings (S89) -- write LIVE numeric values into data
# cells + apply these formats, so a client can SUM / sort / filter / chart the
# deliverable. Never write pre-formatted display strings into summable columns.
#
# FMT_USD0 / FMT_USD2 are fixed-USD formats: reserve them ONLY for genuinely
# US-calibrated benchmark constants (Intl Benchmarks sheet, FRED macro data,
# Joveo Campaign Warehouse) that carry an explicit inline USD marker. Any cell
# holding the PLAN'S OWN budget/spend/CPA/CPH/allocation figure must use
# ``_usd0_fmt()`` / ``_usd2_fmt()`` instead (S3) -- those resolve the ACTIVE
# plan currency at write-time so NZD/GBP/etc. plans render in their own
# currency rather than a hardcoded "$".
# ---------------------------------------------------------------------------
FMT_USD0 = "$#,##0"  # whole-dollar columns -- US-benchmark constants ONLY
FMT_USD2 = '"$"#,##0.00'  # per-unit money -- US-benchmark constants ONLY
FMT_PCT1 = "0.0%"  # percentages (cell stores the fraction, e.g. 0.32)
FMT_INT = "#,##0"  # integer counts (clicks, applications, hires)
# ---------------------------------------------------------------------------
# Brand name casing -- preserves known brand names when title-casing client
# ---------------------------------------------------------------------------
_BRAND_CASING: dict[str, str] = {
    "fedex": "FedEx",
    "linkedin": "LinkedIn",
    "youtube": "YouTube",
    "ibm": "IBM",
    "ups": "UPS",
    "jpmorgan": "JPMorgan",
    "walmart": "Walmart",
    "mcdonalds": "McDonald's",
    "at&t": "AT&T",
    "bmw": "BMW",
    "dhl": "DHL",
    "usps": "USPS",
    "xpo": "XPO",
    "jb hunt": "J.B. Hunt",
    "j.b. hunt": "J.B. Hunt",
    "hca": "HCA",
    "cvs": "CVS",
    "ge": "GE",
    "3m": "3M",
    "bp": "BP",
    "ihg": "IHG",
}


def _proper_client_name(name: str) -> str:
    """Client-facing casing for a client name.

    Known brand overrides (_BRAND_CASING: "AT&T", "J.B. Hunt", ...) win
    first; otherwise delegates to display_format.client_display_name's
    word-wise casing (mirrors ppt_generator._proper_client_name -- the deck
    and workbook must agree). The previous "title-case only if fully upper
    OR fully lower, else leave alone" rule silently passed mixed-case raw
    input straight through: "atria Senior living" (as literally received)
    never became "Atria Senior Living" here because it wasn't ALL lower or
    ALL upper, even though ppt_generator's copy of this same helper had
    already been fixed -- the deck showed the correct casing while the
    workbook showed the raw client-submitted casing verbatim.
    """
    if not name or name == "Client":
        return name
    lower = name.strip().lower()
    if lower in _BRAND_CASING:
        return _BRAND_CASING[lower]
    return (
        display_format.client_display_name(name)
        or name.strip()
    )


GREEN = "16A34A"
GREEN_BG = "DCFCE7"
AMBER = "D97706"
AMBER_BG = "FEF3C7"
RED = "DC2626"
RED_BG = "FEE2E2"
WHITE = "FFFFFF"

# ---------------------------------------------------------------------------
# Reusable openpyxl style objects
# ---------------------------------------------------------------------------
# Joveo brand fonts: Poppins for headings/titles, Inter for body.
FONT_HEAD = "Poppins"
FONT_BODY_NAME = "Inter"
_FONT_SECTION = Font(name=FONT_HEAD, bold=True, size=14, color=WHITE)
_FONT_SUBSECTION = Font(name=FONT_HEAD, bold=True, size=12, color=NAVY)
_FONT_TABLE_HEADER = Font(name=FONT_HEAD, bold=True, size=10, color=WHITE)
_FONT_TABLE_HEADER_ALT = Font(name=FONT_HEAD, bold=True, size=10, color=NAVY)
_FONT_BODY = Font(name=FONT_BODY_NAME, size=10, color=STONE)
_FONT_BODY_BOLD = Font(name=FONT_BODY_NAME, bold=True, size=10, color=STONE)
_FONT_FOOTNOTE = Font(name=FONT_BODY_NAME, italic=True, size=9, color=MUTED)
_FONT_HERO = Font(name=FONT_HEAD, bold=True, size=18, color=NAVY)
_FONT_HERO_VALUE = Font(name=FONT_HEAD, bold=True, size=22, color=SAPPHIRE)
_FONT_METRIC_LABEL = Font(name=FONT_BODY_NAME, size=9, color=MUTED)
_FONT_METRIC_VALUE = Font(name=FONT_HEAD, bold=True, size=14, color=NAVY)
_FONT_GRADE_LARGE = Font(name=FONT_HEAD, bold=True, size=36, color=WHITE)

_FILL_NAVY = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
_FILL_SAPPHIRE = PatternFill(
    start_color=SAPPHIRE, end_color=SAPPHIRE, fill_type="solid"
)
_FILL_BLUE_LIGHT = PatternFill(
    start_color=BLUE_LIGHT, end_color=BLUE_LIGHT, fill_type="solid"
)
_FILL_BLUE_PALE = PatternFill(
    start_color=BLUE_PALE, end_color=BLUE_PALE, fill_type="solid"
)
_FILL_OFF_WHITE = PatternFill(
    start_color=OFF_WHITE, end_color=OFF_WHITE, fill_type="solid"
)
_FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
_FILL_GREEN = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
_FILL_GREEN_BG = PatternFill(
    start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid"
)
_FILL_AMBER_BG = PatternFill(
    start_color=AMBER_BG, end_color=AMBER_BG, fill_type="solid"
)
_FILL_RED_BG = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")

_ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_CENTER_TOP = Alignment(horizontal="center", vertical="top", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=True)

_BORDER_THIN = Border(
    left=Side(style="thin", color=WARM_GRAY),
    right=Side(style="thin", color=WARM_GRAY),
    top=Side(style="thin", color=WARM_GRAY),
    bottom=Side(style="thin", color=WARM_GRAY),
)
_BORDER_BOTTOM_SAPPHIRE = Border(bottom=Side(style="medium", color=SAPPHIRE))

# Content column range (B=2 through H=8)
COL_START = 2  # column B
COL_END = 8  # column H
COL_SPAN = COL_END - COL_START + 1  # 7 columns

# ---------------------------------------------------------------------------
# Industry Niche Channels
# ---------------------------------------------------------------------------
INDUSTRY_NICHE_CHANNELS: Dict[str, List[str]] = {
    "healthcare_medical": [
        "Health eCareers",
        "Vivian Health",
        "Nurse.com",
        "PracticeLink",
        "JAMA Career Center",
        "myCNAjobs.com",
        "SeniorJobBank.com",
        "AllNurses.com",
        "NursingJobs.com",
        "CareListings.com",
    ],
    "tech_engineering": [
        "GitHub Jobs",
        "Stack Overflow",
        "Wellfound (AngelList)",
        "Dice",
        "HackerRank",
    ],
    "blue_collar_trades": [
        "TradeHounds",
        "iHire",
        "Jobcase",
        "WorkStep",
        "Skilled Workers Agency",
    ],
    "finance_banking": [
        "eFinancialCareers",
        "Wall Street Oasis",
        "Financial Job Network",
    ],
    "aerospace_defense": [
        "ClearedJobs.Net",
        "ClearanceJobs",
        "Military.com",
        "USAJOBS",
        "Hire Heroes USA",
    ],
    "logistics_supply_chain": [
        "CDLjobs.com",
        "TruckingJobs.com",
        "WarehouseJobs.com",
        "Supply Chain Online",
    ],
    "pharma_biotech": [
        "BioSpace",
        "MedReps",
        "Science Careers (AAAS)",
        "Nature Careers",
    ],
    "retail_consumer": ["RetailGigs", "Snagajob", "Wonolo", "Instawork"],
    "hospitality_travel": ["Hcareers", "Poached", "Culinary Agents", "Harri"],
    "education": ["HigherEdJobs", "SchoolSpring", "K12JobSpot", "Chronicle Vitae"],
    "energy_utilities": ["EnergyJobline", "Rigzone", "Power Magazine Careers"],
    "insurance": ["Insurance Jobs", "The Institutes", "InsuranceJobs.com"],
    "construction_real_estate": [
        "ConstructionJobs.com",
        "iHireConstruction",
        "Built Hire",
    ],
    "automotive": ["AutoJobs.com", "AutomotiveCrossing", "Automotive News Careers"],
    "food_beverage": ["Poached", "Culinary Agents", "RestaurantJobs.com"],
    "media_entertainment": ["MediaBistro", "ProductionHub", "Mandy.com", "Staff Me Up"],
    "telecommunications": [
        "WirelessEstimator",
        "FierceTelecom Jobs",
        "Light Reading Careers",
    ],
    "mental_health": ["Psychology Today Jobs", "SAMHSA Jobs", "APA PsycCareers"],
    "maritime_marine": ["Maritime Jobs", "Rigzone", "Sea Career"],
}


def get_niche_vendor_availability(industry: str, us_plan: bool) -> Dict[str, bool]:
    """channel_key -> whether real (non-fabricated) niche-vendor data exists.

    ``INDUSTRY_NICHE_CHANNELS`` is a US-domiciled board list -- it is only
    valid vendor data for a US plan targeting a covered industry. Callers
    (e.g. app.py's budget allocation) use this to decide whether the
    ``niche_boards`` channel can carry a data-backed allocation, instead of
    always assuming named vendors are available. Dependency-light: reads
    only the module-level board table, no research/enrichment calls.
    """
    has_vendors = bool(us_plan) and bool(INDUSTRY_NICHE_CHANNELS.get(industry))
    return {"niche_boards": has_vendors}


# ---------------------------------------------------------------------------
# Role-Level Niche Board Matching (Task 4)
# Maps role keywords to specialty job boards for targeted channel recommendations.
# ---------------------------------------------------------------------------
ROLE_NICHE_BOARDS: Dict[str, Dict[str, Any]] = {
    "software_engineer": {
        "industry": "tech_engineering",
        "keywords": [
            "software engineer",
            "developer",
            "full stack",
            "frontend",
            "backend",
            "web developer",
            "mobile developer",
            "ios",
            "android",
            "react",
            "python developer",
            "java developer",
            "golang",
        ],
        "boards": [
            {
                "name": "Dice",
                "url": "dice.com",
                "strength": "Tech-focused aggregator, strong for contract/perm",
            },
            {
                "name": "Stack Overflow Jobs",
                "url": "stackoverflow.com/jobs",
                "strength": "Developer community, passive candidates",
            },
            {
                "name": "GitHub Jobs",
                "url": "github.com/jobs",
                "strength": "Open source community reach",
            },
            {
                "name": "Wellfound (AngelList)",
                "url": "wellfound.com",
                "strength": "Startup ecosystem, equity-motivated candidates",
            },
            {
                "name": "HackerRank",
                "url": "hackerrank.com",
                "strength": "Skill-verified candidates",
            },
        ],
    },
    "data_science": {
        "industry": "tech_engineering",
        "keywords": [
            "data scientist",
            "machine learning",
            "ml engineer",
            "ai engineer",
            "deep learning",
            "nlp",
            "data analyst",
            "analytics engineer",
            "data engineer",
            "mlops",
        ],
        "boards": [
            {
                "name": "Kaggle Jobs",
                "url": "kaggle.com/jobs",
                "strength": "ML/AI community, competition-proven talent",
            },
            {
                "name": "DataJobs.com",
                "url": "datajobs.com",
                "strength": "Data-specific roles",
            },
            {
                "name": "Dice",
                "url": "dice.com",
                "strength": "Tech-focused, strong data science segment",
            },
            {
                "name": "AI Jobs Board",
                "url": "aijobs.net",
                "strength": "AI/ML specialty listings",
            },
        ],
    },
    "cybersecurity": {
        "industry": "tech_engineering",
        "keywords": [
            "cybersecurity",
            "security engineer",
            "security analyst",
            "infosec",
            "penetration tester",
            "soc analyst",
            "ciso",
            "security architect",
        ],
        "boards": [
            {
                "name": "CyberSecJobs",
                "url": "cybersecjobs.com",
                "strength": "Cybersecurity-only board",
            },
            {
                "name": "ClearedJobs.Net",
                "url": "clearedjobs.net",
                "strength": "Clearance-required security roles",
            },
            {
                "name": "Dice",
                "url": "dice.com",
                "strength": "Strong cybersecurity segment",
            },
            {
                "name": "SANS Job Board",
                "url": "sans.org/careers",
                "strength": "SANS-certified professional network",
            },
        ],
    },
    "nursing": {
        "industry": "healthcare_medical",
        "keywords": [
            "nurse",
            "rn",
            "lpn",
            "cna",
            "nurse practitioner",
            "np",
            "registered nurse",
            "travel nurse",
            "icu nurse",
            "or nurse",
            "nursing",
            "bsn",
        ],
        "boards": [
            {
                "name": "Vivian Health",
                "url": "vivian.com",
                "strength": "Travel + staff nursing, transparent pay",
            },
            {
                "name": "Nurse.com",
                "url": "nurse.com",
                "strength": "Largest nursing community, CE integration",
            },
            {
                "name": "Health eCareers",
                "url": "healthecareers.com",
                "strength": "Multi-specialty healthcare",
            },
            {
                "name": "NurseFly",
                "url": "nursefly.com",
                "strength": "Travel nursing marketplace",
            },
            {
                "name": "Incredible Health",
                "url": "incrediblehealth.com",
                "strength": "Employer-applies-to-nurse model",
            },
            {
                "name": "AllNurses.com",
                "url": "allnurses.com",
                "strength": "Largest nursing community, strong employer brand reach",
            },
            {
                "name": "NursingJobs.com",
                "url": "nursingjobs.com",
                "strength": "Nursing-only job board, high-intent candidates",
            },
        ],
    },
    "senior_care": {
        "industry": "healthcare_medical",
        "keywords": [
            "senior care",
            "senior living",
            "assisted living",
            "memory care",
            "long term care",
            "ltc",
            "home health aide",
            "hha",
            "caregiver",
            "geriatric",
            "elder care",
            "residential care",
        ],
        "boards": [
            {
                "name": "myCNAjobs.com",
                "url": "mycnajobs.com",
                "strength": "CNA/caregiver-focused, senior care specialty",
            },
            {
                "name": "SeniorJobBank.com",
                "url": "seniorjobbank.com",
                "strength": "Senior living industry job board",
            },
            {
                "name": "Health eCareers",
                "url": "healthecareers.com",
                "strength": "Multi-specialty healthcare including senior care",
            },
            {
                "name": "CareListings.com",
                "url": "carelistings.com",
                "strength": "Senior care and home health job board",
            },
            {
                "name": "Vivian Health",
                "url": "vivian.com",
                "strength": "Healthcare staffing including senior care facilities",
            },
        ],
    },
    "physician": {
        "industry": "healthcare_medical",
        "keywords": [
            "physician",
            "doctor",
            "md",
            "surgeon",
            "hospitalist",
            "anesthesiologist",
            "radiologist",
            "cardiologist",
            "dermatologist",
            "psychiatrist",
            "pediatrician",
        ],
        "boards": [
            {
                "name": "PracticeLink",
                "url": "practicelink.com",
                "strength": "Physician-only, permanent placement",
            },
            {
                "name": "Doximity",
                "url": "doximity.com",
                "strength": "Physician social network, verified MDs",
            },
            {
                "name": "JAMA Career Center",
                "url": "careers.jamanetwork.com",
                "strength": "Academic/research physicians",
            },
            {
                "name": "Health eCareers",
                "url": "healthecareers.com",
                "strength": "Broad healthcare, physician segment",
            },
        ],
    },
    "allied_health": {
        "industry": "healthcare_medical",
        "keywords": [
            "therapist",
            "physical therapist",
            "occupational therapist",
            "pharmacist",
            "respiratory therapist",
            "radiology tech",
            "medical assistant",
            "lab technician",
            "phlebotomist",
        ],
        "boards": [
            {
                "name": "Health eCareers",
                "url": "healthecareers.com",
                "strength": "Multi-specialty allied health",
            },
            {
                "name": "Vivian Health",
                "url": "vivian.com",
                "strength": "Allied health travel positions",
            },
            {
                "name": "AlliedTravelCareers",
                "url": "alliedtravelcareers.com",
                "strength": "Travel allied health",
            },
        ],
    },
    "executive": {
        "industry": "general",
        "keywords": [
            "ceo",
            "cfo",
            "cto",
            "cio",
            "coo",
            "cmo",
            "chief",
            "president",
            "vice president",
            "vp",
            "svp",
            "evp",
            "managing director",
            "general manager",
            "c-suite",
        ],
        "boards": [
            {
                "name": "LinkedIn Executive Search",
                "url": "linkedin.com/talent",
                "strength": "Executive passive candidate network",
            },
            {
                "name": "ExecuNet",
                "url": "execunet.com",
                "strength": "C-suite and board-level positions",
            },
            {
                "name": "Ladders",
                "url": "theladders.com",
                "strength": "$100K+ positions, executive focus",
            },
            {
                "name": "BlueSteps",
                "url": "bluesteps.com",
                "strength": "AESC-affiliated executive search",
            },
        ],
    },
    "trucking": {
        "industry": "transportation_logistics",
        "keywords": [
            "cdl",
            "truck driver",
            "trucker",
            "otr driver",
            "class a",
            "class b",
            "delivery driver",
            "long haul",
            "local driver",
            "fleet driver",
        ],
        "boards": [
            {
                "name": "CDLjobs.com",
                "url": "cdljobs.com",
                "strength": "CDL-specific, high intent",
            },
            {
                "name": "TruckingJobs.com",
                "url": "truckingjobs.com",
                "strength": "Trucking industry focus",
            },
            {
                "name": "DriveMyWay",
                "url": "drivemyway.com",
                "strength": "Driver-job matching algorithm",
            },
            {
                "name": "TruckersReport Jobs",
                "url": "thetruckersreport.com/jobs",
                "strength": "Active trucker community",
            },
        ],
    },
    "warehouse": {
        "industry": "transportation_logistics",
        "keywords": [
            "warehouse",
            "forklift",
            "picker",
            "packer",
            "shipping",
            "receiving",
            "inventory",
            "distribution",
            "fulfillment",
            "material handler",
        ],
        "boards": [
            {
                "name": "WarehouseJobs.com",
                "url": "warehousejobs.com",
                "strength": "Warehouse-specific board",
            },
            {
                "name": "Wonolo",
                "url": "wonolo.com",
                "strength": "On-demand warehouse staffing",
            },
            {
                "name": "Instawork",
                "url": "instawork.com",
                "strength": "Flexible warehouse/logistics shifts",
            },
            {
                "name": "Jobcase",
                "url": "jobcase.com",
                "strength": "Hourly/blue collar community",
            },
        ],
    },
    "accounting": {
        "industry": "finance_banking",
        "keywords": [
            "accountant",
            "cpa",
            "auditor",
            "tax",
            "bookkeeper",
            "controller",
            "financial analyst",
            "accounts payable",
            "accounts receivable",
        ],
        "boards": [
            {
                "name": "eFinancialCareers",
                "url": "efinancialcareers.com",
                "strength": "Finance/accounting specialty",
            },
            {
                "name": "AccountingJobsToday",
                "url": "accountingjobstoday.com",
                "strength": "Accounting-only listings",
            },
            {
                "name": "Robert Half",
                "url": "roberthalf.com",
                "strength": "Accounting staffing leader",
            },
        ],
    },
    "sales": {
        "industry": "general",
        "keywords": [
            "sales representative",
            "account executive",
            "business development",
            "sales manager",
            "account manager",
            "sdr",
            "bdr",
            "sales engineer",
            "enterprise sales",
        ],
        "boards": [
            {
                "name": "Rainmakers",
                "url": "rainmakers.co",
                "strength": "Sales talent marketplace, verified quotas",
            },
            {
                "name": "SalesJobs.com",
                "url": "salesjobs.com",
                "strength": "Sales-only board",
            },
            {
                "name": "RepVue",
                "url": "repvue.com",
                "strength": "Sales org ratings, compensation data",
            },
        ],
    },
    "marketing": {
        "industry": "general",
        "keywords": [
            "marketing manager",
            "digital marketing",
            "content marketing",
            "seo",
            "social media manager",
            "brand manager",
            "growth marketing",
            "product marketing",
        ],
        "boards": [
            {
                "name": "MarketingHire",
                "url": "marketinghire.com",
                "strength": "Marketing-specific positions",
            },
            {
                "name": "MediaBistro",
                "url": "mediabistro.com",
                "strength": "Media/marketing/creative jobs",
            },
            {
                "name": "Built In",
                "url": "builtin.com",
                "strength": "Tech marketing roles, company profiles",
            },
        ],
    },
    "construction": {
        "industry": "construction_real_estate",
        "keywords": [
            "construction",
            "electrician",
            "plumber",
            "hvac",
            "carpenter",
            "welder",
            "mason",
            "ironworker",
            "heavy equipment operator",
            "project manager construction",
        ],
        "boards": [
            {
                "name": "ConstructionJobs.com",
                "url": "constructionjobs.com",
                "strength": "Construction-only board",
            },
            {
                "name": "iHireConstruction",
                "url": "ihireconstruction.com",
                "strength": "Construction staffing network",
            },
            {
                "name": "TradeHounds",
                "url": "tradehounds.com",
                "strength": "Skilled trades social network",
            },
            {
                "name": "Built Hire",
                "url": "builthire.com",
                "strength": "Construction workforce platform",
            },
        ],
    },
    "education": {
        "industry": "education",
        "keywords": [
            "teacher",
            "professor",
            "instructor",
            "educator",
            "principal",
            "academic",
            "curriculum",
            "dean",
            "superintendent",
        ],
        "boards": [
            {
                "name": "HigherEdJobs",
                "url": "higheredjobs.com",
                "strength": "Higher education positions",
            },
            {
                "name": "SchoolSpring",
                "url": "schoolspring.com",
                "strength": "K-12 teaching positions",
            },
            {
                "name": "K12JobSpot",
                "url": "k12jobspot.com",
                "strength": "K-12 administration and teaching",
            },
            {
                "name": "Chronicle Vitae",
                "url": "chroniclevitae.com",
                "strength": "Academic career network",
            },
        ],
    },
    "legal": {
        "industry": "professional_services",
        "keywords": [
            "attorney",
            "lawyer",
            "paralegal",
            "legal assistant",
            "general counsel",
            "litigation",
            "corporate counsel",
            "compliance officer",
        ],
        "boards": [
            {
                "name": "LawCrossing",
                "url": "lawcrossing.com",
                "strength": "Legal-only job aggregator",
            },
            {
                "name": "Lawjobs.com",
                "url": "lawjobs.com",
                "strength": "Legal staffing marketplace",
            },
            {
                "name": "Robert Half Legal",
                "url": "roberthalf.com/legal",
                "strength": "Legal staffing leader",
            },
        ],
    },
    "localization": {
        "industry": "general",
        "keywords": [
            "translator",
            "interpreter",
            "localization",
            "translation",
            "bilingual",
            "multilingual",
            "language specialist",
            "voice talent",
            "voice actor",
            "voice over",
            "voiceover",
            "narration",
        ],
        "boards": [
            {
                "name": "ProZ.com",
                "url": "proz.com",
                "strength": "Largest translation community, 1M+ translators",
            },
            {
                "name": "TranslatorsCafe.com",
                "url": "translatorscafe.com",
                "strength": "Translation and localization job board",
            },
            {
                "name": "Voices.com",
                "url": "voices.com",
                "strength": "Voice talent marketplace, 4M+ voice actors",
            },
            {
                "name": "Voice123.com",
                "url": "voice123.com",
                "strength": "Voice over talent platform",
            },
        ],
    },
}


def _keyword_matches_role(kw: str, role_lower: str) -> bool:
    """Check if keyword matches role with word-boundary awareness.

    Short keywords (< 4 chars like 'rn', 'np', 'cna', 'md', 'vp') require
    word-boundary matching to prevent false positives where 'rn' matches
    inside 'frontend' or 'learning'.  Longer keywords use substring matching.
    """
    if kw in role_lower:
        # Keyword found in role -- verify word boundary for short keywords
        if len(kw) < 4:
            # Require word boundary: keyword must appear as a standalone word
            # e.g., "rn" matches "rn", "rn supervisor", "icu rn" but NOT "frontend"
            return bool(
                re.search(r"(?<![a-z])" + re.escape(kw) + r"(?![a-z])", role_lower)
            )
        return True
    # Reverse check: role appears in keyword (e.g., role="nurse" matches kw="nurse practitioner")
    # Only allow this for longer role strings to prevent short-token false positives
    if len(role_lower) >= 4 and role_lower in kw:
        return True
    return False


# Industry compatibility matrix: which board industries are allowed for which plan industries.
# "general" boards are always allowed.  Categories not listed here are only matched when
# the plan industry is compatible or unset.
_INDUSTRY_COMPATIBILITY: Dict[str, set] = {
    "healthcare_medical": {
        "healthcare_medical",
        "general",
    },
    "tech_engineering": {
        "tech_engineering",
        "general",
    },
    "finance_banking": {
        "finance_banking",
        "general",
    },
    "professional_services": {
        "professional_services",
        "finance_banking",
        "general",
    },
    "transportation_logistics": {
        "transportation_logistics",
        "general",
    },
    "construction_real_estate": {
        "construction_real_estate",
        "general",
    },
    "education": {
        "education",
        "general",
    },
    "retail_consumer": {
        "retail_consumer",
        "general",
    },
    "hospitality_travel": {
        "hospitality_travel",
        "retail_consumer",
        "general",
    },
}


def _match_roles_to_niche_boards(
    roles: List[str],
    industry: str = "",
) -> Dict[str, List[Dict[str, str]]]:
    """Cross-reference role titles against ROLE_NICHE_BOARDS to find specialty boards.

    Uses word-boundary-aware matching for short keywords to prevent false
    positives (e.g., 'rn' matching 'frontend').  When an industry is provided,
    filters out boards from incompatible industries (e.g., tech boards for
    healthcare plans).

    Args:
        roles: List of role title strings.
        industry: Canonical industry key (e.g., 'healthcare_medical').
            When provided, boards from incompatible industries are excluded.

    Returns:
        Dict mapping role title to list of recommended niche boards.
        Each board entry has keys: name, url, strength.
    """
    if not roles:
        return {}

    # Determine which board industries are allowed for this plan industry
    ind_lower = (industry or "").lower().strip().replace(" ", "_").replace("-", "_")
    allowed_industries: Optional[set] = None
    if ind_lower:
        # Look up compatibility; if the industry has an explicit set, use it.
        # Otherwise, allow boards from the same industry + "general".
        allowed_industries = _INDUSTRY_COMPATIBILITY.get(ind_lower)
        if allowed_industries is None:
            allowed_industries = {ind_lower, "general"}

    results: Dict[str, List[Dict[str, str]]] = {}

    for role in roles:
        role_lower = role.lower().strip()
        matched_boards: List[Dict[str, str]] = []
        matched_categories: set = set()

        for category, config in ROLE_NICHE_BOARDS.items():
            # Industry filter: skip categories whose industry is incompatible
            board_industry = config.get("industry", "general")
            if allowed_industries and board_industry not in allowed_industries:
                continue

            keywords = config.get("keywords") or []
            for kw in keywords:
                if _keyword_matches_role(kw, role_lower):
                    if category not in matched_categories:
                        matched_categories.add(category)
                        for board in config.get("boards") or []:
                            matched_boards.append(dict(board))
                    break

        if matched_boards:
            # Deduplicate by board name
            seen: set = set()
            deduped: List[Dict[str, str]] = []
            for b in matched_boards:
                if b["name"] not in seen:
                    seen.add(b["name"])
                    deduped.append(b)
            results[role] = deduped

    return results


# ---------------------------------------------------------------------------
# Channel Vetting Requirements
# ---------------------------------------------------------------------------
INDUSTRY_CHANNEL_REQUIREMENTS: Dict[str, Dict[str, Any]] = {
    "healthcare_medical": {
        "preferred": [
            "health",
            "nurse",
            "medical",
            "clinical",
            "vivian",
            "doximity",
            "practicelink",
        ],
        "excluded_keywords": ["developer", "github", "stack overflow", "hacker"],
    },
    "tech_engineering": {
        "preferred": [
            "tech",
            "engineer",
            "developer",
            "github",
            "stack overflow",
            "dice",
            "wellfound",
            "hacker",
        ],
        "excluded_keywords": ["nurse", "clinical", "medical staffing"],
    },
    "blue_collar_trades": {
        "preferred": [
            "trade",
            "jobcase",
            "workstep",
            "hourly",
            "skilled",
            "warehouse",
            "cdl",
        ],
        "excluded_keywords": ["executive search", "c-suite"],
    },
    "finance_banking": {
        "preferred": ["finance", "efinancial", "wall street", "banking", "fintech"],
        "excluded_keywords": ["nurse", "clinical", "warehouse"],
    },
    "retail_consumer": {
        "preferred": ["retail", "snagajob", "hourly", "wonolo", "instawork"],
        "excluded_keywords": ["executive search", "c-suite", "clinical"],
    },
    "logistics_supply_chain": {
        "preferred": [
            "logistics",
            "cdl",
            "trucking",
            "warehouse",
            "supply chain",
            "driver",
        ],
        "excluded_keywords": ["clinical", "nurse", "executive search"],
    },
    "hospitality_travel": {
        "preferred": [
            "hospitality",
            "hcareers",
            "poached",
            "culinary",
            "harri",
            "hotel",
        ],
        "excluded_keywords": ["clinical", "developer", "github"],
    },
    "pharma_biotech": {
        "preferred": ["bio", "pharma", "science", "medreps", "clinical research"],
        "excluded_keywords": ["warehouse", "trucking"],
    },
    "aerospace_defense": {
        "preferred": [
            "cleared",
            "clearance",
            "military",
            "defense",
            "usajobs",
            "aerospace",
        ],
        "excluded_keywords": ["retail", "food service"],
    },
    "education": {
        "preferred": [
            "education",
            "highered",
            "schoolspring",
            "k12",
            "academic",
            "teaching",
        ],
        "excluded_keywords": ["warehouse", "trucking", "clinical"],
    },
}

ROLE_CHANNEL_REQUIREMENTS: Dict[str, Dict[str, Any]] = {
    "executive": {
        "preferred": [
            "executive search",
            "linkedin",
            "c-suite",
            "board",
            "spencer stuart",
        ],
        "excluded_keywords": ["hourly", "snagajob", "warehouse", "entry level"],
    },
    "professional": {
        "preferred": ["linkedin", "indeed", "glassdoor", "professional"],
        "excluded_keywords": [],
    },
    "hourly": {
        "preferred": ["snagajob", "wonolo", "instawork", "jobcase", "hourly", "shift"],
        "excluded_keywords": ["executive search", "c-suite", "spencer stuart"],
    },
    "clinical": {
        "preferred": ["vivian", "nurse", "health", "medical", "clinical", "doximity"],
        "excluded_keywords": ["warehouse", "trucking", "developer"],
    },
    "trades": {
        "preferred": ["trade", "ihire", "skilled", "construction", "cdl"],
        "excluded_keywords": ["executive search", "c-suite"],
    },
}


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════


# Mirrors ppt_generator._TITLE_ACRONYMS -- kept as a local copy (this
# codebase's convention: small formatting helpers are per-file, not
# cross-imported) so the deck and workbook always agree on channel-name
# casing. Fixes findings like "Programmatic Dsp" (should be "Programmatic
# DSP") that .replace("_", " ").title() alone can't get right.
_TITLE_ACRONYMS = {
    "Ai": "AI", "Roi": "ROI", "Cpa": "CPA", "Cpc": "CPC", "Cph": "CPH",
    "Dsp": "DSP", "Crm": "CRM", "Kpi": "KPI", "Us": "US", "Uk": "UK",
    "Eu": "EU", "Seo": "SEO", "Ats": "ATS", "Ppc": "PPC", "Qc": "QC",
    "Roas": "ROAS", "Dei": "DEI",
}


_CHANNEL_KEY_RE = re.compile(
    r"\b("
    + "|".join(
        sorted(
            (re.escape(k) for k in display_format.CHANNEL_DISPLAY),
            key=len,
            reverse=True,
        )
    )
    + r")\b"
)


def _delabel_channel_keys_in_text(text: Any) -> Any:
    """Replace any raw snake_case channel key embedded in free-text prose
    (e.g. budget_engine's own recommendation/warning strings, which
    interpolate channel keys directly -- "niche_boards, employer_branding
    may benefit from...") with its client-facing display label. Non-string
    input and channel-agnostic text pass through unchanged."""
    if not isinstance(text, str) or not text:
        return text
    return _CHANNEL_KEY_RE.sub(
        lambda m: display_format.channel_label(m.group(1)), text
    )


def _smart_title(s: str) -> str:
    """Client-facing display name for a channel key (or any string).

    Delegates to :func:`display_format.channel_label` for known channel
    keys (e.g. ``"niche_boards"`` -> "Niche / Industry Boards") so channel
    names are consistent workbook-wide; falls back to acronym-preserving
    title-case (AI, ROI, DSP...) for anything not a recognized channel key.
    """
    key = str(s)
    if key in display_format.CHANNEL_DISPLAY:
        return display_format.CHANNEL_DISPLAY[key]
    return " ".join(_TITLE_ACRONYMS.get(w, w) for w in str(s).title().split())


def _humanize_snake_key(k: Any) -> str:
    """Client-facing label for a raw snake_case dict key (e.g. KB benchmark
    keys like ``"warehouse_hourly"``, ``"cdl_drivers"``, ``"trend_yoy"``) --
    underscores become spaces BEFORE title-casing (unlike ``_smart_title``'s
    fallback branch, which only strips underscores for known channel keys
    via the CHANNEL_DISPLAY short-circuit and would otherwise leave a raw
    ``"Warehouse_Hourly"`` on the page), with acronyms (CPA, ROI, YoY, ...)
    restored via ``_TITLE_ACRONYMS``. Never returns a string containing an
    underscore."""
    words = str(k or "").replace("_", " ").title().split()
    return " ".join(_TITLE_ACRONYMS.get(w, w) for w in words)


# Matches a string that is ENTIRELY a snake_case identifier (all-lowercase
# alnum segments joined by underscores, no spaces/punctuation) -- e.g. an
# internal industry key ("logistics_supply_chain") or a KB/benchmark bucket
# key ("warehousing_logistics", "appcast_benchmark_2023"). Ordinary
# client-facing prose always contains a space or punctuation and can never
# fullmatch this, so this never mistakes real text for a leaked identifier.
_RAW_IDENTIFIER_RE = re.compile(r"^[a-z0-9]+(?:_[a-z0-9]+)+$")


def _looks_like_raw_identifier(s: str) -> bool:
    """True if ``s`` is (almost certainly) a raw internal snake_case
    identifier that should have been resolved to a display label before it
    reached client-facing text, rather than legitimate prose/formatted data."""
    return bool(_RAW_IDENTIFIER_RE.match(s))


def _safe_num(val: Any, default: float = 0.0) -> float:
    """Safely convert a value to float."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        clean = val.replace(",", "").replace("$", "").replace("%", "").strip()
        try:
            return float(clean)
        except (ValueError, TypeError):
            return default
    return default


def _fmt_currency(val: Any, prefix: Optional[str] = None, show_cents: bool = False) -> str:
    """Format a numeric value as currency.

    ``prefix`` defaults to the ACTIVE plan currency symbol (S3) -- pass an
    explicit prefix (e.g. ``"$"``) only for intentionally USD-coded
    benchmark figures that must stay USD regardless of the plan's currency.
    """
    if prefix is None:
        prefix = _cur_symbol()
    num = _safe_num(val)
    if num == 0:
        return f"{prefix}0"
    if abs(num) >= 1_000_000:
        return f"{prefix}{num / 1_000_000:,.1f}M"
    if abs(num) >= 10_000 and not show_cents:
        return f"{prefix}{num:,.0f}"
    if show_cents or abs(num) < 10:
        return f"{prefix}{num:,.2f}"
    return f"{prefix}{num:,.0f}"


def _fmt_number(val: Any, decimals: int = 0) -> str:
    """Format a number with thousand separators."""
    num = _safe_num(val)
    if num == 0:
        return "0"
    if decimals > 0:
        return f"{num:,.{decimals}f}"
    return f"{num:,.0f}"


def _fmt_pct(val: Any, decimals: int = 1) -> str:
    """Format as percentage. If val < 1, treat as fraction (0.05 -> 5.0%)."""
    num = _safe_num(val)
    if num == 0:
        return "0%"
    # If value looks like a fraction (less than 1 but not negative), convert
    if 0 < num < 1:
        num *= 100
    return f"{num:.{decimals}f}%"


def _flatten_value(val: Any, max_depth: int = 3) -> str:
    """Safely flatten a nested dict/list into a readable string.

    CRITICAL: Never call str() on raw nested structures. This iterates
    through dicts and lists to produce human-readable key-value text.
    """
    if val is None:
        return ""
    if isinstance(val, str):
        # Defense-in-depth against the same defect class as the dict-key
        # case below: a raw internal snake_case identifier can also arrive
        # as a VALUE (an industry key, an occupation-bucket key, ...) when a
        # caller forgets to resolve it through its own label map first.
        # Humanizing it here closes the leak class at its single chokepoint
        # instead of relying on every call site remembering to do it.
        if _looks_like_raw_identifier(val):
            return _humanize_snake_key(val)
        return val
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if isinstance(val, (int, float)):
        return str(val)

    if max_depth <= 0:
        return "[nested data]"

    if isinstance(val, list):
        flat_items = []
        for item in val[:10]:  # cap at 10 items
            flat_items.append(_flatten_value(item, max_depth - 1))
        return ", ".join(flat_items)

    if isinstance(val, dict):
        parts = []
        for k, v in list(val.items())[:10]:
            flat_v = _flatten_value(v, max_depth - 1)
            if flat_v:
                # Never emit a raw snake_case dict key into client-facing
                # text (shipped defect: KB benchmark dicts like
                # {"warehouse_hourly": "...", "cdl_drivers": "..."} were
                # flattened with the raw key verbatim).
                parts.append(f"{_humanize_snake_key(k)}: {flat_v}")
        return "; ".join(parts)

    return str(val)[:200]


_PHYSICIAN_ROLE_KEYWORDS = (
    "physician",
    "md ",
    " md",
    "doctor",
    "cardiolog",
    "radiolog",
    "surgeon",
    "psychiatr",
    "anesthesio",
    "oncolog",
    "neurolog",
    "dermatolog",
    "gastroenterolog",
    "urolog",
    "ophthalmolog",
    "otolaryngolog",
    "pulmonolog",
    "hospitalist",
    "obgyn",
    "ob-gyn",
)

# copy:both#6 fix: the shared "healthcare_medical" KB benchmark record mixes
# physician/MD-specific sub-stats into dicts that are otherwise generic
# healthcare figures (e.g. cph.physician_recruitment alongside
# cph.recruitment_marketing_only). Rendered unfiltered for a senior-living
# or allied-health client, "Physician Recruitment: $180,000-$250,000" reads
# as THIS client's own hiring intelligence rather than off-topic context.
_PHYSICIAN_ONLY_BENCHMARK_SUBKEYS = frozenset(
    {"physician_recruitment", "primary_care", "cardiology_psychiatry"}
)


def _plan_targets_physician_roles(roles: List[str]) -> bool:
    """True when this plan's own roles include a physician/MD-type role."""
    text = " ".join(str(r) for r in (roles or [])).lower()
    return any(kw in text for kw in _PHYSICIAN_ROLE_KEYWORDS)


def _scope_benchmark_to_plan_roles(val: Any, roles: List[str]) -> Any:
    """Drop physician/MD-only sub-stats from a shared healthcare benchmark
    dict when this plan's own roles aren't physician/MD roles (finding
    copy:both#6), so a senior-living or allied-health workbook doesn't
    present physician-specialty compensation/time-to-fill figures as its
    own hiring intelligence. No-op for non-dict values, physician-targeting
    plans, or dicts that don't carry any of the known physician-only keys.
    """
    if not isinstance(val, dict) or _plan_targets_physician_roles(roles):
        return val
    filtered = {
        k: v
        for k, v in val.items()
        if k not in _PHYSICIAN_ONLY_BENCHMARK_SUBKEYS
        and "physician" not in str(k).lower()
    }
    return filtered if filtered != val else val


# S: Sub-vertical seasonal override text for the Executive Summary's
# "Seasonal Patterns" benchmark row. gold_standard.build_activation_calendar
# (Gate 7) may match a plan against a narrower sub-vertical whose real
# seasonality differs from its parent industry (e.g. propane/heating-fuel
# delivery vs. generic freight/e-commerce logistics -- see
# data/subvertical_seasonal_overrides.json). When that override fired, the
# raw KB seasonal_patterns benchmark (recruitment_benchmarks_deep.json,
# keyed only by the plan's fixed industry classification) is WRONG for this
# specific client and must not be shown; this renders the override's own
# peak/trough months + rationale + source instead. Returns "" when no
# override applies -- callers fall back to the generic KB text unchanged.
def _subvertical_seasonal_override_text(data: dict) -> str:
    gold = data.get("_gold_standard") or {}
    activation = gold.get("activation_calendar") or {}
    subvertical = activation.get("subvertical")
    if not subvertical:
        return ""
    timeline = activation.get("timeline") or []
    peak_months: list[str] = []
    trough_months: list[str] = []
    for m in timeline:
        if not isinstance(m, dict):
            continue
        _name = str(m.get("month_name") or "")
        if not _name:
            continue
        _intensity = str(m.get("hiring_intensity") or "").lower()
        if _intensity in ("high", "very_high") and _name not in peak_months:
            peak_months.append(_name)
        elif _intensity == "low" and _name not in trough_months:
            trough_months.append(_name)

    label = activation.get("subvertical_label") or _humanize_snake_key(subvertical)
    rationale = activation.get("subvertical_rationale") or ""
    source = activation.get("subvertical_source") or ""

    parts = [f"Profile: {label} (overrides generic industry seasonality)"]
    if peak_months:
        parts.append(f"Peak Months: {', '.join(peak_months)}")
    if trough_months:
        parts.append(f"Trough Months: {', '.join(trough_months)}")
    if rationale:
        parts.append(f"Rationale: {rationale}")
    if source:
        parts.append(f"Source: {source}")
    return "; ".join(parts)


# S5 (2026-07-03, finding 54): a small number of city names shipped from
# upstream enrichment lookups all-lowercase (and "Phoenix" as "pheonix") --
# never client-facing. Title-case is applied at every display call site via
# ``_title_case_city``; known misspellings are corrected in this map before
# the generic ``.title()`` normalization runs.
_CITY_NAME_CORRECTIONS: Dict[str, str] = {
    "pheonix": "Phoenix",
}


def _title_case_city(city_name: Any) -> str:
    """Normalize a city name for client-facing display.

    Corrects known misspellings (case-insensitively) and title-cases the
    result so e.g. "houston" -> "Houston" and "las vegas" -> "Las Vegas".
    Leaves already-correct multi-word/proper names alone via str.title(),
    which is safe for the plain city names this renders (no internal
    hyphenation/apostrophe edge cases in the current city lists).
    """
    if not isinstance(city_name, str) or not city_name.strip():
        return str(city_name or "")
    corrected = _CITY_NAME_CORRECTIONS.get(city_name.strip().lower())
    if corrected:
        return corrected
    return city_name.strip().title()


def _get_roles(data: dict) -> List[str]:
    """Extract normalized role strings from data dict."""
    roles_raw = data.get("target_roles") or data.get("roles") or []
    if isinstance(roles_raw, str):
        return [r.strip() for r in roles_raw.split(",") if r.strip()]
    roles = []
    for r in roles_raw:
        if isinstance(r, dict):
            roles.append(r.get("title") or r.get("role") or str(r))
        elif isinstance(r, str):
            roles.append(r.strip())
    return roles or ["General"]


def _format_roles_stat(roles: List[str]) -> Tuple[str, str]:
    """Return a (label, value) pair summarizing a plan's role list the same
    way the Executive Summary's "Roles" count card does (finding
    data:atria#4) -- e.g. ("Roles", "10 (Memory Care Associate, Nurse, +8
    more)") -- instead of silently truncating a multi-role plan down to a
    single named role.
    """
    n = len(roles)
    if n <= 1:
        return "Role", str(roles[0]) if roles else "Various"
    shown = ", ".join(str(r) for r in roles[:2])
    remaining = n - 2
    value = f"{n} ({shown}, +{remaining} more)" if remaining > 0 else f"{n} ({shown})"
    return "Roles", value


def _largest_remainder_fractions(
    fractions: List[float], decimals: int = 1
) -> List[float]:
    """Round proportions (summing to ~1.0) to ``decimals`` places using the
    largest-remainder method (Hamilton apportionment), so the DISPLAYED
    percentages sum to exactly 100.0% (finding data:atria#3-related
    polish). Independent per-value rounding (e.g. round(28.05,1)=28.1,
    round(26.55,1)=26.5, ...) can drift the printed total to 100.1%/99.9%
    even when the underlying values sum to exactly 1.0; this distributes
    the rounding remainder to the values with the largest fractional part
    so they still sum exactly. Returns fractions on the same 0..1 scale.
    """
    n = len(fractions)
    if n == 0:
        return []
    scale = 10**decimals
    total_units = round(100 * scale)
    raw_units = [max(0.0, f) * 100 * scale for f in fractions]
    floor_units = [int(u) for u in raw_units]
    remainder = total_units - sum(floor_units)
    if remainder > 0:
        order = sorted(
            range(n), key=lambda i: raw_units[i] - floor_units[i], reverse=True
        )
        for i in range(min(remainder, n)):
            floor_units[order[i]] += 1
    elif remainder < 0:
        # Only possible if inputs summed to > 1.0; trim from the smallest
        # fractional-remainder entries so the total still lands exactly.
        order = sorted(range(n), key=lambda i: raw_units[i] - floor_units[i])
        for i in range(min(-remainder, n)):
            floor_units[order[i]] = max(0, floor_units[order[i]] - 1)
    return [u / scale / 100.0 for u in floor_units]


def _corrected_channel_pct_display(channel_allocs: Dict[str, Any]) -> Dict[str, float]:
    """Single-source, largest-remainder-rounded display percentages for a
    channel_allocations dict, keyed by channel name (finding
    data:atria#3-related polish). Percentages are derived from each
    channel's dollar_amount / total dollar_amount -- the SAME dollar
    figures that already foot exactly -- then rounded with
    ``_largest_remainder_fractions`` so the displayed percentages also
    foot to exactly 100.0%. The Executive Summary, Channels & Strategy, and
    Channel Recommendations sheets all call this with the SAME
    channel_allocs so every sheet shows the identical percentage for the
    identical channel. Returns ``{}`` when there's nothing to allocate.
    """
    if not isinstance(channel_allocs, dict) or not channel_allocs:
        return {}
    names = list(channel_allocs.keys())
    dollars = [
        _safe_num(
            (channel_allocs[n] or {}).get(
                "dollar_amount", (channel_allocs[n] or {}).get("dollars") or 0
            )
        )
        for n in names
    ]
    total = sum(dollars)
    if total <= 0:
        return {}
    fractions = [d / total for d in dollars]
    rounded = _largest_remainder_fractions(fractions, decimals=1)
    return dict(zip(names, rounded))


def _get_locations(data: dict) -> List[str]:
    """Extract location strings from data dict."""
    locs = data.get("locations") or []
    if isinstance(locs, str):
        return [locs]
    if isinstance(locs, list):
        return [str(loc) for loc in locs if loc] or ["United States"]
    return ["United States"]


_US_TOKENS = frozenset({"us", "usa", "u.s.", "u.s.a.", "united states", "america"})


def _is_us_plan(data: dict) -> bool:
    """True when the plan's target country is the United States.

    Thin delegate to :func:`plan_geo.is_us_plan` -- the single source of
    truth for locale resolution (handles bare US state names, "City, ST"
    pairs, and target_region, and never hard-returns False on the first
    unresolvable location candidate). Kept as a local name because it is
    called throughout this module; do not re-implement the resolution logic
    here.
    """
    return plan_geo.is_us_plan(data)


def _non_us_signals(data: dict) -> List[str]:
    """Location strings that drove ``_is_us_plan`` to False, for honest
    non-US messaging. Thin delegate to :func:`plan_geo.non_us_signals`."""
    return plan_geo.non_us_signals(data)


def _get_budget_numeric(data: dict) -> float:
    """Parse budget from data dict to numeric value."""
    budget_raw = data.get("budget") or data.get("budget_range") or ""
    return parse_budget(budget_raw, default=100_000.0) if budget_raw else 100_000.0


def _parse_hire_goal(hire_volume: Any) -> int:
    """Parse the client's stated hiring GOAL to a comparable integer (low end).

    Thin delegate to :func:`display_format.parse_hire_goal` (single source
    of truth, shared with the deck). O2 (2026-07-03, findings 42/49/64): the
    hire-goal-vs-projection gap must be stated head-on; this parser feeds the
    Executive Summary gap callout.
    """
    return display_format.parse_hire_goal(hire_volume)


def _get_industry_label(industry_key: str) -> str:
    """Convert industry key to display label."""
    return INDUSTRY_LABEL_MAP.get(industry_key, industry_key.replace("_", " ").title())


def _canonical_duration_from_weeks(weeks: int) -> str:
    """Format a canonical campaign-duration label from a week count.

    Thin delegate to :func:`display_format.weeks_to_duration_label` -- the
    single source of truth (round-trip safe with
    :func:`display_format.parse_duration_to_weeks`: e.g. 78 weeks always
    reads back as 78 weeks, never re-derived to a slightly different value
    like the old 18-month input silently becoming "17 months" after a
    weeks-per-month rounding round-trip). Every duration label this module
    emits -- workbook or (via app.py's own resolver) deck -- must trace back
    to this same function so the same week count never renders two
    different-sounding strings.
    """
    weeks = int(weeks or 0)
    if weeks <= 0:
        return "Not specified"
    return display_format.weeks_to_duration_label(weeks)


def _resolve_campaign_duration(data: dict) -> str:
    """Single source of truth for the campaign-duration string shown anywhere.

    Preference order (O2, findings 58/77):
      1. ``campaign_duration_canonical`` (set once in app.py from campaign_weeks)
      2. derived from ``campaign_weeks`` when present
      3. derived by parsing ``campaign_weeks`` out of the raw duration string
      4. the raw ``campaign_duration`` string as a last resort

    This guarantees every sheet references the SAME duration value instead of
    independently re-wording the raw user input, which previously produced
    contradictions like "1-2 years" on the Executive Summary vs a 12-week
    timeline / 90-day full-budget forecast elsewhere in the same bundle.
    """
    canonical = data.get("campaign_duration_canonical")
    if isinstance(canonical, str) and canonical.strip():
        return canonical.strip()

    weeks = data.get("campaign_weeks")
    try:
        weeks_int = int(weeks) if weeks else 0
    except (ValueError, TypeError):
        weeks_int = 0
    if weeks_int > 0:
        return _canonical_duration_from_weeks(weeks_int)

    # Derive weeks from the raw duration string via the SAME 52/12
    # weeks-per-month ratio display_format.weeks_to_duration_label uses --
    # a locally-hardcoded ladder (e.g. "12 month" -> 48 weeks, a *4
    # shortcut) would silently disagree with the canonical formatter above
    # and reintroduce exactly the kind of duration-string drift this
    # resolver exists to prevent.
    raw = str(data.get("campaign_duration") or data.get("timeline") or "").strip()
    dur_lower = raw.lower()
    if not dur_lower or dur_lower in ("not specified", "tbd", "n/a"):
        return "Not specified"
    derived = display_format.parse_duration_to_weeks(raw)
    if derived > 0:
        return _canonical_duration_from_weeks(derived)
    return raw or "Not specified"


def _grade_from_score(score: float) -> str:
    """Convert a 0-1 confidence score to a letter grade."""
    if score >= 0.9:
        return "A"
    if score >= 0.8:
        return "B"
    if score >= 0.65:
        return "C"
    if score >= 0.5:
        return "D"
    return "F"


def _grade_fill(grade: str) -> PatternFill:
    """Return fill color for a confidence grade."""
    if grade in ("A", "B"):
        return _FILL_GREEN_BG
    if grade == "C":
        return _FILL_AMBER_BG
    return _FILL_RED_BG


def _grade_font(grade: str) -> Font:
    """Return font color for a confidence grade."""
    if grade in ("A", "B"):
        return Font(name=FONT_BODY_NAME, bold=True, size=10, color=GREEN)
    if grade == "C":
        return Font(name=FONT_BODY_NAME, bold=True, size=10, color=AMBER)
    return Font(name=FONT_BODY_NAME, bold=True, size=10, color=RED)


def _fit_fill(fit: str) -> PatternFill:
    """Return fill for channel fit rating."""
    fit_lower = fit.lower() if isinstance(fit, str) else ""
    if "excellent" in fit_lower:
        return _FILL_GREEN_BG
    if "good" in fit_lower:
        return _FILL_BLUE_PALE
    return _FILL_AMBER_BG


def _overall_confidence_score(data: dict) -> float:
    """Single source for the plan's overall data-confidence score (0-1).

    Mirrors the exact calc ``_build_sheet_sources`` (Sheet 4, "Sources &
    Confidence") uses for its headline grade, so any other sheet that needs
    an overall-confidence gate reads the SAME number regardless of sheet
    build order (findings data:manpower#3 / data:atria#3 / strategy:atria#8).
    """
    synthesized = data.get("_synthesized", {}) or {}
    confidence_scores = synthesized.get("confidence_scores", {}) or {}
    return _safe_num(
        confidence_scores.get(
            "overall", confidence_scores.get("overall_confidence", 0.5)
        )
    )


# cpc_source values (set by budget_engine's CPC resolution cascade) that
# indicate the channel's CPC/CPA is grounded in a real benchmark/live/trend
# source rather than the static fallback table.
_REAL_BENCHMARK_CPC_SOURCES = {
    "synthesized",
    "live_benchmark",
    "trend_engine",
    "knowledge_base",
}


def _derive_channel_confidence(data: dict, ch_data: dict) -> str:
    """Single source of truth for per-channel confidence tier.

    Findings data:manpower#3 / data:atria#3 / strategy:atria#8: budget_engine's
    own per-channel ``confidence`` field is unreliable in production -- it is
    designed to downgrade with upstream data quality, but the call site does
    not always pass through the input-quality signal, so every channel can
    come back "high" even when the plan's own Sources & Confidence grade is a
    D at 50%. Re-derive it here from two signals excel_v2 can verify itself:

      1. Whether the channel's CPC/CPA came from a real benchmark source
         (``cpc_source`` in {synthesized, live_benchmark, trend_engine,
         knowledge_base}) rather than the static/estimated fallback table --
         fallback/estimated-sourced metrics are always LOW.
      2. The plan's OVERALL confidence score (the same number the Sources &
         Confidence sheet grades) -- HIGH is only awarded when that overall
         grade is B or better (>=80%); otherwise a real-benchmark metric caps
         out at MEDIUM.

    Used by ROI Projections, Channels & Strategy, Confidence Intervals, and
    Channel Recommendations so all four sheets agree on every channel's tier.
    """
    if not isinstance(ch_data, dict):
        return "LOW"
    cpc_source = str(ch_data.get("cpc_source") or "").strip().lower()
    if cpc_source not in _REAL_BENCHMARK_CPC_SOURCES:
        return "LOW"
    if _overall_confidence_score(data) >= 0.80:
        return "HIGH"
    return "MEDIUM"


def _confidence_variance(confidence: str) -> float:
    """The workbook's one documented confidence-tier variance ladder:
    HIGH = +/-15%, MEDIUM = +/-20%, LOW = +/-25% (this is the exact
    methodology stated on the Confidence Intervals sheet). Single source
    of truth so no other sheet invents its own variance schedule.
    """
    if confidence == "HIGH":
        return 0.15
    if confidence == "MEDIUM":
        return 0.20
    return 0.25


def _confidence_range(
    value: float, confidence: str, cost_metric: bool = False
) -> Optional[Tuple[float, float]]:
    """Single source of truth: (value, confidence_tier) -> (low, high).

    Applies the documented +/-15/20/25% variance ladder (see
    ``_confidence_variance``) and returns a clamped, non-inverted band.
    Used identically by ROI Projections' "Hire Range" column and the
    Confidence Intervals sheet's per-channel Hires row so the same
    channel's projected-hires range can never be stated two different ways
    in the same workbook (S89 FIX, findings data:manpower#1/#2,
    data:atria#1 -- ROI Projections previously computed its own
    HIGH/MEDIUM/LOW = 10/25/40% ladder that disagreed with Confidence
    Intervals' 15/20/25% ladder on every channel).

    ``cost_metric=True`` means "pessimistic" is the HIGHER value (CPA/CPH);
    otherwise (count metrics -- applications/hires) pessimistic is the
    LOWER value. Count metrics are truncated to whole units before
    clamping (matching the Confidence Intervals sheet's own arithmetic) so
    integer displays never drift from this helper by a rounding step.
    Returns None when ``value <= 0`` or the band collapses to a single
    point -- callers should skip the row/column rather than print a fake
    range.
    """
    if value <= 0:
        return None
    variance = _confidence_variance(confidence)
    if cost_metric:
        lo = value * (1 + variance)
        hi = value * (1 - variance)
    else:
        lo = max(0, int(value * (1 - variance)))
        hi = int(value * (1 + variance))
    return _clamped_band(lo, value, hi, cost_metric=cost_metric)


def _parse_cph_point_estimate(raw: Any) -> float:
    """Parse a KB cost-per-hire benchmark value into one numeric estimate.

    Handles a bare number, a range string ("$9,000-$12,000" -> midpoint), or
    an open-ended string ("$5,000+" -> 5000). Returns 0.0 when nothing
    numeric can be parsed.
    """
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return float(raw)
    text = str(raw or "")
    nums = re.findall(r"[\d,]+(?:\.\d+)?", text)
    nums = [float(n.replace(",", "")) for n in nums if n]
    if not nums:
        return 0.0
    if len(nums) >= 2:
        return round((nums[0] + nums[1]) / 2.0, 2)
    return nums[0]


def _kb_industry_cph_benchmark(
    industry: str, load_kb_fn=None, kb: Optional[dict] = None
) -> float:
    """Single getter for an "industry average cost-per-hire" figure cited
    anywhere in the workbook (findings strategy:manpower#12 /
    strategy:atria#2 / consistency:atria#2). Reads the SAME KB section the
    "Recruitment Benchmarks" table on the Executive Summary reads, and the
    deck quotes (``kb['recruitment_benchmarks']['industry_benchmarks']``) --
    never a budget-engine constant like the old $5,525/$7,650/$500-per-opening
    figures, which came from a different table and disagreed with this one.

    Returns 0.0 (meaning "no KB benchmark available") when neither ``kb`` nor
    a working ``load_kb_fn`` is supplied, or the industry isn't in the KB.
    """
    if kb is None:
        if not load_kb_fn:
            return 0.0
        try:
            kb = load_kb_fn()
        except Exception:
            return 0.0
    if not isinstance(kb, dict):
        return 0.0
    kb_benchmarks = (kb.get("recruitment_benchmarks", {}) or {}).get(
        "industry_benchmarks", {}
    )
    if not kb_benchmarks:
        kb_benchmarks = kb.get("benchmarks", {}) or {}
    if not isinstance(kb_benchmarks, dict):
        return 0.0
    ind_bench = kb_benchmarks.get(industry) or kb_benchmarks.get(
        "general_entry_level", {}
    )
    if not isinstance(ind_bench, dict):
        return 0.0
    cph_section = ind_bench.get("cph")
    if isinstance(cph_section, dict):
        for key in ("total_cost_per_hire", "recruitment_marketing_only"):
            parsed = _parse_cph_point_estimate(cph_section.get(key))
            if parsed > 0:
                return parsed
    # Flat fallback for KB variants that store CPH at the top level.
    for key in ("total_cost_per_hire", "cost_per_hire", "cph"):
        raw = ind_bench.get(key)
        if isinstance(raw, (int, float, str)):
            parsed = _parse_cph_point_estimate(raw)
            if parsed > 0:
                return parsed
    return 0.0


def _kb_industry_benchmark_section(
    industry: str, load_kb_fn=None, kb: Optional[dict] = None
) -> Dict[str, Any]:
    """Load the SAME per-industry benchmark record the Executive Summary's
    "Recruitment Benchmarks" table reads
    (``kb['recruitment_benchmarks']['industry_benchmarks'][industry]``), so
    any sheet that needs to compare the plan's own numbers against the
    cited CPA/apply-rate benchmark reads identical source data. Returns
    ``{}`` when no KB is available or the industry isn't present.
    """
    if kb is None:
        if not load_kb_fn:
            return {}
        try:
            kb = load_kb_fn()
        except Exception:
            return {}
    if not isinstance(kb, dict):
        return {}
    kb_benchmarks = (kb.get("recruitment_benchmarks", {}) or {}).get(
        "industry_benchmarks", {}
    )
    if not kb_benchmarks:
        kb_benchmarks = kb.get("benchmarks", {}) or {}
    if not isinstance(kb_benchmarks, dict):
        return {}
    ind_bench = kb_benchmarks.get(industry) or kb_benchmarks.get(
        "general_entry_level", {}
    )
    return ind_bench if isinstance(ind_bench, dict) else {}


def _parse_numeric_range(text: Any) -> Optional[Tuple[float, float]]:
    """Parse the first two numbers out of a benchmark string like
    "$12-$35", "5.5-7.5%", or "25-40 days" into a ``(low, high)`` tuple.
    Returns ``None`` when fewer than two numbers are present.
    """
    if not text:
        return None
    # Require at least one digit per match -- a bare "," in prose (e.g.
    # "24 occupations, declining") otherwise matches ``[\d,]+`` on its own
    # and crashes float("") once commas are stripped.
    nums = re.findall(r"\d[\d,]*(?:\.\d+)?", str(text))
    nums = [float(n.replace(",", "")) for n in nums if n]
    if len(nums) < 2:
        return None
    lo, hi = nums[0], nums[1]
    return (lo, hi) if lo <= hi else (hi, lo)


def _model_vs_benchmark_note(data: dict, ind_bench: Dict[str, Any]) -> Optional[str]:
    """ONE 'Model vs. benchmark' reconciliation note, shared verbatim by the
    Executive Summary benchmark block and ROI Projections (findings
    data:manpower#3/#4, visual:manpower#3, strategy:manpower#3).

    Computes the plan's own blended CPA and apply rate LIVE from
    ``_budget_allocation`` and compares them against the cited industry
    benchmark's CPA floor and most-recent apply-rate ceiling. Returns a
    one-line honest disclosure ONLY when the plan sits outside the cited
    range in the direction that would otherwise look unexplained (modeled
    CPA below the floor, or modeled apply rate above the ceiling). Returns
    ``None`` when the plan is within range, or when there isn't enough data
    to compute either side of the comparison -- never fabricates a note.
    """
    if not isinstance(ind_bench, dict) or not ind_bench:
        return None

    budget_alloc = data.get("_budget_allocation", {})
    channel_allocs = (
        budget_alloc.get("channel_allocations", {})
        if isinstance(budget_alloc, dict)
        else {}
    )
    if not isinstance(channel_allocs, dict) or not channel_allocs:
        return None

    total_dollars = 0.0
    total_apps = 0.0
    total_clicks = 0.0
    for ch_data in channel_allocs.values():
        if not isinstance(ch_data, dict):
            continue
        total_dollars += _safe_num(
            ch_data.get("dollar_amount", ch_data.get("dollars") or 0)
        )
        total_apps += _safe_num(ch_data.get("projected_applications") or 0)
        total_clicks += _safe_num(ch_data.get("projected_clicks") or 0)

    if total_dollars <= 0 or total_apps <= 0:
        return None

    blended_cpa = total_dollars / total_apps
    apply_rate_pct = (
        (total_apps / total_clicks) * 100.0 if total_clicks > 0 else None
    )

    # CPA floor from the benchmark's "range" field (e.g. "$12-$35" -> 12).
    cpa_floor = None
    cpa_field = ind_bench.get("cpa")
    if isinstance(cpa_field, dict):
        cpa_range = _parse_numeric_range(cpa_field.get("range"))
    else:
        cpa_range = _parse_numeric_range(cpa_field)
    if cpa_range:
        cpa_floor = cpa_range[0]

    # Apply-rate ceiling: prefer the most recent (highest) cited year, e.g.
    # {"2024": "5-7%", "2025": "5.5-7.5%"} -> use "2025"'s ceiling (7.5).
    apply_ceiling = None
    apply_field = ind_bench.get("apply_rate")
    apply_source: Any = None
    if isinstance(apply_field, dict):
        year_keys = sorted(
            (k for k in apply_field if str(k).strip().isdigit()), reverse=True
        )
        if year_keys:
            apply_source = apply_field.get(year_keys[0])
        else:
            for v in apply_field.values():
                if isinstance(v, str) and re.search(r"\d", v):
                    apply_source = v
                    break
    else:
        apply_source = apply_field
    apply_range = _parse_numeric_range(apply_source)
    if apply_range:
        apply_ceiling = apply_range[1]

    reasons = []
    if cpa_floor is not None and blended_cpa < cpa_floor:
        reasons.append(
            f"this plan's blended CPA ({_fmt_currency(blended_cpa, show_cents=True)}) "
            f"sits below the cited benchmark floor ({_fmt_currency(cpa_floor)})"
        )
    if (
        apply_ceiling is not None
        and apply_rate_pct is not None
        and apply_rate_pct > apply_ceiling
    ):
        reasons.append(
            f"this plan's blended apply rate ({apply_rate_pct:.1f}%) sits above the "
            f"cited benchmark ceiling ({apply_ceiling:.1f}%)"
        )

    if not reasons:
        return None

    return (
        "Model vs. benchmark: "
        + " and ".join(reasons)
        + ". This is a deliberate modeling position, not an error: programmatic "
        "buying and ML bid optimization price applications below posted-rate "
        "benchmarks; benchmark ranges reflect classic post-and-pray channel pricing."
    )


def _fit_score_fill(score: float) -> PatternFill:
    """Return fill for numeric fit score."""
    if score >= 0.7:
        return _FILL_GREEN_BG
    if score >= 0.4:
        return _FILL_AMBER_BG
    return _FILL_RED_BG


def _fit_score_font(score: float) -> Font:
    """Return font for numeric fit score."""
    if score >= 0.7:
        return Font(name=FONT_BODY_NAME, bold=True, size=10, color=GREEN)
    if score >= 0.4:
        return Font(name=FONT_BODY_NAME, bold=True, size=10, color=AMBER)
    return Font(name=FONT_BODY_NAME, bold=True, size=10, color=RED)


def _detect_role_type(roles: List[str]) -> str:
    """Detect dominant role type from role titles."""
    if not roles:
        return "professional"
    combined = " ".join(r.lower() for r in roles)
    if any(
        kw in combined
        for kw in [
            "nurse",
            "physician",
            "clinical",
            "medical",
            "therapist",
            "rn ",
            "lpn",
            "cna",
        ]
    ):
        return "clinical"
    if any(
        kw in combined
        for kw in ["ceo", "cfo", "vp ", "director", "chief", "president", "executive"]
    ):
        return "executive"
    if any(
        kw in combined
        for kw in [
            "warehouse",
            "driver",
            "assembler",
            "operator",
            "laborer",
            "mechanic",
            "technician",
            "welder",
        ]
    ):
        return "hourly"
    if any(
        kw in combined
        for kw in ["plumber", "electrician", "carpenter", "hvac", "mason", "welder"]
    ):
        return "trades"
    return "professional"


# ═══════════════════════════════════════════════════════════════════════════════
# WORKSHEET HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════


def _set_column_widths(ws, widths: Dict[int, float]):
    """Set column widths. Key = 1-based column number."""
    for col_num, width in widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width


# S89: deliberate per-sheet tab palette (deck tokens only) so the workbook
# reads as a designed set instead of a wall of identical purple tabs. Keyed by
# sheet title; falls back to PURPLE. Status colors (green/amber/red) are never
# used for tabs -- those read as state, not brand.
_TAB_COLORS: Dict[str, str] = {
    "Executive Summary": NAVY,
    "Channels & Strategy": SAPPHIRE,
    "Market Intelligence": TEAL_HEX,
    "Sources & Confidence": PURPLE_LIGHT_HEX,
    "ROI Projections": SAPPHIRE,
    "Quality Intelligence": TEAL_HEX,
    "90-Day Forecast": NAVY,
    "Confidence Intervals": MAGENTA_HEX,
    "Niche Board Matching": MAGENTA_HEX,
    "Channel Recommendations": SAPPHIRE,
    # S5 (2026-07-03, finding 45): this key never matched the sheet's actual
    # title ("Intl Benchmarks", set in _build_sheet_international_benchmarks),
    # so that sheet silently got no tab color at all.
    "Intl Benchmarks": TEAL_HEX,
}


def _finalize_workbook(wb) -> None:
    """S89: apply cross-cutting polish to every sheet just before save.

    1. Freeze the primary table header (or the title block) so it stays visible
       while the client scrolls long tables.
    2. Assign a deliberate brand tab color per sheet.
    3. Hide gridlines on every sheet (S5, 2026-07-03 finding 45) so the brand
       card/canvas fills read as a designed surface instead of a gridded
       spreadsheet with near-white cards floating on visible cell borders.
    Fully defensive: any per-sheet failure is logged and skipped so finalization
    never blocks the workbook from saving.
    """
    for ws in wb.worksheets:
        try:
            ws.freeze_panes = getattr(ws, "_nova_freeze_row", None) or "B3"
        except Exception as exc:  # noqa: BLE001 - cosmetic, never block save
            logger.debug("freeze_panes skipped for %s: %s", ws.title, exc)
        try:
            color = _TAB_COLORS.get(ws.title)
            if color:
                ws.sheet_properties.tabColor = color
        except Exception as exc:  # noqa: BLE001
            logger.debug("tabColor skipped for %s: %s", ws.title, exc)
        try:
            ws.sheet_view.showGridLines = False
        except Exception as exc:  # noqa: BLE001
            logger.debug("showGridLines skipped for %s: %s", ws.title, exc)


def _write_section_header(ws, row: int, title: str) -> int:
    """Write a full-width section header (navy background, white text).
    Returns the row AFTER the header (header row + 1).
    """
    ws.merge_cells(
        start_row=row,
        start_column=COL_START,
        end_row=row,
        end_column=COL_END,
    )
    cell = ws.cell(row=row, column=COL_START, value=title.upper())
    cell.font = _FONT_SECTION
    cell.fill = _FILL_NAVY
    cell.alignment = _ALIGN_LEFT
    # Set row height for header prominence
    ws.row_dimensions[row].height = 28
    return row + 1


def _write_subsection_header(ws, row: int, title: str) -> int:
    """Write a sub-section header (navy text, sapphire bottom border).
    Returns the row AFTER the header (header row + 1).
    """
    ws.merge_cells(
        start_row=row,
        start_column=COL_START,
        end_row=row,
        end_column=COL_END,
    )
    cell = ws.cell(row=row, column=COL_START, value=title)
    cell.font = _FONT_SUBSECTION
    cell.alignment = _ALIGN_LEFT
    cell.border = _BORDER_BOTTOM_SAPPHIRE
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_table_header(
    ws,
    row: int,
    headers: List[str],
    col_start: int = COL_START,
    fill: PatternFill = None,
) -> int:
    """Write a table header row. Returns the next row."""
    use_fill = fill or _FILL_SAPPHIRE
    use_font = _FONT_TABLE_HEADER if fill is None else _FONT_TABLE_HEADER_ALT
    if fill == _FILL_BLUE_LIGHT:
        use_font = _FONT_TABLE_HEADER_ALT
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=header)
        cell.font = use_font
        cell.fill = use_fill
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN
    ws.row_dimensions[row].height = 22
    # S89: record the first table header near the top of a sheet so the
    # finalizer can freeze just below it (keeps the header visible while the
    # client scrolls). Only the first qualifying header per sheet wins, and
    # only if it sits near the top (else freezing would hide too much).
    if getattr(ws, "_nova_freeze_row", None) is None and row <= 8:
        ws._nova_freeze_row = f"B{row + 1}"
    return row + 1


def _write_num(cell, raw_value: Any, number_format: str) -> None:
    """Write a LIVE numeric value into a cell + apply an Excel number_format.

    Percentages use ``FMT_PCT1`` and expect the cell value to be the *fraction*
    (0.32), so callers pass the already-divided value. Everything else stores
    the raw number so the client can SUM / sort / chart it.
    """
    cell.value = _safe_num(raw_value)
    cell.number_format = number_format


def _write_table_row(
    ws,
    row: int,
    values: List[Any],
    col_start: int = COL_START,
    alternate: bool = False,
    fonts: List[Optional[Font]] = None,
    fills: List[Optional[PatternFill]] = None,
    aligns: List[Optional[Alignment]] = None,
    number_formats: List[Optional[str]] = None,
) -> int:
    """Write a single table data row. Returns the next row.

    ``number_formats`` (S89): a parallel list; where an entry is a format
    string, the corresponding value is written as a LIVE number with that
    Excel ``number_format`` (summable/sortable) instead of as text. Entries
    that are ``None`` keep the legacy string-write behaviour, so existing
    callers are unaffected.
    """
    row_fill = _FILL_BLUE_PALE if alternate else _FILL_WHITE
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i)
        fmt = (
            number_formats[i]
            if number_formats and i < len(number_formats)
            else None
        )
        if fmt:
            _write_num(cell, val, fmt)
        else:
            cell.value = val
        cell.font = fonts[i] if fonts and i < len(fonts) and fonts[i] else _FONT_BODY
        cell.fill = fills[i] if fills and i < len(fills) and fills[i] else row_fill
        cell.alignment = (
            aligns[i] if aligns and i < len(aligns) and aligns[i] else _ALIGN_WRAP
        )
        cell.border = _BORDER_THIN
    return row + 1


def _write_metric_card(ws, row: int, col: int, label: str, value: str):
    """Write a metric card (label above, value below) in a 2-row, 2-col block."""
    # Value cell
    cell_val = ws.cell(row=row, column=col, value=value)
    cell_val.font = _FONT_METRIC_VALUE
    cell_val.fill = _FILL_OFF_WHITE
    cell_val.alignment = _ALIGN_CENTER
    cell_val.border = _BORDER_THIN
    # Merge value across 2 columns if space allows
    if col + 1 <= COL_END:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    # Label cell (row below)
    cell_lbl = ws.cell(row=row + 1, column=col, value=label)
    cell_lbl.font = _FONT_METRIC_LABEL
    cell_lbl.fill = _FILL_OFF_WHITE
    cell_lbl.alignment = _ALIGN_CENTER
    cell_lbl.border = _BORDER_THIN
    if col + 1 <= COL_END:
        ws.merge_cells(
            start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1
        )


def _write_kv_row(
    ws, row: int, key: str, value: str, col_start: int = COL_START
) -> int:
    """Write a key-value pair spanning 2 + 5 columns. Returns next row."""
    # Key cell (B:C)
    ws.merge_cells(
        start_row=row, start_column=col_start, end_row=row, end_column=col_start + 1
    )
    cell_k = ws.cell(row=row, column=col_start, value=key)
    cell_k.font = _FONT_BODY_BOLD
    cell_k.alignment = _ALIGN_LEFT
    cell_k.border = _BORDER_THIN
    cell_k.fill = _FILL_OFF_WHITE
    # Value cell (D:H)
    ws.merge_cells(
        start_row=row, start_column=col_start + 2, end_row=row, end_column=COL_END
    )
    cell_v = ws.cell(row=row, column=col_start + 2, value=value)
    cell_v.font = _FONT_BODY
    cell_v.alignment = _ALIGN_WRAP
    cell_v.border = _BORDER_THIN
    return row + 1


def _write_footnote(ws, row: int, text: str) -> int:
    """Write a footnote row spanning full width. Returns next row."""
    ws.merge_cells(
        start_row=row,
        start_column=COL_START,
        end_row=row,
        end_column=COL_END,
    )
    cell = ws.cell(row=row, column=COL_START, value=text)
    cell.font = _FONT_FOOTNOTE
    cell.alignment = _ALIGN_LEFT
    return row + 1


def _write_attribution_footer(ws, row: int) -> int:
    """Write data attribution footer. Returns next row."""
    row = _write_footnote(
        ws,
        row,
        f"Generated by Nova AI Media Plan Generator | {datetime.date.today().strftime('%B %d, %Y')} | "
        "Created by Shubham Singh Chandel | Joveo Global Supply Team",
    )
    return row


# ═══════════════════════════════════════════════════════════════════════════════
# CHANNEL VETTING FUNCTION
# ═══════════════════════════════════════════════════════════════════════════════


def vet_channels(
    channels: List[Dict[str, Any]],
    industry: str,
    roles: List[str],
    locations: List[str],
) -> List[Dict[str, Any]]:
    """Vet and score channels against industry, role, and location requirements.

    Args:
        channels: List of channel dicts (from _channels_db or budget allocations).
                  Each must have at least a 'name' key.
        industry: Industry key (e.g. "healthcare_medical").
        roles: List of role title strings.
        locations: List of location strings.

    Returns:
        Sorted list of channel dicts with added 'fit' and 'fit_score' keys.
        Sorted by fit_score descending (Excellent > Good > Fair).
    """
    if not channels:
        return []

    ind_reqs = INDUSTRY_CHANNEL_REQUIREMENTS.get(industry, {})
    ind_preferred = [kw.lower() for kw in ind_reqs.get("preferred") or []]
    ind_excluded = [kw.lower() for kw in ind_reqs.get("excluded_keywords") or []]

    role_type = _detect_role_type(roles)
    role_reqs = ROLE_CHANNEL_REQUIREMENTS.get(role_type, {})
    role_preferred = [kw.lower() for kw in role_reqs.get("preferred") or []]
    role_excluded = [kw.lower() for kw in role_reqs.get("excluded_keywords") or []]

    # Deduplicate by normalized name
    seen_names = set()
    deduped = []
    for ch in channels:
        name = ch.get("name") or "" if isinstance(ch, dict) else str(ch)
        # Strip common suffixes BEFORE removing non-alphanumeric chars
        norm_name = name.lower().strip()
        for suffix in [".com", ".net", ".org", ".io", " jobs", " job"]:
            norm_name = norm_name.replace(suffix, "")
        norm_name = re.sub(r"[^a-z0-9]", "", norm_name)
        if norm_name and norm_name not in seen_names:
            seen_names.add(norm_name)
            if isinstance(ch, dict):
                deduped.append(ch)
            else:
                deduped.append({"name": str(ch)})

    # Detect if locations are US-only
    location_lower = " ".join(loc.lower() for loc in locations)
    _us_states = {
        "california",
        "new york",
        "texas",
        "florida",
        "illinois",
        "ohio",
        "georgia",
        "michigan",
        "pennsylvania",
        "virginia",
        "washington",
        "arizona",
        "massachusetts",
        "colorado",
        "minnesota",
        "oregon",
        "nevada",
        "tennessee",
        "indiana",
        "north carolina",
        "south carolina",
        "new jersey",
        "maryland",
        "missouri",
        "wisconsin",
        "connecticut",
        "iowa",
        "utah",
        "kansas",
        "kentucky",
        "louisiana",
        "alabama",
        "oklahoma",
        "nebraska",
        "mississippi",
        "arkansas",
        "montana",
        "new mexico",
        "new hampshire",
        "idaho",
        "hawaii",
        "maine",
        "rhode island",
        "delaware",
        "south dakota",
        "north dakota",
        "alaska",
        "vermont",
        "wyoming",
        "west virginia",
    }

    def _is_us_location(loc_str: str) -> bool:
        ll = loc_str.lower().strip()
        if "united states" in ll or ll == "usa" or ll == "us":
            return True
        # Check against state names (exact or in comma-separated parts)
        parts = [p.strip().lower() for p in ll.split(",")]
        return any(p in _us_states for p in parts)

    is_us_only = all(_is_us_location(loc) for loc in locations) if locations else True

    vetted = []
    for ch in deduped:
        name = ch.get("name") or ""
        name_lower = name.lower()

        # Check exclusions -- remove if channel matches industry OR role exclusions
        excluded = False
        for excl_kw in ind_excluded + role_excluded:
            if excl_kw and excl_kw in name_lower:
                excluded = True
                break
        if excluded:
            continue

        # Check geographic fit -- skip international-only boards if US-only
        intl_only_keywords = ["apac", "emea", "latam"]
        if is_us_only and any(kw in name_lower for kw in intl_only_keywords):
            continue

        cat = _roi_category_for_channel(name)

        # S92 FIX (bundle-quality findings data:manpower#4 / data:atria#1 /
        # strategy:manpower#4 / strategy:atria#4): when this channel came
        # from the live budget allocation, budget_engine has already scored
        # it with an authoritative, post-rebalance roi_score-derived
        # fit_score (see budget_engine._finalize_channel_ranking). Using
        # that here -- instead of the independent industry-keyword
        # heuristic below -- is what eliminates the standing
        # self-contradiction where this table ranked a channel #1 "fit"
        # while the Channel Strategy Overview table (built from the SAME
        # allocation's roi_score) ranked the identical channel near-last.
        # Channels with no allocation data (pulled in from the channels DB
        # only) still use the heuristic -- there's no ROI signal for them.
        _be_fit_score = ch.get("fit_score")
        _be_vetted_tier = ch.get("vetted_tier")
        if _be_fit_score is not None:
            score = max(0.0, min(1.0, float(_be_fit_score)))
        else:
            # Score the channel -- start with a category-based baseline
            # so different channel types get differentiated scores even
            # without exact keyword matches.
            _category_baselines: Dict[str, float] = {
                "niche_board": 0.75,
                "referral": 0.80,
                "career_site": 0.70,
                "events": 0.65,
                "staffing": 0.65,
                "job_board": 0.60,
                "social": 0.55,
                "programmatic": 0.50,
                "search": 0.55,
                "display": 0.45,
                "email": 0.55,
                "employer_branding": 0.60,
                "regional": 0.60,
            }
            score = _category_baselines.get(cat, 0.50)

            # Industry preference match
            for pref in ind_preferred:
                if pref in name_lower:
                    score += 0.20
                    break

            # Role preference match
            for pref in role_preferred:
                if pref in name_lower:
                    score += 0.10
                    break

            # Major boards always get a baseline boost (broad fit)
            major_boards = [
                "indeed",
                "linkedin",
                "glassdoor",
                "ziprecruiter",
                "google",
                "meta",
                "facebook",
            ]
            if any(mb in name_lower for mb in major_boards):
                score = max(score, 0.65)

            # Niche board for the industry = excellent
            niche_for_industry = INDUSTRY_NICHE_CHANNELS.get(industry, [])
            if any(
                niche.lower() in name_lower or name_lower in niche.lower()
                for niche in niche_for_industry
            ):
                score = max(score, 0.85)

            # Industry-specific channel type bonus: niche boards score
            # higher for matching industries (e.g., healthcare niche boards
            # for healthcare)
            if cat == "niche_board" and ind_preferred:
                score = max(score, 0.80)

        # Determine fit label. Brand channels get their own labeled tier
        # (S92) instead of Excellent/Good/Fair -- a 0-hire projection is BY
        # DESIGN for reach/awareness spend, not a poor "fit" (see
        # budget_engine._BRAND_RATIONALE).
        if _be_vetted_tier == "Brand & Awareness":
            fit = "Brand/Awareness"
        elif score >= 0.8:
            fit = "Excellent"
        elif score >= 0.6:
            fit = "Good"
        else:
            fit = "Fair"

        ch_copy = dict(ch)
        ch_copy["fit"] = fit
        ch_copy["fit_score"] = round(min(score, 1.0), 2)
        vetted.append(ch_copy)

    # Sort by fit_score descending
    vetted.sort(key=lambda x: x.get("fit_score") or 0, reverse=True)
    return vetted


# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE CONFIDENCE / BIAS ASSESSMENT
# ═══════════════════════════════════════════════════════════════════════════════


def assess_source_bias(source_name: str) -> Dict[str, Any]:
    """Categorize a data source and assess potential bias.

    Returns dict with:
        category: str — the source category
        confidence_modifier: float — multiplier for confidence (0.6 - 1.0)
        bias: str — bias assessment text
    """
    if not source_name or not isinstance(source_name, str):
        return {
            "category": "Unknown",
            "confidence_modifier": 0.60,
            "bias": "Unknown - Unable to verify independence",
        }

    name = source_name.lower().strip()

    # Government / Academic (highest trust)
    gov_academic_keywords = [
        "bls",
        "bureau of labor",
        "census",
        "acs",
        "o*net",
        "onet",
        "careeronestop",
        "fred",
        "imf",
        "world bank",
        "worldbank",
        "sec",
        "edgar",
        "usda",
        "nih",
        "cdc",
        "shrm",
        "university",
        "academic",
        "aaas",
        "government",
        "federal",
        "datausa",
        "geonames",
        "rest countries",
        "restcountries",
    ]
    if any(kw in name for kw in gov_academic_keywords):
        return {
            "category": "Government / Academic",
            "confidence_modifier": 1.0,
            "bias": "Low - Independent",
        }

    # Industry Analyst
    analyst_keywords = [
        "gartner",
        "forrester",
        "deloitte",
        "mckinsey",
        "bain",
        "bcg",
        "korn ferry",
        "mercer",
        "aon",
        "pwc",
        "ernst & young",
        "ey ",
        "kpmg",
        "accenture",
    ]
    if any(kw in name for kw in analyst_keywords):
        return {
            "category": "Industry Analyst",
            "confidence_modifier": 0.95,
            "bias": "Low - Independent analyst",
        }

    # Platform / Publisher (may promote their own platform)
    platform_keywords = [
        "indeed",
        "linkedin",
        "glassdoor",
        "ziprecruiter",
        "monster",
        "careerbuilder",
        "google ads",
        "google trends",
        "meta ads",
        "facebook ads",
        "bing ads",
        "tiktok ads",
        "snap ads",
        "clearbit",
        "wikipedia",
        "teleport",
    ]
    if any(kw in name for kw in platform_keywords):
        return {
            "category": "Platform / Publisher",
            "confidence_modifier": 0.75,
            "bias": "Medium - May promote own platform",
        }

    # Vendor / Marketer
    vendor_keywords = [
        "appcast",
        "recruitics",
        "icims",
        "phenom",
        "radancy",
        "pandologic",
        "talroo",
        "programmatic",
        "vendor",
        "marketer",
        "recruitology",
        "nexxt",
        "jovian",
    ]
    if any(kw in name for kw in vendor_keywords):
        return {
            "category": "Vendor / Marketer",
            "confidence_modifier": 0.70,
            "bias": "Medium-High - Promotes own services",
        }

    # Internal / First-Party (Joveo)
    internal_keywords = ["joveo", "mojo", "first-party", "internal", "campaign data"]
    if any(kw in name for kw in internal_keywords):
        return {
            "category": "Internal / First-Party",
            "confidence_modifier": 0.85,
            "bias": "Low - First-party campaign data",
        }

    # Unknown
    return {
        "category": "Unknown",
        "confidence_modifier": 0.60,
        "bias": "Unknown - Unable to verify independence",
    }


def _funded_brand_channel_names(channel_allocs: dict) -> List[str]:
    """Names of funded 'brand' channels (e.g. Employer Branding) -- these
    always carry ``projected_hires == 0`` by design (see budget_engine's
    brand-channel handling), the identical zero-hire profile a named
    performance channel gets flagged for below.
    """
    return sorted(
        _smart_title(str(name))
        for name, ch in (channel_allocs or {}).items()
        if isinstance(ch, dict)
        and ch.get("channel_role") == "brand"
        and (ch.get("dollar_amount", ch.get("dollars") or 0) or 0) > 0
    )


def _brand_asymmetry_clause(channel_allocs: dict) -> str:
    """data:atria#6 fix: when a low-ROI/low-efficiency recommendation names
    a performance channel (e.g. Social Media) for its zero-hire profile, a
    funded brand channel (e.g. Employer Branding) with the SAME zero-hire
    profile sits right next to it in the budget and is silently never
    named -- a reader is left to wonder why one zero-hire channel is
    flagged and the other isn't. Returns a short clause explaining the
    asymmetry (brand channels aren't scored on hires by design), or "" when
    there's no funded brand channel to explain.
    """
    brand_names = _funded_brand_channel_names(channel_allocs)
    if not brand_names:
        return ""
    return (
        f" {', '.join(brand_names)} also shows 0 hires but is held as brand "
        f"investment by design -- it is measured on reach and pipeline "
        f"influence, not direct CPA/hires, so it is not a reallocation "
        f"candidate here."
    )


def _rewrite_low_efficiency_recommendation(channel_allocs: dict) -> Optional[str]:
    """Rebuild the "Low Efficiency alert" recommendation from budget_engine.

    Two problems with the raw budget_engine text (finding strategy:
    manpower#7): it names 'brand' channels (employer branding etc.) whose
    zero-hire projection is BY DESIGN -- they already carry their own
    rationale note elsewhere on this sheet, not a defect to flag again here
    -- and it tells the client to "reallocate" budget the plan itself just
    committed to those channels this same run, which is self-contradictory.
    Zero-hire PERFORMANCE channels get the vetted-tier action instead: hold
    at pilot level, scale only on observed conversion.

    S89A FIX (finding data:atria#6): when a funded brand channel (e.g.
    Employer Branding) shares the identical zero-hire profile as the named
    channel(s) here, say so in the SAME sentence instead of silently
    omitting it -- see ``_brand_asymmetry_clause``.

    Returns ``None`` when no non-brand zero-hire channel remains (drop the
    recommendation entirely rather than show an empty alert).
    """
    perf_zero_hire = sorted(
        _smart_title(str(name))
        for name, ch in (channel_allocs or {}).items()
        if isinstance(ch, dict)
        and ch.get("efficiency_flag") == "Low Efficiency"
        and ch.get("channel_role") != "brand"
    )
    if not perf_zero_hire:
        return None
    return (
        f"Low Efficiency alert: {', '.join(perf_zero_hire)} projected 0 hires "
        f"despite >$1,000 spend. Hold at pilot level; scale only on observed "
        f"conversion rather than reallocating budget already committed to "
        f"this plan." + _brand_asymmetry_clause(channel_allocs)
    )


def _rewrite_low_roi_recommendation(channel_allocs: dict) -> Optional[str]:
    """Rebuild the "Channels with low ROI scores" recommendation from
    budget_engine.

    Same problem as the Low Efficiency alert (see
    ``_rewrite_low_efficiency_recommendation``): budget_engine's raw list
    names 'brand' channels (employer branding etc.) whose low ROI score is
    BY DESIGN -- brand channels aren't scored/optimized for ROI the way
    performance channels are, and they already carry their own rationale
    note elsewhere on this sheet. Naming them here as reallocation
    candidates is a defect, not a finding.

    S89A FIX (finding strategy:manpower#2): a channel that ALSO carries the
    "Low Efficiency" flag gets its own, more specific pilot-hold
    recommendation from ``_rewrite_low_efficiency_recommendation`` --
    "hold at pilot level; scale only on observed conversion rather than
    reallocating". Leaving that same channel in THIS list too produced two
    adjacent Key Recommendations bullets giving opposite instructions on
    the same channel (reallocate its budget away vs. don't reallocate it).
    One channel, one recommendation: when both generators would fire for
    the same channel, the pilot-hold phrasing wins and this list drops it.

    Mirrors budget_engine's own filter (``roi_score <= 3`` and
    ``dollar_amount > 0``) so the named channels agree with what
    budget_engine actually flagged, then drops 'brand' channels and
    Low-Efficiency-flagged channels from that list.

    S89A FIX (finding data:atria#6): a funded 'brand' channel (e.g.
    Employer Branding) routinely shares the IDENTICAL zero-hire,
    ROI-Score-1 profile as a named performance channel here (e.g. Social
    Media) -- omitting it silently reads as an inconsistency when a reader
    cross-checks the ROI Score column. Name the asymmetry instead of
    hiding it -- see ``_brand_asymmetry_clause``.

    Returns ``None`` when no channel remains (drop the recommendation
    entirely rather than show an empty alert).
    """
    low_roi = sorted(
        _smart_title(str(name))
        for name, ch in (channel_allocs or {}).items()
        if isinstance(ch, dict)
        and (ch.get("roi_score", 5) or 0) <= 3
        and (ch.get("dollar_amount", ch.get("dollars") or 0) or 0) > 0
        and ch.get("channel_role") != "brand"
        and ch.get("efficiency_flag") != "Low Efficiency"
    )
    if not low_roi:
        return None
    return (
        f"Channels with low ROI scores ({', '.join(low_roi)}) "
        f"may benefit from budget reallocation to higher-performing channels."
        + _brand_asymmetry_clause(channel_allocs)
    )


def _clean_budget_alloc_narrative(
    warnings: list, recommendations: list, channel_allocs: dict
) -> Tuple[list, list]:
    """De-duplicate/rewrite budget_engine's warnings & recommendations before
    they reach the Executive Summary.

    budget_engine's ``assess_budget_sufficiency`` computes its "Budget of $X
    for N openings... industry average of $Y/hire" warnings and top-up
    recommendation against ``total_openings`` (role headcount) and its own
    ``avg_cph``/``industry_min_cph`` constants -- a DIFFERENT goal and a
    DIFFERENT cost-per-hire from the ones the sheet's own goal-gap callout
    (built from ``display_format.parse_hire_goal``/``goal_gap`` and this
    plan's actual blended CPH) uses. Left in place, the sheet ends up
    stating two irreconcilable goals and CPH benchmarks in the same
    workbook (findings data:manpower#1/#2, strategy:manpower#1/#12,
    consistency:manpower#1, data:atria#2/#4, strategy:atria#2/#3,
    consistency:atria#2). Rather than reconstruct a second, parallel
    narrative, this drops the budget_engine-sourced duplicates so the
    goal-gap callout is the ONE goal/CPH/top-up figure in the workbook, and
    rewrites (or drops) the recommendations that reference a nonexistent
    'optimized' section or recommend reallocating budget the plan itself
    just allocated.
    """
    _cleaned_warnings = [
        w
        for w in warnings
        if not (
            isinstance(w, str)
            and "/opening)" in w
            and "industry average of $" in w
        )
        and not (isinstance(w, str) and "target openings by" in w)
    ]

    _cleaned_recommendations: list = []
    for rec in recommendations:
        if not isinstance(rec, str):
            _cleaned_recommendations.append(rec)
            continue
        _low = rec.lower()
        # Duplicate of the goal-gap callout's own top-up figure, but using
        # budget_engine's role-headcount "openings" total instead of the
        # client's stated goal -- drop it (findings data:manpower#1/#2,
        # data:atria#2/#4).
        if "to fully fund all" in _low and "openings at industry-average" in _low:
            continue
        # Dangling reference to a nonexistent 'optimized' section backing an
        # unverifiable improvement % (findings data:manpower#2, data:atria#2,
        # strategy:manpower#2/atria#3). Prefer dropping over guessing at a
        # "corrected" percentage -- the underlying optimizer comparison has
        # already been shown (in these findings) to produce wildly unstable
        # numbers (383%/467%) that don't survive a back-of-envelope check.
        if "could improve projected hires by" in _low and "section" in _low:
            continue
        if rec.strip().lower().startswith("low efficiency alert:"):
            _rewritten = _rewrite_low_efficiency_recommendation(channel_allocs)
            if _rewritten:
                _cleaned_recommendations.append(_rewritten)
            continue
        if _low.startswith("channels with low roi scores"):
            _rewritten = _rewrite_low_roi_recommendation(channel_allocs)
            if _rewritten:
                _cleaned_recommendations.append(_rewritten)
            continue
        _cleaned_recommendations.append(rec)

    return _cleaned_warnings, _cleaned_recommendations


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: EXECUTIVE SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════


def _build_sheet_executive_summary(
    ws, data: dict, research_mod=None, load_kb_fn=None, classify_tier_fn=None
):
    """Build Sheet 1: Executive Summary."""
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = SAPPHIRE

    # Column widths
    _set_column_widths(
        ws,
        {
            1: 3,  # A: margin
            2: 18,  # B
            3: 18,  # C
            4: 18,  # D
            5: 18,  # E
            6: 18,  # F
            7: 18,  # G
            8: 18,  # H
            9: 14,  # I: CPA column (budget allocation table extends to 9 cols)
            10: 14,  # J: ROI Score column
        },
    )

    client_name = data.get("client_name", "Client")
    industry = data.get("industry", "general_entry_level")
    industry_label = _get_industry_label(industry)
    locations = _get_locations(data)
    roles = _get_roles(data)
    budget_num = _get_budget_numeric(data)
    # O2 (findings 58/77): single source of truth for duration — resolve the
    # canonical label (from campaign_weeks) so every sheet shows the SAME value.
    duration = _resolve_campaign_duration(data)
    hire_volume = data.get("hire_volume") or ""
    _stated_work_env = data.get("work_environment", "hybrid")
    work_env, _work_env_note = gs_lib.effective_work_model(_stated_work_env, roles)

    budget_alloc = data.get("_budget_allocation", {})
    total_proj = budget_alloc.get("total_projected", {})
    sufficiency = budget_alloc.get("sufficiency", {})
    channel_allocs = budget_alloc.get("channel_allocations", {})
    warnings = budget_alloc.get("warnings") or []
    recommendations = budget_alloc.get("recommendations") or []
    warnings, recommendations = _clean_budget_alloc_narrative(
        warnings, recommendations, channel_allocs
    )

    # S49 P2-20: Append research-backed recommendations from shared constants
    try:
        from research_constants import get_plan_recommendations_text

        _research_recs = get_plan_recommendations_text()
        recommendations = list(recommendations) + _research_recs
    except ImportError:
        pass

    synthesized = data.get("_synthesized", {})
    enriched = data.get("_enriched", {})
    tier_groups = data.get("_tier_groups", {})

    row = 2

    # ── 1. Campaign Overview ──
    # Hero banner
    ws.merge_cells(
        start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
    )
    hero_cell = ws.cell(
        row=row, column=COL_START, value=f"Recruitment Media Plan: {client_name}"
    )
    hero_cell.font = _FONT_HERO
    hero_cell.alignment = _ALIGN_LEFT
    ws.row_dimensions[row].height = 36
    row += 1

    # Subtitle
    ws.merge_cells(
        start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
    )
    sub_cell = ws.cell(
        row=row,
        column=COL_START,
        value=f"{industry_label} | {work_env.title()} | "
        f"Generated {datetime.date.today().strftime('%B %d, %Y')}",
    )
    sub_cell.font = _FONT_FOOTNOTE
    sub_cell.alignment = _ALIGN_LEFT
    row += 2

    row = _write_section_header(ws, row, "Campaign Overview")

    # 2x3 metric cards grid
    metrics = [
        ("Budget", _fmt_currency(budget_num)),
        ("Duration", str(duration)),
        ("Locations", str(len(locations))),
        ("Roles", str(len(roles))),
        ("Industry", industry_label),
        ("Hire Volume", str(hire_volume) if hire_volume else "TBD"),
    ]
    card_row = row
    for idx, (label, value) in enumerate(metrics):
        col_offset = (idx % 3) * 2  # 0, 2, 4
        col = COL_START + col_offset
        _write_metric_card(ws, card_row + (idx // 3) * 3, col, label, value)
    row = card_row + 6  # 2 rows of cards * 3 height each
    row += 1  # gap

    # Work-model correction note (gold_standard.effective_work_model): a
    # stated "Remote" model gets flagged when the plan's roles are
    # inherently site-based (care, culinary, logistics, etc.).
    if _work_env_note:
        row = _write_footnote(ws, row, _work_env_note)
        row += 1

    # ── 2. Company Intelligence ──
    company_intel = {}
    if research_mod:
        try:
            company_intel = research_mod.get_company_intelligence(client_name)
        except Exception as exc:
            logger.warning("Company intelligence lookup failed: %s", exc)

    # Also pull from synthesized competitive intelligence
    comp_intel = synthesized.get("competitive_intelligence", {})
    company_profile = comp_intel.get("company_profile", {})

    if company_intel.get("matched") or company_profile:
        row = _write_section_header(ws, row, "Company Intelligence")

        # Merge company_intel and company_profile, preferring company_intel
        display_fields = [
            (
                "Industry",
                company_intel.get(
                    "industry", company_profile.get("industry", industry_label)
                ),
            ),
            ("Size", company_intel.get("size", company_profile.get("size") or "")),
            (
                "Employer Brand",
                company_intel.get(
                    "employer_brand", company_profile.get("employer_brand") or ""
                ),
            ),
            (
                "Hiring Channels",
                company_intel.get(
                    "hiring_channels", company_profile.get("hiring_channels") or ""
                ),
            ),
            (
                "Known Strategies",
                company_intel.get(
                    "known_strategies", company_profile.get("known_strategies") or ""
                ),
            ),
            (
                "Glassdoor Rating",
                company_intel.get(
                    "glassdoor_rating", company_profile.get("glassdoor_rating") or ""
                ),
            ),
            (
                "Talent Focus",
                company_intel.get(
                    "talent_focus", company_profile.get("talent_focus") or ""
                ),
            ),
        ]
        for key, val in display_fields:
            val_str = _flatten_value(val) if val else ""
            if val_str:
                row = _write_kv_row(ws, row, key, val_str)
        row += 1

    # ── 3. Budget Allocation ──
    row = _write_section_header(ws, row, "Budget Allocation")

    # S48 FIX: Compute header hires as SUM of per-channel hires to guarantee
    # the header matches the channel rows below.  Derive cost_per_hire from
    # that same total so all three numbers are internally consistent.
    _header_hires = sum(
        int(ch.get("projected_hires") or 0) for ch in channel_allocs.values()
    )
    # Fall back to budget engine total_projected only if channel_allocs is empty
    if _header_hires == 0:
        _header_hires = int(total_proj.get("hires") or 0)
    _header_cph = (
        round(budget_num / max(_header_hires, 1), 2) if _header_hires > 0 else 0
    )

    # Hero metrics row: Total Budget | Projected Hires | Cost/Hire
    hero_metrics = [
        ("Total Budget", _fmt_currency(budget_num)),
        ("Projected Hires", _fmt_number(_header_hires)),
        ("Cost / Hire", _fmt_currency(_header_cph)),
    ]
    for idx, (label, value) in enumerate(hero_metrics):
        col = COL_START + idx * 2
        _write_metric_card(ws, row, col, label, value)
    row += 3  # 2-row cards + gap

    # ── Hiring-goal vs projection reconciliation (O2, findings 42/49/64) ──
    # The client's stated hiring goal must be addressed head-on, never silently
    # ignored. When the plan projects materially fewer hires than the stated
    # goal, state the gap plainly and quantify the budget that would close it.
    #
    # S89 (findings data:manpower#1 / strategy:manpower#1 / consistency:
    # manpower#1 / data:atria#4 / strategy:atria#2 / consistency:atria#2):
    # this used to be duplicated further down the sheet by budget_engine's
    # own "sufficiency" warnings -- computed against a DIFFERENT goal (total
    # role headcount, not the client's stated goal) and a DIFFERENT CPH
    # constant, producing a second, contradictory "Budget of $X for N
    # openings... industry average of $Y/hire" narrative on the same sheet.
    # That duplicate is now dropped below (see the warnings/recommendations
    # filter) so this callout -- routed through the SAME
    # display_format.parse_hire_goal / goal_gap shared functions the deck
    # uses -- is the ONE goal, ONE CPH, ONE top-up figure in the workbook.
    _goal = _parse_hire_goal(hire_volume)
    if _goal > 0 and _header_hires >= 0:
        # Basis for "budget to close the gap": this plan's own realized cost per
        # hire (most defensible — it's what THIS mix actually achieves), falling
        # back to the KB's own industry CPH benchmark -- the SAME
        # "Recruitment Benchmarks" KB section the table below (and the deck)
        # read, never a budget_engine constant -- when the plan projects zero
        # hires.
        _cph_basis = _header_cph if _header_cph and _header_cph > 0 else 0
        if _cph_basis <= 0:
            _cph_basis = _kb_industry_cph_benchmark(industry, load_kb_fn=load_kb_fn)
        _gap_result = display_format.goal_gap(_header_hires, _goal, _cph_basis)
        # Only call out a gap when it's material (>10% short of goal).
        if _gap_result and (100 - _gap_result["pct_of_goal"]) > 10:
            _gap = _gap_result["goal"] - _gap_result["projected"]
            _extra_budget = _gap_result["additional_budget"]
            _pct_of_goal = round(_gap_result["pct_of_goal"])
            _gap_msg = (
                f"Hiring-goal gap: this plan projects {_header_hires:,} hires "
                f"against a stated goal of {_goal:,} "
                f"({_pct_of_goal}% of goal). "
            )
            if _extra_budget > 0:
                _gap_msg += (
                    f"Closing the ~{_gap:,}-hire gap at this plan's "
                    f"{_fmt_currency(_cph_basis)}/hire would need roughly "
                    f"{_fmt_currency(_extra_budget)} of additional budget "
                    f"(total ~{_fmt_currency(budget_num + _extra_budget)}). "
                    f"Alternatively, phase the goal across multiple cycles or "
                    f"prioritise the highest-ROI roles within this budget."
                )
            else:
                _gap_msg += (
                    "Consider increasing budget, phasing the goal across "
                    "multiple hiring cycles, or narrowing to the highest-ROI "
                    "roles within this budget."
                )
            ws.merge_cells(
                start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
            )
            _gcell = ws.cell(row=row, column=COL_START, value=_gap_msg)
            _gcell.font = _FONT_BODY_BOLD
            _gcell.alignment = _ALIGN_LEFT
            _gcell.fill = _FILL_AMBER_BG
            ws.row_dimensions[row].height = 46
            row += 2

    # Sufficiency grade
    grade_str = sufficiency.get("grade") or ""
    grade_msg = sufficiency.get(
        "message", sufficiency.get("budget_reality_check", {}).get("message") or ""
    )
    if grade_str:
        ws.merge_cells(
            start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
        )
        cell = ws.cell(
            row=row,
            column=COL_START,
            value=f"Budget Sufficiency: {grade_str} — {grade_msg}",
        )
        cell.font = _FONT_BODY_BOLD
        cell.alignment = _ALIGN_LEFT
        if "A" in str(grade_str) or "B" in str(grade_str):
            cell.fill = _FILL_GREEN_BG
        elif "C" in str(grade_str):
            cell.fill = _FILL_AMBER_BG
        else:
            cell.fill = _FILL_RED_BG
        row += 1

    row += 1

    # Channel allocation table
    if channel_allocs:
        headers = [
            "Channel",
            "%",
            "Amount",
            "Proj. Clicks",
            "Proj. Apps",
            "Proj. Hires",
            "CPC",
            "CPA",
            "ROI Score",
        ]
        # S89: use the canonical header helper (single source of truth for
        # font/fill/height + finalizer freeze-row tracking).
        row = _write_table_header(ws, row, headers)

        sorted_channels = sorted(
            channel_allocs.items(),
            key=lambda x: x[1].get("dollar_amount", x[1].get("dollars") or 0),
            reverse=True,
        )
        # data:atria#3-related polish: ONE largest-remainder-rounded
        # percentage per channel, derived from the same dollar amounts this
        # table already displays -- shared verbatim with the Channels &
        # Strategy and Channel Recommendations sheets so the % column foots
        # to exactly 100.0% everywhere it's shown instead of drifting to
        # 100.1% from independent per-cell rounding.
        _corrected_pct = _corrected_channel_pct_display(channel_allocs)
        # S82: collect numeric (name, dollars) pairs for a native pie chart.
        _chart_pairs: List[tuple] = []
        # S89: write LIVE numeric values + Excel number_formats so the client can
        # SUM/sort/filter/chart this table. Column order matches `headers`:
        # Channel(text) | %(pct) | Amount($) | Clicks | Apps | Hires | CPC | CPA | ROI
        # S3: these are the plan's OWN budget/CPC/CPA figures -- localize to the
        # active plan currency instead of a hardcoded USD format.
        _col_formats = [
            None, FMT_PCT1, _usd0_fmt(), FMT_INT, FMT_INT, FMT_INT,
            _usd2_fmt(), _usd2_fmt(), "0.0",
        ]
        _first_data_row = row
        _row_idx = 0
        for ch_name, ch_data in sorted_channels[:15]:
            # Bug 23: Skip garbage rows where all metrics are zero/empty
            _ch_cpc = ch_data.get("cpc") or 0
            _ch_cpa = ch_data.get("cpa") or 0
            _ch_dollars = ch_data.get("dollar_amount", ch_data.get("dollars") or 0) or 0
            _ch_roi = ch_data.get("roi_score") or 0
            _ch_pct = ch_data.get("percentage") or 0
            if not any([_ch_cpc, _ch_cpa, _ch_dollars, _ch_roi, _ch_pct]):
                continue
            idx = _row_idx
            _row_idx += 1
            _display_name = _smart_title(ch_name)
            if _safe_num(_ch_dollars) > 0:
                _chart_pairs.append((_display_name, round(_safe_num(_ch_dollars), 2)))
            # Percentage source is on a 0-100 scale; FMT_PCT1 expects a fraction.
            _pct_frac = _corrected_pct.get(ch_name, _safe_num(_ch_pct) / 100.0)
            values = [
                _display_name,
                _pct_frac,
                _safe_num(_ch_dollars),
                int(_safe_num(ch_data.get("projected_clicks") or 0)),
                int(_safe_num(ch_data.get("projected_applications") or 0)),
                int(_safe_num(ch_data.get("projected_hires") or 0)),
                _safe_num(_ch_cpc),
                _safe_num(_ch_cpa),
                round(_safe_num(_ch_roi), 1),
            ]
            _aligns = [_ALIGN_LEFT] + [_ALIGN_CENTER] * (len(values) - 1)
            row = _write_table_row(
                ws,
                row,
                values,
                alternate=bool(idx % 2),
                aligns=_aligns,
                number_formats=_col_formats,
            )
        _last_data_row = row - 1

        # S89: live totals row (self-verifying, audit-ready footer). Sum the
        # money + count columns with native Excel formulas so the client sees a
        # total that recomputes if they edit the data.
        if _last_data_row >= _first_data_row:
            _tot_cells = [
                ("Total", None),
                (None, None),  # % column — a SUM of shares is ~100%, skip for clarity
                (f"=SUM(D{_first_data_row}:D{_last_data_row})", _usd0_fmt()),
                (f"=SUM(E{_first_data_row}:E{_last_data_row})", FMT_INT),
                (f"=SUM(F{_first_data_row}:F{_last_data_row})", FMT_INT),
                (f"=SUM(G{_first_data_row}:G{_last_data_row})", FMT_INT),
                (None, None),  # CPC — a sum is meaningless
                (None, None),  # CPA — a sum is meaningless
                (None, None),  # ROI — a sum is meaningless
            ]
            for i, (cval, cfmt) in enumerate(_tot_cells):
                cell = ws.cell(row=row, column=COL_START + i)
                if cval is not None:
                    cell.value = cval
                    if cfmt:
                        cell.number_format = cfmt
                cell.font = _FONT_BODY_BOLD
                cell.fill = _FILL_BLUE_LIGHT
                cell.alignment = _ALIGN_LEFT if i == 0 else _ALIGN_CENTER
                cell.border = _BORDER_THIN
            row += 1

            # S89: conditional formatting now that cells are numeric -- evaluates
            # live as the client edits. 3-color scale on ROI Score (red->amber->
            # green) and a purple data bar on Amount.
            try:
                _roi_col = get_column_letter(COL_START + 8)
                _amt_col = get_column_letter(COL_START + 2)
                _rng_roi = f"{_roi_col}{_first_data_row}:{_roi_col}{_last_data_row}"
                _rng_amt = f"{_amt_col}{_first_data_row}:{_amt_col}{_last_data_row}"
                ws.conditional_formatting.add(
                    _rng_roi,
                    ColorScaleRule(
                        start_type="min", start_color="DC2626",
                        mid_type="percentile", mid_value=50, mid_color="D97706",
                        end_type="max", end_color="16A34A",
                    ),
                )
                ws.conditional_formatting.add(
                    _rng_amt,
                    DataBarRule(
                        start_type="min", end_type="max",
                        color=SAPPHIRE, showValue=True,
                    ),
                )
            except Exception as _cf_exc:  # noqa: BLE001 - cosmetic
                logger.debug("Channel CF skipped: %s", _cf_exc)

        # S82: native, editable pie chart of budget share by channel. Upgrades
        # the Excel deliverable from tables-only -- the client can restyle/edit
        # it (unlike an embedded PNG). Numeric source data lives in a helper
        # block parked to the right of the main layout (COL_END+2 onward, top
        # 8 channels). Fully isolated: any failure logs and skips the chart.
        try:
            if len(_chart_pairs) >= 2:
                _chart_pairs = _chart_pairs[:8]
                _helper_col = COL_END + 2  # park well clear of the B..H layout
                _helper_top = row + 1
                _hcat = get_column_letter(_helper_col)
                _hval = get_column_letter(_helper_col + 1)
                _hc1 = ws.cell(row=_helper_top, column=_helper_col, value="Channel")
                _hc1.font = _FONT_BODY_BOLD
                # S3: this is the plan's OWN per-channel budget -- header/format
                # must reflect the active plan currency, not a hardcoded "(USD)".
                _hc2 = ws.cell(
                    row=_helper_top,
                    column=_helper_col + 1,
                    value=f"Budget ({_get_active_currency()})",
                )
                _hc2.font = _FONT_BODY_BOLD
                for _ci, (_cn, _cv) in enumerate(_chart_pairs):
                    _hk = ws.cell(row=_helper_top + 1 + _ci, column=_helper_col, value=_cn)
                    _hk.font = _FONT_BODY
                    _hv = ws.cell(
                        row=_helper_top + 1 + _ci,
                        column=_helper_col + 1,
                        value=_cv,
                    )
                    _hv.font = _FONT_BODY
                    _hv.number_format = _usd0_fmt()
                # The source-data table is internal plumbing for the pie chart --
                # hide its columns so the client sees only the chart, not a stray
                # raw-number table beside the layout.
                ws.column_dimensions[_hcat].hidden = True
                ws.column_dimensions[_hval].hidden = True
                _pie = PieChart()
                _pie.title = "Budget Allocation by Channel"
                _data_ref = Reference(
                    ws,
                    min_col=_helper_col + 1,
                    min_row=_helper_top,
                    max_row=_helper_top + len(_chart_pairs),
                )
                _cats_ref = Reference(
                    ws,
                    min_col=_helper_col,
                    min_row=_helper_top + 1,
                    max_row=_helper_top + len(_chart_pairs),
                )
                _pie.add_data(_data_ref, titles_from_data=True)
                _pie.set_categories(_cats_ref)
                _pie.height = 7.5
                _pie.width = 12.5
                ws.add_chart(_pie, f"{get_column_letter(COL_START)}{row + 1}")
                # Reserve vertical space so following sections don't overlap the chart.
                row += 16
        except Exception as _chart_err:  # pragma: no cover — never break the workbook
            logger.warning("Excel budget pie chart skipped: %s", _chart_err)
    row += 1

    # ── 4. Recruitment Benchmarks ──
    # Load benchmarks from knowledge base (NOT hardcoded)
    kb_benchmarks = {}
    if load_kb_fn:
        try:
            kb = load_kb_fn()
            kb_benchmarks = kb.get("recruitment_benchmarks", {}).get(
                "industry_benchmarks", {}
            )
            if not kb_benchmarks:
                # Try alternative paths
                kb_benchmarks = kb.get("benchmarks", {})
        except Exception as exc:
            logger.warning("Knowledge base load failed for benchmarks: %s", exc)

    if kb_benchmarks:
        row = _write_section_header(ws, row, "Recruitment Benchmarks")

        # Determine client's relevant region(s) from locations
        def _detect_region(loc: str) -> str:
            loc_lower = loc.lower()
            us_indicators = [
                "united states",
                "usa",
                "california",
                "new york",
                "texas",
                "florida",
                "chicago",
                "los angeles",
                "houston",
                "phoenix",
            ]
            if any(kw in loc_lower for kw in us_indicators):
                return "North America"
            eu_indicators = [
                "uk",
                "united kingdom",
                "germany",
                "france",
                "spain",
                "italy",
                "netherlands",
                "europe",
            ]
            if any(kw in loc_lower for kw in eu_indicators):
                return "Europe"
            apac_indicators = [
                "india",
                "china",
                "japan",
                "singapore",
                "australia",
                "asia",
                "pacific",
            ]
            if any(kw in loc_lower for kw in apac_indicators):
                return "APAC"
            latam_indicators = [
                "brazil",
                "mexico",
                "colombia",
                "argentina",
                "latin america",
            ]
            if any(kw in loc_lower for kw in latam_indicators):
                return "LATAM"
            return "North America"  # default

        client_regions = list(set(_detect_region(loc) for loc in locations))

        # Try to find industry-specific benchmarks
        ind_bench = kb_benchmarks.get(
            industry, kb_benchmarks.get("general_entry_level", {})
        )
        if isinstance(ind_bench, dict):
            row = _write_subsection_header(
                ws, row, f"Industry Benchmarks: {industry_label}"
            )

            # If benchmarks have regional breakdown, filter to client regions
            regional = ind_bench.get("regional", ind_bench.get("by_region", {}))
            if regional and isinstance(regional, dict):
                filtered_regional = {
                    k: v
                    for k, v in regional.items()
                    if any(r.lower() in k.lower() for r in client_regions)
                }
                if filtered_regional:
                    headers = ["Region", "CPA", "CPC", "Cost/Hire", "Apply Rate"]
                    row = _write_table_header(ws, row, headers)
                    for idx, (region, rdata) in enumerate(filtered_regional.items()):
                        if isinstance(rdata, dict):
                            values = [
                                region,
                                _flatten_value(
                                    rdata.get(
                                        "cpa", rdata.get("cost_per_application") or ""
                                    )
                                ),
                                _flatten_value(
                                    rdata.get("cpc", rdata.get("cost_per_click") or "")
                                ),
                                _flatten_value(
                                    rdata.get("cph", rdata.get("cost_per_hire") or "")
                                ),
                                _flatten_value(rdata.get("apply_rate") or ""),
                            ]
                        else:
                            values = [region, _flatten_value(rdata), "", "", ""]
                        row = _write_table_row(ws, row, values, alternate=idx % 2 == 1)
            else:
                # Flat benchmarks (no regional breakdown)
                for key, val in ind_bench.items():
                    if key not in ("regional", "by_region", "metadata"):
                        if key == "seasonal_patterns":
                            # S: prefer a matched sub-vertical's own seasonal
                            # profile (gold_standard Gate 7) over the generic
                            # KB benchmark for this row -- see
                            # _subvertical_seasonal_override_text for why.
                            val_str = _subvertical_seasonal_override_text(
                                data
                            ) or _flatten_value(val)
                        else:
                            # copy:both#6 fix: strip physician/MD-only
                            # sub-stats (e.g. cph.physician_recruitment) out
                            # of this shared healthcare_medical benchmark
                            # dict when this plan isn't hiring physicians --
                            # otherwise a senior-living/allied-health client
                            # sees physician compensation presented as its
                            # own hiring intelligence.
                            val_str = _flatten_value(
                                _scope_benchmark_to_plan_roles(val, roles)
                            )
                        if val_str:
                            row = _write_kv_row(
                                ws, row, _humanize_snake_key(key), val_str
                            )

            # S89A FIX (findings data:manpower#3/#4, visual:manpower#3,
            # strategy:manpower#3): when this plan's own blended CPA/apply
            # rate falls outside the benchmark range just cited above,
            # explain why honestly instead of leaving it unreconciled.
            _bm_note = _model_vs_benchmark_note(
                data, ind_bench if isinstance(ind_bench, dict) else {}
            )
            if _bm_note:
                row = _write_footnote(ws, row, _bm_note)
        row += 1

    # ── 5. Executive Strategic Narrative (LLM-generated) ──
    # Generate a C-suite quality narrative using Claude Haiku via the LLM router directly
    # (avoids circular import with app.py)
    exec_narrative = ""
    try:
        from llm_router import LLMRouter, TASK_PLAN_NARRATIVE

        _exec_router = LLMRouter()
        _narrative_prompt = (
            f"Write a 4-5 sentence executive summary for a recruitment media plan.\n\n"
            f"Client: {client_name}\n"
            f"Industry: {industry_label}\n"
            f"Budget: {_fmt_currency(budget_num)}\n"
            f"Locations: {', '.join(str(l) for l in locations[:5])}\n"
            f"Roles: {', '.join(str(r) for r in roles[:5])}\n"
            f"Hire Volume: {hire_volume}\n"
            f"Duration: {duration}\n"
            f"Projected Hires: {_header_hires or 'TBD'}\n"
            f"Cost/Hire: {_fmt_currency(_header_cph)}\n"
            f"Budget Grade: {sufficiency.get('grade') or 'N/A'}\n"
            f"Top Channels: {', '.join(list(channel_allocs.keys())[:5])}\n\n"
            f"Write as a senior recruitment strategist presenting to a VP of Talent Acquisition. "
            f"Include: (1) market thesis — why this plan will succeed, "
            f"(2) ROI projection summary with specific numbers, "
            f"(3) key risks to monitor, "
            f"(4) recommended next steps with timeline. "
            f"Be specific, cite data from above, no generic statements."
        )
        # S50: 10s timeout for plan-gen LLM calls to avoid blocking Excel generation.
        _exec_result = _exec_router.call_llm(
            messages=[{"role": "user", "content": _narrative_prompt}],
            system_prompt=(
                "You are a senior recruitment marketing strategist presenting to "
                "C-suite executives. Write with authority, cite specific data points, "
                "and explain causal reasoning. Every sentence must contain a number "
                "or specific insight. No fluff, no platitudes."
            ),
            task_type=TASK_PLAN_NARRATIVE,  # S48: Route narratives to Groq (fast prose)
            max_tokens=600,
            timeout_budget=10.0,
        )
        exec_narrative = _exec_result.get("text") or ""
    except ImportError:
        logger.warning("LLM router not available for executive narrative")
    except Exception as exc:
        logger.warning("Executive narrative generation failed (non-fatal): %s", exc)

    if exec_narrative:
        row = _write_section_header(ws, row, "Executive Strategic Summary")
        # Wrap the narrative in a merged cell
        ws.merge_cells(
            start_row=row, start_column=COL_START, end_row=row + 3, end_column=COL_END
        )
        cell = ws.cell(row=row, column=COL_START, value=exec_narrative)
        cell.font = Font(name=FONT_BODY_NAME, size=11, color=NAVY)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.fill = _FILL_BLUE_PALE
        for r in range(row, row + 4):
            ws.row_dimensions[r].height = 20
        row += 5

    # ── 5b. Risk Analysis ──
    row = _write_section_header(ws, row, "Risk Analysis")
    risk_items: list[tuple[str, str, str]] = []  # (risk, impact, mitigation)

    # Budget risk -- use consistent header values (S48)
    proj_hires = _header_hires
    cph = _header_cph
    if proj_hires > 0 and cph > 0:
        hires_at_20_pct_increase = int(budget_num / (cph * 1.2)) if cph > 0 else 0
        risk_items.append(
            (
                "Budget Risk: CPA Inflation",
                f"If CPA rises 20%, projected hires drop from {proj_hires:,.0f} to {hires_at_20_pct_increase:,.0f} "
                f"({proj_hires - hires_at_20_pct_increase:,.0f} fewer hires)",
                "Build 10-15% budget contingency; diversify to lower-CPA channels",
            )
        )

    # Market timing risk
    import datetime as _dt_risk

    current_month = _dt_risk.date.today().month
    _q2_months = {4, 5, 6}
    _q1_months = {1, 2, 3}
    campaign_start = data.get("campaign_start_month") or current_month
    if isinstance(campaign_start, int) and campaign_start in _q2_months:
        risk_items.append(
            (
                "Market Timing: Q2 Competition",
                "Q2 hiring is 15-20% more competitive than Q4 due to fiscal year budget cycles",
                "Front-load spend in first 4 weeks; lock in niche channel inventory early",
            )
        )
    elif isinstance(campaign_start, int) and campaign_start in _q1_months:
        risk_items.append(
            (
                "Market Timing: New Year Surge",
                "Q1 sees 25% increase in job seeker activity but also 20% more employer competition",
                "Leverage higher candidate supply with aggressive apply-rate optimization",
            )
        )

    # Channel dependency risk
    if channel_allocs:
        sorted_ch = sorted(
            channel_allocs.items(),
            key=lambda x: x[1].get("percentage", 0),
            reverse=True,
        )
        top_2_pct = sum(ch[1].get("percentage", 0) for ch in sorted_ch[:2])
        if top_2_pct > 55:
            ch_names = ", ".join(
                display_format.channel_label(ch[0]) for ch in sorted_ch[:2]
            )
            # S89: hires-at-risk = the ACTUAL projected hires of the named
            # top-2 channels (never a top_2_pct * total-hires estimate --
            # that overstates risk whenever the named channels' hire mix
            # differs from their budget-share mix).
            _named_hires = sum(
                int(ch[1].get("projected_hires") or 0) for ch in sorted_ch[:2]
            )
            risk_items.append(
                (
                    "Channel Dependency",
                    f"{top_2_pct:.0f}% of budget concentrated on {ch_names} — "
                    f"single-channel disruption could impact {_named_hires:,} projected hires",
                    "Diversify to 4+ channels; maintain 3 backup channels on standby",
                )
            )

    # Competitive pressure risk
    gold_standard_data = data.get("_gold_standard") or {}
    competitor_map = gold_standard_data.get("competitor_mapping") or {}
    n_competitive_cities = sum(
        1
        for city_key, info in competitor_map.items()
        if not str(city_key).startswith("_")
        and str(info.get("hiring_intensity") or "").lower() in ("high", "very_high")
    )
    if n_competitive_cities > 0:
        _market_noun = "market has" if n_competitive_cities == 1 else "markets have"
        risk_items.append(
            (
                "Competitive Pressure",
                f"{n_competitive_cities} {_market_noun} high competitive intensity — "
                f"Fortune 500+ companies actively hiring similar roles",
                "Differentiate with employer brand messaging; emphasize career growth, culture, flexibility",
            )
        )

    if risk_items:
        headers = ["Risk Factor", "Impact Assessment", "Mitigation Strategy"]
        _risk_fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
        row_h = row
        for i, h in enumerate(headers):
            col_start = COL_START + i * 2
            ws.merge_cells(
                start_row=row_h,
                start_column=col_start,
                end_row=row_h,
                end_column=col_start + 1,
            )
            cell = ws.cell(row=row_h, column=col_start, value=h)
            cell.font = _FONT_TABLE_HEADER
            cell.fill = _risk_fill
            cell.alignment = _ALIGN_CENTER
            cell.border = _BORDER_THIN
        ws.row_dimensions[row_h].height = 22
        row = row_h + 1

        for idx, (risk, impact, mitigation) in enumerate(risk_items):
            bg_fill = _FILL_RED_BG if idx % 2 == 0 else _FILL_WHITE
            for col_idx, val in enumerate([risk, impact, mitigation]):
                col_start = COL_START + col_idx * 2
                ws.merge_cells(
                    start_row=row,
                    start_column=col_start,
                    end_row=row,
                    end_column=col_start + 1,
                )
                cell = ws.cell(row=row, column=col_start, value=val)
                cell.font = _FONT_BODY if col_idx > 0 else _FONT_BODY_BOLD
                cell.fill = bg_fill
                cell.alignment = _ALIGN_WRAP
                cell.border = _BORDER_THIN
            ws.row_dimensions[row].height = 40
            row += 1
    else:
        ws.merge_cells(
            start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
        )
        ws.cell(
            row=row,
            column=COL_START,
            value="Insufficient data to generate risk analysis. Add locations and budget for detailed risk assessment.",
        ).font = _FONT_FOOTNOTE
        row += 1

    row += 1

    # ── 6. Key Recommendations ──
    all_recommendations = list(recommendations)

    # Add industry-specific recommendations from tier groups
    if tier_groups:
        for tier_name, tier_data in tier_groups.items():
            tier_info = tier_data.get("tier_info", {})
            tier_roles = tier_data.get("roles") or []
            strategy = tier_info.get("sourcing_strategy") or ""
            if strategy and tier_roles:
                all_recommendations.append(
                    f"{tier_name} roles ({', '.join(tier_roles[:3])}): {strategy}"
                )

    # Curated regulatory/certification callouts (Hazmat/Tanker for propane
    # CDL roles, licensure for nursing roles, etc.) -- factual, non-fabricated.
    # Prepended (not appended) so they always survive the top-8 cap below --
    # a safety/compliance callout is exactly the kind of item that must
    # never get silently truncated off the end of a longer recommendations
    # list.
    _role_callouts = insight_composer.role_requirements_callout(industry, roles)
    all_recommendations = _role_callouts + all_recommendations

    if all_recommendations or warnings:
        row = _write_section_header(ws, row, "Key Recommendations")

        if warnings:
            row = _write_subsection_header(ws, row, "Warnings")
            for w in warnings[:5]:
                ws.merge_cells(
                    start_row=row,
                    start_column=COL_START,
                    end_row=row,
                    end_column=COL_END,
                )
                cell = ws.cell(
                    row=row,
                    column=COL_START,
                    value=f"  {_delabel_channel_keys_in_text(w)}",
                )
                cell.font = Font(name=FONT_BODY_NAME, size=10, color=RED)
                cell.fill = _FILL_RED_BG
                cell.alignment = _ALIGN_WRAP
                row += 1
            row += 1

        if all_recommendations:
            row = _write_subsection_header(ws, row, "Recommendations")
            for idx, rec in enumerate(all_recommendations[:8]):
                ws.merge_cells(
                    start_row=row,
                    start_column=COL_START,
                    end_row=row,
                    end_column=COL_END,
                )
                cell = ws.cell(
                    row=row,
                    column=COL_START,
                    value=f"  {idx + 1}. {_delabel_channel_keys_in_text(rec)}",
                )
                cell.font = _FONT_BODY
                cell.alignment = _ALIGN_WRAP
                cell.fill = _FILL_BLUE_PALE if idx % 2 else _FILL_WHITE
                row += 1

    # ── 7. Creative Quality Score (P1-16) ──
    # S1 (2026-07-03): internal QA artifact -- gated OFF by default so a
    # client-facing workbook never shows a "Grade F" style badge. Only
    # renders when internal_qc_mode(data) is explicitly enabled.
    cqs = data.get("_creative_quality_score")
    if (
        _internal_qc_mode(data)
        and cqs
        and isinstance(cqs, dict)
        and cqs.get("score") is not None
        and not cqs.get("degenerate")
    ):
        row += 1
        row = _write_section_header(ws, row, "Creative Quality Score")

        cqs_score = cqs.get("score", 0)
        cqs_grade = cqs.get("grade", "—")
        ws.merge_cells(
            start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
        )
        badge_cell = ws.cell(
            row=row,
            column=COL_START,
            value=f"  Overall: {cqs_score}/100 (Grade {cqs_grade})",
        )
        badge_cell.font = Font(name=FONT_BODY_NAME, size=12, bold=True, color=SAPPHIRE)
        badge_cell.alignment = _ALIGN_LEFT
        row += 1

        cqs_factors = cqs.get("factors", {})
        for factor_name, factor_data in cqs_factors.items():
            if isinstance(factor_data, dict):
                label = factor_name.replace("_", " ").title()
                pts = factor_data.get("score", 0)
                mx = factor_data.get("max", 0)
                ws.merge_cells(
                    start_row=row,
                    start_column=COL_START,
                    end_row=row,
                    end_column=COL_END,
                )
                ws.cell(
                    row=row,
                    column=COL_START,
                    value=f"    {label}: {pts}/{mx}",
                ).font = _FONT_BODY
                row += 1

        cqs_recs = cqs.get("recommendations") or []
        if cqs_recs:
            row = _write_subsection_header(ws, row, "Creative Recommendations")
            for idx, rec in enumerate(cqs_recs[:5]):
                ws.merge_cells(
                    start_row=row,
                    start_column=COL_START,
                    end_row=row,
                    end_column=COL_END,
                )
                cell = ws.cell(row=row, column=COL_START, value=f"  {idx + 1}. {rec}")
                cell.font = _FONT_BODY
                cell.alignment = _ALIGN_WRAP
                cell.fill = _FILL_BLUE_PALE if idx % 2 else _FILL_WHITE
                row += 1

    row += 2
    _write_attribution_footer(ws, row)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: CHANNELS & CHANNEL STRATEGY
# ═══════════════════════════════════════════════════════════════════════════════


def _build_sheet_channels(ws, data: dict, research_mod=None, load_kb_fn=None):
    """Build Sheet 2: Channels & Channel Strategy."""
    ws.title = "Channels & Strategy"
    ws.sheet_properties.tabColor = SAPPHIRE

    _set_column_widths(
        ws,
        {
            1: 3,
            2: 22,
            3: 14,
            4: 14,
            5: 14,
            6: 14,
            7: 14,
            8: 18,
        },
    )

    industry = data.get("industry", "general_entry_level")
    industry_label = _get_industry_label(industry)
    roles = _get_roles(data)
    locations = _get_locations(data)

    budget_alloc = data.get("_budget_allocation", {})
    channel_allocs = budget_alloc.get("channel_allocations", {})
    synthesized = data.get("_synthesized", {})
    channels_db = data.get("_channels_db", {})
    collar_type = data.get("_collar_type", "mixed")

    row = 2

    # ── 1. Channel Strategy Overview ──
    row = _write_section_header(ws, row, "Channel Strategy Overview")

    if channel_allocs:
        # Top channels sorted by budget allocation
        sorted_channels = sorted(
            channel_allocs.items(),
            key=lambda x: x[1].get("dollar_amount", x[1].get("dollars") or 0),
            reverse=True,
        )

        headers = [
            "Channel",
            "Budget %",
            "Amount",
            "Category",
            "CPC",
            "Confidence",
            "Fit",
        ]
        row = _write_table_header(ws, row, headers)

        # S89: live numeric cells (summable/sortable). Cols: Channel | Budget%
        # | Amount($) | Category | CPC($) | Confidence | Fit
        # S3: Amount/CPC are the plan's OWN figures -- active plan currency.
        _ch_formats = [None, FMT_PCT1, _usd0_fmt(), None, _usd2_fmt(), None, "0.0"]
        _ch_first = row
        # data:atria#3-related polish: same largest-remainder-rounded
        # percentage map the Executive Summary and Channel Recommendations
        # sheets use, so this sheet's Budget % column foots to exactly
        # 100.0% and agrees cell-for-cell with the other two sheets.
        _corrected_pct = _corrected_channel_pct_display(channel_allocs)
        for idx, (ch_name, ch_data) in enumerate(sorted_channels[:15]):
            roi = ch_data.get("roi_score") or ""
            confidence = _derive_channel_confidence(data, ch_data)
            category = ch_data.get("category") or ""
            _fit_val = _safe_num(roi) if not isinstance(roi, str) else None

            values = [
                display_format.channel_label(ch_name),
                _corrected_pct.get(
                    ch_name, _safe_num(ch_data.get("percentage") or 0) / 100.0
                ),
                _safe_num(ch_data.get("dollar_amount", ch_data.get("dollars") or 0)),
                display_format.channel_label(category) if category else "",
                _safe_num(ch_data.get("cpc") or 0),
                confidence.title() if isinstance(confidence, str) else str(confidence),
                round(_fit_val, 1) if _fit_val is not None else (roi or ""),
            ]
            # Keep the Fit column as text when the source isn't numeric.
            _row_formats = list(_ch_formats)
            if _fit_val is None:
                _row_formats[6] = None
            row = _write_table_row(
                ws, row, values, alternate=idx % 2 == 1, number_formats=_row_formats
            )
        _ch_last = row - 1
        # S89: live totals row for the money column.
        if _ch_last >= _ch_first:
            _amt_letter = get_column_letter(COL_START + 2)
            _tot = ws.cell(row=row, column=COL_START, value="Total")
            _tot.font = _FONT_BODY_BOLD
            _tot.fill = _FILL_BLUE_LIGHT
            _tot.alignment = _ALIGN_LEFT
            _tot.border = _BORDER_THIN
            for _ci in range(1, len(headers)):
                _c = ws.cell(row=row, column=COL_START + _ci)
                if _ci == 2:
                    _c.value = f"=SUM({_amt_letter}{_ch_first}:{_amt_letter}{_ch_last})"
                    _c.number_format = _usd0_fmt()
                _c.font = _FONT_BODY_BOLD
                _c.fill = _FILL_BLUE_LIGHT
                _c.alignment = _ALIGN_CENTER
                _c.border = _BORDER_THIN
            row += 1
            try:
                _amt_rng = f"{_amt_letter}{_ch_first}:{_amt_letter}{_ch_last}"
                ws.conditional_formatting.add(
                    _amt_rng,
                    DataBarRule(
                        start_type="min", end_type="max",
                        color=SAPPHIRE, showValue=True,
                    ),
                )
            except Exception as _cf_exc:  # noqa: BLE001
                logger.debug("Channels CF skipped: %s", _cf_exc)
    else:
        ws.merge_cells(
            start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
        )
        ws.cell(
            row=row, column=COL_START, value="No channel allocation data available."
        ).font = _FONT_BODY
        row += 1

    row += 2

    # ── 2. Recommended Channels (Vetted) ──
    row = _write_section_header(ws, row, "Recommended Channels (Vetted)")

    # Build channel list from multiple sources
    raw_channels = []

    # From channel allocations
    # S92: thread through budget_engine's authoritative, post-rebalance
    # roi_score-derived ranking (fit_score/vetted_tier -- see
    # budget_engine._finalize_channel_ranking) so vet_channels() can use it
    # instead of an independent industry-keyword heuristic that used to
    # contradict this same allocation's own ROI ranking.
    for ch_name, ch_data in channel_allocs.items():
        raw_channels.append(
            {
                "name": ch_name,
                "category": ch_data.get("category") or "",
                "budget_pct": ch_data.get("percentage") or 0,
                "cpc": ch_data.get("cpc") or 0,
                "fit_score": ch_data.get("fit_score"),
                "vetted_tier": ch_data.get("vetted_tier"),
            }
        )

    # From channels DB (add channels not already in allocations)
    alloc_names_lower = {n.lower() for n in channel_allocs.keys()}
    if isinstance(channels_db, dict):
        for cat_key in [
            "traditional",
            "non_traditional",
            "programmatic",
            "social_media",
            "niche",
            "regional",
        ]:
            cat_channels = channels_db.get(cat_key, {})
            if isinstance(cat_channels, dict):
                for ch_name, ch_info in cat_channels.items():
                    if ch_name.lower() not in alloc_names_lower:
                        entry = {"name": ch_name, "category": cat_key}
                        if isinstance(ch_info, dict):
                            entry.update(ch_info)
                        raw_channels.append(entry)
            elif isinstance(cat_channels, list):
                for ch_item in cat_channels:
                    ch_name = (
                        ch_item.get("name", str(ch_item))
                        if isinstance(ch_item, dict)
                        else str(ch_item)
                    )
                    if ch_name.lower() not in alloc_names_lower:
                        entry = {"name": ch_name, "category": cat_key}
                        if isinstance(ch_item, dict):
                            entry.update(ch_item)
                        raw_channels.append(entry)

    # Vet the channels
    vetted = vet_channels(raw_channels, industry, roles, locations)

    if vetted:
        headers = [
            "Channel",
            "Category",
            "Fit",
            "CPC",
            "Budget %",
            "Strategic Rationale",
            "Fit Score",
        ]
        row = _write_table_header(ws, row, headers)

        for idx, ch in enumerate(vetted[:20]):  # cap at 20
            fit = ch.get("fit", "Fair")
            fit_score = ch.get("fit_score", 0.5)
            ch_name = display_format.channel_label(ch.get("name") or "")
            ch_category = display_format.channel_label(ch.get("category") or "")
            ch_cpc = ch.get("cpc") or 0
            ch_pct = ch.get("budget_pct") or 0
            notes = ch.get("description", ch.get("notes") or "")
            if isinstance(notes, dict):
                notes = _flatten_value(notes)

            # Build strategic rationale with WHY reasoning
            rationale_parts: list[str] = []
            if fit == "Strong" and fit_score >= 0.7:
                rationale_parts.append(
                    f"High-fit ({fit_score:.0%}) for {industry_label}"
                )
            elif fit == "Good":
                rationale_parts.append(f"Good industry alignment ({fit_score:.0%})")
            if ch_cpc > 0:
                rationale_parts.append(f"CPC {_fmt_currency(ch_cpc, show_cents=True)}")
            if ch_pct > 15:
                rationale_parts.append(
                    f"Primary channel — {ch_pct:.0f}% of budget for volume"
                )
            elif ch_pct > 5:
                rationale_parts.append(f"Supporting channel at {ch_pct:.0f}%")
            # Add role/location context
            if roles and len(roles) <= 3:
                rationale_parts.append(f"targets {', '.join(roles[:2])}")
            if locations and len(locations) <= 3:
                rationale_parts.append(
                    f"in {', '.join(str(l).split(',')[0] for l in locations[:2])}"
                )
            if notes and len(notes) > 10:
                rationale_parts.append(notes[:60])

            # S5 (2026-07-03, findings 39/52): never hard-truncate the
            # rationale mid-word -- write it in full and let the row wrap
            # (matching the risk-table convention elsewhere in this file),
            # sizing the row taller so the wrapped text is actually readable.
            rationale = (
                "; ".join(rationale_parts)
                if rationale_parts
                else (notes if notes else "")
            )

            values = [
                ch_name,
                ch_category,
                fit,
                (ch_cpc if ch_cpc else ""),
                (_safe_num(ch_pct) / 100.0 if ch_pct else ""),
                rationale,
                f"{fit_score:.2f}",
            ]

            # S5: CPC/Budget % are the plan's own live numbers, not text.
            _vetted_formats = [
                None,
                None,
                None,
                _usd2_fmt() if ch_cpc else None,
                FMT_PCT1 if ch_pct else None,
                None,
                None,
            ]

            # Custom fills for fit column
            fit_fills = [
                None,
                None,
                _fit_fill(fit),
                None,
                None,
                None,
                _fit_score_fill(fit_score),
            ]
            fit_fonts = [None, None, None, None, None, None, _fit_score_font(fit_score)]
            row = _write_table_row(
                ws,
                row,
                values,
                alternate=idx % 2 == 1,
                fills=fit_fills,
                fonts=fit_fonts,
                number_formats=_vetted_formats,
            )
            if len(rationale) > 90:
                ws.row_dimensions[row - 1].height = 40
    else:
        ws.merge_cells(
            start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
        )
        ws.cell(
            row=row, column=COL_START, value="No vetted channels available."
        ).font = _FONT_BODY
        row += 1

    row += 2

    # ── 3. Ad Platform Analysis ──
    ad_platforms = synthesized.get("ad_platform_analysis", {})

    row = _write_section_header(ws, row, "Ad Platform Analysis")

    if ad_platforms:
        # Build headers dynamically -- exclude ROI Projection
        # Include Audience Reach only if at least one platform has non-zero reach
        has_reach = any(
            _safe_num(p.get("audience_reach", p.get("estimated_reach") or 0)) > 0
            for p in ad_platforms.values()
            if isinstance(p, dict)
        )

        headers = ["Platform", "CPC", "CPM", "CPA", "Fit Score"]
        if has_reach:
            headers.insert(4, "Audience Reach")

        row = _write_table_header(ws, row, headers)

        _plat_idx = 0
        for plat_key, plat_data in ad_platforms.items():
            if not isinstance(plat_data, dict):
                continue
            plat_name = plat_data.get(
                "platform_name", plat_key.replace("_", " ").title()
            )
            fit_score = _safe_num(plat_data.get("fit_score") or 0)
            _p_cpc = plat_data.get("avg_cpc", plat_data.get("cpc") or 0) or 0
            _p_cpm = plat_data.get("avg_cpm", plat_data.get("cpm") or 0) or 0
            _p_cpa = plat_data.get("avg_cpa", plat_data.get("cpa") or 0) or 0
            if not any([_p_cpc, _p_cpm, _p_cpa, fit_score]):
                continue
            idx = _plat_idx
            _plat_idx += 1

            # S5 (2026-07-03, finding 44/51): CPC/CPM/CPA are the plan's own
            # live figures -- write them as numbers with a currency
            # number_format instead of pre-formatted text strings.
            values = [
                plat_name,
                _p_cpc,
                _p_cpm,
                _p_cpa,
            ]
            _plat_formats = [None, _usd2_fmt(), _usd2_fmt(), _usd2_fmt()]

            if has_reach:
                reach = _safe_num(
                    plat_data.get(
                        "audience_reach", plat_data.get("estimated_reach") or 0
                    )
                )
                values.append(reach if reach > 0 else "")
                _plat_formats.append(FMT_INT if reach > 0 else None)

            values.append(f"{fit_score:.2f}")
            _plat_formats.append(None)

            # Color-code fit scores
            fit_col_idx = len(values) - 1
            row_fills = [None] * len(values)
            row_fonts = [None] * len(values)
            row_fills[fit_col_idx] = _fit_score_fill(fit_score)
            row_fonts[fit_col_idx] = _fit_score_font(fit_score)

            row = _write_table_row(
                ws,
                row,
                values,
                alternate=idx % 2 == 1,
                fills=row_fills,
                fonts=row_fonts,
                number_formats=_plat_formats,
            )
        # S46: Per-role breakdown sub-table when multiple roles have different metrics
        roles_input = data.get("roles") or data.get("target_roles") or []
        if isinstance(roles_input, list) and len(roles_input) > 1:
            # Check if any platform has per_role_metrics
            _has_role_data = False
            for _pk, _pd in ad_platforms.items():
                if isinstance(_pd, dict) and _pd.get("per_role_metrics"):
                    _has_role_data = True
                    break

            if _has_role_data:
                row += 1
                row = _write_section_header(ws, row, "Ad Platform Metrics by Role")
                _role_headers = ["Platform", "Role", "CPC", "CPM", "CPA"]
                row = _write_table_header(ws, row, _role_headers)

                _role_idx = 0
                for plat_key, plat_data in ad_platforms.items():
                    if not isinstance(plat_data, dict):
                        continue
                    per_role = plat_data.get("per_role_metrics") or {}
                    if not per_role:
                        continue
                    plat_name = plat_data.get(
                        "platform_name", plat_key.replace("_", " ").title()
                    )
                    for role_name, role_metrics in per_role.items():
                        if not isinstance(role_metrics, dict):
                            continue
                        r_cpc = role_metrics.get("avg_cpc") or 0
                        r_cpm = role_metrics.get("avg_cpm") or 0
                        r_cpa = role_metrics.get("avg_cpa") or 0
                        if not any([r_cpc, r_cpm, r_cpa]):
                            continue
                        vals = [
                            plat_name,
                            str(role_name),
                            r_cpc if r_cpc else "",
                            r_cpm if r_cpm else "",
                            r_cpa if r_cpa else "",
                        ]
                        _role_formats = [
                            None,
                            None,
                            _usd2_fmt() if r_cpc else None,
                            _usd2_fmt() if r_cpm else None,
                            _usd2_fmt() if r_cpa else None,
                        ]
                        row = _write_table_row(
                            ws,
                            row,
                            vals,
                            alternate=_role_idx % 2 == 1,
                            number_formats=_role_formats,
                        )
                        _role_idx += 1

    else:
        # Fallback: show a "data pending" note with general guidance
        ws.merge_cells(
            start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
        )
        ws.cell(
            row=row,
            column=COL_START,
            value=(
                "Ad platform performance data pending — live integration "
                "will populate this section once campaign data is available. "
                "In the interim, refer to the Channel Benchmarks table on the "
                "Executive Summary sheet for estimated CPC/CPA ranges."
            ),
        ).font = _FONT_FOOTNOTE
        ws.cell(row=row, column=COL_START).alignment = _ALIGN_WRAP
        row += 1

    row += 2

    # ── 4. Industry Niche Channels ──
    # S4: INDUSTRY_NICHE_CHANNELS is a US-domiciled board list (ClearedJobs.Net,
    # USAJOBS, Military.com, etc. for aerospace_defense; similar US-only boards
    # for other industries). Never ship it on a non-US plan -- fall back to a
    # clearly-labeled reference-framework note instead of fabricating local
    # board data we don't have.
    niche_channels = INDUSTRY_NICHE_CHANNELS.get(industry, []) if _is_us_plan(data) else []

    if niche_channels:
        row = _write_section_header(ws, row, f"Niche Channels: {industry_label}")

        headers = ["Channel", "Type", "Relevance"]
        row = _write_table_header(ws, row, headers)

        for idx, ch_name in enumerate(niche_channels):
            values = [
                ch_name,
                "Industry Niche Board",
                "High - Specialized for " + industry_label,
            ]
            row = _write_table_row(ws, row, values, alternate=idx % 2 == 1)
    elif INDUSTRY_NICHE_CHANNELS.get(industry) and not _is_us_plan(data):
        # Reference-framework note: we know US niche boards for this industry
        # but have no local-market equivalent data -- disclose rather than
        # silently drop or fabricate.
        row = _write_section_header(ws, row, f"Niche Channels: {industry_label}")
        _signals = _non_us_signals(data)
        _signal_txt = (
            f" (targets {', '.join(_signals[:3])})" if _signals else ""
        )
        row = _write_kv_row(
            ws,
            row,
            "Note",
            "US-domiciled niche boards for this industry are not shown because "
            f"this plan targets a non-US market{_signal_txt}. Local specialty "
            "board data was not available for this campaign; consult the Intl "
            "Benchmarks sheet and regional job boards for this market.",
        )

    row += 2
    _write_attribution_footer(ws, row)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: MARKET INTELLIGENCE
# ═══════════════════════════════════════════════════════════════════════════════


def _build_sheet_market_intelligence(ws, data: dict, research_mod=None):
    """Build Sheet 3: Market Intelligence."""
    ws.title = "Market Intelligence"
    ws.sheet_properties.tabColor = SAPPHIRE

    _set_column_widths(
        ws,
        {
            1: 3,
            2: 22,
            3: 16,
            4: 16,
            5: 14,
            6: 14,
            7: 14,
            8: 18,
        },
    )

    industry = data.get("industry", "general_entry_level")
    industry_label = _get_industry_label(industry)
    locations = _get_locations(data)
    roles = _get_roles(data)
    client_name = data.get("client_name", "Client")
    competitors = data.get("competitors") or []
    if isinstance(competitors, str):
        competitors = [c.strip() for c in competitors.split(",") if c.strip()]

    synthesized = data.get("_synthesized", {})
    enriched = data.get("_enriched", {})

    row = 2

    # ── 1. Labour Market Overview ──
    row = _write_section_header(ws, row, "Labour Market Overview")

    labour_data = {}
    if research_mod:
        try:
            labour_data = research_mod.get_labour_market_intelligence(
                industry, locations
            )
        except Exception as exc:
            logger.warning("Labour market intelligence lookup failed: %s", exc)

    national = labour_data.get("national_summary", {})
    ind_metrics = labour_data.get("industry_metrics", {})

    # National Economic Snapshot -- use live data or hardcoded fallback
    row = _write_subsection_header(ws, row, "National Economic Snapshot")
    if national:
        display_fields = [
            ("Unemployment Rate", national.get("unemployment_rate") or ""),
            ("Job Openings", national.get("job_openings") or ""),
            ("Hires Rate", national.get("hires_rate") or ""),
            ("Quits Rate", national.get("quits_rate") or ""),
            (
                "Labor Force Participation",
                national.get("labor_force_participation") or "",
            ),
        ]
    else:
        # Fallback: latest available government figures (updated quarterly)
        display_fields = [
            ("Unemployment Rate", "4.0% (Q1 2026 est.)"),
            ("Job Openings", "~8.0M (latest available)"),
            ("Hires Rate", "3.4% (latest available)"),
            ("Quits Rate", "2.2% (latest available)"),
            ("Labor Force Participation", "62.5% (latest available)"),
            (
                "Note",
                "Live data unavailable; figures are latest published government estimates",
            ),
        ]
    for key, val in display_fields:
        val_str = _flatten_value(val)
        if val_str:
            row = _write_kv_row(ws, row, key, val_str)
    row += 1

    if ind_metrics:
        row = _write_subsection_header(ws, row, f"Industry Metrics: {industry_label}")
        for key, val in ind_metrics.items():
            if key in ("metadata", "source", "sources"):
                continue
            val_str = _flatten_value(val)
            if val_str:
                display_key = key.replace("_", " ").title()
                row = _write_kv_row(ws, row, display_key, val_str)
        row += 1

    # Location contexts from labour market data
    loc_contexts = labour_data.get("location_contexts") or []
    if loc_contexts:
        row = _write_subsection_header(ws, row, "Location Economic Context")
        headers = ["Location", "Country", "Unemployment", "Median Salary", "Context"]
        row = _write_table_header(ws, row, headers)
        for idx, lc in enumerate(loc_contexts):
            if isinstance(lc, dict):
                values = [
                    lc.get("location") or "",
                    lc.get("country") or "",
                    _flatten_value(lc.get("unemployment_rate") or ""),
                    _flatten_value(lc.get("median_salary") or ""),
                    lc.get("context_note") or ""[:80],
                ]
                row = _write_table_row(ws, row, values, alternate=idx % 2 == 1)
        row += 1

    row += 1

    # ── 2. Location Intelligence ──
    row = _write_section_header(ws, row, "Location Intelligence")

    loc_profiles = synthesized.get("location_profiles", {})
    loc_demographics = enriched.get("location_demographics", {})

    # S3: median_income here traces to US Census/DataUSA/METRO_DATA fallbacks
    # only (research.py METRO_DATA is US-metro-only) -- a fixed US-benchmark
    # constant, not the plan's own figure, and it never has a match for
    # non-US locations (the column renders "--" for them). Only label it
    # "(USD)" for a USD plan; for a non-USD plan the column is empty anyway,
    # so a USD marker here would misleadingly imply US data exists.
    _income_header = (
        "Median Income (USD)" if _get_active_currency() == "USD" else "Median Income"
    )

    # S89 (finding strategy:atria#5): build all row values FIRST so we know
    # which optional metric columns (Population/Unemployment/Income/Key
    # Industries) actually have data for this run's locations before
    # writing headers -- a header promising a field that's blank for every
    # location reads as "we looked and found nothing," not "never sourced."
    _loc_rows: List[Dict[str, str]] = []
    for idx, loc in enumerate(locations):
        # Data cascade: synthesized > enriched > research fallback
        loc_data = {}

        # Try synthesized location profiles
        if loc_profiles:
            for loc_key, profile in loc_profiles.items():
                if isinstance(profile, dict) and (
                    loc.lower() in loc_key.lower() or loc_key.lower() in loc.lower()
                ):
                    loc_data = profile
                    break

        # Try enriched demographics
        if not loc_data and loc_demographics:
            if isinstance(loc_demographics, dict):
                for demo_key, demo_data in loc_demographics.items():
                    if isinstance(demo_data, dict) and (
                        loc.lower() in demo_key.lower()
                        or demo_key.lower() in loc.lower()
                    ):
                        loc_data = demo_data
                        break
            elif isinstance(loc_demographics, list):
                for demo_item in loc_demographics:
                    if isinstance(demo_item, dict):
                        demo_loc = demo_item.get(
                            "location", demo_item.get("name") or ""
                        )
                        if loc.lower() in str(demo_loc).lower():
                            loc_data = demo_item
                            break

        # Try research module as fallback
        if not loc_data and research_mod:
            try:
                loc_data = research_mod.get_location_info(loc) or {}
            except Exception:
                loc_data = {}

        # Extract values with fallback chain
        country = loc_data.get("country") or ""
        if not country:
            # Try to infer from location string ("City, Country" -> last
            # token). S92 residual fix: for the common 2-part "City, ST"
            # form (e.g. "Denver, CO"), the trailing token is a US state
            # abbreviation/name, not a country -- naive parts[-1] used to
            # put "CO" in the Country field. Check plan_geo's state table
            # first (mirrors app.py's _split_city_state_country) before
            # treating the trailing token as a literal country name.
            parts = loc.split(",")
            if len(parts) == 2:
                _trailing = parts[-1].strip()
                _is_us_state = _trailing.lower() in plan_geo.US_STATE_NAME_TO_ABBR or (
                    len(_trailing) == 2
                    and _trailing.isalpha()
                    and _trailing.upper() in plan_geo.US_STATE_ABBR
                )
                country = "United States" if _is_us_state else _trailing
            elif len(parts) > 2:
                country = parts[-1].strip()
            else:
                # No comma: the location string itself may already BE the
                # country (e.g. loc == "New Zealand"), or the plan carries an
                # explicit top-level country. Only default to "United States"
                # when neither signal is available -- never hardcode it over
                # a non-US plan (S4: findings Market Intelligence B30).
                plan_country = data.get("country") or ""
                if isinstance(plan_country, str) and plan_country.strip():
                    country = plan_country.strip()
                else:
                    country = "United States"

        # Prefer metro/city population over state-level population
        population = (
            loc_data.get("metro_population")
            or loc_data.get("city_population")
            or loc_data.get("population")
            or loc_data.get("pop")
            or ""
        )
        # Guard against state-level populations leaking through:
        # if the number is > 20M and location is a city, it's likely state-level
        if isinstance(population, (int, float)) and population > 20_000_000:
            # Use known metro populations for major US cities
            _metro_pop_fallback: Dict[str, str] = {
                "los angeles": "13.2M metro",
                "new york": "20.1M metro",
                "chicago": "9.5M metro",
                "dallas": "7.6M metro",
                "houston": "7.1M metro",
                "phoenix": "4.9M metro",
                "philadelphia": "6.2M metro",
                "san antonio": "2.6M metro",
                "san diego": "3.3M metro",
                "san jose": "2.0M metro",
                "san francisco": "4.7M metro",
                "seattle": "4.0M metro",
                "denver": "2.9M metro",
                "boston": "4.9M metro",
                "atlanta": "6.1M metro",
                "miami": "6.2M metro",
                "detroit": "4.3M metro",
                "minneapolis": "3.6M metro",
                "portland": "2.5M metro",
            }
            loc_lower = loc.lower()
            for city_key, metro_val in _metro_pop_fallback.items():
                if city_key in loc_lower:
                    population = metro_val
                    break
        unemployment = loc_data.get(
            "unemployment", loc_data.get("unemployment_rate") or ""
        )
        median_income = loc_data.get(
            "median_income",
            loc_data.get(
                "median_salary", loc_data.get("median_household_income") or ""
            ),
        )
        key_industries = loc_data.get(
            "key_industries",
            loc_data.get("major_employers", loc_data.get("top_industries") or ""),
        )

        pop_str = (
            _fmt_number(population)
            if isinstance(population, (int, float))
            else _flatten_value(population)
        )
        unemp_str = _flatten_value(unemployment)
        # US Census/METRO_DATA source only (see header note) -- always USD.
        income_str = (
            _fmt_currency(median_income, prefix="$")
            if isinstance(median_income, (int, float))
            else _flatten_value(median_income)
        )
        industry_str = _flatten_value(key_industries)

        _rationale = insight_composer.geography_rationale(loc, loc_data)

        _loc_rows.append(
            {
                "location": loc,
                "country": country,
                "population": pop_str,
                "unemployment": unemp_str,
                "income": income_str,
                "industries": industry_str[:80] if industry_str else "",
                "rationale": _rationale,
            }
        )

    if _loc_rows:
        _optional_loc_cols = [
            ("population", "Population"),
            ("unemployment", "Unemployment"),
            ("income", _income_header),
            ("industries", "Key Industries"),
        ]
        _populated_loc_cols = [
            (key, label)
            for key, label in _optional_loc_cols
            if any(str(r.get(key) or "").strip() for r in _loc_rows)
        ]
        headers = (
            ["Location", "Country"]
            + [label for _, label in _populated_loc_cols]
            + ["Why This Market"]
        )
        row = _write_table_header(ws, row, headers)

        for idx, r in enumerate(_loc_rows):
            values = (
                [r["location"], r["country"]]
                + [r[key] or "—" for key, _ in _populated_loc_cols]
                + [r["rationale"]]
            )
            row = _write_table_row(ws, row, values, alternate=idx % 2 == 1)

    # ── 2b. Macro Economic Context (FRED indicators) ──
    _fred_macro = {}
    # Try synthesized macro_economic from first role's job_market_demand
    _jmd = synthesized.get("job_market_demand", {})
    if isinstance(_jmd, dict):
        for _jmd_v in _jmd.values():
            if isinstance(_jmd_v, dict) and _jmd_v.get("macro_economic"):
                _fred_macro = _jmd_v["macro_economic"]
                break
    # KB fallback: fred_indicators.json
    if not _fred_macro:
        _kb = data.get("_knowledge_base", {})
        _fred_kb = _kb.get("fred_indicators", {}) if isinstance(_kb, dict) else {}
        _fred_data_raw = (
            _fred_kb.get("data", _fred_kb) if isinstance(_fred_kb, dict) else {}
        )
        if isinstance(_fred_data_raw, dict):
            for _fk, _fv in _fred_data_raw.items():
                if _fk in ("source", "_refreshed_at", "_refreshed_iso"):
                    continue
                if isinstance(_fv, dict) and "value" in _fv:
                    _fred_macro[_fk] = _fv["value"]
                elif isinstance(_fv, (int, float)):
                    _fred_macro[_fk] = _fv

    if _fred_macro:
        row += 1
        row = _write_subsection_header(ws, row, "Macro Economic Context")
        # S3: FRED (US Federal Reserve) macro data is US-only by definition --
        # keep it USD but mark it explicitly so it never reads as a bare $
        # beside a non-USD plan's own figures elsewhere on this sheet.
        _fred_display = [
            ("Unemployment Rate", "unemployment_rate", "%"),
            ("Job Openings (000s)", "job_openings", "K"),
            ("Avg Hourly Earnings (USD)", "avg_hourly_earnings", "$"),
            ("Fed Funds Rate", "fed_funds_rate", "%"),
            ("CPI Index", "cpi_inflation", ""),
        ]
        for _label, _key, _unit in _fred_display:
            _val = _fred_macro.get(_key)
            if _val is not None:
                if _unit == "%":
                    _val_str = f"{_val}%"
                elif _unit == "$":
                    _val_str = (
                        f"${_val:,.2f}" if isinstance(_val, (int, float)) else str(_val)
                    )
                elif _unit == "K":
                    _val_str = (
                        f"{_val:,.0f}" if isinstance(_val, (int, float)) else str(_val)
                    )
                else:
                    _val_str = f"{_val:,.2f}" if isinstance(_val, float) else str(_val)
                row = _write_kv_row(ws, row, _label, _val_str)

    row += 2

    # ── 3. Competitive Landscape ──
    row = _write_section_header(ws, row, "Competitive Landscape")

    comp_intel = synthesized.get("competitive_intelligence", {})
    company_profile = comp_intel.get("company_profile", {})
    sec_data = enriched.get("sec_data", {})

    # Company profile section
    if company_profile or sec_data:
        row = _write_subsection_header(ws, row, f"Company Profile: {client_name}")
        profile_fields = {}

        # Merge from sec_data and company_profile
        if isinstance(sec_data, dict):
            profile_fields.update(
                {
                    "Company Name": sec_data.get(
                        "name", sec_data.get("company_name", client_name)
                    ),
                    "CIK": sec_data.get("cik") or "",
                    "SIC Code": sec_data.get("sic", sec_data.get("sic_code") or ""),
                    "SIC Description": sec_data.get("sic_description") or "",
                    "State": sec_data.get(
                        "state", sec_data.get("state_of_incorporation") or ""
                    ),
                    "Fiscal Year End": sec_data.get("fiscal_year_end") or "",
                }
            )

        if isinstance(company_profile, dict):
            for k, v in company_profile.items():
                if k not in ("metadata", "source") and v:
                    profile_fields[k.replace("_", " ").title()] = v

        for key, val in profile_fields.items():
            val_str = _flatten_value(val)
            if val_str:
                row = _write_kv_row(ws, row, key, val_str)
        row += 1

    # Competitors table
    comp_analysis = comp_intel.get(
        "competitors", comp_intel.get("competitor_analysis") or []
    )
    if not comp_analysis and competitors:
        # Build minimal competitor entries from names list
        comp_analysis = [{"name": c} for c in competitors]

    # Fallback: use industry top employers from knowledge base
    if not comp_analysis:
        _industry_top_employers: Dict[str, List[str]] = {
            "healthcare_medical": [
                "HCA Healthcare",
                "UnitedHealth Group",
                "Ascension",
                "CommonSpirit Health",
                "Kaiser Permanente",
            ],
            "tech_engineering": ["Google", "Amazon", "Microsoft", "Meta", "Apple"],
            "finance_banking": [
                "JPMorgan Chase",
                "Bank of America",
                "Goldman Sachs",
                "Citigroup",
                "Wells Fargo",
            ],
            "retail_consumer": ["Walmart", "Amazon", "Costco", "Target", "Home Depot"],
            "aerospace_defense": [
                "Lockheed Martin",
                "Boeing",
                "Raytheon",
                "Northrop Grumman",
                "General Dynamics",
            ],
            "logistics_supply_chain": [
                "UPS",
                "FedEx",
                "Amazon Logistics",
                "XPO Logistics",
                "C.H. Robinson",
            ],
            "pharma_biotech": [
                "Pfizer",
                "Johnson & Johnson",
                "AbbVie",
                "Merck",
                "Amgen",
            ],
            "hospitality_travel": ["Marriott", "Hilton", "Hyatt", "IHG", "Airbnb"],
            "education": [
                "Pearson",
                "McGraw-Hill",
                "Chegg",
                "Coursera",
                "University Systems",
            ],
            "energy_utilities": [
                "ExxonMobil",
                "Chevron",
                "NextEra Energy",
                "Duke Energy",
                "Southern Company",
            ],
            "trucking": [
                "Werner Enterprises",
                "Schneider National",
                "J.B. Hunt",
                "Knight-Swift",
                "Swift Transportation",
            ],
            "transportation": [
                "Werner Enterprises",
                "Schneider National",
                "J.B. Hunt",
                "Knight-Swift",
                "UPS",
            ],
            "manufacturing": [
                "General Electric",
                "3M",
                "Honeywell",
                "Caterpillar",
                "Deere & Co",
            ],
            "construction": [
                "Turner Construction",
                "Bechtel",
                "Fluor",
                "Skanska",
                "AECOM",
            ],
            "staffing": [
                "Robert Half",
                "Adecco",
                "ManpowerGroup",
                "Kelly Services",
                "Randstad",
            ],
            "government": [
                "Lockheed Martin",
                "Raytheon",
                "Northrop Grumman",
                "General Dynamics",
                "Boeing",
            ],
        }
        # Industry-aware fallback: try exact key, then substring match
        fallback_names = _industry_top_employers.get(industry, [])
        if not fallback_names:
            _ind_lower = str(industry).lower()
            for _fb_key, _fb_list in _industry_top_employers.items():
                if _fb_key in _ind_lower or _ind_lower in _fb_key:
                    fallback_names = _fb_list
                    break
        if fallback_names:
            comp_analysis = [
                {
                    "name": n,
                    "industry": industry_label,
                    "size": "",
                    "hiring_activity": "Active (est.)",
                    "overlap_score": "",
                }
                for n in fallback_names
            ]

    if comp_analysis:
        comp_list = comp_analysis if isinstance(comp_analysis, list) else []
        if isinstance(comp_analysis, dict):
            comp_list = [
                {"name": k, **v} if isinstance(v, dict) else {"name": k}
                for k, v in comp_analysis.items()
            ]

        _first_role = roles[0] if roles else ""
        _first_city = locations[0] if locations else ""

        # S89 (finding data:manpower#7): build all row values FIRST so we
        # know which optional columns (Industry/Size/Hiring Activity/Overlap
        # Score) actually have data for this run before writing headers --
        # a header promising a field that's 100% blank across every row
        # reads as "we looked and found nothing" rather than "this was
        # never sourced."
        _comp_rows: List[Dict[str, str]] = []
        for idx, comp in enumerate(comp_list[:10]):
            if isinstance(comp, dict):
                _comp_name = comp.get("name", comp.get("company") or "")
                _counter = insight_composer.compose_counter_strategy(
                    _comp_name,
                    {
                        "role": _first_role,
                        "city": _first_city,
                        "industry": industry_label,
                        "competitor_type": comp.get("competitor_type") or "",
                        "intensity": comp.get("hiring_activity") or "",
                        "ordinal": idx,
                    },
                )
                _comp_rows.append(
                    {
                        "name": _comp_name,
                        "industry": _flatten_value(comp.get("industry") or ""),
                        "size": _flatten_value(
                            comp.get("size", comp.get("employee_count") or "")
                        ),
                        "hiring_activity": _flatten_value(
                            comp.get(
                                "hiring_activity", comp.get("hiring_channels") or ""
                            )
                        ),
                        "overlap_score": _flatten_value(
                            comp.get("overlap_score", comp.get("overlap") or "")
                        ),
                        "counter": _counter,
                    }
                )
            elif isinstance(comp, str):
                _counter = insight_composer.compose_counter_strategy(
                    comp,
                    {
                        "role": _first_role,
                        "city": _first_city,
                        "industry": industry_label,
                        "ordinal": idx,
                    },
                )
                _comp_rows.append(
                    {
                        "name": comp,
                        "industry": "",
                        "size": "",
                        "hiring_activity": "",
                        "overlap_score": "",
                        "counter": _counter,
                    }
                )

        if _comp_rows:
            row = _write_subsection_header(ws, row, "Competitor Analysis")

            _optional_cols = [
                ("industry", "Industry"),
                ("size", "Size"),
                ("hiring_activity", "Hiring Activity"),
                ("overlap_score", "Overlap Score"),
            ]
            _populated_cols = [
                (key, label)
                for key, label in _optional_cols
                if any(str(r.get(key) or "").strip() for r in _comp_rows)
            ]

            headers = (
                ["Name"] + [label for _, label in _populated_cols] + ["Counter-Strategy"]
            )
            row = _write_table_header(ws, row, headers)

            for idx, r in enumerate(_comp_rows):
                values = (
                    [r["name"]]
                    + [r[key] for key, _ in _populated_cols]
                    + [r["counter"]]
                )
                row = _write_table_row(ws, row, values, alternate=idx % 2 == 1)

            row += 1

    # Market positioning summary
    market_pos = comp_intel.get("market_positioning", comp_intel.get("summary") or "")
    if market_pos:
        row = _write_subsection_header(ws, row, "Market Positioning")
        ws.merge_cells(
            start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
        )
        _market_pos_display = market_pos
        if isinstance(market_pos, dict) and market_pos.get("industry_sector"):
            # data_synthesizer.fuse_competitive_intelligence stores the raw
            # internal industry key here (e.g. "logistics_supply_chain") --
            # route it through the same canonical shared_utils.INDUSTRY_LABEL_MAP
            # every other industry label on this workbook resolves through,
            # rather than leaking the key verbatim ("Industry Sector:
            # logistics_supply_chain").
            _market_pos_display = dict(market_pos)
            _market_pos_display["industry_sector"] = INDUSTRY_LABEL_MAP.get(
                market_pos["industry_sector"],
                _humanize_snake_key(market_pos["industry_sector"]),
            )
        cell = ws.cell(
            row=row, column=COL_START, value=_flatten_value(_market_pos_display)
        )
        cell.font = _FONT_BODY
        cell.alignment = _ALIGN_WRAP
        row += 1

    row += 2

    # ── 4. Salary Intelligence ──
    salary_intel = synthesized.get("salary_intelligence", {})

    if salary_intel:
        row = _write_section_header(ws, row, "Salary Intelligence")

        headers = ["Role", "Min", "P25", "Median", "P75", "Max", "Confidence"]
        row = _write_table_header(ws, row, headers)

        salary_items = salary_intel
        if isinstance(salary_intel, dict):
            # Could be keyed by role name or be a list
            if isinstance(next(iter(salary_intel.values()), None), dict):
                salary_items = salary_intel
            else:
                salary_items = {"All Roles": salary_intel}

        for idx, (role_key, sal_data) in enumerate(
            salary_items.items()
            if isinstance(salary_items, dict)
            else enumerate(salary_items)
        ):
            if isinstance(sal_data, dict):
                role_name = sal_data.get("role", sal_data.get("title", role_key))
                confidence = _safe_num(
                    sal_data.get("confidence", sal_data.get("confidence_score", 0.5))
                )
                is_low_conf = confidence < 0.5

                values = [
                    role_name if isinstance(role_name, str) else str(role_name),
                    _fmt_currency(sal_data.get("min", sal_data.get("p10") or 0)),
                    _fmt_currency(sal_data.get("p25") or 0),
                    _fmt_currency(sal_data.get("median", sal_data.get("p50") or 0)),
                    _fmt_currency(sal_data.get("p75") or 0),
                    _fmt_currency(sal_data.get("max", sal_data.get("p90") or 0)),
                    f"{confidence:.0%}",
                ]

                # Highlight low-confidence rows
                row_fill = _FILL_AMBER_BG if is_low_conf else None
                fills_list = [row_fill] * len(values) if is_low_conf else None
                conf_font = [None] * 6 + [
                    _grade_font("C" if confidence >= 0.5 else "D")
                ]

                row = _write_table_row(
                    ws,
                    row,
                    values,
                    alternate=idx % 2 == 1,
                    fills=fills_list,
                    fonts=conf_font,
                )

        # Sources footnote
        sources = set()
        for sal_data in (
            salary_intel.values() if isinstance(salary_intel, dict) else []
        ):
            if isinstance(sal_data, dict):
                src = sal_data.get("sources", sal_data.get("source") or "")
                if src:
                    if isinstance(src, list):
                        sources.update(src)
                    else:
                        sources.add(str(src))
        if sources:
            row = _write_footnote(ws, row, f"Sources: {', '.join(sorted(sources))}")

    row += 2

    # ── 5. Market Demand ──
    market_demand = synthesized.get("job_market_demand", {})

    # S50: Load Google Trends from KB for search interest enrichment
    _gt_roles_for_demand: Dict[str, dict] = {}
    _kb_for_gt = data.get("_knowledge_base", {})
    if isinstance(_kb_for_gt, dict):
        _gt_kb_raw = _kb_for_gt.get("google_trends", {})
        if isinstance(_gt_kb_raw, dict):
            _gt_data_raw = _gt_kb_raw.get("data", _gt_kb_raw)
            if isinstance(_gt_data_raw, dict):
                _gt_roles_for_demand = _gt_data_raw.get("roles", {})

    if market_demand:
        row = _write_section_header(ws, row, "Market Demand by Role")

        headers = [
            "Role",
            "Postings",
            "Talent Pool",
            "Competition",
            "Temperature",
            "Trend",
            "Search Interest",
        ]
        row = _write_table_header(ws, row, headers)

        demand_items = market_demand
        if isinstance(market_demand, dict) and not all(
            isinstance(v, dict) for v in market_demand.values()
        ):
            demand_items = {"All": market_demand}

        for idx, (role_key, demand) in enumerate(
            demand_items.items()
            if isinstance(demand_items, dict)
            else enumerate(demand_items)
        ):
            if isinstance(demand, dict):
                role_name = demand.get("role", demand.get("title", role_key))

                # S50: Extract search interest from Google Trends
                _search_interest_str = ""
                _st = demand.get("search_trend", {})
                if isinstance(_st, dict) and _st.get("current_interest"):
                    _ci = _st["current_interest"]
                    _td = _st.get("trend_direction", "")
                    _tc = _st.get("trend_change_pct")
                    _search_interest_str = f"{_ci}/100"
                    if _td:
                        _search_interest_str += f" ({_td}"
                        if _tc is not None:
                            _search_interest_str += f" {_tc:+.1f}%"
                        _search_interest_str += ")"
                # KB fallback: match role to GT roles dict
                if not _search_interest_str and _gt_roles_for_demand:
                    _rn_str = str(role_name) if role_name else str(role_key)
                    _rn_lower = _rn_str.lower()
                    for _gtk, _gtv in _gt_roles_for_demand.items():
                        if isinstance(_gtv, dict) and (
                            _rn_lower in _gtk.lower() or _gtk.lower() in _rn_lower
                        ):
                            _ci = _gtv.get("current_interest")
                            _td = _gtv.get("trend_direction", "")
                            _tc = _gtv.get("trend_change_pct")
                            if _ci is not None:
                                _search_interest_str = f"{_ci}/100"
                                if _td:
                                    _search_interest_str += f" ({_td}"
                                    if _tc is not None:
                                        _search_interest_str += f" {_tc:+.1f}%"
                                    _search_interest_str += ")"
                            break

                # S4: fuse_job_market_demand() (data_synthesizer.py) writes
                # "total_postings"/"talent_pool_estimate" -- this table was
                # reading "postings"/"job_postings" and "talent_pool"/"supply",
                # none of which exist in that dict, so it always rendered 0
                # regardless of the real (or fabricated-fallback) value the
                # deck showed for the same role. Read the actual keys, with
                # the old names kept as a compatibility fallback only.
                _postings_val = demand.get(
                    "total_postings",
                    demand.get("postings", demand.get("job_postings") or 0),
                )
                _talent_pool_val = demand.get(
                    "talent_pool_estimate",
                    demand.get("talent_pool", demand.get("supply") or 0),
                )
                # Never present a fabricated industry-benchmark fallback number
                # as if it were a measured "Live" postings count (S4: no
                # fabricated stats over empty data).
                _posting_sources = demand.get("posting_sources") or []
                _is_fabricated_postings = "Industry Benchmark" in _posting_sources
                values = [
                    role_name if isinstance(role_name, str) else str(role_name),
                    "Data not available"
                    if (not _postings_val or _is_fabricated_postings)
                    else _fmt_number(_postings_val),
                    _fmt_number(_talent_pool_val),
                    _flatten_value(
                        demand.get("competition", demand.get("competition_level") or "")
                    ),
                    _flatten_value(
                        demand.get(
                            "temperature", demand.get("market_temperature") or ""
                        )
                    ),
                    _flatten_value(
                        demand.get("trend", demand.get("trend_direction") or "")
                    ),
                    _search_interest_str or "—",
                ]
                row = _write_table_row(ws, row, values, alternate=idx % 2 == 1)

    row += 2

    # ── 6. Workforce Trends ──
    workforce = synthesized.get("workforce_insights", {})

    if workforce:
        row = _write_section_header(ws, row, "Workforce Trends")

        # CRITICAL: Properly flatten nested structures -- never use str() on dicts
        for section_key, section_val in workforce.items():
            if section_key in ("metadata", "source", "sources", "confidence"):
                continue

            section_label = section_key.replace("_", " ").title()

            if isinstance(section_val, dict):
                row = _write_subsection_header(ws, row, section_label)
                for k, v in section_val.items():
                    if k in ("metadata", "source"):
                        continue
                    val_str = _flatten_value(v)
                    if val_str:
                        row = _write_kv_row(
                            ws, row, k.replace("_", " ").title(), val_str
                        )
                row += 1

            elif isinstance(section_val, list):
                row = _write_subsection_header(ws, row, section_label)
                for item in section_val[:8]:
                    _item_for_display = item
                    if (
                        section_key == "relevant_research"
                        and isinstance(item, dict)
                        and "key" in item
                    ):
                        # data_synthesizer.fuse_workforce_insights carries a
                        # "key" field (the raw KB source key, e.g.
                        # "appcast_benchmark_2023") purely as an internal
                        # identifier -- it duplicates "title"/"publisher"/
                        # "year", which is what a client should actually
                        # see, so drop it rather than leak "Key:
                        # appcast_benchmark_2023" into the workbook.
                        _item_for_display = {
                            k: v for k, v in item.items() if k != "key"
                        }
                    val_str = _flatten_value(_item_for_display)
                    if val_str:
                        ws.merge_cells(
                            start_row=row,
                            start_column=COL_START,
                            end_row=row,
                            end_column=COL_END,
                        )
                        cell = ws.cell(
                            row=row, column=COL_START, value=f"  - {val_str}"
                        )
                        cell.font = _FONT_BODY
                        cell.alignment = _ALIGN_WRAP
                        row += 1
                row += 1

            elif isinstance(section_val, (str, int, float, bool)):
                row = _write_kv_row(ws, row, section_label, _flatten_value(section_val))

    # ── 7. LinkedIn Benchmarks (SlotOps 108K dataset) ──
    li_intel = (data.get("_gold_standard") or {}).get("linkedin_intelligence", {})
    if not li_intel:
        # Fallback: check direct injection from slotops_engine
        li_intel = data.get("_slotops_linkedin_benchmarks", {})

    if li_intel and li_intel.get("country_apply_rate"):
        row = _write_section_header(ws, row, "LinkedIn Benchmarks")
        row = _write_footnote(
            ws,
            row,
            f"Based on {_fmt_number(li_intel.get('total_jobs_analyzed', li_intel.get('sample_size', 108871)))} "
            f"LinkedIn job postings across {li_intel.get('countries_covered', 76)} countries "
            f"(Joveo SlotOps dataset)",
        )
        row += 1

        # Country-level apply rates
        country_ar = li_intel.get("country_apply_rate", {})
        country_name = li_intel.get("country", "United States")
        row = _write_subsection_header(ws, row, f"Apply Rates: {country_name}")
        ar_fields = [
            ("Average Apply Rate", f"{country_ar.get('avg', 0):.1f}%"),
            ("Median Apply Rate", f"{country_ar.get('median', 0):.1f}%"),
            ("75th Percentile", f"{country_ar.get('p75', 0):.1f}%"),
        ]
        p90 = country_ar.get("p90", 0)
        if p90:
            ar_fields.append(("90th Percentile", f"{p90:.1f}%"))
        sample = li_intel.get("sample_size", 0)
        if sample:
            ar_fields.append(("Sample Size", _fmt_number(sample)))
        avg_views = li_intel.get("avg_views", 0)
        if avg_views:
            ar_fields.append(("Avg Views per Posting", _fmt_number(avg_views)))
        avg_days = li_intel.get("avg_days_open", 0)
        if avg_days:
            ar_fields.append(("Avg Days Open", f"{avg_days:.1f}"))
        for key, val in ar_fields:
            row = _write_kv_row(ws, row, key, val)
        row += 1

        # Easy Apply vs ATS
        ea_ats = li_intel.get("ea_vs_ats", {})
        if ea_ats and ea_ats.get("easy_apply_rate"):
            ea_scope = ea_ats.get("scope", "global")
            scope_label = f" ({country_name})" if ea_scope == "country" else " (Global)"
            row = _write_subsection_header(ws, row, f"Easy Apply vs ATS{scope_label}")
            headers = ["Apply Type", "Apply Rate", "Sample Size", "Lift Factor"]
            row = _write_table_header(ws, row, headers)

            ea_rate = ea_ats.get("easy_apply_rate", 0)
            ats_rate = ea_ats.get("ats_rate", 0)
            lift = ea_ats.get("lift_factor", 0)

            row = _write_table_row(
                ws,
                row,
                [
                    "Easy Apply",
                    f"{ea_rate:.1f}%",
                    _fmt_number(ea_ats.get("easy_apply_sample", 0)),
                    f"{lift:.2f}x" if lift else "",
                ],
                alternate=False,
            )
            row = _write_table_row(
                ws,
                row,
                [
                    "ATS (Standard)",
                    f"{ats_rate:.1f}%",
                    _fmt_number(ea_ats.get("ats_sample", 0)),
                    "1.00x (baseline)",
                ],
                alternate=True,
            )

            rec = ea_ats.get("recommendation", "")
            if rec:
                row = _write_footnote(ws, row + 1, f"Recommendation: {rec}")
            row += 1

        # Best posting days
        best_days = li_intel.get("best_posting_days", [])
        if best_days:
            row = _write_subsection_header(ws, row, "Optimal Posting Schedule")
            row = _write_kv_row(ws, row, "Best Posting Days", ", ".join(best_days))
            refresh = li_intel.get("refresh_cadence_days", [])
            if refresh:
                row = _write_kv_row(
                    ws,
                    row,
                    "Recommended Refresh Cadence",
                    f"Every {refresh[0]}-{refresh[-1]} days",
                )
            row += 1

        # Role-specific benchmarks
        role_benchmarks = li_intel.get("role_benchmarks", [])
        if role_benchmarks:
            row = _write_subsection_header(
                ws, row, "Role-Specific LinkedIn Performance"
            )
            headers = [
                "Target Role",
                "Matched Title",
                "Apply Rate",
                "Avg Views",
                "Sample",
            ]
            row = _write_table_header(ws, row, headers)
            for idx, rb in enumerate(role_benchmarks):
                values = [
                    rb.get("role", ""),
                    rb.get("matched_title", ""),
                    f"{rb.get('apply_rate_avg', 0):.1f}%",
                    _fmt_number(rb.get("avg_views", 0)),
                    _fmt_number(rb.get("sample_size", 0)),
                ]
                row = _write_table_row(ws, row, values, alternate=idx % 2 == 1)
            row += 1

        row += 1

    # ── 8. Geographic CPC Variance ──
    if len(locations) > 1:
        try:
            from feature_store import get_feature_store

            fs = get_feature_store()
            row += 2
            row = _write_section_header(ws, row, "Geographic Cost Variance")

            geo_headers = ["Market", "Cost Index", "CPC Adjustment", "Impact"]
            row = _write_table_header(ws, row, geo_headers)

            # S89A FIX (finding strategy:manpower#7): ``get_geo_cost_index``
            # returns exactly 1.0 as its flat DEFAULT when a location has no
            # real metro match (feature_store.py, "Default to national
            # average") -- 2+ locations landing on that same default and
            # being presented as distinct "metro-area" rows overstates the
            # sourcing this table claims. Collapse the default-index
            # locations into one honest row and soften the sourcing claim
            # when any are present; locations with a real metro match still
            # get their own row and keep the sourced claim.
            _geo_idx_by_loc = [(loc, fs.get_geo_cost_index(loc)) for loc in locations]
            _fallback_geo_locs = [
                loc for loc, idx in _geo_idx_by_loc if abs(idx - 1.0) < 1e-9
            ]
            _collapse_geo = len(_fallback_geo_locs) >= 2

            _write_idx = 0
            _collapsed_geo_written = False
            for loc, geo_idx in _geo_idx_by_loc:
                is_fallback = abs(geo_idx - 1.0) < 1e-9
                if is_fallback and _collapse_geo:
                    if _collapsed_geo_written:
                        continue
                    loc_label = (
                        "All listed markets (default index — market-level "
                        "data pending enrichment)"
                    )
                    _collapsed_geo_written = True
                else:
                    loc_label = loc
                if geo_idx >= 1.2:
                    impact = "Premium market (+20%+ costs)"
                elif geo_idx >= 1.05:
                    impact = "Above-average costs"
                elif geo_idx >= 0.95:
                    impact = "Average market rate"
                else:
                    impact = "Below-average costs"
                values = [
                    loc_label,
                    f"{geo_idx:.2f}x",
                    f"{(geo_idx - 1) * 100:+.0f}%",
                    impact,
                ]
                row = _write_table_row(
                    ws, row, values, alternate=_write_idx % 2 == 1
                )
                _write_idx += 1

            _geo_footnote = "Cost indices are relative to the national average (1.00x)."
            if _collapse_geo:
                _geo_footnote += (
                    ' "All listed markets" is the national-default index '
                    "(confidence: estimated) -- no metro-specific cost data was "
                    "available to differentiate these locations."
                )
            else:
                _geo_footnote += (
                    " Based on metro-area hiring cost data from Joveo and "
                    "validated industry sources."
                )
            row = _write_footnote(ws, row + 1, _geo_footnote)
        except ImportError:
            logger.warning("feature_store not available; skipping geographic variance")
        except Exception as exc:
            logger.warning("Geographic CPC variance section failed: %s", exc)

    row += 2
    _write_attribution_footer(ws, row)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: SOURCES & DATA CONFIDENCE
# ═══════════════════════════════════════════════════════════════════════════════

# Provenance confidence labels -> 0-1 score (for grade/colour mapping).
_PROVENANCE_CONFIDENCE_SCORE: Dict[str, float] = {
    "measured": 0.95,
    "verified": 0.9,
    "high": 0.9,
    "good": 0.8,
    "medium": 0.65,
    "moderate": 0.65,
    "estimated": 0.55,
    "low": 0.45,
    "fallback": 0.4,
    "unknown": 0.5,
}


def _provenance_confidence_label(raw: Any) -> str:
    """Normalize a provenance confidence value to a display label.

    Accepts strings ("measured", "high", ...) or numeric 0-1 scores and
    returns a Title-cased label. Defensive: unknown/empty -> "Estimated".
    """
    if raw is None or raw == "":
        return "Estimated"
    if isinstance(raw, (int, float)):
        score = float(raw)
        if score > 1:  # tolerate 0-100 scale
            score /= 100.0
        return _label_from_score(score)
    return str(raw).strip().title()


def _label_from_score(score: float) -> str:
    """0-1 score -> coarse confidence label."""
    if score >= 0.9:
        return "Verified"
    if score >= 0.65:
        return "High"
    if score >= 0.5:
        return "Medium"
    return "Low"


def _provenance_conf_font(label: str) -> Font:
    """Colour-code a provenance confidence label (reuses grade colours)."""
    score = _PROVENANCE_CONFIDENCE_SCORE.get(label.strip().lower(), 0.55)
    return _grade_font(_grade_from_score(score))


def _collect_kb_provenance(data: dict) -> List[Dict[str, str]]:
    """Collect (source, vintage, confidence) provenance rows from the data dict.

    The S89 KB tags figures with source / vintage / confidence. This reads
    those defensively from several shapes that may appear in the pipeline:

    - ``data["_enriched"]["provenance"]`` (explicit list or section->meta dict)
    - ``data["_synthesized"]["provenance"]`` (same shape)
    - per-section ``metadata`` blocks carrying source/vintage/confidence under
      ``_enriched`` / ``_synthesized``

    Returns a de-duplicated list of ``{"source","vintage","confidence"}`` dicts.
    Never raises — on any unexpected shape it simply yields what it can.
    """
    rows: List[Dict[str, str]] = []
    seen: set = set()

    _vintage_fields = (
        "vintage",
        "data_year",
        "benchmark_year",
        "data_coverage_period",
        "data_period",
        "as_of",
        "last_updated",
        "year",
    )

    def _vintage_of(meta: dict) -> str:
        for f in _vintage_fields:
            v = meta.get(f)
            if v not in (None, ""):
                return str(v)
        return "—"

    def _add(source: str, vintage: str, confidence: Any) -> None:
        source = (source or "").strip()
        if not source:
            return
        key = source.lower()
        if key in seen:
            return
        seen.add(key)
        rows.append(
            {
                "source": source,
                "vintage": str(vintage or "—"),
                "confidence": _provenance_confidence_label(confidence),
            }
        )

    def _ingest(container: Any) -> None:
        """Ingest an explicit provenance container (list or dict)."""
        if isinstance(container, list):
            for item in container:
                if isinstance(item, dict):
                    _add(
                        item.get("source")
                        or item.get("name")
                        or item.get("dataset")
                        or "",
                        _vintage_of(item) if isinstance(item, dict) else "—",
                        item.get("confidence"),
                    )
                elif isinstance(item, str):
                    _add(item, "—", None)
        elif isinstance(container, dict):
            for sect_key, meta in container.items():
                if not isinstance(meta, dict):
                    # source -> confidence string mapping
                    _add(str(sect_key), "—", meta)
                    continue
                _add(
                    meta.get("source") or str(sect_key),
                    _vintage_of(meta),
                    meta.get("confidence"),
                )

    for top in ("_enriched", "_synthesized"):
        block = data.get(top)
        if not isinstance(block, dict):
            continue
        # Explicit provenance container, if the pipeline supplied one.
        _ingest(block.get("provenance"))
        _ingest(block.get("sources"))
        # Per-section metadata blocks that carry a data vintage / source.
        for sect_key, sect_val in block.items():
            if sect_key.startswith("_") or sect_key in ("provenance", "sources"):
                continue
            if not isinstance(sect_val, dict):
                continue
            meta = sect_val.get("metadata")
            if isinstance(meta, dict) and (meta.get("source") or meta.get("vintage")):
                _add(
                    meta.get("source") or sect_key.replace("_", " ").title(),
                    _vintage_of(meta),
                    meta.get("confidence"),
                )

    return rows


def _build_provenance_section(ws, data: dict, row: int) -> int:
    """Additive S89 provenance block: data sources, vintage, confidence.

    Surfaces, on the "Sources & Confidence" sheet:
      1. KB / enrichment provenance — which data sources fed the plan, their
         data vintage, and the confidence tag the KB attached.
      2. Live API enrichment summary — how many real-time sources succeeded.
      3. The "Joveo measured" warehouse signal (cg_benchmarks real outcomes),
         when ``_budget_allocation.metadata.real_outcomes`` is present.

    Purely additive — it appends rows; it does not alter any existing table's
    numbers. Read defensively; never raises (caller already wraps in try/except,
    but we also guard here so one bad shape never blanks the section).
    """
    try:
        return _build_provenance_section_inner(ws, data, row)
    except Exception as exc:  # pragma: no cover - defensive
        logger.warning("Provenance section skipped: %s", exc)
        return row


def _build_provenance_section_inner(ws, data: dict, row: int) -> int:
    row = _write_section_header(ws, row, "Data Provenance & Vintage")

    intro = ws.cell(
        row=row,
        column=COL_START,
        value=(
            "Every figure in this plan is traceable to a source. The table below "
            "lists the data sources used, the vintage (period the data describes), "
            "and the confidence tag attached at ingestion."
        ),
    )
    ws.merge_cells(
        start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
    )
    intro.font = _FONT_FOOTNOTE
    intro.alignment = _ALIGN_WRAP
    ws.row_dimensions[row].height = 28
    row += 2

    # ── 1. Data sources / vintage / confidence ──
    prov_rows = _collect_kb_provenance(data)

    # Always surface the live-enrichment summary as a source line too.
    enriched = data.get("_enriched", {}) if isinstance(data.get("_enriched"), dict) else {}
    summary = enriched.get("enrichment_summary", {})
    if isinstance(summary, dict):
        succeeded = summary.get("apis_succeeded") or []
        if isinstance(succeeded, list) and succeeded:
            n_ok = len(succeeded)
            n_called = len(summary.get("apis_called") or []) or n_ok
            prov_rows.append(
                {
                    "source": f"Live market APIs ({n_ok}/{n_called} responded)",
                    "vintage": str(datetime.date.today().year),
                    "confidence": "High" if n_ok >= max(1, n_called * 0.6) else "Medium",
                }
            )

    # KB age / freshness (from kb_loader) as a freshness signal.
    synthesized = data.get("_synthesized", {}) if isinstance(data.get("_synthesized"), dict) else {}
    kb_age_days = synthesized.get("_kb_age_days")
    if isinstance(kb_age_days, (int, float)):
        prov_rows.append(
            {
                "source": "Joveo Knowledge Base (benchmarks)",
                "vintage": f"{kb_age_days:.0f} days old",
                "confidence": "High" if kb_age_days <= 60 else (
                    "Medium" if kb_age_days <= 90 else "Low"
                ),
            }
        )

    if prov_rows:
        headers = ["Data Source", "Vintage", "Confidence"]
        row = _write_table_header(ws, row, headers)
        for idx, pr in enumerate(prov_rows):
            conf_label = pr["confidence"]
            values = [pr["source"][:70], pr["vintage"][:30], conf_label]
            fonts_list = [None, None, _provenance_conf_font(conf_label)]
            row = _write_table_row(
                ws,
                row,
                values,
                alternate=idx % 2 == 1,
                fonts=fonts_list,
            )
        row += 1
    else:
        row = _write_footnote(
            ws,
            row,
            "Source provenance metadata was not attached to this plan's data. "
            "Figures rely on the curated Knowledge Base and validated benchmarks.",
        )
        row += 1

    # ── 2. Joveo Campaign Warehouse (measured outcomes) ──
    row = _build_warehouse_provenance(ws, data, row)

    return row


def _build_warehouse_provenance(ws, data: dict, row: int) -> int:
    """Surface the cg_benchmarks "Joveo measured" signal, if present.

    Reads ``data["_budget_allocation"]["metadata"]["real_outcomes"]`` defensively
    — it may be absent (no warehouse coverage) in which case nothing is written.
    When present, shows the matched title, measured cost-per-apply and the
    sample size behind it, attributed to "Joveo Campaign Warehouse".
    """
    ba = data.get("_budget_allocation", {})
    if not isinstance(ba, dict):
        return row
    meta = ba.get("metadata", {})
    if not isinstance(meta, dict):
        return row
    ro = meta.get("real_outcomes")
    # Accept a single dict or a list of measured-outcome dicts.
    if isinstance(ro, dict):
        outcomes = [ro]
    elif isinstance(ro, list):
        outcomes = [o for o in ro if isinstance(o, dict)]
    else:
        return row

    # Keep only matched, usable rows.
    matched = [o for o in outcomes if o.get("matched") and o.get("title")]
    if not matched:
        return row

    row = _write_subsection_header(
        ws, row, "Joveo Measured Outcomes (Campaign Warehouse)"
    )

    note = ws.cell(
        row=row,
        column=COL_START,
        value=(
            "First-party measured performance from real Joveo campaigns "
            "(cg_benchmarks). Where a role matches, these measured figures take "
            "precedence over modelled estimates."
        ),
    )
    ws.merge_cells(
        start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
    )
    note.font = _FONT_FOOTNOTE
    note.alignment = _ALIGN_WRAP
    ws.row_dimensions[row].height = 28
    row += 2

    # S3: cg_benchmarks is Joveo's first-party USD campaign warehouse (not the
    # current plan's own currency) -- mark it explicitly rather than a bare $.
    headers = [
        "Role / Title",
        "Cost per Apply (USD)",
        "Sample Size",
        "Last Updated",
        "Source",
    ]
    row = _write_table_header(ws, row, headers)

    # Per-column number formats: cost-per-apply as $ per-unit, sample as int.
    _fmts = [None, FMT_USD2, FMT_INT, None, None]

    for idx, o in enumerate(matched):
        cpa = o.get("cost_per_apply")
        sample = o.get("sample_size")
        values = [
            str(o.get("title") or "")[:60],
            cpa if cpa is not None else "—",
            sample if sample not in (None, "") else "—",
            str(o.get("last_updated") or "—"),
            "Joveo Campaign Warehouse",
        ]
        # Only attach a numeric format where the cell holds a real number;
        # the "—" placeholders stay as text.
        row_fmts = [
            _fmts[i] if (i in (1, 2) and not isinstance(values[i], str)) else None
            for i in range(len(values))
        ]
        row = _write_table_row(
            ws,
            row,
            values,
            alternate=idx % 2 == 1,
            number_formats=row_fmts,
        )

    row += 1
    return row


def _build_sheet_sources(ws, data: dict):
    """Build Sheet 4: Sources & Data Confidence."""
    ws.title = "Sources & Confidence"
    ws.sheet_properties.tabColor = SAPPHIRE

    _set_column_widths(
        ws,
        {
            1: 3,
            2: 22,
            3: 14,
            4: 14,
            5: 14,
            6: 14,
            7: 14,
            8: 18,
        },
    )

    synthesized = data.get("_synthesized", {})
    enriched = data.get("_enriched", {})
    confidence_scores = synthesized.get("confidence_scores", {})
    data_quality = synthesized.get("data_quality", {})
    enrichment_summary = enriched.get("enrichment_summary", {})

    row = 2

    # ── 1. Overall Confidence Grade ──
    row = _write_section_header(ws, row, "Data Confidence Assessment")

    overall_score = _safe_num(
        confidence_scores.get(
            "overall", confidence_scores.get("overall_confidence", 0.5)
        )
    )
    overall_grade = _grade_from_score(overall_score)

    # Store computed confidence so PPT uses the same value
    data["_computed_confidence_pct"] = round(overall_score * 100)

    # Large grade display
    ws.merge_cells(
        start_row=row, start_column=COL_START, end_row=row + 2, end_column=COL_START + 1
    )
    grade_cell = ws.cell(row=row, column=COL_START, value=overall_grade)
    grade_cell.font = _FONT_GRADE_LARGE
    if overall_grade in ("A", "B"):
        grade_cell.fill = PatternFill(
            start_color=GREEN, end_color=GREEN, fill_type="solid"
        )
    elif overall_grade == "C":
        grade_cell.fill = PatternFill(
            start_color=AMBER, end_color=AMBER, fill_type="solid"
        )
    else:
        grade_cell.fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    grade_cell.alignment = _ALIGN_CENTER

    # Grade description next to it
    ws.merge_cells(
        start_row=row, start_column=COL_START + 2, end_row=row, end_column=COL_END
    )
    desc_cell = ws.cell(
        row=row, column=COL_START + 2, value=f"Overall Confidence: {overall_score:.0%}"
    )
    desc_cell.font = _FONT_HERO
    desc_cell.alignment = _ALIGN_LEFT

    ws.merge_cells(
        start_row=row + 1,
        start_column=COL_START + 2,
        end_row=row + 1,
        end_column=COL_END,
    )
    quality_msg = data_quality.get("message", data_quality.get("summary") or "")
    if not quality_msg:
        if overall_grade in ("A", "B"):
            quality_msg = "High-quality data from multiple verified sources"
        elif overall_grade == "C":
            quality_msg = "Moderate data quality — some sections rely on benchmarks"
        else:
            quality_msg = "Limited data availability — results should be validated"

    qual_cell = ws.cell(row=row + 1, column=COL_START + 2, value=quality_msg)
    qual_cell.font = _FONT_BODY
    qual_cell.alignment = _ALIGN_WRAP

    # KB data freshness indicator
    kb_age_days = synthesized.get("_kb_age_days")
    freshness_warning = synthesized.get("_data_freshness_warning")
    if kb_age_days is not None:
        ws.merge_cells(
            start_row=row + 2,
            start_column=COL_START + 2,
            end_row=row + 2,
            end_column=COL_END,
        )
        age_label = f"Knowledge Base Age: {kb_age_days:.0f} days"
        if freshness_warning:
            age_label += f"  —  {freshness_warning}"
        age_cell = ws.cell(row=row + 2, column=COL_START + 2, value=age_label)
        age_cell.alignment = _ALIGN_WRAP
        if kb_age_days > 90:
            age_cell.font = Font(name=FONT_BODY_NAME, size=10, color=RED, italic=True)
        elif kb_age_days > 60:
            age_cell.font = Font(name=FONT_BODY_NAME, size=10, color=AMBER, italic=True)
        else:
            age_cell.font = Font(name=FONT_BODY_NAME, size=10, color=GREEN, italic=True)

    row += 4

    # ── 2. Per-Section Confidence ──
    section_scores = confidence_scores.get(
        "sections", confidence_scores.get("per_section", {})
    )

    if section_scores and isinstance(section_scores, dict):
        row = _write_section_header(ws, row, "Per-Section Confidence")

        headers = ["Section", "Score", "Grade", "Sources"]
        row = _write_table_header(ws, row, headers)

        for idx, (section, score_data) in enumerate(section_scores.items()):
            if isinstance(score_data, dict):
                score = _safe_num(
                    score_data.get("score", score_data.get("confidence") or 0)
                )
                sources = score_data.get(
                    "sources", score_data.get("data_sources") or []
                )
                sources_str = _flatten_value(sources) if sources else ""
            elif isinstance(score_data, (int, float)):
                score = float(score_data)
                sources_str = ""
            else:
                continue

            grade = _grade_from_score(score)
            values = [
                section.replace("_", " ").title(),
                f"{score:.0%}",
                grade,
                sources_str[:60],
            ]

            grade_f = _grade_fill(grade)
            g_font = _grade_font(grade)
            fills_list = [None, None, grade_f, None]
            fonts_list = [None, None, g_font, None]
            row = _write_table_row(
                ws,
                row,
                values,
                alternate=idx % 2 == 1,
                fills=fills_list,
                fonts=fonts_list,
            )

    row += 2

    # ── 3. Source Assessment ── REMOVED from client output (S50)
    # API names, bias analysis, and source lists are internal-only.

    # ── 3/4. Source Assessment & API Status Report ── REMOVED (S50)
    # API names, source lists, and status reports are internal-only.
    # Client sees confidence grades and methodology only.

    # ── 4b. Location Plausibility Warnings (S50) ──
    loc_warnings = synthesized.get("_validation", {}).get("location_warnings") or []
    if loc_warnings:
        row = _write_section_header(ws, row, "Location Plausibility Warnings")

        # Explanation note
        ws.merge_cells(
            start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
        )
        note = ws.cell(
            row=row,
            column=COL_START,
            value=(
                "The following locations may not align with the company's known "
                "operating area. These are advisory warnings — they do not block "
                "plan generation. Please verify before finalizing."
            ),
        )
        note.font = _FONT_FOOTNOTE
        note.alignment = _ALIGN_WRAP
        row += 2

        headers = ["Location", "Severity", "Reason", "Known Locations", "Suggestion"]
        row = _write_table_header(ws, row, headers)

        for idx, warn in enumerate(loc_warnings):
            severity = (warn.get("severity") or "medium").capitalize()

            if severity == "High":
                sev_fill = _FILL_RED_BG
                sev_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=RED)
            elif severity == "Medium":
                sev_fill = _FILL_AMBER_BG
                sev_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=AMBER)
            else:
                sev_fill = _FILL_BLUE_PALE
                sev_font = Font(name=FONT_BODY_NAME, size=10, color=SAPPHIRE)

            # S5 (2026-07-03, finding 39): never hard-truncate mid-word --
            # write the full text and wrap/tallen the row instead.
            reason_val = warn.get("reason") or ""
            values = [
                warn.get("location", ""),
                severity,
                reason_val,
                warn.get("known_states_display") or "—",
                warn.get("suggestion") or "",
            ]
            fills_list = [None, sev_fill, None, None, None]
            fonts_list = [None, sev_font, None, None, None]
            row = _write_table_row(
                ws,
                row,
                values,
                alternate=idx % 2 == 1,
                fills=fills_list,
                fonts=fonts_list,
            )
            if len(reason_val) > 60:
                ws.row_dimensions[row - 1].height = 40

        # Company HQ line
        first_hq = loc_warnings[0].get("company_hq") or "Unknown"
        row += 1
        ws.merge_cells(
            start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
        )
        hq_cell = ws.cell(
            row=row,
            column=COL_START,
            value=f"Company HQ (verified): {first_hq}",
        )
        hq_cell.font = _FONT_FOOTNOTE
        hq_cell.alignment = _ALIGN_WRAP
        row += 2

    # ── 5. Plan Validation Results ──
    # S1 (2026-07-03): this validation banner ("N checks run | N passed | ...")
    # is an internal pipeline-QA artifact, not a client-facing statement --
    # gate it OFF by default so it never ships in a client bundle.
    validation = data.get("_validation", {})
    val_findings = validation.get("findings") or []
    val_checks_run = validation.get("checks_run", 0)
    val_checks_failed = validation.get("checks_failed", 0)
    val_auto_corrections = validation.get("auto_corrections", 0)

    if _internal_qc_mode(data) and (val_checks_run > 0 or val_checks_failed > 0):
        row = _write_section_header(ws, row, "Plan Validation Results")

        # Summary line: X checks, Y passed, Z findings, W auto-corrected
        total_checks = val_checks_run + val_checks_failed
        passed = val_checks_run - min(
            val_checks_run,
            len([f for f in val_findings if f.get("severity") == "error"]),
        )
        summary_text = (
            f"{total_checks} checks run  |  "
            f"{passed} passed  |  "
            f"{len(val_findings)} finding(s)  |  "
            f"{val_auto_corrections} auto-corrected"
        )
        if val_checks_failed > 0:
            summary_text += f"  |  {val_checks_failed} check(s) errored"

        ws.merge_cells(
            start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
        )
        sum_cell = ws.cell(row=row, column=COL_START, value=summary_text)
        sum_cell.font = _FONT_BODY_BOLD
        sum_cell.alignment = _ALIGN_LEFT
        row += 2

        if val_findings:
            headers = ["Check", "Severity", "Description", "Auto-Corrected"]
            row = _write_table_header(ws, row, headers)

            for idx, finding in enumerate(val_findings):
                sev = (finding.get("severity") or "info").capitalize()
                auto = "Yes" if finding.get("auto_corrected") else "No"
                values = [
                    (finding.get("check") or "").replace("_", " ").title(),
                    sev,
                    (finding.get("message") or "")[:80],
                    auto,
                ]

                # Color-code severity
                if sev in ("Error", "High"):
                    sev_fill = _FILL_RED_BG
                    sev_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=RED)
                elif sev in ("Warning", "Medium"):
                    sev_fill = _FILL_AMBER_BG
                    sev_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=AMBER)
                else:
                    sev_fill = _FILL_GREEN_BG
                    sev_font = Font(name=FONT_BODY_NAME, size=10, color=GREEN)

                fills_list = [None, sev_fill, None, None]
                fonts_list = [None, sev_font, None, None]
                row = _write_table_row(
                    ws,
                    row,
                    values,
                    alternate=idx % 2 == 1,
                    fills=fills_list,
                    fonts=fonts_list,
                )

            row += 1

    # ── 5b. Data Provenance & Vintage (S89: end-to-end provenance) ──
    row = _build_provenance_section(ws, data, row)
    row += 1

    # ── 6. Methodology Notes ──
    row = _write_section_header(ws, row, "Methodology & Data Hierarchy")

    methodology_items = [
        (
            "Priority 1: Client Data",
            "Client-provided data (uploaded briefs, historical campaign data) takes highest precedence.",
        ),
        (
            "Priority 2: Real-Time Market Data",
            "Real-time data from multiple validated government, industry, and market sources "
            "provides current market signals.",
        ),
        (
            "Priority 3: Industry Benchmarks",
            "Curated industry benchmarks and validated reports provide "
            "baseline data for cost and performance estimates.",
        ),
        (
            "Priority 4: Curated Fallbacks",
            "Embedded fallback data ensures coverage when real-time sources "
            "are temporarily unavailable.",
        ),
    ]

    for key, desc in methodology_items:
        row = _write_kv_row(ws, row, key, desc)

    row += 1

    # Data quality note
    ws.merge_cells(
        start_row=row, start_column=COL_START, end_row=row, end_column=COL_END
    )
    note_cell = ws.cell(
        row=row,
        column=COL_START,
        value="Note: Data is sourced from government agencies, independent research bodies, "
        "and validated industry benchmarks. Vendor-originated data receives lower "
        "confidence weighting to reduce potential bias.",
    )
    note_cell.font = _FONT_FOOTNOTE
    note_cell.alignment = _ALIGN_WRAP
    row += 2

    # Geopolitical context (if available)
    geo_context = synthesized.get("geopolitical_context", {})
    if geo_context and isinstance(geo_context, dict):
        row = _write_subsection_header(ws, row, "Geopolitical Context")
        for key, val in geo_context.items():
            if key in ("metadata", "source"):
                continue
            val_str = _flatten_value(val)
            if val_str:
                row = _write_kv_row(ws, row, key.replace("_", " ").title(), val_str)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: ROI PROJECTIONS
# ═══════════════════════════════════════════════════════════════════════════════

# Channel-type conversion rates (application-to-hire %)
_ROI_CONVERSION_RATES: Dict[str, Tuple[float, float]] = {
    "job_board": (0.08, 0.12),
    "programmatic": (0.05, 0.08),
    "social": (0.03, 0.06),
    "niche_board": (0.10, 0.15),
    "regional": (0.06, 0.10),
    "search": (0.05, 0.08),
    "display": (0.03, 0.06),
    "employer_branding": (0.04, 0.08),
    "career_site": (0.08, 0.12),
    "referral": (0.15, 0.25),
    "events": (0.10, 0.18),
    "staffing": (0.12, 0.20),
    "email": (0.05, 0.10),
}

# Time-to-fill estimates by channel type (days)
_ROI_TIME_TO_FILL: Dict[str, Tuple[int, int]] = {
    "programmatic": (25, 35),
    "job_board": (30, 45),
    "social": (35, 50),
    "niche_board": (20, 30),
    "regional": (30, 45),
    "search": (30, 40),
    "display": (35, 50),
    "employer_branding": (40, 60),
    "career_site": (25, 40),
    "referral": (15, 25),
    "events": (30, 50),
    "staffing": (20, 35),
    "email": (30, 45),
}


def _roi_category_for_channel(channel_name: str) -> str:
    """Map a channel name to its ROI category key for conversion/time estimates."""
    name_lower = channel_name.lower()
    mapping: Dict[str, str] = {
        "programmatic": "programmatic",
        "dsp": "programmatic",
        "global_boards": "job_board",
        "global job": "job_board",
        "job board": "job_board",
        "indeed": "job_board",
        "niche": "niche_board",
        "specialty": "niche_board",
        "social": "social",
        "linkedin": "social",
        "meta": "social",
        "facebook": "social",
        "regional": "regional",
        "local": "regional",
        "employer_branding": "employer_branding",
        "employer brand": "employer_branding",
        "career_site": "career_site",
        "career site": "career_site",
        "referral": "referral",
        "event": "events",
        "staffing": "staffing",
        "agency": "staffing",
        "search": "search",
        "sem": "search",
        "display": "display",
        "banner": "display",
        "email": "email",
        "apac": "regional",
        "emea": "regional",
    }
    for keyword, category in mapping.items():
        if keyword in name_lower:
            return category
    return "job_board"


def _niche_board_implied_rate(data: dict) -> Tuple[Optional[float], int, int]:
    """This plan's own modeled niche-board apply-to-hire rate.

    Reads the SAME channel allocation ROI Projections' "Implied App-to-Hire
    Rate" table reads (never hardcoded) so the Niche Board Matching sheet's
    narrative always agrees with the plan's own modeled numbers, whatever
    the hire-distribution model currently produces for niche channels
    (findings strategy:manpower#8 / data:atria#7).

    Returns ``(rate_or_None, total_applications, total_hires)`` -- rate is
    ``None`` when the plan has no niche-board applications to compute a
    rate from.
    """
    budget_alloc = data.get("_budget_allocation", {})
    channel_allocs = (
        budget_alloc.get("channel_allocations", {})
        if isinstance(budget_alloc, dict)
        else {}
    )
    total_apps = 0
    total_hires = 0
    for ch_name, ch_data in (channel_allocs or {}).items():
        if not isinstance(ch_data, dict):
            continue
        category = ch_data.get("category") or _roi_category_for_channel(
            str(ch_name)
        )
        if category != "niche_board":
            continue
        total_apps += int(
            _safe_num(
                ch_data.get("projected_applications")
                or ch_data.get("projected_apps")
                or 0
            )
        )
        total_hires += int(_safe_num(ch_data.get("projected_hires") or 0))
    if total_apps <= 0:
        return None, total_apps, total_hires
    return total_hires / total_apps, total_apps, total_hires


# ---------------------------------------------------------------------------
# Role difficulty -> base time-to-fill adjustments (days)
# ---------------------------------------------------------------------------
_ROLE_DIFFICULTY_TTF: dict[str, tuple[int, int]] = {
    "executive": (60, 90),
    "c-suite": (60, 90),
    "vp": (60, 90),
    "director": (45, 75),
    "tech": (45, 60),
    "engineering": (45, 60),
    "software": (45, 60),
    "data_science": (45, 60),
    "cybersecurity": (45, 60),
    "nursing": (30, 45),
    "healthcare": (30, 45),
    "medical": (30, 45),
    "rn": (30, 45),
    "lpn": (30, 45),
    "cna": (25, 35),
    "hourly": (14, 21),
    "entry": (14, 21),
    "retail": (14, 21),
    "warehouse": (14, 21),
    "food_service": (14, 21),
    "cdl": (21, 30),
    "trucking": (21, 30),
    "driver": (21, 30),
    "logistics": (21, 30),
}


def _compute_dynamic_ttf(channel_base_ttf: int, data: dict) -> int:
    """Compute dynamic time-to-fill by adjusting channel base with role/volume/market factors.

    Factors applied:
    - Role difficulty: the plan's own roles, scored by
      gold_standard.classify_difficulty() (the same model driving the
      Quality Intelligence sheet's Role Difficulty Classification table) --
      falls back to a legacy industry-keyword table only when the plan has
      no roles for classify_difficulty() to score.
    - Volume: >50 hires adds 15-30 days proportionally
    - Market conditions: 'drought' adds 10 days, 'surplus' subtracts 5

    Args:
        channel_base_ttf: Midpoint time-to-fill for the channel type (days).
        data: Full plan data dict with industry, roles, hire_volume, etc.

    Returns:
        Adjusted time-to-fill in days (minimum 10).
    """
    # strategy:manpower#4 fix: derive the role-driven component from
    # gold_standard.classify_difficulty() -- the SAME per-role
    # time-to-fill model that feeds the Quality Intelligence sheet's Role
    # Difficulty Classification table -- instead of a second, independently
    # maintained keyword table (_ROLE_DIFFICULTY_TTF). Before this fix,
    # "CDL A Driver" priced differently here (bare "cdl"/"driver" keyword
    # range, midpoint 25.5d) than on Quality Intelligence (role-profile
    # map), producing two disagreeing time-to-fill figures for one role.
    try:
        role_results = gs_lib.classify_difficulty(data)
    except Exception:
        role_results = []
    role_ttfs = [
        r.get("avg_time_to_fill_days")
        for r in role_results
        if isinstance(r, dict) and r.get("avg_time_to_fill_days")
    ]
    role_midpoint: float | None = sum(role_ttfs) / len(role_ttfs) if role_ttfs else None

    if role_midpoint is None:
        # Defensive fallback (e.g. no target_roles/roles on data at all) --
        # legacy industry-keyword matching against _ROLE_DIFFICULTY_TTF.
        industry = str(data.get("industry") or "").lower()
        roles_raw = data.get("target_roles") or data.get("roles") or []
        role_texts: list[str] = []
        for r in (roles_raw if isinstance(roles_raw, list) else []):
            if isinstance(r, str):
                role_texts.append(r.lower())
            elif isinstance(r, dict):
                role_texts.append(str(r.get("title") or "").lower())
        combined_role_text = " ".join(role_texts) + f" {industry}"
        for keyword, ttf_range in _ROLE_DIFFICULTY_TTF.items():
            if keyword in combined_role_text:
                role_midpoint = (ttf_range[0] + ttf_range[1]) / 2.0
                break

    if role_midpoint is not None:
        # Scale channel TTF toward the role-appropriate range
        # Blend: 60% role-driven, 40% channel-driven
        adjusted_ttf = int(role_midpoint * 0.6 + channel_base_ttf * 0.4)
    else:
        adjusted_ttf = channel_base_ttf

    # ── Volume adjustment: >50 hires extends timeline ──
    try:
        hire_vol_str = str(data.get("hire_volume") or "0")
        hire_vol = int(hire_vol_str.replace(",", "").replace("+", "").strip() or "0")
    except (ValueError, TypeError):
        hire_vol = 0

    if hire_vol > 200:
        adjusted_ttf += 30
    elif hire_vol > 100:
        adjusted_ttf += 22
    elif hire_vol > 50:
        adjusted_ttf += 15

    # ── Market condition adjustment ──
    synthesized = data.get("_synthesized", {})
    market_condition = str(synthesized.get("market_condition") or "").lower()
    if "drought" in market_condition or "tight" in market_condition:
        adjusted_ttf += 10
    elif "surplus" in market_condition or "favorable" in market_condition:
        adjusted_ttf -= 5

    return max(10, adjusted_ttf)


def _build_sheet_roi_projections(ws, data: dict, load_kb_fn=None) -> None:
    """Build Sheet 5: ROI Projections with per-channel hire projections and efficiency scores.

    Reads channel allocation data from _budget_allocation and computes:
    - Projected applications and hires per channel
    - Cost per hire and time-to-fill estimates
    - ROI efficiency scores (1-10)
    - Summary totals row
    """
    ws.title = "ROI Projections"
    ws.sheet_properties.tabColor = GREEN

    _set_column_widths(
        ws,
        {
            1: 3,  # margin
            2: 24,  # Channel Name
            3: 16,  # Budget Allocated
            4: 18,  # Projected Applications
            5: 16,  # Projected Hires
            6: 14,  # Confidence
            7: 20,  # Hire Range
            8: 16,  # Cost Per Hire
            9: 18,  # Est. Time to Fill
            10: 12,  # ROI Score
            11: 40,  # Notes (brand rationale)
        },
    )

    budget_alloc = data.get("_budget_allocation", {})
    channel_allocs = budget_alloc.get("channel_allocations", {})

    row = 2

    # ── Section Header ──
    row = _write_section_header(ws, row, "ROI Projections & Hire Forecast")

    # ── Summary Cards (computed after channel loop, written first) ──
    summary_row_start = row
    row += 2  # reserve 2 rows for summary

    # ── Gather ROI data per channel ──
    roi_rows: List[Dict[str, Any]] = []
    total_budget = 0.0
    total_projected_hires = 0
    total_projected_apps = 0
    sum_cph = 0.0
    sum_ttf = 0.0
    channels_with_hires = 0

    sorted_channels = sorted(
        channel_allocs.items(),
        key=lambda x: x[1].get("dollar_amount", x[1].get("dollars") or 0),
        reverse=True,
    )

    for ch_name, ch_data in sorted_channels:
        try:
            dollars = ch_data.get("dollar_amount", ch_data.get("dollars") or 0)
            if not dollars or dollars <= 0:
                continue

            category = _roi_category_for_channel(ch_name)

            # S24: CPA estimate with role + location difficulty multipliers.
            # S23 CPA floors ($35-75) produced unrealistically low cost/hire
            # for professional roles ($732/hire for SWE in NYC, real is $5K-15K).
            existing_cpa = ch_data.get("cpa") or 0
            if existing_cpa and existing_cpa > 0:
                cpa_estimate = existing_cpa
            else:
                # Base CPA floors by channel category
                _CPA_FLOORS = {
                    "programmatic": 45.0,
                    "job_board": 35.0,
                    "social": 65.0,
                    "niche_board": 50.0,
                    "search": 55.0,
                    "display": 40.0,
                    "employer_branding": 75.0,
                    "career_site": 30.0,
                    "referral": 20.0,
                    "regional": 40.0,
                }
                _cpa_floor = _CPA_FLOORS.get(category, 40.0)

                # Role difficulty multiplier -- professional roles have much higher CPAs
                _role_lower = str(
                    data.get("role") or data.get("job_title") or ""
                ).lower()
                _ROLE_CPA_MULTIPLIER = 1.0
                if any(
                    k in _role_lower
                    for k in (
                        "engineer",
                        "developer",
                        "architect",
                        "devops",
                        "sre",
                        "data scientist",
                        "machine learning",
                    )
                ):
                    _ROLE_CPA_MULTIPLIER = 3.0
                elif any(
                    k in _role_lower
                    for k in (
                        "director",
                        "vp",
                        "vice president",
                        "head of",
                        "chief",
                        "executive",
                        "cto",
                        "cfo",
                        "cio",
                    )
                ):
                    _ROLE_CPA_MULTIPLIER = 4.0
                elif any(
                    k in _role_lower
                    for k in ("manager", "lead", "senior", "principal", "staff")
                ):
                    _ROLE_CPA_MULTIPLIER = 2.0
                elif any(
                    k in _role_lower
                    for k in (
                        "nurse",
                        "physician",
                        "pharmacist",
                        "therapist",
                        "surgeon",
                    )
                ):
                    _ROLE_CPA_MULTIPLIER = 2.5

                # Location cost multiplier -- high-cost metros
                _loc_lower = str(data.get("location") or "").lower()
                _LOC_CPA_MULTIPLIER = 1.0
                if any(
                    c in _loc_lower
                    for c in (
                        "new york",
                        "nyc",
                        "san francisco",
                        "sf",
                        "silicon valley",
                        "seattle",
                        "boston",
                        "washington dc",
                        "los angeles",
                    )
                ):
                    _LOC_CPA_MULTIPLIER = 1.5
                elif any(
                    c in _loc_lower
                    for c in (
                        "chicago",
                        "denver",
                        "dallas",
                        "atlanta",
                        "austin",
                        "miami",
                        "portland",
                    )
                ):
                    _LOC_CPA_MULTIPLIER = 1.2

                cpa_estimate = max(
                    _cpa_floor * _ROLE_CPA_MULTIPLIER * _LOC_CPA_MULTIPLIER, 40.0
                )

            projected_apps = (
                max(1, int(dollars / cpa_estimate)) if cpa_estimate > 0 else 0
            )

            # Use existing projected apps if available and reasonable
            existing_apps = ch_data.get("projected_applications") or 0
            if existing_apps > 0:
                projected_apps = existing_apps

            # Conversion rate: midpoint of channel-type range
            conv_lo, conv_hi = _ROI_CONVERSION_RATES.get(category, (0.05, 0.10))
            conversion_rate = (conv_lo + conv_hi) / 2.0

            # S48 FIX: Use upstream projected_hires as THE source of truth
            # to ensure ROI Projections total matches Executive Summary header.
            # Only fall back to conversion-rate estimation when the budget
            # engine truly did not set a value (key missing or None).
            existing_hires = ch_data.get("projected_hires")
            if existing_hires is not None and existing_hires >= 0:
                projected_hires = int(existing_hires)
            else:
                projected_hires = max(0, int(projected_apps * conversion_rate))

            # S89: a zero-hire channel has NO cost-per-hire -- dollars/1 was
            # silently printing the full channel budget as a fake CPH
            # (e.g. "$15,000/hire" for a channel that projects 0 hires).
            cost_per_hire = (
                round(dollars / projected_hires, 2) if projected_hires > 0 else None
            )

            # Time to fill: channel midpoint adjusted for role/volume/market
            ttf_lo, ttf_hi = _ROI_TIME_TO_FILL.get(category, (30, 45))
            base_ttf = (ttf_lo + ttf_hi) // 2
            est_time_to_fill = _compute_dynamic_ttf(base_ttf, data)

            # ROI Score (1-10): inversely proportional to cost-per-hire
            # Uses realistic recruitment industry thresholds:
            #   <$300 CPH = 10, $300-600 = 9, $600-1000 = 8, $1000-1500 = 7,
            #   $1500-2500 = 6, $2500-4000 = 5, $4000-6000 = 4, $6000-10000 = 3,
            #   $10000-20000 = 2, >$20000 = 1
            existing_roi = ch_data.get("roi_score") or 0
            if existing_roi and 1 <= existing_roi <= 10:
                roi_score = existing_roi
            elif cost_per_hire is None:
                # Zero-hire channel -- no CPH to tier the score against.
                roi_score = 1
            else:
                if cost_per_hire <= 300:
                    roi_score = 10
                elif cost_per_hire <= 600:
                    roi_score = 9
                elif cost_per_hire <= 1000:
                    roi_score = 8
                elif cost_per_hire <= 1500:
                    roi_score = 7
                elif cost_per_hire <= 2500:
                    roi_score = 6
                elif cost_per_hire <= 4000:
                    roi_score = 5
                elif cost_per_hire <= 6000:
                    roi_score = 4
                elif cost_per_hire <= 10000:
                    roi_score = 3
                elif cost_per_hire <= 20000:
                    roi_score = 2
                else:
                    roi_score = 1

            # Determine data confidence level for this channel.
            # S89 FIX (findings data:manpower#3/atria#3, strategy:atria#8):
            # budget_engine's own per-channel `confidence` field comes back
            # "high" for every channel regardless of the plan's overall
            # Sources & Confidence grade, so it's no longer authoritative --
            # re-derive from the channel's actual CPC/CPA data tier plus the
            # plan's overall confidence score (single source of truth, also
            # used by Channels & Strategy, Confidence Intervals, and Channel
            # Recommendations).
            hire_confidence = _derive_channel_confidence(data, ch_data)
            # S89A FIX (findings data:manpower#1/#2, data:atria#1): route
            # through the same _confidence_range() helper the Confidence
            # Intervals sheet uses, with the same 15/20/25% ladder -- this
            # column previously computed its own 10/25/40% variance and
            # disagreed with Confidence Intervals on every channel.
            _hire_band = _confidence_range(
                projected_hires, hire_confidence, cost_metric=False
            )
            if _hire_band is not None:
                hire_lo, hire_hi = (int(v) for v in _hire_band)
                hire_range_str = f"{hire_lo} - {hire_hi}"
            else:
                # S89: never show a fabricated "0 - 0" (or nonzero) range
                # for a channel that isn't projected to produce any hires.
                hire_range_str = "0"

            # S91: surface the budget_engine brand rationale (employer
            # branding etc. are measured on reach/pipeline influence, not
            # CPA -- their zero-hire projection is BY DESIGN, not a defect).
            _is_brand_channel = ch_data.get("channel_role") == "brand"
            _brand_note = ch_data.get("rationale") if _is_brand_channel else ""

            roi_rows.append(
                {
                    "name": _smart_title(ch_name),
                    "budget": dollars,
                    "projected_apps": projected_apps,
                    "projected_hires": projected_hires,
                    "cost_per_hire": cost_per_hire,
                    "time_to_fill": est_time_to_fill,
                    "roi_score": roi_score,
                    "category": category,
                    "conversion_rate": conversion_rate,
                    "hire_confidence": hire_confidence,
                    "hire_range": hire_range_str,
                    "is_brand": _is_brand_channel,
                    "brand_note": _brand_note,
                }
            )

            total_budget += dollars
            total_projected_hires += projected_hires
            total_projected_apps += projected_apps
            if projected_hires > 0:
                sum_cph += cost_per_hire
                sum_ttf += est_time_to_fill
                channels_with_hires += 1

        except Exception as exc:
            logger.warning("ROI projection failed for channel %s: %s", ch_name, exc)
            continue

    # Cost/Hire = total_budget / total_hires (consistent with Executive Summary)
    avg_cph = round(total_budget / max(total_projected_hires, 1), 2)
    avg_ttf = round(sum_ttf / max(channels_with_hires, 1))

    # ── Write summary row at reserved position ──
    # S5 (2026-07-03, findings 44/51): these headline KPIs (ROI Projections
    # B4:E4) were pre-formatted text strings -- write live numbers with a
    # number_format so they stay summable/sortable like the per-channel table.
    summary_labels = [
        "Total Budget",
        "Total Proj. Hires",
        "Avg Cost/Hire",
        "Avg Time to Fill",
    ]
    summary_values = [
        _safe_num(total_budget),
        int(total_projected_hires),
        _safe_num(avg_cph),
        avg_ttf,
    ]
    summary_formats = [_usd0_fmt(), FMT_INT, _usd2_fmt(), '0" days"']

    for i, (label, value, fmt) in enumerate(
        zip(summary_labels, summary_values, summary_formats)
    ):
        col = COL_START + i
        # Label row
        cell_l = ws.cell(row=summary_row_start, column=col, value=label)
        cell_l.font = _FONT_METRIC_LABEL
        cell_l.alignment = _ALIGN_CENTER
        cell_l.fill = _FILL_BLUE_PALE
        # Value row
        cell_v = ws.cell(row=summary_row_start + 1, column=col, value=value)
        cell_v.number_format = fmt
        cell_v.font = _FONT_METRIC_VALUE
        cell_v.alignment = _ALIGN_CENTER
        cell_v.fill = _FILL_WHITE
        cell_v.border = _BORDER_THIN

    row = summary_row_start + 3

    # ── Channel ROI Table ──
    row = _write_subsection_header(ws, row, "Per-Channel ROI Analysis")

    # S3: this is the plan's OWN per-channel budget -- header must reflect the
    # active plan currency symbol, not a hardcoded "$".
    headers = [
        "Channel Name",
        f"Budget ({_cur_symbol()})",
        "Proj. Applications",
        "Proj. Hires",
        "Confidence",
        "Hire Range",
        "Cost Per Hire",
        "Time to Fill",
        "ROI Score",
        "Notes",
    ]
    row = _write_table_header(ws, row, headers)

    for idx, roi_data in enumerate(roi_rows):
        roi_score = roi_data["roi_score"]
        # Color-code ROI score
        if roi_score >= 7:
            score_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=GREEN)
            score_fill = _FILL_GREEN_BG
        elif roi_score >= 4:
            score_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=AMBER)
            score_fill = _FILL_AMBER_BG
        else:
            score_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=RED)
            score_fill = _FILL_RED_BG

        # Color-code confidence level
        hire_conf = roi_data.get("hire_confidence", "LOW")
        if hire_conf == "HIGH":
            conf_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=GREEN)
            conf_fill = _FILL_GREEN_BG
        elif hire_conf == "MEDIUM":
            conf_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=AMBER)
            conf_fill = _FILL_AMBER_BG
        else:
            conf_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=RED)
            conf_fill = _FILL_RED_BG

        # S89: live numeric cells so the client can SUM/sort the ROI table.
        # Units are carried by the number_format (e.g. 0" days", 0"/10").
        # S89: a zero-hire channel has no cost-per-hire -- show "—" as text
        # rather than a fabricated budget-as-CPH number.
        _cph = roi_data["cost_per_hire"]
        _cph_val = _safe_num(_cph) if _cph is not None else "—"
        _cph_fmt = _usd0_fmt() if _cph is not None else None
        _notes = roi_data.get("brand_note") or ""
        values = [
            roi_data["name"],
            _safe_num(roi_data["budget"]),
            int(_safe_num(roi_data["projected_apps"])),
            int(_safe_num(roi_data["projected_hires"])),
            hire_conf,
            roi_data.get("hire_range", ""),
            _cph_val,
            _safe_num(roi_data["time_to_fill"]),
            round(_safe_num(roi_score), 1),
            _notes,
        ]
        # S3: budget/cost-per-hire are the plan's OWN figures -- active currency.
        _roi_fmts = [
            None, _usd0_fmt(), FMT_INT, FMT_INT, None, None,
            _cph_fmt, '0" days"', '0"/10"', None,
        ]
        row = _write_table_row(
            ws, row, values, alternate=(idx % 2 == 0), number_formats=_roi_fmts
        )

        # Override confidence cell styling
        conf_cell = ws.cell(row=row - 1, column=COL_START + 4)
        conf_cell.font = conf_font
        conf_cell.fill = conf_fill

        # Override ROI score cell styling
        roi_cell = ws.cell(row=row - 1, column=COL_START + 8)
        roi_cell.font = score_font
        roi_cell.fill = score_fill

    row += 1

    # ── Recruitment Funnel (S93: funnel-calibration model) ──
    # HARD INVARIANT: Clicks/Applications/Hires below are the SAME numbers
    # printed in the "Per-Channel ROI Analysis" table above -- Qualified and
    # Interviews are ADDED explanatory stages read verbatim from
    # metadata.funnel (never recomputed here), fitted so each channel's own
    # hires/raw_apps rate is reproduced exactly by the three stage rates.
    # This replaces the old "Implied App-to-Hire Rate" table, which showed
    # hires sitting directly next to raw applications and implied a
    # 0.2-0.5% apply-to-hire rate that reads as 10-40x below the 5-15%
    # funnel benchmarks this workbook cites elsewhere.
    row = _write_subsection_header(ws, row, "Recruitment Funnel")

    _funnel = budget_alloc.get("metadata", {}).get("funnel", {})
    _funnel_per_channel = (
        _funnel.get("per_channel", {}) if isinstance(_funnel, dict) else {}
    )
    _funnel_totals = _funnel.get("totals", {}) if isinstance(_funnel, dict) else {}
    _funnel_bands = _funnel.get("stage_rates", {}) if isinstance(_funnel, dict) else {}

    funnel_headers = [
        "Channel Name",
        "Clicks",
        "Applications",
        "Qualified",
        "Interviews",
        "Hires",
        "App→Qualified",
        "Qualified→Interview",
        "Interview→Hire",
    ]
    row = _write_table_header(ws, row, funnel_headers)

    _funnel_fmts = [
        None, FMT_INT, FMT_INT, FMT_INT, FMT_INT, FMT_INT,
        FMT_PCT1, FMT_PCT1, FMT_PCT1,
    ]

    _funnel_total_clicks = 0
    for idx, (ch_name, ch_data) in enumerate(sorted_channels):
        dollars = ch_data.get("dollar_amount", ch_data.get("dollars") or 0)
        if not dollars or dollars <= 0:
            continue
        f_row = _funnel_per_channel.get(ch_name)
        if not isinstance(f_row, dict):
            continue
        _hires = int(_safe_num(f_row.get("hires")))
        _clicks = int(_safe_num(ch_data.get("projected_clicks") or 0))
        _funnel_total_clicks += _clicks
        _rates = f_row.get("rates") or {}
        # S93: zero-hire channels (brand spend, or a performance channel
        # that landed on 0 hires) show '—' for Hires and the
        # interview->hire rate rather than a literal 0 -- they aren't
        # CPA-scored, and a printed "0" reads as a defect next to a real
        # Interviews count. When the value is text ("—"), the parallel
        # number_format entry must be None -- _write_num() would otherwise
        # coerce the "—" string to 0.0 via _safe_num().
        _hires_val: Any = _hires if _hires > 0 else "—"
        _int_to_hire_val: Any = (
            _rates.get("interview_to_hire") if _hires > 0 else "—"
        )
        values = [
            _smart_title(ch_name),
            _clicks,
            int(_safe_num(f_row.get("raw_apps"))),
            int(_safe_num(f_row.get("qualified_apps"))),
            int(_safe_num(f_row.get("interviews"))),
            _hires_val,
            _rates.get("raw_to_qualified"),
            _rates.get("qualified_to_interview"),
            _int_to_hire_val,
        ]
        _row_fmts = list(_funnel_fmts)
        if _hires_val == "—":
            _row_fmts[5] = None
        if _int_to_hire_val == "—":
            _row_fmts[8] = None
        row = _write_table_row(
            ws, row, values, alternate=(idx % 2 == 0), number_formats=_row_fmts
        )

    # TOTAL row -- sum of the channel rows above (funnel invariant guard).
    _t_hires = int(_safe_num(_funnel_totals.get("hires")))
    _t_rates = _funnel_totals.get("rates") or {}
    total_row_values = [
        "TOTAL",
        _funnel_total_clicks,
        int(_safe_num(_funnel_totals.get("raw_apps"))),
        int(_safe_num(_funnel_totals.get("qualified_apps"))),
        int(_safe_num(_funnel_totals.get("interviews"))),
        _t_hires,
        _t_rates.get("raw_to_qualified"),
        _t_rates.get("qualified_to_interview"),
        _t_rates.get("interview_to_hire"),
    ]
    row = _write_table_row(
        ws,
        row,
        total_row_values,
        alternate=False,
        number_formats=_funnel_fmts,
        fonts=[_FONT_BODY_BOLD] * len(total_row_values),
    )

    row += 1
    if _funnel_bands:
        _rq = _funnel_bands.get("raw_to_qualified", (0.08, 0.30))
        _qi = _funnel_bands.get("qualified_to_interview", (0.20, 0.50))
        _ih = _funnel_bands.get("interview_to_hire", (0.15, 0.35))
        row = _write_footnote(
            ws,
            row,
            "Planning-assumption bands: App→Qualified "
            f"{_rq[0]:.0%}-{_rq[1]:.0%}, Qualified→Interview "
            f"{_qi[0]:.0%}-{_qi[1]:.0%}, Interview→Hire "
            f"{_ih[0]:.0%}-{_ih[1]:.0%} (programmatic/high-volume channels "
            "skew toward the low end of App→Qualified; niche/referral "
            "channels skew toward the high end).",
        )
    row = _write_footnote(
        ws,
        row,
        "Methodology: hire totals are anchored to industry cost-per-hire "
        "benchmarks (not to this funnel); the stage rates above are "
        "planning assumptions fitted within the stated bands so their "
        "product reproduces this plan's own hires ÷ applications rate "
        "exactly per channel -- they explain the math, they do not drive it.",
    )
    row = _write_footnote(
        ws,
        row,
        "ROI Score: 9-10 = Excellent, 7-8 = Good, 4-6 = Average, 1-3 = Below Average.",
    )

    # S89A FIX (findings data:manpower#3/#4, visual:manpower#3,
    # strategy:manpower#3): same reconciliation note as the Executive
    # Summary benchmark block, computed via the identical shared helper --
    # when this plan's blended CPA/apply rate sits outside the cited
    # industry benchmark, say why instead of leaving it unexplained.
    _industry = data.get("industry", "general_entry_level")
    _ind_bench = _kb_industry_benchmark_section(_industry, load_kb_fn=load_kb_fn)
    _bm_note = _model_vs_benchmark_note(data, _ind_bench)
    if _bm_note:
        row = _write_footnote(ws, row, _bm_note)

    row += 1
    _write_attribution_footer(ws, row)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN GENERATOR FUNCTION
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6: Quality Intelligence (Gold Standard Gates)
# ═══════════════════════════════════════════════════════════════════════════════


def _collapse_fallback_market_rows(
    city_data: Dict[str, Any],
) -> List[Tuple[str, Dict[str, Any], bool]]:
    """Collapse 2+ markets that share byte-identical fallback data into ONE
    representative row (finding data:manpower#5).

    ``gold_standard.enrich_city_level_data`` flags a market ``fallback_uniform``
    when it had zero city/state/country/metro signal and bottomed out on the
    same flat generic default (1.00x multiplier, 5.5/10 difficulty) every
    other such market also gets -- presenting N of those as N distinct
    "location-specific" rows reads as fabricated per-market precision.
    Markets with real (non-fallback) data, or a lone fallback market with no
    peer to collapse with, pass through individually unchanged.

    Returns a list of ``(label, info, is_collapsed)`` tuples in original
    order (the collapsed group appears where its first member would have).
    """
    if not isinstance(city_data, dict):
        return []
    fallback_count = sum(
        1
        for info in city_data.values()
        if isinstance(info, dict) and info.get("fallback_uniform")
    )
    out: List[Tuple[str, Dict[str, Any], bool]] = []
    collapsed_written = False
    for name, info in city_data.items():
        if not isinstance(info, dict):
            continue
        if info.get("fallback_uniform") and fallback_count >= 2:
            if not collapsed_written:
                out.append(
                    (
                        "All listed markets (default index — market-level "
                        "data pending enrichment)",
                        info,
                        True,
                    )
                )
                collapsed_written = True
            continue
        out.append((_title_case_city(name), info, False))
    return out


def _salary_range_from_per_role(info: Dict[str, Any]) -> str | None:
    """Derive a market's headline Salary Range from its own per-role salary
    rows (min of mins, max of maxes) instead of an independently-computed
    figure, so the City-Level Supply-Demand table's Salary Range column
    always agrees with the Salary Intelligence table's Min/Max columns for
    the same market (finding data:manpower#1). Returns None when the market
    has no per-role salary data to derive from.
    """
    per_role: Dict[str, Any] = info.get("per_role_salary") or {}
    mins = [
        r.get("min")
        for r in per_role.values()
        if isinstance(r, dict) and isinstance(r.get("min"), (int, float))
    ]
    maxes = [
        r.get("max")
        for r in per_role.values()
        if isinstance(r, dict) and isinstance(r.get("max"), (int, float))
    ]
    if not mins or not maxes:
        return None
    return f"${min(mins):,.0f} - ${max(maxes):,.0f}"


def _build_sheet_quality_intelligence(
    ws, data: dict, gold_standard: dict[str, Any]
) -> None:
    """Build the Quality Intelligence worksheet from Gold Standard gate outputs.

    Renders 7 sections corresponding to the quality gates:
    1. City-level supply-demand data
    2. Security clearance segmentation (if applicable)
    3. Competitor mapping per city
    4. Difficulty classification per role
    5. Channel strategy (traditional vs non-traditional)
    6. Budget tier breakdowns
    7. Activation event calendar

    Args:
        ws: The openpyxl worksheet to populate.
        data: The full enriched data dict.
        gold_standard: The ``data["_gold_standard"]`` dict from apply_all_quality_gates.
    """
    ws.title = "Quality Intelligence"
    ws.sheet_properties.tabColor = SAPPHIRE

    # Column widths (B-H)
    _set_column_widths(ws, {1: 2, 2: 22, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18, 8: 18})

    row = 2

    # ── Title banner ──
    row = _write_section_header(ws, row, "QUALITY INTELLIGENCE — GOLD STANDARD GATES")
    row += 1

    client_name = data.get("client_name") or "Client"
    industry_label = data.get("industry_label") or (
        (data.get("industry") or "").replace("_", " ").title()
    )
    row = _write_footnote(
        ws,
        row,
        f"Gold Standard quality gate analysis for {client_name} | "
        f"Industry: {industry_label} | "
        f"Generated {datetime.date.today().strftime('%B %d, %Y')}",
    )
    row += 1

    # ── Section 1: City-Level Supply-Demand Data ──
    city_data: dict = gold_standard.get("city_level_data") or {}
    try:
        if city_data:
            row = _write_subsection_header(ws, row, "City-Level Supply-Demand Data")
            # S89A FIX (finding data:manpower#5): header was "City" even
            # though this table also carries state/region-level fallback
            # rows -- "Market" is accurate for both.
            row = _write_table_header(
                ws,
                row,
                [
                    "Market",
                    "Salary Multiplier",
                    "Estimated Salary",
                    "Hiring Difficulty",
                    "Supply Tier",
                    "COL Index",
                    "Salary Range",
                ],
            )
            # S5 (2026-07-03, findings 44/51/54): title-case the city name for
            # display and write Estimated Salary as a live number instead of
            # a pre-formatted string.
            _city_money_fmts = [None, None, _usd0_fmt(), None, None, None, None]
            # S89A FIX (finding data:manpower#5): collapse 2+ markets that
            # share byte-identical fallback data into one honest row instead
            # of repeating it once per market.
            _display_rows = _collapse_fallback_market_rows(city_data)
            _has_collapsed_market_row = any(is_c for _, _, is_c in _display_rows)
            for idx, (market_label, info, _is_collapsed) in enumerate(_display_rows):
                # W4A FIX (finding data:manpower#1): this row's "Salary Range"
                # used to be computed independently as a flat
                # est_salary-$10k/+$15k offset, which contradicts the
                # Min/Max columns of the "Salary Intelligence" table below
                # (same market, same roles) since that table's per-role
                # bands are multiplier/tier-scaled, not a flat offset.
                # Derive this cell from the SAME per_role_salary rows the
                # Salary Intelligence table renders -- min of mins, max of
                # maxes -- so the two tables can't disagree about one market.
                row = _write_table_row(
                    ws,
                    row,
                    [
                        market_label,
                        f"{info.get('salary_multiplier', 1.0):.2f}x",
                        _safe_num(info.get("estimated_salary", 0)),
                        f"{info.get('hiring_difficulty', 0):.1f}/10",
                        str(info.get("supply_tier") or "balanced")
                        .replace("_", " ")
                        .title(),
                        f"{info.get('cost_of_living_index', 100):.1f}",
                        str(
                            _salary_range_from_per_role(info)
                            or info.get("salary_range")
                            or "—"
                        ),
                    ],
                    alternate=idx % 2 == 1,
                    number_formats=_city_money_fmts,
                )
            _city_footnote = (
                "Salary multipliers relative to national average. "
                "Hiring difficulty: 1 (easy) to 10 (hardest)."
            )
            if _has_collapsed_market_row:
                _city_footnote += (
                    ' "All listed markets" is a national-default estimate '
                    "(confidence: estimated) -- no market-specific salary/cost "
                    "data was available to differentiate these locations."
                )
            row = _write_footnote(ws, row, _city_footnote)
            row += 1

            # ── Per-Role Salary Breakdown (additive section) ──
            # Check if any city has per_role_salary data
            has_role_salary = any(
                info.get("per_role_salary") for info in city_data.values()
            )
            if has_role_salary:
                # S89: honest section title -- the old "Per-Role Salary
                # Breakdown by City" title implied every row was sourced
                # data; a large share are tier-scaled estimates for roles
                # with no keyword match. Retitle + badge those rows instead
                # of presenting them at the same confidence as a benchmark.
                row = _write_subsection_header(
                    ws, row, "Salary Intelligence — estimated where noted"
                )
                row = _write_table_header(
                    ws,
                    row,
                    [
                        "City",
                        "Role",
                        "Min Salary",
                        "P25",
                        "Median Salary",
                        "P75",
                        "Max Salary",
                        "City Multiplier",
                        "Source",
                    ],
                )
                alt_idx = 0
                _role_sal_fmts = [
                    None,
                    None,
                    _usd0_fmt(),
                    _usd0_fmt(),
                    _usd0_fmt(),
                    _usd0_fmt(),
                    _usd0_fmt(),
                    None,
                    None,
                ]
                _estimated_fill = PatternFill(
                    start_color=AMBER, end_color=AMBER, fill_type="solid"
                )
                # S92 fix: gold_standard.enrich_city_level_data's internal
                # ``source`` value for the tier-scaled fallback branch is the
                # raw code-facing token "generic_enrichment" (its sibling
                # matched-keyword branch already uses a display-ready
                # "Industry Benchmark" string) -- map it to a client-facing
                # label at render time rather than leaking the internal
                # token into this cell. Internal ``source`` value itself is
                # left untouched (tests/test_gold_standard_taxonomy.py
                # asserts it literally).
                _SOURCE_DISPLAY_LABELS = {"generic_enrichment": "Tier-Scaled Estimate"}
                _all_salary_rows: List[Dict[str, Any]] = []
                for city_name, info in city_data.items():
                    role_salary: dict = info.get("per_role_salary") or {}
                    for role_name, sal in role_salary.items():
                        _all_salary_rows.append(sal)
                        _is_estimated = sal.get("confidence") == "estimated"
                        _role_display = (
                            f"{role_name} (est.)" if _is_estimated else role_name
                        )
                        _source_raw = sal.get("source") or ""
                        _source_display = _SOURCE_DISPLAY_LABELS.get(
                            _source_raw, _source_raw
                        )
                        row = _write_table_row(
                            ws,
                            row,
                            [
                                _title_case_city(city_name),
                                _role_display,
                                _safe_num(sal.get("min", 0)),
                                _safe_num(sal.get("p25", sal.get("min", 0))),
                                _safe_num(sal.get("median", 0)),
                                _safe_num(sal.get("p75", sal.get("max", 0))),
                                _safe_num(sal.get("max", 0)),
                                f"{sal.get('multiplier', 1.0):.2f}x",
                                str(_source_display or "—"),
                            ],
                            number_formats=_role_sal_fmts,
                            alternate=alt_idx % 2 == 1,
                            fills=(
                                [_estimated_fill] * 9 if _is_estimated else None
                            ),
                        )
                        alt_idx += 1
                _conf = gs_lib.confidence_summary(_all_salary_rows)
                row = _write_footnote(
                    ws,
                    row,
                    "Per-role salaries adjusted by city multiplier; P25/P75 "
                    "bands are ordered (min<=P25<=median<=P75<=max). Rows "
                    "marked (est.) use a tier-scaled generic estimate rather "
                    "than a specific industry benchmark -- "
                    f"{_conf['benchmark_count']} of {_conf['total_rows']} rows "
                    f"({_conf['pct_benchmark']:.0f}%) are benchmark-sourced.",
                )
                row += 1
    except Exception as exc:
        logger.error(
            "Quality Intelligence: city-level section failed: %s", exc, exc_info=True
        )
        row += 1

    # ── Section 2: Security Clearance Segmentation ──
    clearance: Optional[dict] = gold_standard.get("clearance_segmentation")
    try:
        if clearance:
            row = _write_subsection_header(ws, row, "Security Clearance Segmentation")

            primary = clearance.get("primary_clearance") or {}
            # S4: non-US plans get a disclosed reference-framework note instead
            # of a fabricated local clearance tier (see detect_clearance_requirements).
            if clearance.get("us_framework_only"):
                row = _write_kv_row(
                    ws,
                    row,
                    "Defense Related",
                    "Yes — clearance requirements detected (non-US plan; "
                    "US clearance tiers below are reference-only, not local data)",
                )
            else:
                row = _write_kv_row(
                    ws, row, "Defense Related", "Yes — clearance requirements detected"
                )
            if not clearance.get("us_framework_only"):
                row = _write_kv_row(
                    ws, row, "Primary Clearance", str(primary.get("level") or "—")
                )
            row = _write_kv_row(
                ws,
                row,
                "Detected Keywords",
                ", ".join(clearance.get("detected_keywords") or []),
            )
            row += 1

            # Clearance tiers table
            all_tiers: list = clearance.get("all_clearance_tiers") or []
            if all_tiers:
                row = _write_table_header(
                    ws,
                    row,
                    [
                        "Clearance Level",
                        "Salary Premium",
                        "Time-to-Fill (wks)",
                        "Pool Reduction",
                        "Budget Multiplier",
                        "Recommended Channels",
                    ],
                )
                for idx, tier in enumerate(all_tiers):
                    row = _write_table_row(
                        ws,
                        row,
                        [
                            str(tier.get("level") or ""),
                            f"+{tier.get('salary_premium_pct', 0)}%",
                            str(tier.get("time_to_fill_weeks") or ""),
                            f"{tier.get('candidate_pool_reduction_pct', 0)}%",
                            f"{tier.get('budget_multiplier', 1.0):.1f}x",
                            ", ".join(tier.get("channels") or []),
                        ],
                        alternate=idx % 2 == 1,
                    )

            # Recommendations
            recs: list = clearance.get("recommendations") or []
            for rec in recs:
                row = _write_kv_row(ws, row, "Recommendation", str(rec))
            row += 1
    except Exception as exc:
        logger.error(
            "Quality Intelligence: clearance section failed: %s", exc, exc_info=True
        )
        row += 1

    # ── Section 3: Competitor Mapping with Counter-Strategies ──
    competitor_map: dict = gold_standard.get("competitor_mapping") or {}
    try:
        if competitor_map:
            row = _write_subsection_header(
                ws, row, "Competitive Landscape & Counter-Strategies"
            )
            row = _write_table_header(
                ws,
                row,
                [
                    "City",
                    "Top Employers",
                    "Hiring Intensity",
                    "Est. Competing Postings",
                    "Why They Matter",
                    "Counter-Strategy",
                ],
            )
            client_name_qs = data.get("client_name") or "Client"
            industry_label_qs = data.get("industry_label") or (
                (data.get("industry") or "").replace("_", " ").title()
            )
            _roles_qs = _get_roles(data)
            _first_role_qs = _roles_qs[0] if _roles_qs else ""
            for idx, (city_name_raw, info) in enumerate(competitor_map.items()):
                if city_name_raw.startswith("_"):
                    continue  # skip internal keys like _national
                # S5 (2026-07-03, finding 54): source data sometimes carries
                # lowercase city names (e.g. "houston", "atlanta") -- display
                # them Title Case everywhere, including inside the generated
                # why_matter/counter prose below.
                city_name = _title_case_city(city_name_raw)
                employers = info.get("top_employers") or []
                intensity = str(info.get("hiring_intensity") or "moderate").lower()
                est_postings = info.get("estimated_competing_postings") or "—"

                # Generate WHY each competitor group matters
                if intensity in ("high", "very_high"):
                    why_matter = (
                        f"High hiring volume in {city_name} — "
                        f"these employers compete for the same {industry_label_qs} talent pool"
                    )
                elif intensity == "moderate":
                    why_matter = (
                        f"Active but not dominant — opportunity to capture market share "
                        f"with targeted positioning in {city_name}"
                    )
                else:
                    why_matter = (
                        f"Lower competition in {city_name} — favorable market for "
                        f"{client_name_qs}'s talent acquisition"
                    )

                # Generate counter-strategy.
                # S92 fix: the 3 fixed boilerplate sentences below (keyed
                # ONLY off the 3-value `intensity` bucket, city/employer name
                # interpolated in the "high" branch only) rendered
                # BYTE-IDENTICAL text for every city that landed in the same
                # bucket -- e.g. 5 different low-intensity cities in one
                # plan all got the exact same "Capitalize on low
                # competition..." sentence, which bundle_qa's
                # counter_strategy_near_duplicate check correctly flags as
                # "competitors must not read as interchangeable". Use the
                # same insight_composer.compose_counter_strategy skeleton
                # bank the Market Intelligence sheet's competitor table
                # (line ~4470 above) and the deck's competitor cards
                # (ppt_generator.py) already use for exactly this reason --
                # ordinal-indexed skeleton selection guarantees adjacent
                # rows never share a sentence, and the top employer + city
                # are interpolated into every row, not just the "high" one.
                top_employer = employers[0] if employers else "competitors"
                counter = insight_composer.compose_counter_strategy(
                    top_employer,
                    {
                        "role": _first_role_qs,
                        "city": city_name,
                        "industry": industry_label_qs,
                        "intensity": intensity,
                        "ordinal": idx,
                    },
                )

                # S5 (2026-07-03, finding 39): never hard-truncate why_matter/
                # counter mid-word -- wrap + tallen the row instead.
                row = _write_table_row(
                    ws,
                    row,
                    [
                        city_name,
                        ", ".join(employers[:4]),
                        intensity.title(),
                        str(est_postings),
                        why_matter,
                        counter,
                    ],
                    alternate=idx % 2 == 1,
                )
                if len(why_matter) > 80 or len(counter) > 80:
                    ws.row_dimensions[row - 1].height = 40

            # National competitors row
            national: dict = competitor_map.get("_national") or {}
            if national:
                national_employers = national.get("top_employers") or []
                row = _write_table_row(
                    ws,
                    row,
                    [
                        "National (All Markets)",
                        ", ".join(national_employers[:5]),
                        str(national.get("hiring_intensity") or "moderate").title(),
                        "",
                        "National competitors set salary and benefits benchmarks",
                        "Match or exceed top benefits; lead with mission and impact",
                    ],
                    fonts=[
                        _FONT_BODY_BOLD,
                        _FONT_BODY,
                        _FONT_BODY,
                        _FONT_BODY,
                        _FONT_BODY,
                        _FONT_BODY,
                    ],
                )
            row += 1
    except Exception as exc:
        logger.error(
            "Quality Intelligence: competitor section failed: %s", exc, exc_info=True
        )
        row += 1

    # ── Section 4: Difficulty Classification ──
    difficulty_framework: list = gold_standard.get("difficulty_framework") or []
    try:
        if difficulty_framework:
            row = _write_subsection_header(ws, row, "Role Difficulty Classification")
            row = _write_table_header(
                ws,
                row,
                [
                    "Role Title",
                    "Seniority Level",
                    "Difficulty (1-10)",
                    "Supply Level",
                    "Avg Time-to-Fill",
                    "Location Modifier",
                    "Budget Weight",
                    "Channel Emphasis",
                    "Description",
                ],
            )
            for idx, role_info in enumerate(difficulty_framework):
                loc_mod = role_info.get("location_modifier", 0.0)
                loc_name = role_info.get("location_matched") or ""
                loc_display = (
                    f"+{loc_mod:.1f} ({loc_name})"
                    if loc_mod > 0 and loc_name
                    else (
                        f"{loc_mod:.1f} ({loc_name})"
                        if loc_mod < 0 and loc_name
                        else "0 (baseline)"
                    )
                )
                supply_raw = str(role_info.get("supply_level") or "moderate")
                supply_display = supply_raw.replace("_", " ").title()
                row = _write_table_row(
                    ws,
                    row,
                    [
                        str(role_info.get("role_title") or ""),
                        str(role_info.get("seniority_level") or "mid").title(),
                        str(role_info.get("complexity_score") or ""),
                        supply_display,
                        f"{role_info.get('avg_time_to_fill_days', 0)} days",
                        loc_display,
                        f"{role_info.get('budget_weight', 1.0):.1f}x",
                        str(role_info.get("channel_emphasis") or "")
                        .replace("_", " ")
                        .title(),
                        str(role_info.get("description") or ""),
                    ],
                    alternate=idx % 2 == 1,
                )
            row += 1
    except Exception as exc:
        logger.error(
            "Quality Intelligence: difficulty section failed: %s", exc, exc_info=True
        )
        row += 1

    # ── Section 5: Channel Strategy ──
    channel_strategy: dict = gold_standard.get("channel_strategy") or {}
    try:
        if channel_strategy:
            row = _write_subsection_header(
                ws, row, "Channel Strategy — Traditional vs Non-Traditional"
            )

            split = channel_strategy.get("recommended_split") or {}
            trad_pct = split.get("traditional_pct", 65)
            nontrad_pct = split.get("non_traditional_pct", 35)
            avg_complexity = channel_strategy.get("avg_role_complexity", 0)

            row = _write_kv_row(
                ws,
                row,
                "Recommended Split",
                f"{trad_pct}% Traditional / {nontrad_pct}% Non-Traditional",
            )
            row = _write_kv_row(ws, row, "Avg Role Complexity", f"{avg_complexity}/10")
            strategy_note = channel_strategy.get("strategy_note") or ""
            if strategy_note:
                row = _write_kv_row(ws, row, "Strategy Note", strategy_note)
            row += 1

            # Traditional channels
            trad_channels: list = channel_strategy.get("traditional_channels") or []
            if trad_channels:
                row = _write_table_header(
                    ws,
                    row,
                    ["Traditional Channel", "Type", "Reach", "Relevance Score"],
                    fill=_FILL_BLUE_LIGHT,
                )
                for idx, ch in enumerate(trad_channels):
                    row = _write_table_row(
                        ws,
                        row,
                        [
                            str(ch.get("name") or ""),
                            str(ch.get("type") or "").replace("_", " ").title(),
                            str(ch.get("reach") or "").replace("_", " ").title(),
                            str(ch.get("relevance_score") or ""),
                        ],
                        alternate=idx % 2 == 1,
                    )
                row += 1

            # Non-traditional channels
            nontrad_channels: list = (
                channel_strategy.get("non_traditional_channels") or []
            )
            if nontrad_channels:
                row = _write_table_header(
                    ws,
                    row,
                    ["Non-Traditional Channel", "Type", "Reach"],
                    fill=_FILL_BLUE_LIGHT,
                )
                for idx, ch in enumerate(nontrad_channels):
                    row = _write_table_row(
                        ws,
                        row,
                        [
                            str(ch.get("name") or ""),
                            str(ch.get("type") or "").replace("_", " ").title(),
                            str(ch.get("reach") or "").replace("_", " ").title(),
                        ],
                        alternate=idx % 2 == 1,
                    )
                row += 1
    except Exception as exc:
        logger.error(
            "Quality Intelligence: channel strategy section failed: %s",
            exc,
            exc_info=True,
        )
        row += 1

    # ── Section 6: Recommended Program Structure ──
    # S89: gold_standard.compute_budget_tiers splits the plan's OWN media
    # budget into media/creative/contingency percentages of ONE total --
    # but that total IS this plan's full media budget already (100% of it
    # is allocated across channels in _budget_allocation). Presenting
    # "creative 17%" and "contingency 11%" as slices of that SAME total
    # double-claims dollars that are already spoken for as media spend.
    # Reframe: media = 100% of the plan's media budget (no re-slicing);
    # creative/contingency are RECOMMENDED INCREMENTAL reserves sized as a
    # % of media, on TOP of the media budget -- never carved out of it.
    budget_tiers: dict = gold_standard.get("budget_tiers") or {}
    try:
        if budget_tiers and "error" not in budget_tiers:
            row = _write_subsection_header(ws, row, "Recommended Program Structure")
            media_budget = _safe_num(budget_tiers.get("total_budget", 0))
            row = _write_kv_row(
                ws, row, "This Plan's Media Budget", _fmt_currency(media_budget)
            )
            row = _write_footnote(
                ws,
                row,
                "The media figure below IS this plan's full budget (already "
                "100% allocated across channels elsewhere in this workbook). "
                "Creative and contingency are RECOMMENDED INCREMENTAL "
                "reserves on top of it, not a second claim on the same "
                "dollars -- fund them only if additional budget is available.",
            )
            row += 1

            tier_breakdown: dict = budget_tiers.get("tier_breakdown") or {}
            row = _write_table_header(
                ws,
                row,
                ["Program Component", "Amount", "% of Media Budget", "Description"],
            )
            _tier_formats = [None, _usd0_fmt(), FMT_PCT1, None]

            # Media = 100% of the plan's own budget, never re-sliced.
            row = _write_table_row(
                ws,
                row,
                [
                    "Media (this plan's full budget)",
                    media_budget,
                    1.0,
                    "Direct job advertising, programmatic, boards, social ads "
                    "-- the plan's entire media spend, detailed elsewhere in "
                    "this workbook.",
                ],
                fonts=[_FONT_BODY_BOLD] * 4,
                fills=[_FILL_GREEN_BG] * 4,
                number_formats=_tier_formats,
            )

            _reserve_total = media_budget
            idx = 0
            for tier_key, tier_info in tier_breakdown.items():
                if tier_key == "media_spend":
                    continue  # already written above at its true 100%
                tier_label = (
                    display_format.channel_label(tier_key) + " — suggested addition"
                )
                _amount = _safe_num(tier_info.get("amount", 0))
                _reserve_total += _amount
                row = _write_table_row(
                    ws,
                    row,
                    [
                        tier_label,
                        _amount,
                        _safe_num(tier_info.get("pct", 0)) / 100.0,
                        str(tier_info.get("description") or ""),
                    ],
                    alternate=idx % 2 == 1,
                    number_formats=_tier_formats,
                )
                idx += 1

                # Sub-allocations
                sub_alloc: dict = tier_info.get("sub_allocation") or {}
                if sub_alloc:
                    for sub_key, sub_amount in sub_alloc.items():
                        sub_label = f"  — {display_format.channel_label(sub_key)}"
                        row = _write_table_row(
                            ws,
                            row,
                            [sub_label, _safe_num(sub_amount), "", ""],
                            fonts=[_FONT_FOOTNOTE, _FONT_FOOTNOTE, None, None],
                            number_formats=[None, _usd0_fmt(), None, None],
                        )

            # Total Program = media + reserves (never == media_budget alone).
            row = _write_table_row(
                ws,
                row,
                [
                    "Total Program (media + suggested reserves)",
                    _reserve_total,
                    _reserve_total / media_budget if media_budget > 0 else 0,
                    "Fully-funded program if the suggested reserves are added.",
                ],
                fonts=[_FONT_BODY_BOLD] * 4,
                fills=[_FILL_BLUE_PALE] * 4,
                number_formats=_tier_formats,
            )
            row += 1

            # S89: locally-composed recommendations reflecting the media
            # (100%) + reserves (incremental) reframe above -- NOT
            # gold_standard.compute_budget_tiers' raw recommendation
            # strings, which describe the pre-reframe "72/17/11 split of
            # one total" framing and would contradict this section's own
            # table if printed verbatim.
            _reframed_recs = [
                f"Media: {_fmt_currency(media_budget)} is this plan's full "
                "budget -- already committed to the channel mix detailed "
                "elsewhere in this workbook; not available to re-slice.",
            ]
            for tier_key, tier_info in tier_breakdown.items():
                if tier_key == "media_spend":
                    continue
                _amount = _safe_num(tier_info.get("amount", 0))
                _label = display_format.channel_label(tier_key)
                _reframed_recs.append(
                    f"{_label}: {_fmt_currency(_amount)} suggested addition "
                    "if incremental budget becomes available -- "
                    f"{str(tier_info.get('description') or '').rstrip('.')}."
                )
            for rec in _reframed_recs:
                row = _write_kv_row(ws, row, "Recommendation", str(rec))
            row += 1
    except Exception as exc:
        logger.error(
            "Quality Intelligence: budget tiers section failed: %s", exc, exc_info=True
        )
        row += 1

    # ── Section 7: Activation Event Calendar ──
    activation: dict = gold_standard.get("activation_calendar") or {}
    try:
        if activation:
            row = _write_subsection_header(ws, row, "Activation Event Calendar")
            start_month = activation.get("campaign_start_month", 0)
            if start_month:
                row = _write_kv_row(
                    ws,
                    row,
                    "Campaign Start",
                    datetime.date(2026, start_month, 1).strftime("%B %Y"),
                )
            phasing_note = activation.get("budget_phasing_note") or ""
            if phasing_note:
                row = _write_kv_row(ws, row, "Budget Phasing", phasing_note)
            row += 1

            timeline: list = activation.get("timeline") or []
            if timeline:
                row = _write_table_header(
                    ws,
                    row,
                    [
                        "Month",
                        "Season",
                        "Hiring Intensity",
                        "Budget Weight",
                        "Key Events",
                        "Recommendation",
                    ],
                )
                for idx, month_info in enumerate(timeline):
                    events = month_info.get("key_events") or []
                    row = _write_table_row(
                        ws,
                        row,
                        [
                            str(month_info.get("month_name") or ""),
                            str(month_info.get("season") or ""),
                            str(month_info.get("hiring_intensity") or "")
                            .replace("_", " ")
                            .title(),
                            f"{month_info.get('budget_weight', 1.0):.1f}x",
                            "; ".join(events),
                            str(month_info.get("recommendation") or ""),
                        ],
                        alternate=idx % 2 == 1,
                    )
                row += 1

            # Industry-specific events
            industry_events: list = activation.get("industry_events") or []
            if industry_events:
                row = _write_kv_row(
                    ws,
                    row,
                    "Industry Events",
                    "; ".join(industry_events),
                )
                row += 1

            # strategy:atria#8 fix: for campaigns longer than 12 months, say
            # explicitly that the annual calendar above repeats rather than
            # leaving a client to assume it only covers part of the term.
            _repeats_note = activation.get("repeats_annually_note") or ""
            if _repeats_note:
                row = _write_footnote(ws, row, _repeats_note)
                row += 1
    except Exception as exc:
        logger.error(
            "Quality Intelligence: activation calendar section failed: %s",
            exc,
            exc_info=True,
        )
        row += 1

    # ── Attribution footer ──
    row += 1
    _write_attribution_footer(ws, row)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7: 90-Day Rolling Forecast
# ═══════════════════════════════════════════════════════════════════════════════


def _build_sheet_rolling_forecast(ws, data: dict) -> None:
    """Build Sheet 7: 90-Day Rolling Forecast with monthly spend, applications, hires, and CPA trend.

    Breaks the campaign timeline into 3 monthly periods showing projected metrics
    with a ramp-up curve (base shape 25/35/40, seasonally adjusted per industry
    and start month by ``_seasonal_monthly_phasing`` -- the rendered narrative
    always quotes the actual computed split, never a hardcoded percentage).
    """
    ws.title = "90-Day Forecast"
    ws.sheet_properties.tabColor = SAPPHIRE

    _set_column_widths(
        ws,
        {
            1: 3,  # margin
            2: 22,  # Metric / Channel
            3: 18,  # Month 1
            4: 18,  # Month 2
            5: 18,  # Month 3
            6: 18,  # 90-Day Total
            7: 16,  # Trend
        },
    )

    budget_alloc = data.get("_budget_allocation", {})
    if not isinstance(budget_alloc, dict):
        budget_alloc = {}
    ba_total_proj = budget_alloc.get("total_projected", {})
    if not isinstance(ba_total_proj, dict):
        ba_total_proj = {}
    ba_metadata = budget_alloc.get("metadata", {})
    if not isinstance(ba_metadata, dict):
        ba_metadata = {}
    ba_channel_alloc = budget_alloc.get("channel_allocations", {})
    if not isinstance(ba_channel_alloc, dict):
        ba_channel_alloc = {}

    total_budget = _safe_num(ba_metadata.get("total_budget") or 0)
    if total_budget <= 0:
        from shared_utils import parse_budget as _pb

        total_budget = _safe_num(_pb(data.get("budget") or ""))

    total_apps = int(_safe_num(ba_total_proj.get("applications") or 0))
    # S48 FIX: Compute total_hires from per-channel sum (source of truth)
    # to stay consistent with Executive Summary and ROI Projections sheets.
    total_hires = sum(
        int(ch.get("projected_hires") or 0) for ch in ba_channel_alloc.values()
    )
    if total_hires == 0:
        total_hires = int(_safe_num(ba_total_proj.get("hires") or 0))

    today = datetime.date.today()

    # S49 FIX (Issue 18): Use campaign_start_month from form data instead of
    # today's month so the forecast aligns with the activation timeline.
    # e.g. if campaign starts May, forecast = May/Jun/Jul (not Apr/May/Jun).
    _csm_raw = data.get("campaign_start_month") or 0
    try:
        _campaign_start_month = int(_csm_raw) if _csm_raw else 0
    except (ValueError, TypeError):
        _campaign_start_month = 0
    if _campaign_start_month < 1 or _campaign_start_month > 12:
        _campaign_start_month = today.month  # fallback to current month

    # S50: Seasonal-aware budget phasing replaces the flat 25/35/40 ramp-up.
    # Uses seasonal_hiring_trends.json to shift budget toward peak hiring months
    # for the campaign's industry, while preserving the ramp-up base shape.
    # Falls back to [0.25, 0.35, 0.40] when no seasonal data is available.
    _industry_raw = str(data.get("industry") or "")
    monthly_pcts = _seasonal_monthly_phasing(_industry_raw, _campaign_start_month)

    # Determine the forecast start year: if campaign month is in the past
    # relative to current date, assume it starts this year anyway (form input);
    # otherwise use current year.
    _forecast_year = today.year

    month_labels = []
    for i in range(3):
        m = _campaign_start_month + i
        y = _forecast_year
        if m > 12:
            m -= 12
            y += 1
        month_labels.append(datetime.date(y, m, 1).strftime("%B %Y"))

    # Compute forecast period start/end from campaign start month
    _forecast_start = datetime.date(_forecast_year, _campaign_start_month, 1)
    _forecast_end_m = _campaign_start_month + 2
    _forecast_end_y = _forecast_year
    if _forecast_end_m > 12:
        _forecast_end_m -= 12
        _forecast_end_y += 1
    # Last day of the 3rd month
    if _forecast_end_m == 12:
        _forecast_end = datetime.date(_forecast_end_y, 12, 31)
    else:
        _forecast_end = datetime.date(
            _forecast_end_y, _forecast_end_m + 1, 1
        ) - datetime.timedelta(days=1)

    row = 2

    # ── Section Header ──
    row = _write_section_header(ws, row, "90-Day Rolling Forecast")

    # ── Campaign Period ──
    row = _write_kv_row(
        ws,
        row,
        "Forecast Period",
        f"{_forecast_start.strftime('%b %d, %Y')} - {_forecast_end.strftime('%b %d, %Y')}",
    )
    # O2 (findings 58/77): reconcile this 90-day view with the plan's stated
    # duration so a longer campaign never reads as a contradiction. The
    # duration string comes from the SAME resolver every other sheet uses.
    _plan_duration = _resolve_campaign_duration(data)
    _cw = data.get("campaign_weeks")
    try:
        _cw_int = int(_cw) if _cw else 0
    except (ValueError, TypeError):
        _cw_int = 0
    if _cw_int <= 0:
        _cw_int = display_format.parse_duration_to_weeks(_plan_duration)
    row = _write_kv_row(ws, row, "Campaign Duration", _plan_duration)

    # S89: for a campaign longer than ~13 weeks (~90 days), this forecast
    # must burn only the ramp-weighted FIRST-90-DAYS share of budget, not
    # the plan's entire budget -- a plan can't spend 100% of an 18-month
    # budget inside the first 90 days and also claim to sustain hiring
    # for the other 15 months on the same dollars.
    if _cw_int > 13:
        _ninety_day_scale = 13.0 / _cw_int
    else:
        _ninety_day_scale = 1.0
    _remaining_budget = total_budget * (1 - _ninety_day_scale)

    # S89 (finding data:manpower#5/atria#5): state the ACTUAL computed
    # monthly split (which may be seasonally shifted away from the base
    # 25/35/40 ramp-up curve by _seasonal_monthly_phasing), never a
    # hardcoded "25/35/40" that can silently drift out of sync with the
    # monthly spend row below it.
    _ramp_pct_str = "/".join(f"{p * 100:.0f}" for p in monthly_pcts)

    if _cw_int > 13:
        row = _write_footnote(
            ws,
            row,
            f"Stated campaign duration is {_plan_duration}. This forecast shows "
            f"only the first-90-days share of the plan's budget -- "
            f"{_ninety_day_scale * 100:.0f}% ({_fmt_currency(total_budget * _ninety_day_scale)}) "
            f"of the total {_fmt_currency(total_budget)} budget, ramp-weighted "
            f"{_ramp_pct_str} across the first three months. The remaining "
            f"{(1 - _ninety_day_scale) * 100:.0f}% ({_fmt_currency(_remaining_budget)}) is "
            "deployed across the rest of the campaign at the same channel "
            "economics -- it is not spent, and not reflected, in the 90-day "
            "figures below.",
        )
    row += 1

    # ── Summary Forecast Table ──
    row = _write_subsection_header(ws, row, "Monthly Projections Overview")

    headers = ["Metric"] + month_labels + ["90-Day Total", "Trend"]
    row = _write_table_header(ws, row, headers)

    # Calculate monthly values -- scaled to the first-90-days budget share
    # (S89: never the full-campaign total for a campaign > 13 weeks).
    _budget_90d = total_budget * _ninety_day_scale
    _apps_90d = total_apps * _ninety_day_scale
    _hires_90d = total_hires * _ninety_day_scale

    monthly_spend = [_budget_90d * p for p in monthly_pcts]
    # S89: reconcile_monthly_to_total guarantees the printed monthly ints
    # foot EXACTLY to the printed total (largest-remainder rounding) --
    # plain int() truncation per month can under-count the total by 1-2.
    monthly_apps = display_format.reconcile_monthly_to_total(
        [_apps_90d * p for p in monthly_pcts], _apps_90d
    )
    monthly_hires = display_format.reconcile_monthly_to_total(
        [_hires_90d * p for p in monthly_pcts], _hires_90d
    )
    total_apps_90d = sum(monthly_apps)
    total_hires_90d = sum(monthly_hires)

    # CPA = spend/applications DERIVED per month (never a flat multiplier
    # applied to a single base rate) -- total CPA = total 90-day spend /
    # total 90-day applications, consistent with the monthly rows.
    monthly_cpa = [
        (monthly_spend[i] / monthly_apps[i]) if monthly_apps[i] > 0 else 0
        for i in range(3)
    ]
    base_cpa = _budget_90d / max(total_apps_90d, 1) if total_apps_90d > 0 else 0

    # S89: carry raw numbers + a per-row Excel number_format so the forecast
    # is summable/sortable; units live in the format, not the cell text.
    # S3: Spend/CPA are the plan's OWN figures -- active plan currency.
    forecast_rows = [
        ("Spend", monthly_spend, _budget_90d, "—", _usd0_fmt()),
        (
            "Applications",
            monthly_apps,
            total_apps_90d,
            "Increasing" if total_apps_90d > 0 else "—",
            FMT_INT,
        ),
        (
            "Hires",
            monthly_hires,
            total_hires_90d,
            "Increasing" if total_hires_90d > 0 else "—",
            FMT_INT,
        ),
        (
            "CPA (Cost Per Application)",
            monthly_cpa,
            base_cpa,
            "Decreasing" if base_cpa > 0 else "—",
            _usd0_fmt(),
        ),
    ]

    for idx, (metric, monthly_vals, total_val, trend, row_fmt) in enumerate(
        forecast_rows
    ):
        numeric_vals = [_safe_num(v) for v in monthly_vals] + [_safe_num(total_val)]
        values = [metric] + numeric_vals + [trend]
        fonts_list = [_FONT_BODY_BOLD] + [_FONT_BODY] * (len(values) - 1)
        # Metric label (text) | month cols (row_fmt) | total (row_fmt) | trend (text)
        number_formats = [None] + [row_fmt] * len(numeric_vals) + [None]

        # Color-code trend
        trend_font = _FONT_BODY
        if trend == "Increasing":
            trend_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=GREEN)
        elif trend == "Decreasing" and "CPA" in metric:
            trend_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=GREEN)
        elif trend == "Decreasing":
            trend_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=RED)
        fonts_list[-1] = trend_font

        row = _write_table_row(
            ws,
            row,
            values,
            alternate=(idx % 2 == 0),
            fonts=fonts_list,
            number_formats=number_formats,
        )

    row += 1

    # ── Per-Channel Monthly Breakdown ──
    if ba_channel_alloc:
        row = _write_subsection_header(ws, row, "Per-Channel Monthly Spend Forecast")

        ch_headers = ["Channel"] + month_labels + ["Total", "% of Budget"]
        row = _write_table_header(ws, row, ch_headers)

        sorted_channels = sorted(
            ba_channel_alloc.items(),
            key=lambda x: _safe_num(
                x[1].get("dollar_amount", x[1].get("dollars") or 0)
                if isinstance(x[1], dict)
                else 0
            ),
            reverse=True,
        )

        for idx, (ch_name, ch_data) in enumerate(sorted_channels):
            if not isinstance(ch_data, dict):
                continue
            ch_dollars = _safe_num(
                ch_data.get("dollar_amount", ch_data.get("dollars") or 0)
            )
            if ch_dollars <= 0:
                continue

            # S89: scale to the same first-90-days budget share as the
            # summary table above -- a channel can't show its FULL
            # campaign allocation deployed inside 90 days on a longer plan.
            _ch_dollars_90d = ch_dollars * _ninety_day_scale
            ch_monthly = [_ch_dollars_90d * p for p in monthly_pcts]
            ch_frac = (ch_dollars / total_budget) if total_budget > 0 else 0

            # S89: numeric month/total spend + fractional % (FMT_PCT1).
            values = (
                [_smart_title(ch_name)]
                + [_safe_num(m) for m in ch_monthly]
                + [_safe_num(_ch_dollars_90d), ch_frac]
            )
            number_formats = (
                [None] + [_usd0_fmt()] * len(ch_monthly) + [_usd0_fmt(), FMT_PCT1]
            )
            row = _write_table_row(
                ws,
                row,
                values,
                alternate=(idx % 2 == 0),
                number_formats=number_formats,
            )

    row += 1

    # ── Optimization Milestones ──
    row = _write_subsection_header(ws, row, "Optimization Milestones")

    milestones = [
        (
            "Week 1-2",
            "Campaign launch, initial bid calibration, creative A/B testing begins",
        ),
        (
            "Week 3-4",
            "First optimization cycle: pause underperforming channels, reallocate budget",
        ),
        ("Week 5-6", "Conversion tracking validated, CPA benchmarks established"),
        ("Week 7-8", "Second optimization: refine targeting, scale winning channels"),
        (
            "Week 9-10",
            "Quality-of-hire feedback loop, adjust for retention correlation",
        ),
        (
            "Week 11-12",
            "Final optimization, prepare renewal recommendations, ROI summary",
        ),
    ]

    for idx, (period, action) in enumerate(milestones):
        row = _write_table_row(
            ws,
            row,
            [period, action],
            alternate=(idx % 2 == 0),
            fonts=[_FONT_BODY_BOLD, _FONT_BODY],
        )

    row += 1
    # S89 (finding data:manpower#5/atria#5): quote this plan's ACTUAL
    # computed monthly split (from _seasonal_monthly_phasing, which may
    # shift budget toward peak hiring months for this industry), not a
    # hardcoded 25/35/40 that can drift out of sync with the monthly Spend
    # row above.
    row = _write_footnote(
        ws,
        row,
        f"This forecast phases budget {monthly_pcts[0] * 100:.0f}% Month 1 "
        f"(learning), {monthly_pcts[1] * 100:.0f}% Month 2 (optimizing), "
        f"{monthly_pcts[2] * 100:.0f}% Month 3 (peak performance) -- this "
        "plan's own computed split, seasonally adjusted for its industry and "
        "start month. Actual distribution may vary based on channel mix and "
        "market conditions.",
    )
    row += 1
    _write_attribution_footer(ws, row)


def _clamped_band(
    lo: float, expected: float, hi: float, cost_metric: bool = False
) -> Optional[Tuple[float, float]]:
    """Clamp a (low, expected, high) confidence band to a valid, non-inverted
    order and report whether it is degenerate.

    ``cost_metric=True`` means "pessimistic" is the HIGHER value (CPA/CPH):
    the returned order is lo >= expected >= hi. Otherwise (count metrics --
    applications/hires) lo <= expected <= hi. Returns ``None`` when the band
    collapses to a single point (low == expected == high) -- callers should
    skip the row entirely rather than print a fake +/-X% range that isn't one.
    """
    if cost_metric:
        lo = max(lo, expected)
        hi = min(hi, expected)
    else:
        lo = min(lo, expected)
        hi = max(hi, expected)
    if lo == expected == hi:
        return None
    return lo, hi


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8: Confidence Intervals
# ═══════════════════════════════════════════════════════════════════════════════


def _build_sheet_confidence_intervals(ws, data: dict) -> None:
    """Build Sheet 8: Confidence Intervals showing low/expected/high estimates.

    Instead of point estimates, shows ranges for CPA, CPH, applications, and hires
    with variance based on data confidence level per channel.
    """
    ws.title = "Confidence Intervals"
    ws.sheet_properties.tabColor = AMBER

    _set_column_widths(
        ws,
        {
            1: 3,  # margin
            2: 22,  # Channel
            3: 14,  # Metric
            4: 16,  # Low (Pessimistic)
            5: 16,  # Expected
            6: 16,  # High (Optimistic)
            7: 14,  # Variance %
            8: 14,  # Confidence
        },
    )

    budget_alloc = data.get("_budget_allocation", {})
    if not isinstance(budget_alloc, dict):
        budget_alloc = {}
    ba_channel_alloc = budget_alloc.get("channel_allocations", {})
    if not isinstance(ba_channel_alloc, dict):
        ba_channel_alloc = {}

    row = 2

    # ── Section Header ──
    row = _write_section_header(ws, row, "Confidence Intervals & Metric Ranges")

    row = _write_kv_row(
        ws,
        row,
        "Methodology",
        "Ranges derived from data confidence levels. HIGH confidence = +/-15% variance, "
        "MEDIUM = +/-20%, LOW = +/-25%. Based on source count and KB validation.",
    )
    row += 1

    # ── Variance explanation ──
    row = _write_subsection_header(ws, row, "Variance Scale")

    var_headers = [
        "Confidence Level",
        "Variance Applied",
        "Description",
        "Typical Sources",
    ]
    row = _write_table_header(ws, row, var_headers)

    var_data = [
        (
            "HIGH",
            "+/- 15%",
            "Multiple validated data sources",
            "2+ independent sources",
        ),
        (
            "MEDIUM",
            "+/- 20%",
            "Single source or benchmark-validated",
            "1 validated source",
        ),
        ("LOW", "+/- 25%", "Estimated or insufficient data", "No direct data sources"),
    ]
    for idx, (level, variance, desc, sources) in enumerate(var_data):
        conf_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=GREEN)
        if level == "MEDIUM":
            conf_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=AMBER)
        elif level == "LOW":
            conf_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=RED)

        row = _write_table_row(
            ws,
            row,
            [level, variance, desc, sources],
            alternate=(idx % 2 == 0),
            fonts=[conf_font, _FONT_BODY, _FONT_BODY, _FONT_BODY],
        )

    row += 1

    # ── Per-Channel Confidence Intervals ──
    row = _write_subsection_header(ws, row, "Per-Channel Metric Ranges")

    headers = [
        "Channel",
        "Metric",
        "Low (Pessimistic)",
        "Expected",
        "High (Optimistic)",
        "Variance",
        "Confidence",
    ]
    row = _write_table_header(ws, row, headers)

    sorted_channels = sorted(
        ba_channel_alloc.items(),
        key=lambda x: _safe_num(
            x[1].get("dollar_amount", x[1].get("dollars") or 0)
            if isinstance(x[1], dict)
            else 0
        ),
        reverse=True,
    )

    idx = 0
    for ch_name, ch_data in sorted_channels:
        if not isinstance(ch_data, dict):
            continue
        dollars = _safe_num(ch_data.get("dollar_amount", ch_data.get("dollars") or 0))
        if dollars <= 0:
            continue

        # Determine confidence and variance.
        # S89 FIX (findings data:manpower#3/atria#3, strategy:atria#8):
        # budget_engine's own per-channel `confidence` field comes back
        # "high" for every channel regardless of the plan's overall Sources
        # & Confidence grade, so it's no longer authoritative -- re-derive
        # from the channel's actual CPC/CPA data tier plus the plan's
        # overall confidence score (single source of truth, also used by
        # Channels & Strategy, ROI Projections, and Channel Recommendations).
        confidence = _derive_channel_confidence(data, ch_data)
        # S89A FIX: variance ladder now sourced from the single shared
        # _confidence_variance() helper (also used by ROI Projections'
        # Hire Range column) instead of a locally duplicated if/elif, so
        # the two sheets can never drift apart again.
        variance = _confidence_variance(confidence)

        conf_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=GREEN)
        if confidence == "MEDIUM":
            conf_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=AMBER)
        elif confidence == "LOW":
            conf_font = Font(name=FONT_BODY_NAME, bold=True, size=10, color=RED)

        ch_label = _smart_title(ch_name)

        # S5 (2026-07-03, findings 44/51): Low/Expected/High are the plan's
        # own live figures -- write as numbers with number_format instead of
        # pre-formatted text strings so the client can sum/sort/chart them.
        _money_fmts = [None, None, _usd2_fmt(), _usd2_fmt(), _usd2_fmt(), None, None]
        _int_fmts = [None, None, FMT_INT, FMT_INT, FMT_INT, None, None]

        # CPA
        cpa = _safe_num(ch_data.get("cpa") or 0)
        if cpa > 0:
            cpa_lo = cpa * (1 + variance)  # Pessimistic = higher CPA
            cpa_hi = cpa * (1 - variance)  # Optimistic = lower CPA
            _cpa_band = _clamped_band(cpa_lo, cpa, cpa_hi, cost_metric=True)
        else:
            _cpa_band = None
        if cpa > 0 and _cpa_band is not None:
            cpa_lo, cpa_hi = _cpa_band
            row = _write_table_row(
                ws,
                row,
                [
                    ch_label,
                    "CPA",
                    cpa_lo,
                    cpa,
                    cpa_hi,
                    f"+/-{int(variance * 100)}%",
                    confidence,
                ],
                alternate=(idx % 2 == 0),
                fonts=[
                    _FONT_BODY_BOLD,
                    _FONT_BODY,
                    _FONT_BODY,
                    _FONT_BODY_BOLD,
                    _FONT_BODY,
                    _FONT_BODY,
                    conf_font,
                ],
                number_formats=_money_fmts,
            )
            idx += 1

        # Applications
        apps = int(_safe_num(ch_data.get("projected_applications") or 0))
        if apps > 0:
            apps_lo = max(0, int(apps * (1 - variance)))
            apps_hi = int(apps * (1 + variance))
            _apps_band = _clamped_band(apps_lo, apps, apps_hi, cost_metric=False)
        else:
            _apps_band = None
        if apps > 0 and _apps_band is not None:
            apps_lo, apps_hi = (int(v) for v in _apps_band)
            row = _write_table_row(
                ws,
                row,
                [
                    ch_label,
                    "Applications",
                    apps_lo,
                    apps,
                    apps_hi,
                    f"+/-{int(variance * 100)}%",
                    confidence,
                ],
                alternate=(idx % 2 == 0),
                fonts=[
                    _FONT_BODY_BOLD,
                    _FONT_BODY,
                    _FONT_BODY,
                    _FONT_BODY_BOLD,
                    _FONT_BODY,
                    _FONT_BODY,
                    conf_font,
                ],
                number_formats=_int_fmts,
            )
            idx += 1

        # Hires
        hires = int(_safe_num(ch_data.get("projected_hires") or 0))
        # Raw (unclamped) low/high used to derive the CPH band below --
        # keep these separate from the (possibly row-skipping) display band.
        hires_lo_raw = max(0, int(hires * (1 - variance))) if hires > 0 else 0
        hires_hi_raw = int(hires * (1 + variance)) if hires > 0 else 0
        # S89A FIX (findings data:manpower#1/#2, data:atria#1): display band
        # routed through the same shared _confidence_range() helper ROI
        # Projections' Hire Range column uses, so the two sheets can never
        # show two different ranges for the same channel again.
        _hires_band = _confidence_range(hires, confidence, cost_metric=False)
        if hires > 0 and _hires_band is not None:
            hires_lo, hires_hi = (int(v) for v in _hires_band)
            row = _write_table_row(
                ws,
                row,
                [
                    ch_label,
                    "Hires",
                    hires_lo,
                    hires,
                    hires_hi,
                    f"+/-{int(variance * 100)}%",
                    confidence,
                ],
                alternate=(idx % 2 == 0),
                fonts=[
                    _FONT_BODY_BOLD,
                    _FONT_BODY,
                    _FONT_BODY,
                    _FONT_BODY_BOLD,
                    _FONT_BODY,
                    _FONT_BODY,
                    conf_font,
                ],
                number_formats=_int_fmts,
            )
            idx += 1

        # CPH (Cost Per Hire)
        if hires > 0 and dollars > 0:
            cph = dollars / hires
            cph_lo = dollars / max(
                hires_lo_raw, 1
            )  # Pessimistic = fewer hires = higher CPH
            cph_hi = dollars / max(
                hires_hi_raw, 1
            )  # Optimistic = more hires = lower CPH
            _cph_band = _clamped_band(cph_lo, cph, cph_hi, cost_metric=True)
        else:
            _cph_band = None
        if hires > 0 and dollars > 0 and _cph_band is not None:
            cph_lo, cph_hi = _cph_band
            row = _write_table_row(
                ws,
                row,
                [
                    ch_label,
                    "Cost Per Hire",
                    cph_lo,
                    cph,
                    cph_hi,
                    f"+/-{int(variance * 100)}%",
                    confidence,
                ],
                alternate=(idx % 2 == 0),
                fonts=[
                    _FONT_BODY_BOLD,
                    _FONT_BODY,
                    _FONT_BODY,
                    _FONT_BODY_BOLD,
                    _FONT_BODY,
                    _FONT_BODY,
                    conf_font,
                ],
                number_formats=_money_fmts,
            )
            idx += 1

    row += 1
    row = _write_footnote(
        ws,
        row,
        "Note: Pessimistic/Optimistic estimates reflect the range of likely outcomes. "
        "For cost metrics (CPA, CPH), pessimistic = higher cost, optimistic = lower cost. "
        "For volume metrics (Applications, Hires), pessimistic = lower volume, optimistic = higher volume.",
    )
    row += 1
    row = _write_footnote(
        ws,
        row,
        "Confidence levels are determined by the number and quality of data sources: "
        "HIGH (2+ independent sources), MEDIUM (1 validated source), LOW (estimated).",
    )
    row += 1
    _write_attribution_footer(ws, row)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 9: Role-Level Niche Board Recommendations
# ═══════════════════════════════════════════════════════════════════════════════


# O2 (2026-07-03, findings 35/48/63/70): a category -> representative ad
# platform map, so the Channel Recommendations sheet can suggest WHICH
# platform to buy each budget_engine channel category on -- WITHOUT ever
# recomputing the plan's spend/apps/hires. All numbers on this sheet now come
# from the same `_budget_allocation.channel_allocations` object the Executive
# Summary / ROI Projections / 90-Day Forecast use, so the two-plans
# contradiction (e.g. 1,637 apps @ $91.63 vs 4,418 @ $33.95 on the SAME
# $150,000) is eliminated by construction and Spend foots to the budget.
_CATEGORY_TO_PLATFORM: Dict[str, str] = {
    "programmatic": "Programmatic Display (DSP)",
    "job_board": "Indeed Sponsored Jobs",
    "niche_board": "Specialty / niche job boards",
    "social": "LinkedIn / Meta / TikTok Ads",
    "regional": "Regional & aggregator boards",
    "employer_branding": "Employer-brand content & career site",
    "career_site": "Career site & organic",
    "referral": "Employee referral program",
    "events": "Career fairs & hiring events",
    "staffing": "Staffing / agency partners",
    "search": "Google / Bing Ads (SEM)",
    "display": "Programmatic Display (DSP)",
    "email": "Email / talent-community nurture",
}


def _recommended_platform_for_channel(channel_name: str) -> str:
    """Suggest a representative ad platform for a budget_engine channel category."""
    return _CATEGORY_TO_PLATFORM.get(
        _roi_category_for_channel(channel_name),
        "Best-fit platform for this channel",
    )


def _build_sheet_channel_recommendations(ws, data: dict) -> None:
    """Build Sheet 10: Channel Recommendations (S48; reconciled in O2).

    O2 (2026-07-03): This sheet now renders the SAME media plan as every other
    sheet. Its numbers are read directly from the plan's single
    ``_budget_allocation`` object (the one produced by
    ``budget_engine.calculate_budget_allocation`` and consumed by the Executive
    Summary, ROI Projections, 90-Day Forecast, and Confidence Intervals
    sheets), NOT recomputed by an independent engine. It re-frames that one plan
    by recommended ad platform and tiers the channels by investment weight, but
    every Spend / Apps / Hires / CPA figure is identical to the rest of the
    workbook, and the Spend column foots to exactly the stated budget.

    Resolves findings 35, 48, 63, 70 (two contradictory plans in one bundle;
    Spend column not summing to the budget).
    """
    ws.title = "Channel Recommendations"
    ws.sheet_properties.tabColor = SAPPHIRE  # PURPLE 5A54BE

    budget_alloc = data.get("_budget_allocation") or {}
    if not isinstance(budget_alloc, dict):
        budget_alloc = {}
    channel_allocs = budget_alloc.get("channel_allocations") or {}
    if not isinstance(channel_allocs, dict):
        channel_allocs = {}

    # Only include funded channels (>$0). Zero-budget channels are surfaced as
    # a "Consider / Test & Learn" note instead of projecting phantom outcomes.
    funded = [
        (name, ch)
        for name, ch in channel_allocs.items()
        if isinstance(ch, dict)
        and _safe_num(ch.get("dollar_amount", ch.get("dollars") or 0)) > 0
    ]

    if not funded:
        # Defensive fallback: no shared allocation available. Rather than emit a
        # second, independently-computed plan (the old behaviour, which caused
        # the two-plans contradiction), state plainly that this view mirrors the
        # Executive Summary and there is nothing extra to show.
        c = ws.cell(row=1, column=2, value="Channel Recommendations")
        c.font = _FONT_SECTION
        c.fill = _FILL_SAPPHIRE
        for col in range(2, 10):
            ws.cell(row=1, column=col).fill = _FILL_SAPPHIRE
        _write_kv_row(
            ws,
            3,
            "Status",
            "Channel-level recommendations mirror the plan shown on the "
            "Executive Summary and ROI Projections sheets. No separate channel "
            "allocation was available for this campaign.",
        )
        for col, w in {2: 28, 3: 22, 4: 12, 5: 14, 6: 12, 7: 12, 8: 12, 9: 12}.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        return

    # ── Totals — summed from the SAME allocation the rest of the plan uses ──
    total_spend = sum(
        _safe_num(ch.get("dollar_amount", ch.get("dollars") or 0)) for _, ch in funded
    )
    total_apps = sum(
        int(_safe_num(ch.get("projected_applications") or 0)) for _, ch in funded
    )
    total_hires = sum(
        int(_safe_num(ch.get("projected_hires") or 0)) for _, ch in funded
    )
    total_clicks = sum(
        int(_safe_num(ch.get("projected_clicks") or 0)) for _, ch in funded
    )
    avg_cpa = round(total_spend / total_apps, 2) if total_apps > 0 else 0.0

    industry = data.get("industry") or "general_entry_level"
    industry_label = _get_industry_label(industry)
    roles = _get_roles(data)
    # data:atria#4 fix: this header block used to show only roles[0]
    # ("Role: Memory Care Associate") for a plan with all 10 roles enumerated
    # elsewhere in the bundle. Use the SAME role list + summary format the
    # Executive Summary's "Roles" count card is built from.
    _role_stat_label, _role_stat_value = _format_roles_stat(roles)

    row = 1

    # ── Title ──
    c = ws.cell(row=row, column=2, value="Channel Recommendations")
    c.font = _FONT_SECTION
    c.fill = _FILL_SAPPHIRE
    for col in range(2, 13):
        ws.cell(row=row, column=col).fill = _FILL_SAPPHIRE
    row += 1

    # ── Reconciliation note (single-plan guarantee) ──
    _rec_note = (
        "This sheet re-frames the SAME plan shown on the Executive Summary, ROI "
        "Projections, and 90-Day Forecast sheets — organised by recommended ad "
        "platform and investment tier. Every Spend, Applications, Hires, and CPA "
        "figure below is identical to those sheets, and the Spend column foots to "
        "the total budget. It is not an alternative scenario."
    )
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
    _nc = ws.cell(row=row, column=2, value=_rec_note)
    _nc.font = _FONT_FOOTNOTE
    _nc.alignment = _ALIGN_WRAP
    ws.row_dimensions[row].height = 42
    row += 2

    # ── Summary stats (all live numbers, currency-localized) ──
    _stat_rows = [
        ("Industry", industry_label, None),
        (_role_stat_label, _role_stat_value, None),
        ("Budget", total_spend, _usd0_fmt()),
        ("Proj. Applications", total_apps, FMT_INT),
        ("Proj. Hires", total_hires, FMT_INT),
        ("Avg CPA", avg_cpa, _usd2_fmt()),
    ]
    for label, val, fmt in _stat_rows:
        ws.cell(row=row, column=2, value=label).font = _FONT_BODY_BOLD
        vcell = ws.cell(row=row, column=3)
        if fmt:
            _write_num(vcell, val, fmt)
        else:
            vcell.value = val
        vcell.font = _FONT_BODY
        row += 1
    row += 1

    # ── Tier the SAME channels by investment weight (share of total spend) ──
    # This is a presentation of the one plan, not a re-allocation: MUST HAVE =
    # the channels carrying the bulk of the budget, SHOULD HAVE = supporting
    # channels. No number changes between tiers.
    ranked = sorted(
        funded,
        key=lambda x: _safe_num(x[1].get("dollar_amount", x[1].get("dollars") or 0)),
        reverse=True,
    )
    must_have: List[tuple] = []
    should_have: List[tuple] = []
    _cum = 0.0
    for name, ch in ranked:
        _cum += _safe_num(ch.get("dollar_amount", ch.get("dollars") or 0))
        # Channels that together make up the first ~80% of spend are MUST HAVE.
        if _cum <= total_spend * 0.80 or not must_have:
            must_have.append((name, ch))
        else:
            should_have.append((name, ch))

    # data:atria#3-related polish: same largest-remainder-rounded
    # percentage map the Executive Summary and Channels & Strategy sheets
    # use, so this sheet's Alloc % column foots to exactly 100.0% and
    # agrees cell-for-cell with the other two sheets.
    _corrected_pct = _corrected_channel_pct_display(channel_allocs)

    headers = [
        "Channel",
        "Recommended Platform",
        "Alloc %",
        "Spend",
        "CPC",
        "CPA",
        "Clicks",
        "Apps",
        "Hires",
        "Confidence",
        "Rationale",
    ]
    # Column number_formats parallel to `headers` (currency-localized).
    _row_formats = [
        None, None, FMT_PCT1, _usd0_fmt(), _usd2_fmt(), _usd2_fmt(),
        FMT_INT, FMT_INT, FMT_INT, None, None,
    ]

    for tier_title, tier_channels, fill in [
        ("MUST HAVE", must_have, _FILL_GREEN_BG),
        ("SHOULD HAVE", should_have, _FILL_BLUE_LIGHT),
    ]:
        if not tier_channels:
            continue

        # Tier header
        c = ws.cell(row=row, column=2, value=tier_title)
        c.font = _FONT_SUBSECTION
        c.fill = fill
        for col in range(2, 13):
            ws.cell(row=row, column=col).fill = fill
        row += 1

        # Column headers
        for ci, hdr in enumerate(headers, start=2):
            c = ws.cell(row=row, column=ci, value=hdr)
            c.font = _FONT_TABLE_HEADER
            c.fill = _FILL_NAVY
            c.alignment = _ALIGN_CENTER
        row += 1

        # Channel rows — values are the plan's own live figures.
        for name, ch in tier_channels:
            _display = _smart_title(str(name))
            _dollars = _safe_num(ch.get("dollar_amount", ch.get("dollars") or 0))
            if name in _corrected_pct:
                _pct = _corrected_pct[name] * 100.0
            else:
                _pct = _safe_num(ch.get("percentage") or 0)
                if _pct <= 0 and total_spend > 0:
                    _pct = round(_dollars / total_spend * 100, 1)
            _cpc = _safe_num(ch.get("cpc") or 0)
            _cpa = _safe_num(ch.get("cpa") or 0)
            # S89 (findings data:manpower#3/atria#3, strategy:atria#8):
            # single-sourced confidence tier -- see _derive_channel_confidence.
            _conf = _derive_channel_confidence(data, ch)
            _rationale = (
                f"Buy via {_recommended_platform_for_channel(name)}. "
                f"{_pct:.1f}% of budget; part of the plan's core channel mix."
            )
            values = [
                _display,
                _recommended_platform_for_channel(name),
                _pct / 100.0,
                _dollars,
                _cpc,
                _cpa,
                int(_safe_num(ch.get("projected_clicks") or 0)),
                int(_safe_num(ch.get("projected_applications") or 0)),
                int(_safe_num(ch.get("projected_hires") or 0)),
                _conf,
                _rationale,
            ]
            _fonts = [_FONT_BODY_BOLD] + [_FONT_BODY] * (len(values) - 2) + [
                _FONT_FOOTNOTE
            ]
            row = _write_table_row(
                ws,
                row,
                values,
                fonts=_fonts,
                number_formats=_row_formats,
            )
            if len(_rationale) > 70:
                ws.row_dimensions[row - 1].height = 34
        row += 1

    # ── Total row — proves the sheet foots to the budget ──
    total_values = [
        "TOTAL",
        "",
        1.0,
        total_spend,
        "",
        avg_cpa,
        total_clicks,
        total_apps,
        total_hires,
        "",
        "",
    ]
    _total_fonts = [_FONT_BODY_BOLD] * len(total_values)
    row = _write_table_row(
        ws,
        row,
        total_values,
        fonts=_total_fonts,
        fills=[_FILL_BLUE_PALE] * len(total_values),
        number_formats=[
            None, None, FMT_PCT1, _usd0_fmt(), None, _usd2_fmt(),
            FMT_INT, FMT_INT, FMT_INT, None, None,
        ],
    )
    row += 1

    # ── Zero-budget / consider channels (surfaced, not projected) ──
    _unfunded = [
        _smart_title(str(name))
        for name, ch in channel_allocs.items()
        if isinstance(ch, dict)
        and _safe_num(ch.get("dollar_amount", ch.get("dollars") or 0)) <= 0
    ]
    if _unfunded:
        row = _write_kv_row(
            ws,
            row,
            "Test & Learn (unfunded)",
            "Not allocated budget in this plan; consider for future phases: "
            + ", ".join(_unfunded[:8]),
        )
        row += 1

    # ── Source line ──
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
    ws.cell(
        row=row,
        column=2,
        value=(
            "Source: same Nova AI budget allocation as the Executive Summary "
            "and ROI Projections sheets. Platform suggestions map each channel "
            "category to a representative ad platform; the plan's economics are "
            "unchanged."
        ),
    ).font = _FONT_FOOTNOTE
    ws.cell(row=row, column=2).alignment = _ALIGN_WRAP

    # ── Column widths ──
    widths = {
        2: 24,
        3: 26,
        4: 10,
        5: 14,
        6: 10,
        7: 10,
        8: 10,
        9: 10,
        10: 10,
        11: 12,
        12: 46,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _build_sheet_niche_board_matching(ws, data: dict) -> None:
    """Build Sheet 9: Role-Level Niche Board Matching.

    Cross-references target roles against ROLE_NICHE_BOARDS to recommend
    specialty job boards tailored to each role type.
    """
    ws.title = "Niche Board Matching"
    ws.sheet_properties.tabColor = MAGENTA_HEX  # Joveo MAGENTA

    _set_column_widths(
        ws,
        {
            1: 3,  # margin
            2: 24,  # Role Title
            3: 22,  # Recommended Board
            4: 22,  # Board URL
            5: 40,  # Why This Board
            6: 16,  # Match Type
        },
    )

    roles = _get_roles(data)
    industry = data.get("industry", "general_entry_level")

    row = 2

    # ── Section Header ──
    row = _write_section_header(ws, row, "Role-Level Niche Board Recommendations")

    # S89 (findings strategy:manpower#8 / data:atria#7): quote THIS plan's
    # own modeled niche-board apply-to-hire rate (computed live from the
    # SAME channel allocation ROI Projections reads) as the headline
    # number, with the generic industry benchmark as secondary context --
    # never hardcode the benchmark as if it were this plan's result. Pulled
    # live so this reads correctly whether the hire-distribution model
    # gives niche boards 0 hires or a nonzero allocation.
    _niche_rate, _niche_apps, _niche_hires = _niche_board_implied_rate(data)
    if _niche_rate is not None:
        # S89A FIX (findings strategy:atria#1 residual / strategy:manpower
        # niche framing): the bare "0.5% vs industry benchmark 10-15%" read
        # as a self-indictment. Reframe against the plan's own disclosed
        # CPH-anchored methodology (ROI Projections!B24: hires are anchored
        # to cost-per-hire benchmarks, not a funnel) -- the low implied rate
        # is an expected side effect of that methodology, a conservative
        # lower bound, not an error -- and quantify the honest upside if
        # niche boards convert at even the general-board floor (5%),
        # computed live so it's never fabricated.
        _upside_rate = 0.05  # floor of the cited "5-8% on general boards" range
        _upside_hires = max(0, int(_niche_apps * _upside_rate) - _niche_hires)
        _rate_str = display_format.fmt_pct(_niche_rate, decimals=1, is_fraction=True)
        if _upside_hires > 0:
            _upside_clause = (
                f"if niche boards convert at even 5% (the general-board floor), "
                f"expect {_upside_hires} additional hire"
                f"{'s' if _upside_hires != 1 else ''} from this channel -- "
                "treat as upside, not plan-of-record."
            )
        else:
            _upside_clause = (
                "this plan's modeled niche-board hires already meet or exceed "
                "what a 5% general-board floor rate would imply."
            )
        _niche_purpose = (
            "Specialty job boards matched to your target roles for higher-quality, "
            "lower-CPA applicants. This plan anchors total hires to cost-per-hire "
            "benchmarks rather than a funnel model (see ROI Projections, "
            "Methodology), so the implied "
            f"{_rate_str} apply-to-hire rate shown here ({_niche_hires} hires / "
            f"{_niche_apps} applications) is a conservative lower bound of this "
            "channel's modeled contribution, not a literal conversion assumption. "
            "Industry benchmark is 10-15% apply-to-hire vs. 5-8% on general "
            f"boards -- {_upside_clause}"
        )
    else:
        _niche_purpose = (
            "Specialty job boards matched to your target roles for higher-quality, "
            "lower-CPA applicants. Industry benchmark: 10-15% apply-to-hire vs. "
            "5-8% on general boards; this plan currently has no niche-board "
            "applications modeled to compare against."
        )
    row = _write_kv_row(ws, row, "Purpose", _niche_purpose)
    row += 1

    # S4: ROLE_NICHE_BOARDS / INDUSTRY_NICHE_CHANNELS are US-domiciled board
    # lists (Dice, ClearedJobs.Net, USAJOBS, Vivian Health, etc.). Suppress
    # both on a non-US plan rather than shipping US boards to a local market.
    _is_us = _is_us_plan(data)

    # ── Role-Based Matches (industry-aware to prevent cross-industry mismatches) ──
    role_matches = (
        _match_roles_to_niche_boards(roles, industry=industry) if _is_us else {}
    )

    if role_matches:
        row = _write_subsection_header(ws, row, "Role-Specific Specialty Boards")

        headers = [
            "Role Title",
            "Recommended Board",
            "URL",
            "Why This Board",
            "Match Type",
        ]
        row = _write_table_header(ws, row, headers)

        idx = 0
        for role, boards in role_matches.items():
            for board in boards:
                values = [
                    role,
                    board.get("name", ""),
                    board.get("url", ""),
                    board.get("strength", ""),
                    "Role-Matched",
                ]
                row = _write_table_row(
                    ws,
                    row,
                    values,
                    alternate=(idx % 2 == 0),
                    fonts=[
                        _FONT_BODY_BOLD,
                        _FONT_BODY,
                        _FONT_BODY,
                        _FONT_BODY,
                        _FONT_BODY,
                    ],
                )
                idx += 1

        row += 1

    # ── Industry-Based Matches ──
    industry_boards = INDUSTRY_NICHE_CHANNELS.get(industry, []) if _is_us else []
    if industry_boards:
        row = _write_subsection_header(ws, row, "Industry-Specific Boards")

        ind_headers = ["Board Name", "Match Type", "Notes"]
        row = _write_table_header(ws, row, ind_headers)

        industry_label = INDUSTRY_LABEL_MAP.get(
            industry, industry.replace("_", " ").title()
        )
        for idx, board_name in enumerate(industry_boards):
            values = [
                board_name,
                "Industry-Matched",
                f"Recommended for {industry_label} roles",
            ]
            row = _write_table_row(ws, row, values, alternate=(idx % 2 == 0))

        row += 1

    # ── No matches fallback ──
    if not role_matches and not industry_boards:
        if not _is_us:
            _signals = _non_us_signals(data)
            _signal_txt = (
                f" (targets {', '.join(_signals[:3])})" if _signals else ""
            )
            row = _write_kv_row(
                ws,
                row,
                "Status",
                "US-domiciled specialty job boards are not shown because this "
                f"plan targets a non-US market{_signal_txt} and no local "
                "niche-board data was available for this campaign. Consider "
                "general-purpose boards available in-market (e.g. Seek, Trade "
                "Me Jobs, LinkedIn) with targeted ad copy and audience filters.",
            )
        else:
            row = _write_kv_row(
                ws,
                row,
                "Status",
                "No specialty board matches found for the specified roles. "
                "Consider general-purpose boards (Indeed, LinkedIn, ZipRecruiter) "
                "with targeted ad copy and audience filters.",
            )
        row += 1

    # ── Niche Board Best Practices ──
    row = _write_subsection_header(ws, row, "Niche Board Best Practices")

    practices = [
        (
            "Budget Allocation",
            "Allocate 10-20% of total budget to niche boards for quality volume",
        ),
        (
            "Job Posting Optimization",
            "Use role-specific keywords and certifications in titles",
        ),
        (
            "Employer Branding",
            "Many niche boards offer enhanced profiles — invest in brand presence",
        ),
        (
            "Tracking",
            "Set up UTM parameters per niche board to measure quality of applicants",
        ),
        (
            "Refresh Cadence",
            "Re-post or refresh listings every 14-21 days for visibility",
        ),
    ]

    for idx, (practice, detail) in enumerate(practices):
        row = _write_table_row(
            ws,
            row,
            [practice, detail],
            alternate=(idx % 2 == 0),
            fonts=[_FONT_BODY_BOLD, _FONT_BODY],
        )

    row += 1
    row = _write_footnote(
        ws,
        row,
        "Niche boards are matched based on role title keyword analysis. "
        "Board availability and pricing may vary. Verify current offerings before purchasing.",
    )
    row += 1
    _write_attribution_footer(ws, row)


def _build_sheet_international_benchmarks(
    ws, data: dict, intl_benchmarks: dict
) -> None:
    """Build the International Benchmarks sheet showing country-level recruitment data.

    Columns: Country, Region, Top Platforms, CPC Range (USD), CPA Range (USD),
    CPH by Tier (USD), Regulatory Notes.
    """
    ws.title = "Intl Benchmarks"

    # Column widths
    for col_idx, width in [
        (1, 3),
        (2, 18),
        (3, 10),
        (4, 30),
        (5, 16),
        (6, 16),
        (7, 22),
        (8, 40),
    ]:
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row = 2
    # Section header
    for col in range(COL_START, COL_END + 1):
        c = ws.cell(row=row, column=col)
        c.fill = _FILL_NAVY
    ws.cell(
        row=row, column=COL_START, value="International Recruitment Benchmarks"
    ).font = _FONT_SECTION
    ws.cell(row=row, column=COL_START).alignment = _ALIGN_LEFT
    row += 1

    # Subtitle
    _region_label = ", ".join(
        r.get("name", k.upper())
        for k, r in (intl_benchmarks.get("regions") or {}).items()
    )
    ws.cell(
        row=row,
        column=COL_START,
        value=f"Regions: {_region_label or 'Global'} | Source: {intl_benchmarks.get('source', 'International Benchmarks 2026')}",
    ).font = _FONT_FOOTNOTE
    row += 2

    # Table header
    headers = [
        "Country",
        "Region",
        "Top Platforms",
        "CPC Range (USD)",
        "CPA Range (USD)",
        "CPH by Tier (USD)",
        "Regulatory Notes",
    ]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=COL_START + i, value=h)
        c.font = _FONT_TABLE_HEADER
        c.fill = _FILL_SAPPHIRE
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_THIN
    row += 1

    countries = intl_benchmarks.get("countries", {})
    for _ck, _cv in sorted(countries.items(), key=lambda x: x[1].get("name", x[0])):
        _name = _cv.get("name", _ck.replace("_", " ").title())
        _region = (_cv.get("region") or "").upper()

        # Top platforms (top 3)
        _platforms = _cv.get("platforms", [])
        _plat_names = ", ".join(p.get("name", "") for p in _platforms[:3])

        # CPC range
        _cpc_parts = []
        for p in _platforms[:3]:
            _cpc = p.get("cpc_usd", {})
            if isinstance(_cpc, dict) and _cpc.get("min") is not None:
                _cpc_parts.append(
                    f"${_cpc.get('min', 0):.2f}-${_cpc.get('max', 0):.2f}"
                )
        _cpc_str = "; ".join(_cpc_parts[:2]) if _cpc_parts else "—"

        # CPA range
        _cpa_parts = []
        for p in _platforms[:3]:
            _cpa = p.get("cpa_usd", {})
            if isinstance(_cpa, dict) and _cpa.get("min") is not None:
                _cpa_parts.append(
                    f"${_cpa.get('min', 0):.0f}-${_cpa.get('max', 0):.0f}"
                )
        _cpa_str = "; ".join(_cpa_parts[:2]) if _cpa_parts else "—"

        # CPH by tier
        _cph = _cv.get("cph_by_tier", {})
        _cph_parts = []
        for tier_key in ("entry_level", "professional", "senior", "executive"):
            tier_data = _cph.get(tier_key, {})
            if isinstance(tier_data, dict) and tier_data.get("usd"):
                _label = tier_key.replace("_", " ").title()
                _cph_parts.append(f"{_label}: ${tier_data['usd']:,}")
        _cph_str = " | ".join(_cph_parts) if _cph_parts else "—"

        # Regulatory
        _reg = _cv.get("regulatory", {})
        _reg_notes = []
        _notice = _reg.get("notice_period_typical_days", {})
        if isinstance(_notice, dict) and _notice.get("common"):
            _reg_notes.append(f"Notice: {_notice['common']}d")
        _key_regs = _reg.get("key_regulations", [])
        if _key_regs:
            _reg_notes.append(_key_regs[0][:80])
        _reg_str = "; ".join(_reg_notes) if _reg_notes else "—"

        # Write row
        is_alt = (row % 2) == 0
        _fill = _FILL_BLUE_PALE if is_alt else _FILL_WHITE
        values = [_name, _region, _plat_names, _cpc_str, _cpa_str, _cph_str, _reg_str]
        for i, val in enumerate(values):
            c = ws.cell(row=row, column=COL_START + i, value=val)
            c.font = _FONT_BODY
            c.alignment = _ALIGN_WRAP
            c.border = _BORDER_THIN
            c.fill = _fill
        row += 1

    row += 1
    # Footnote
    ws.cell(
        row=row,
        column=COL_START,
        value="All USD figures use March 2026 mid-market exchange rates. "
        "CPC/CPA from top 3 platforms per country. CPH = Cost-Per-Hire by role tier. "
        "Source: 28 industry reports aggregated in international_benchmarks_2026.json.",
    ).font = _FONT_FOOTNOTE
    ws.merge_cells(
        start_row=row,
        start_column=COL_START,
        end_row=row,
        end_column=COL_END,
    )
    row += 1
    _write_attribution_footer(ws, row)


def generate_excel_v2(
    data: dict,
    research_mod=None,
    load_kb_fn=None,
    classify_tier_fn=None,
    fetch_logo_fn=None,
) -> bytes:
    """Generate a consolidated 5-sheet media plan Excel file.

    Args:
        data: The enriched data dict (same as generate_excel receives).
        research_mod: The research module for live data calls.
        load_kb_fn: Function to load knowledge base.
        classify_tier_fn: Function to classify role tiers.
        fetch_logo_fn: Function to fetch client logo.

    Returns:
        bytes: The Excel file as bytes.
    """
    try:
        return _generate_excel_v2_inner(
            data, research_mod, load_kb_fn, classify_tier_fn, fetch_logo_fn
        )
    except Exception as exc:
        logger.error("generate_excel_v2 top-level crash: %s", exc, exc_info=True)
        # Return a minimal error workbook so the caller always gets valid bytes
        try:
            err_wb = Workbook()
            err_ws = err_wb.active
            err_ws.title = "Error"
            err_ws.cell(row=1, column=1, value="Media Plan Generation Error")
            err_ws.cell(
                row=3,
                column=1,
                value=f"An error occurred while generating the Excel report: {exc}",
            )
            err_ws.cell(
                row=5,
                column=1,
                value="Please try again or contact support if the issue persists.",
            )
            err_ws.column_dimensions["A"].width = 80
            err_buf = io.BytesIO()
            err_wb.save(err_buf)
            err_buf.seek(0)
            return err_buf.getvalue()
        except Exception as inner_exc:
            logger.error(
                "generate_excel_v2: even error workbook creation failed: %s",
                inner_exc,
                exc_info=True,
            )
            raise RuntimeError(f"Excel generation failed: {exc}") from exc


def _generate_excel_v2_inner(
    data: dict,
    research_mod=None,
    load_kb_fn=None,
    classify_tier_fn=None,
    fetch_logo_fn=None,
) -> bytes:
    """Inner implementation of generate_excel_v2 (wrapped by top-level try/except)."""
    # S3: resolve and stash the plan's currency FIRST (thread-local) so every
    # _fmt_currency() / _usd0_fmt() / _usd2_fmt() call below -- across every
    # sheet builder -- renders the plan's own budget/spend/CPA/CPH figures in
    # its local currency instead of a hardcoded USD "$".
    _set_active_currency(data)

    # ── Input normalization (mirrors generate_excel for compatibility) ──
    if data.get("budget_range") and not data.get("budget"):
        data["budget"] = data["budget_range"]

    for key, default in [
        ("client_name", "Client"),
        ("company_name", "Client"),
        ("industry", "general_entry_level"),
        ("budget", "Not specified"),
        ("work_environment", "hybrid"),
    ]:
        if not data.get(key):
            data[key] = default

    # Normalize client name casing (preserves known brands)
    data["client_name"] = _proper_client_name(data["client_name"] or "Client")
    data["company_name"] = _proper_client_name(data["company_name"] or "Client")

    for key in ["locations", "roles", "target_roles", "campaign_goals", "competitors"]:
        val = data.get(key)
        if val is None:
            data[key] = []
        elif isinstance(val, str):
            data[key] = [val]

    # Normalize work_environment: frontend sends array, we need a string
    we = data.get("work_environment", "hybrid")
    if isinstance(we, list):
        data["work_environment"] = we[0] if we else "hybrid"
    elif not isinstance(we, str):
        data["work_environment"] = str(we) if we else "hybrid"

    # Normalize role titles
    roles = _get_roles(data)
    data["roles"] = roles
    data["target_roles"] = roles

    # Ensure tier data exists
    if not data.get("_role_tiers") and classify_tier_fn:
        role_tiers = {}
        for role in roles:
            try:
                role_tiers[role] = classify_tier_fn(role)
            except Exception:
                role_tiers[role] = {"tier": "Professional", "sourcing_strategy": ""}
        data["_role_tiers"] = role_tiers

        tier_groups = {}
        for role, tier_info in role_tiers.items():
            tier_name = tier_info.get("tier", "Professional")
            if tier_name not in tier_groups:
                tier_groups[tier_name] = {
                    "count": 0,
                    "roles": [],
                    "tier_info": tier_info,
                }
            tier_groups[tier_name]["count"] += 1
            tier_groups[tier_name]["roles"].append(role)
        data["_tier_groups"] = tier_groups

    # Ensure enriched/synthesized dicts exist
    if not data.get("_enriched"):
        data["_enriched"] = {}
    if not data.get("_synthesized"):
        data["_synthesized"] = {}
    if not data.get("_budget_allocation"):
        data["_budget_allocation"] = {}

    # ── Create workbook ──
    wb = Workbook()

    client_name = data.get("client_name", "Client")
    wb.properties.title = f"Recruitment Media Plan - {client_name}"
    wb.properties.creator = "Nova AI by Joveo"
    wb.properties.subject = f"AI-generated recruitment media plan for {client_name}"
    wb.properties.keywords = (
        f"recruitment media plan, "
        f"{data.get('industry') or ''.replace('_', ' ').title()}, "
        "job advertising"
    )
    wb.properties.description = (
        "Generated by Nova AI Media Plan Generator. "
        "Consolidated 5-sheet format with ROI projections."
    )
    wb.properties.category = "Recruitment Advertising"
    wb.properties.lastModifiedBy = "Nova AI by Joveo"
    # Bug #17 fix: Strip application metadata that leaks server tech (openpyxl version)
    wb.properties.application = "Nova AI Suite"
    wb.properties.appVersion = ""

    # ── Sheet 1: Executive Summary ──
    ws1 = wb.active  # Use the default first sheet
    try:
        _build_sheet_executive_summary(
            ws1,
            data,
            research_mod=research_mod,
            load_kb_fn=load_kb_fn,
            classify_tier_fn=classify_tier_fn,
        )
    except Exception as exc:
        logger.error("Executive Summary sheet failed: %s", exc, exc_info=True)
        # Critical sheet -- re-raise to fail the generation
        raise RuntimeError(f"Failed to build Executive Summary: {exc}") from exc

    # ── Sheet 2: Channels & Strategy ──
    ws2 = wb.create_sheet()
    try:
        _build_sheet_channels(
            ws2,
            data,
            research_mod=research_mod,
            load_kb_fn=load_kb_fn,
        )
    except Exception as exc:
        logger.error("Channel Strategy sheet failed: %s", exc, exc_info=True)
        # Critical sheet -- re-raise to fail the generation
        raise RuntimeError(f"Failed to build Channel Strategy: {exc}") from exc

    # ── Sheet 3: Market Intelligence ──
    ws3 = wb.create_sheet()
    try:
        _build_sheet_market_intelligence(
            ws3,
            data,
            research_mod=research_mod,
        )
    except Exception as exc:
        logger.error("Sheet 3 (Market Intelligence) failed: %s", exc, exc_info=True)
        ws3.title = "Market Intelligence"
        ws3.cell(
            row=2, column=2, value=f"Error generating Market Intelligence sheet: {exc}"
        ).font = _FONT_BODY

    # ── Sheet 4: Sources & Confidence ──
    ws4 = wb.create_sheet()
    try:
        _build_sheet_sources(ws4, data)
    except Exception as exc:
        logger.error("Sheet 4 (Sources & Confidence) failed: %s", exc, exc_info=True)
        ws4.title = "Sources & Confidence"
        ws4.cell(
            row=2, column=2, value=f"Error generating Sources sheet: {exc}"
        ).font = _FONT_BODY

    # ── Sheet 5: ROI Projections ──
    ws5 = wb.create_sheet()
    try:
        _build_sheet_roi_projections(ws5, data, load_kb_fn=load_kb_fn)
    except Exception as exc:
        logger.error("Sheet 5 (ROI Projections) failed: %s", exc, exc_info=True)
        ws5.title = "ROI Projections"
        ws5.cell(
            row=2, column=2, value=f"Error generating ROI Projections sheet: {exc}"
        ).font = _FONT_BODY

    # ── Sheet 6: Quality Intelligence (Gold Standard gates) ──
    gold_standard = data.get("_gold_standard") or {}
    if gold_standard:
        ws6 = wb.create_sheet()
        try:
            _build_sheet_quality_intelligence(ws6, data, gold_standard)
        except Exception as exc:
            logger.error(
                "Sheet 6 (Quality Intelligence) failed: %s", exc, exc_info=True
            )
            ws6.title = "Quality Intelligence"
            ws6.cell(
                row=2,
                column=2,
                value=f"Error generating Quality Intelligence sheet: {exc}",
            ).font = _FONT_BODY

    # ── Sheet 7: 90-Day Rolling Forecast ──
    ws7 = wb.create_sheet()
    try:
        _build_sheet_rolling_forecast(ws7, data)
    except Exception as exc:
        logger.error("Sheet 7 (90-Day Forecast) failed: %s", exc, exc_info=True)
        ws7.title = "90-Day Forecast"
        ws7.cell(
            row=2, column=2, value=f"Error generating 90-Day Forecast sheet: {exc}"
        ).font = _FONT_BODY

    # ── Sheet 8: Confidence Intervals ──
    ws8 = wb.create_sheet()
    try:
        _build_sheet_confidence_intervals(ws8, data)
    except Exception as exc:
        logger.error("Sheet 8 (Confidence Intervals) failed: %s", exc, exc_info=True)
        ws8.title = "Confidence Intervals"
        ws8.cell(
            row=2, column=2, value=f"Error generating Confidence Intervals sheet: {exc}"
        ).font = _FONT_BODY

    # ── Sheet 9: Niche Board Matching ──
    ws9 = wb.create_sheet()
    try:
        _build_sheet_niche_board_matching(ws9, data)
    except Exception as exc:
        logger.error("Sheet 9 (Niche Board Matching) failed: %s", exc, exc_info=True)
        ws9.title = "Niche Board Matching"
        ws9.cell(
            row=2, column=2, value=f"Error generating Niche Board Matching sheet: {exc}"
        ).font = _FONT_BODY

    # ── Sheet 10: Channel Recommendations (S48; reconciled in O2) ──
    # O2 (2026-07-03): this sheet now derives entirely from the shared
    # `_budget_allocation`, so it no longer requires the channel_recommender
    # module. Build it whenever a budget allocation exists.
    if _HAS_CHANNEL_RECOMMENDER or (data.get("_budget_allocation") or {}).get(
        "channel_allocations"
    ):
        ws10 = wb.create_sheet()
        try:
            _build_sheet_channel_recommendations(ws10, data)
        except Exception as exc:
            logger.error(
                "Sheet 10 (Channel Recommendations) failed: %s", exc, exc_info=True
            )
            ws10.title = "Channel Recommendations"
            ws10.cell(
                row=2,
                column=2,
                value=f"Error generating Channel Recommendations sheet: {exc}",
            ).font = _FONT_BODY

    # ── Sheet 11: International Benchmarks (conditional -- only when intl data present) ──
    intl_benchmarks = data.get("_intl_benchmarks")
    if intl_benchmarks and intl_benchmarks.get("countries"):
        ws11 = wb.create_sheet()
        try:
            _build_sheet_international_benchmarks(ws11, data, intl_benchmarks)
        except Exception as exc:
            logger.error(
                "Sheet 11 (International Benchmarks) failed: %s", exc, exc_info=True
            )
            ws11.title = "Intl Benchmarks"
            ws11.cell(
                row=2,
                column=2,
                value=f"Error generating International Benchmarks sheet: {exc}",
            ).font = _FONT_BODY

    # ── S89: cross-cutting polish (freeze panes + brand tab colors) ──
    try:
        _finalize_workbook(wb)
    except Exception as exc:  # noqa: BLE001 - never block save on cosmetics
        logger.error("Workbook finalize failed (non-fatal): %s", exc, exc_info=True)

    # ── Write to bytes ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
