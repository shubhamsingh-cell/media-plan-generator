"""archive/excel_legacy.py end-to-end fallback-generation regression test (2026-08-04).

Companion to tests/test_excel_legacy_containment.py, which documented (but,
per its own explicit scope, did not fix) a NameError chain that made
archive/excel_legacy.py's generate_excel(data) -- the sync-path /api/generate
fallback wired at app.py ~line 5199, invoked at app.py ~19317 when
excel_v2.generate_excel_v2 raises mid-generation, and at app.py ~17162 on the
async X-Async:true path when excel_v2 fails to import entirely -- crash on
every single invocation since this file was extracted from app.py in commit
0f6b70a7 (2026-03-25, "Session 12"). The names it depended on (defined in
app.py before extraction) were never carried into this module's namespace.

Fixed on branch fix/excel-legacy-fallback-crash. The chain went deeper than
the four names first spotted (load_channels_db, classify_role_tier,
fetch_client_logo, load_joveo_publishers/global_supply_data, and the
`research` module) -- iterating the repro to a clean return also turned up
load_knowledge_base and INDUSTRY_NICHE_CHANNELS, plus
_HAS_COLLAR_INTEL/_HAS_TREND_ENGINE (collar_intelligence/trend_engine were
never imported into this module either). Wiring:

  - app.py-sourced names (load_channels_db, load_joveo_publishers,
    global_supply_data, classify_role_tier, fetch_client_logo,
    load_knowledge_base, INDUSTRY_NICHE_CHANNELS): imported via a *deferred*
    `from app import ...` inside generate_excel()'s own body, not at module
    level. A module-level import would recurse into a partially-initialized
    `app` module whenever archive.excel_legacy is imported before app.py --
    exactly what this test file (and the original repro command) does --
    since app.py itself imports archive.excel_legacy at its own load time
    (~line 5199). Deferring to call time breaks the cycle: by the time the
    import actually executes, archive.excel_legacy is already fully
    initialized (this function object already exists), so app.py's own
    `from archive.excel_legacy import generate_excel` succeeds cleanly
    on that inner circular pass. Same deferred-import pattern already used by
    data_orchestrator.py's `_register_default_handlers`.
  - research: a standalone module with no app.py dependency, so an ordinary
    top-level `import research` (matching app.py's own `import research`).
  - collar_intelligence / trend_engine (aliased _collar_intel_mod /
    _trend_engine_mod to match the names this file's generate_excel() body
    already referenced): ordinary top-level optional imports, same
    try/except ImportError pattern already used by budget_engine.py,
    budget_simulator.py, ppt_generator.py, etc.

This test calls generate_excel() for real end-to-end (unlike
test_excel_legacy_containment.py's source-level checks, written when the
function could not run at all) and proves the fallback now actually
produces a genuinely valid, non-trivial xlsx workbook.

Runs under pytest, or standalone: ``python3 tests/test_excel_legacy_fallback_generation.py``.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import load_workbook  # noqa: E402

from archive.excel_legacy import generate_excel  # noqa: E402


def _sample_payload() -> dict:
    return {
        "client_name": "Acme Corp",
        "industry": "healthcare_medical",
        "locations": ["United States"],
        "roles": ["Registered Nurse"],
        "budget": "100000",
        "hire_volume": "20",
    }


def test_generate_excel_returns_substantial_bytes_without_raising():
    """The core regression this test guards against: generate_excel() used
    to crash with ``NameError: name 'load_channels_db' is not defined`` (and,
    once that was patched in isolation, a further chain of NameErrors) on
    every single invocation, before rendering a single cell. This call must
    now complete and return real xlsx bytes."""
    result = generate_excel(_sample_payload())
    assert isinstance(result, (bytes, bytearray)), (
        f"expected generate_excel() to return raw xlsx bytes, got {type(result)!r}"
    )
    assert len(result) > 1000, (
        "returned byte string is implausibly small for a real xlsx workbook "
        f"({len(result)} bytes) -- likely an empty/corrupt file, not a real plan"
    )


def test_generate_excel_output_round_trips_through_openpyxl_with_real_content():
    """Proves the returned bytes are a genuinely valid xlsx (not just
    non-empty bytes) by loading them back through openpyxl.load_workbook,
    and that the workbook carries real rendered content across multiple
    worksheets -- not an empty shell. A hollow "fix" that merely swallowed
    the NameError and returned truncated/corrupt bytes, or an empty
    workbook, would fail here."""
    result = generate_excel(_sample_payload())
    wb = load_workbook(io.BytesIO(result))

    assert len(wb.sheetnames) >= 3, (
        f"expected the full multi-sheet legacy workbook (executive summary, "
        f"channel strategy, etc.), got only {wb.sheetnames!r}"
    )

    total_non_empty_cells = 0
    sheets_with_content = 0
    for ws in wb.worksheets:
        non_empty = sum(
            1 for row in ws.iter_rows() for cell in row if cell.value is not None
        )
        total_non_empty_cells += non_empty
        if non_empty > 0:
            sheets_with_content += 1

    assert sheets_with_content >= 3, (
        f"expected at least 3 worksheets with real cell content, found "
        f"{sheets_with_content} of {len(wb.sheetnames)} sheets non-empty"
    )
    assert total_non_empty_cells > 100, (
        f"workbook has implausibly little content ({total_non_empty_cells} "
        "non-empty cells total) for a full recruitment media plan"
    )

    found_client_name = any(
        isinstance(cell.value, str) and "Acme Corp" in cell.value
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
    )
    assert found_client_name, (
        "client_name 'Acme Corp' from the input payload was not found "
        "anywhere in the rendered workbook -- suggests input data isn't "
        "actually flowing through to the output"
    )


def test_generate_excel_works_with_multiple_roles_and_locations():
    """Broader-input smoke test: exercises the collar-mix comparison sheet
    branch (guarded by _HAS_COLLAR_INTEL, requires >= 2 roles) and
    multi-location research lookups that the single-role/single-location
    happy path above doesn't reach, guarding against a NameError surviving
    in a less-common branch of this 12,000+ line function."""
    payload = {
        "client_name": "Acme Corp",
        "industry": "warehouse_logistics",
        "locations": ["United States", "United Kingdom"],
        "roles": [
            "Warehouse Associate",
            "Senior Software Engineer",
            "Registered Nurse",
        ],
        "budget": "250000",
        "hire_volume": "50",
        "client_website": "acme.com",
    }
    result = generate_excel(payload)
    assert isinstance(result, (bytes, bytearray)) and len(result) > 1000

    wb = load_workbook(io.BytesIO(result))
    assert len(wb.sheetnames) >= 3
    total_non_empty_cells = sum(
        1
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert total_non_empty_cells > 100


def test_app_py_still_imports_the_real_generate_excel_not_none():
    """Guards against a silent regression specific to this fix's approach:
    app.py wraps its own `from archive.excel_legacy import generate_excel`
    (~line 5199) in a bare `try/except ImportError`, and a circular-import
    failure raises exactly that exception type -- so a reintroduced cycle
    would silently set app.generate_excel = None (fallback quietly disabled)
    rather than crash loudly. Importing archive.excel_legacy BEFORE app (as
    this test module's own top-level import already does, and as the
    original repro command does) is the specific ordering that would surface
    such a cycle, since it forces app.py's own inner import of this module to
    happen while this module is only partially initialized on the outer
    pass -- if that inner import fails, this checks it didn't leave app.py's
    own reference null."""
    import app

    assert app.generate_excel is not None, (
        "app.generate_excel is None -- the archive.excel_legacy import at "
        "app.py ~line 5199 failed (likely a circular import), silently "
        "disabling the legacy Excel fallback"
    )
    assert app.generate_excel is generate_excel, (
        "app.generate_excel resolved to a different function object than "
        "archive.excel_legacy.generate_excel -- investigate module identity"
    )


if __name__ == "__main__":
    import pytest

    raise SystemExit(pytest.main([__file__, "-v"]))
