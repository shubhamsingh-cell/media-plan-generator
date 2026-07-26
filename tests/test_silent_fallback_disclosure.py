"""Tests for making SILENT industry/role/market fallbacks VISIBLE (research.py).

The system repeatedly resolves missing industry/role/market data by quietly
substituting a generic value that looks like a real, specific answer -- the
shipped defect this generalises: a rideshare plan (industry=
hospitality_travel) showed retail/food-service BLS/JOLTS statistics because
research.get_labour_market_intelligence's INDUSTRY_LABOUR_MARKET dict only
covers 12 of ~22 industry keys and silently fell back to
"general_entry_level" for the rest.

This file:
  1. Verifies that fix still holds (industry_metrics_is_generic_fallback).
  2. Applies the SAME treatment to two siblings with the IDENTICAL shape:
     - research.get_competitors' INDUSTRY_COMPETITORS dict (same 12 keys,
       same "general_entry_level" catch-all).
     - research.get_location_info's final "no metro, no state" branch,
       which returned a flat median_salary=60000 / unemployment="~3.5%"
       national placeholder with NOTHING marking it as generic, right
       alongside "population"/"major_employers" fields that DID already
       say "Data not available" / "Varies by area" in the exact same dict.
  3. Verifies excel_v2's Market Intelligence "Location Intelligence" table
     actually discloses the get_location_info substitution to the client
     (asterisk + shared footnote), not just that research.py computes a
     flag nobody reads.

Per the brief: the fix is disclosure, never invented per-industry data --
none of these tests require (or check for) any new curated content.

VACUOUSNESS: run against a throwaway pre-fix worktree (git worktree add
--detach HEAD at the parent commit) -- see the task report for the observed
pre-fix failures (AttributeError / KeyError: 'is_generic_fallback' /
AssertionError: no disclosure text found).
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl  # noqa: E402

import excel_v2  # noqa: E402
import research  # noqa: E402
import shared_utils  # noqa: E402


# The canonical ~22-key industry taxonomy (shared_utils.INDUSTRY_LABEL_MAP)
# vs. the 12 keys INDUSTRY_LABOUR_MARKET / INDUSTRY_COMPETITORS cover.
_ALL_INDUSTRIES = set(shared_utils.INDUSTRY_LABEL_MAP)
_COVERED = set(research.INDUSTRY_LABOUR_MARKET)
_UNCOVERED = sorted(_ALL_INDUSTRIES - _COVERED)


def test_uncovered_industry_list_is_the_documented_ten():
    """Lock the exact coverage gap so a future data addition is a
    deliberate edit to this test, not a silent drift."""
    assert _UNCOVERED == sorted(
        {
            "energy_utilities",
            "insurance",
            "telecommunications",
            "automotive",
            "food_beverage",
            "logistics_supply_chain",
            "hospitality_travel",
            "media_entertainment",
            "construction_real_estate",
            "education",
        }
    )
    # INDUSTRY_COMPETITORS is the IDENTICAL 12-key set -- true "siblings".
    assert set(research.INDUSTRY_COMPETITORS) == _COVERED


# ---------------------------------------------------------------------------
# 1. get_labour_market_intelligence -- verify the prior-wave fix still holds
# ---------------------------------------------------------------------------


def test_labour_market_flag_false_for_covered_industries():
    for industry in sorted(_COVERED - {"general_entry_level"}):
        result = research.get_labour_market_intelligence(industry, ["Dallas, TX"])
        assert result["industry_metrics_is_generic_fallback"] is False, industry
        assert result["industry_metrics"] == research.INDUSTRY_LABOUR_MARKET[industry]


def test_labour_market_flag_true_for_every_uncovered_industry():
    """Every one of the ~10 uncovered industries -- not just the one
    (hospitality_travel) that shipped -- must be flagged AND must actually
    fall back to general_entry_level's data, never its own fabricated
    content."""
    for industry in _UNCOVERED:
        result = research.get_labour_market_intelligence(industry, ["Dallas, TX"])
        assert result["industry_metrics_is_generic_fallback"] is True, industry
        assert (
            result["industry_metrics"]
            == research.INDUSTRY_LABOUR_MARKET["general_entry_level"]
        ), industry


def test_market_intelligence_sheet_discloses_the_substitution_for_uber_shape():
    """End-to-end: the exact real-incident shape (hospitality_travel /
    rideshare, a US plan so the Industry Metrics subsection actually
    renders) must show the disclosure note on the workbook itself."""
    data = {
        "client_name": "Uber",
        "company_name": "Uber",
        "industry": "hospitality_travel",
        "budget": "$500,000",
        "locations": ["Chicago, IL"],
        "roles": ["Rideshare Driver"],
        "target_roles": ["Rideshare Driver"],
        "campaign_duration": "3 months",
        "hire_volume": "500",
        "work_environment": "onsite",
        "_enriched": {},
        "_synthesized": {},
        "_budget_allocation": {},
    }
    raw = excel_v2.generate_excel_v2(data, research_mod=research)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb["Market Intelligence"]
    text = "\n".join(
        str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None
    )
    assert (
        "No BLS/JOLTS data is curated specifically for" in text
    ), "expected the generic-fallback disclosure note on the Industry Metrics subsection"
    assert "general/entry-level workforce" in text


# ---------------------------------------------------------------------------
# 2. get_competitors -- sibling of the same shape (INDUSTRY_COMPETITORS)
# ---------------------------------------------------------------------------


def test_get_competitors_flag_false_for_covered_industry():
    result = research.get_competitors("tech_engineering", ["Austin, TX"], "Acme")
    assert result, "expected at least one competitor category"
    assert all(r["is_generic_fallback"] is False for r in result)


def test_get_competitors_flag_true_for_every_uncovered_industry():
    for industry in _UNCOVERED:
        result = research.get_competitors(industry, ["Austin, TX"], "Acme")
        assert result, f"{industry}: expected fallback categories, got none"
        assert all(r["is_generic_fallback"] is True for r in result), industry
        # Still general_entry_level's OWN categories, verbatim -- never a
        # fabricated per-industry category invented to paper over the gap.
        expected_categories = set(research.INDUSTRY_COMPETITORS["general_entry_level"])
        assert {r["category"] for r in result} <= expected_categories, industry


# ---------------------------------------------------------------------------
# 3. get_location_info -- an unresolvable US location's flat national
#    median_salary/unemployment must be flagged (sibling of #1, market-
#    keyed rather than industry-keyed).
# ---------------------------------------------------------------------------


def test_location_info_flag_absent_for_known_metro():
    info = research.get_location_info("Dallas, TX")
    assert "is_generic_fallback" not in info


def test_location_info_flag_absent_for_known_international_market():
    info = research.get_location_info("London, United Kingdom")
    assert "is_generic_fallback" not in info


def test_location_info_flag_true_for_unresolvable_us_location():
    """A location string research.py has no metro AND no state entry for
    at all -- population/major_employers already said "Data not
    available"/"Varies by area" honestly; median_salary/unemployment must
    now carry the SAME honesty instead of looking like real, specific
    figures for this town."""
    info = research.get_location_info("Smallville")
    assert info["is_generic_fallback"] is True
    assert info["median_salary"] == 60000  # value unchanged -- disclosure, not invention
    assert info["population"] == "Data not available"


def test_location_intelligence_table_discloses_generic_market_fallback():
    """End-to-end: a plan with an unresolvable location must show the
    asterisk + shared footnote on the Market Intelligence sheet, not just
    compute a flag nobody surfaces."""
    data = {
        "client_name": "Acme",
        "company_name": "Acme",
        "industry": "tech_engineering",
        "budget": "$100,000",
        "locations": ["Smallville"],
        "roles": ["Software Engineer"],
        "target_roles": ["Software Engineer"],
        "campaign_duration": "3 months",
        "hire_volume": "20",
        "work_environment": "onsite",
        "_enriched": {},
        "_synthesized": {},
        "_budget_allocation": {},
    }
    raw = excel_v2.generate_excel_v2(data, research_mod=research)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb["Market Intelligence"]
    text = "\n".join(
        str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None
    )
    assert "No location-specific data on file for this market" in text
    # The flagged row's own cells carry the asterisk tying them to that note.
    assert "3.5%*" in text or "~3.5%*" in text


def test_location_intelligence_table_no_asterisk_for_known_metro():
    """False-positive guard: a plan on a well-covered metro must render
    with NO asterisk/disclosure at all."""
    data = {
        "client_name": "Acme",
        "company_name": "Acme",
        "industry": "tech_engineering",
        "budget": "$100,000",
        "locations": ["Dallas, TX"],
        "roles": ["Software Engineer"],
        "target_roles": ["Software Engineer"],
        "campaign_duration": "3 months",
        "hire_volume": "20",
        "work_environment": "onsite",
        "_enriched": {},
        "_synthesized": {},
        "_budget_allocation": {},
    }
    raw = excel_v2.generate_excel_v2(data, research_mod=research)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb["Market Intelligence"]
    text = "\n".join(
        str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None
    )
    assert "No location-specific data on file for this market" not in text


if __name__ == "__main__":
    import pytest

    raise SystemExit(pytest.main([__file__, "-v"]))
