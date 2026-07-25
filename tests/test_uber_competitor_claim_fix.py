"""Regression tests for the Uber-bundle unsourced-competitor-claim incident.

Real shipped defect (Uber, GBP 2M, "commercial cab driver", 6 non-US
markets): because the industry was misclassified, the competitor list was a
static hotel-chain roster, so the workbook told a rideshare client that
Marriott recruits cab drivers --

  Market Intelligence Competitor Analysis: Marriott / Hilton / Hyatt / IHG /
  Airbnb, each stamped "Active (est.)" -- invented precision with no source.

  Quality Intelligence Competitive Landscape: the same names repeated per
  market with a fabricated "Est. Competing Postings" figure (795/750/300/
  ...), and the internal "(National)" scope-tag prefix leaking into ~24
  client-facing cells ("(National) Marriott, (National) Hilton, ...").

  Both sheets' Counter-Strategy prose (insight_composer.compose_counter_
  strategy) asserted specific, never-observed hiring BEHAVIOUR as fact --
  "Marriott is actively competing for ...", "Expect Hilton to keep
  pressure on ...".

  The workbook's OWN Sources & Confidence sheet grades this exact
  Competitive Intelligence at 20%/grade F -- confidence was measured and
  displayed, but gated nothing.

Covers, in excel_v2.py:
  FIX A -- Market Intelligence Competitor Analysis fallback no longer stamps
           "Active (est.)" (invented precision) and discloses the
           competitor set as inferred, wired to the actual fallback path
           (not shown when comp_intel/brief data is real).
  FIX B -- Quality Intelligence Competitive Landscape strips the internal
           "(National)" scope tag before it reaches any client-facing cell,
           drops the invented "Est. Competing Postings" figure entirely
           (gold_standard._estimate_competing_postings is a difficulty x
           role-count formula, never a real posting count), and discloses
           the competitor set as inferred when the brief supplied none.

And in insight_composer.py:
  FIX C -- every compose_counter_strategy skeleton (all 4 competitor_type
           buckets) and the intensity-escalation append no longer assert
           unverified third-party behaviour ("is actively competing for",
           "keeps pressure on", "is drawing from the same", "has been
           especially aggressive", ...) -- only presence/capability framing
           ("X is a plausible competitor for ...") and unconditional advice.

bundle_qa.py's new `unsourced_competitor_claim` rule (RULE 6) is covered
separately in tests/test_bundle_qa_regression.py, including the real
shipped-fixture proof and the false-positive guard against this exact
corrected generator output.

Runs under pytest, or standalone: ``python3 tests/test_uber_competitor_claim_fix.py``.
"""

from __future__ import annotations

import os
import sys

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

import openpyxl  # noqa: E402

import excel_v2  # noqa: E402
import insight_composer as ic  # noqa: E402
import research  # noqa: E402

# ---------------------------------------------------------------------------
# Shared fixtures / helpers
# ---------------------------------------------------------------------------
_INFERRED_LABEL = (
    "Competitor set inferred from industry classification; not verified "
    "against live posting data."
)


def _sheet_text(ws) -> str:
    parts = []
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if val is not None:
                parts.append(str(val))
    return "\n".join(parts)


def _build_market_intel_ws(data: dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    excel_v2._build_sheet_market_intelligence(ws, data, research_mod=research)
    return ws


def _build_quality_intel_ws(data: dict, gold_standard: dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    excel_v2._build_sheet_quality_intelligence(ws, data, gold_standard)
    return ws


def _uber_data(**overrides) -> dict:
    data = {
        "client_name": "Uber",
        "industry": "hospitality_travel",
        "locations": ["UK"],
        "roles": ["Commercial Cab Driver"],
        "target_roles": ["Commercial Cab Driver"],
        "budget": "GBP 2,000,000",
        "competitors": [],
        "_synthesized": {},
        "_enriched": {},
    }
    data.update(overrides)
    return data


# ===========================================================================
# FIX A -- Market Intelligence Competitor Analysis (excel_v2.py, ~:6491-6749)
# ===========================================================================
def test_fixA_static_fallback_omits_active_est_stamp_and_discloses_inferred():
    """Real bug precondition: no comp_intel/brief competitors at all, so the
    static per-industry roster (Marriott/Hilton/Hyatt/IHG/Airbnb for
    hospitality_travel) is the ONLY source -- the old "Active (est.)"
    hiring-activity stamp (invented precision) must be gone, and the sheet
    must disclose that the set is inferred, not verified."""
    text = _sheet_text(_build_market_intel_ws(_uber_data()))
    assert "Active (est.)" not in text
    # Confirms the fallback path actually fired (not a vacuous pass).
    for name in ("Marriott", "Hilton", "Hyatt", "IHG", "Airbnb"):
        assert name in text, f"expected fallback competitor {name!r} missing"
    assert _INFERRED_LABEL in text


def test_fixA_real_enrichment_keeps_legitimate_content_no_inferred_label():
    """False-positive guard: when comp_intel carries REAL competitor
    enrichment (not the static fallback), that legitimate data must render
    unchanged and the "inferred" disclosure must NOT appear -- the fallback
    branch never executes for this plan."""
    data = _uber_data(
        industry="rideshare_gig_economy",
        _synthesized={
            "competitive_intelligence": {
                "competitors": [
                    {
                        "name": "Bolt",
                        "industry": "Rideshare & Gig Economy",
                        "size": "5,000+",
                        "hiring_activity": "42 open reqs (LinkedIn, live)",
                        "overlap_score": "High",
                    },
                    {
                        "name": "Lyft",
                        "industry": "Rideshare & Gig Economy",
                        "size": "3,000+",
                        "hiring_activity": "18 open reqs (LinkedIn, live)",
                        "overlap_score": "Moderate",
                    },
                ],
            },
        },
    )
    text = _sheet_text(_build_market_intel_ws(data))
    assert "Bolt" in text
    assert "Lyft" in text
    assert "42 open reqs (LinkedIn, live)" in text
    # The static hospitality fallback must never have fired for this plan.
    assert "Marriott" not in text
    assert _INFERRED_LABEL not in text


# ===========================================================================
# FIX B -- Quality Intelligence Competitive Landscape (excel_v2.py, ~:9267-9460)
# ===========================================================================
def test_fixB_strips_national_tag_and_drops_invented_posting_count():
    """Real bug shape: gold_standard's competitor synthesizer prepends
    "(National) " to employer names it only has industry-level data for,
    and stamps a formula-derived "estimated_competing_postings" number
    (never a real posting count). Both must be gone from client-facing
    cells; the "Est. Competing Postings" column must be dropped entirely
    rather than render a blank/invented figure."""
    data = _uber_data()
    gold_standard = {
        "competitor_mapping": {
            "london": {
                "top_employers": ["(National) Marriott", "(National) Hilton"],
                "hiring_intensity": "high",
                "estimated_competing_postings": 795,
            },
            "_national": {
                "top_employers": ["(National) Hyatt", "(National) IHG"],
                "hiring_intensity": "moderate",
            },
        }
    }
    text = _sheet_text(_build_quality_intel_ws(data, gold_standard))
    assert "(National)" not in text
    for name in ("Marriott", "Hilton", "Hyatt", "IHG"):
        assert name in text, f"expected competitor {name!r} missing"
    assert "Est. Competing Postings" not in text
    assert "795" not in text
    assert _INFERRED_LABEL in text


def test_fixB_no_inferred_label_when_brief_supplied_real_competitors():
    """False-positive guard: when the client's OWN brief named real
    competitors (data["competitors"]), the "inferred, not verified"
    disclosure must NOT appear even though the same table-rendering code
    path runs -- the legitimate brief-sourced names must still render."""
    data = _uber_data(
        industry="rideshare_gig_economy",
        competitors=["Bolt", "Lyft"],
    )
    gold_standard = {
        "competitor_mapping": {
            "london": {
                "top_employers": ["Bolt", "Lyft"],
                "hiring_intensity": "high",
                "estimated_competing_postings": 500,
            },
        }
    }
    text = _sheet_text(_build_quality_intel_ws(data, gold_standard))
    assert "Bolt" in text
    assert "Lyft" in text
    assert _INFERRED_LABEL not in text
    # Still no invented precision, regardless of provenance -- this figure
    # is never real data (see gold_standard._estimate_competing_postings).
    assert "Est. Competing Postings" not in text
    assert "500" not in text


def test_fixB_national_tag_stripped_from_counter_strategy_prose_too():
    """The scope tag must not leak into the generated Counter-Strategy
    sentence either (only the top employer name is interpolated into it)."""
    data = _uber_data()
    gold_standard = {
        "competitor_mapping": {
            "london": {
                "top_employers": ["(National) Marriott"],
                "hiring_intensity": "moderate",
            },
        }
    }
    text = _sheet_text(_build_quality_intel_ws(data, gold_standard))
    assert "(National) Marriott" not in text
    assert "Marriott" in text


# ===========================================================================
# FIX C -- insight_composer.py skeleton banks + intensity escalation
# ===========================================================================
_BANNED_SUBSTRINGS = (
    "actively recruit",
    "actively staffing",
    "actively competing",
    "keep pressure on",
    "keeps pressure on",
    "puts direct pressure on",
    "put direct pressure on",
    "drawing from the same",
    "is slower to respond",
    "has been especially aggressive",
)


def test_fixC_skeleton_banks_contain_no_banned_asserted_behavior_verbs():
    for bucket, bank in ic._SKELETON_BANKS.items():
        assert len(bank) >= 4, f"{bucket} has fewer than 4 skeletons"
        for sentence in bank:
            rendered = sentence.format(
                competitor="Marriott", angle="commercial cab driver candidates in UK"
            )
            lowered = rendered.lower()
            for banned in _BANNED_SUBSTRINGS:
                assert banned not in lowered, f"{bucket}: {banned!r} in {rendered!r}"


def test_fixC_skeletons_stay_distinct_per_bucket():
    """Regression guard on the pre-existing near-duplicate-prose fix (S92):
    the rewrite must not have accidentally collapsed any bucket's 10
    skeletons down to fewer distinct strings."""
    for bucket, bank in ic._SKELETON_BANKS.items():
        assert len(bank) == len(set(bank)), f"{bucket} has duplicate skeletons"


def test_fixC_intensity_escalation_no_longer_asserts_recent_behavior():
    sentence = ic.compose_counter_strategy(
        "Marriott",
        {
            "role": "commercial cab driver",
            "city": "UK",
            "ordinal": 0,
            "intensity": "high",
        },
    )
    assert "has been especially aggressive" not in sentence
    assert "flagged high-intensity" in sentence


def test_fixC_allowed_presence_framing_still_names_the_competitor():
    """The fix must not have gone so far as to stop naming the competitor
    at all -- presence/capability framing ("is a plausible/likely
    competitor") is explicitly allowed and expected."""
    for bucket in ic._SKELETON_BANKS:
        sentence = ic.compose_counter_strategy(
            "Marriott",
            {
                "role": "commercial cab driver",
                "city": "UK",
                "competitor_type": bucket,
                "ordinal": 0,
            },
        )
        assert "Marriott" in sentence


def test_fixC_output_never_uses_double_hyphen():
    """A prior wave converted this file's raw "--" to em dashes; the
    rewrite must preserve that, not reintroduce "--" in generated text."""
    for bucket, bank in ic._SKELETON_BANKS.items():
        for i in range(len(bank)):
            for intensity in ("", "high"):
                sentence = ic.compose_counter_strategy(
                    "Marriott",
                    {
                        "role": "commercial cab driver",
                        "city": "UK",
                        "competitor_type": bucket,
                        "ordinal": i,
                        "intensity": intensity,
                    },
                )
                assert "--" not in sentence


if __name__ == "__main__":
    import pytest as _pytest

    raise SystemExit(_pytest.main([__file__, "-v"]))
