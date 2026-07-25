"""Regression tests for the Uber-bundle geo-integrity / data-contradiction
incident: a plan with SIX non-US markets (UK, Australia, Mexico, Argentina,
Canada, New Zealand) and ZERO US markets shipped presenting US labour data
as local. Covers all six fixes in excel_v2.py / research.py:

  FIX 1 -- Country column hardcoded "United States" for every location on a
           non-US plan (excel_v2._build_sheet_market_intelligence, the
           per-location Country column). Root cause:
           data_synthesizer.fuse_location_profiles never sets a flat
           "country" key (nests it at loc_profile["country_info"]["name"]),
           so the fallback chain fell straight through to a literal default.
  FIX 2 -- "Macro Economic Context" (FRED US indicators) rendered with no
           geo gate at all.
  FIX 3 -- US BLS/JOLTS "Industry Metrics" rendered with no geo gate, dict
           keys rendered via a naive .replace("_"," ").title() that mangled
           acronyms/country suffixes ("Total Employment Us", "Bls Sector
           Code", "Job Openings Rate Jolts"), and a silent
           general_entry_level fallback for industries missing from
           INDUSTRY_LABOUR_MARKET (10 of ~22 legacy keys, including BOTH
           hospitality_travel and logistics_supply_chain).
  FIX 4 -- research.get_labour_market_intelligence hardcoded a literal "$"
           glyph on non-USD median_salary strings ("$42,000 (GBP)").
  FIX 5 -- Market Positioning summary printed a stale
           comp_intel["market_positioning"]["competitor_count"] (snapshotted
           at synthesis time from an empty competitors dict) that
           contradicted the 5-row Competitor Analysis table rendered just
           above it from a later industry-fallback list.
  FIX 6 -- Channels & Strategy niche-board note asserted "Local specialty
           board data was not available for this campaign" -- false; the
           same workbook's Intl Benchmarks sheet (and research's
           international_benchmarks_2026.json) lists real per-country
           platforms for all six markets.

Each test pairs a non-US six-market plan (matching the shipped incident)
with a US-plan false-positive guard, per the standing rule that regressions
must never break what already works for US plans (FRED macro, BLS sector
data, "United States" countries must render EXACTLY as before).

Runs under pytest, or standalone: ``python3 tests/test_uber_audit_geo_integrity.py``.
"""

from __future__ import annotations

import os
import sys

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

import openpyxl  # noqa: E402

import data_synthesizer  # noqa: E402
import excel_v2  # noqa: E402
import research  # noqa: E402

# ---------------------------------------------------------------------------
# Shared fixtures / helpers
# ---------------------------------------------------------------------------

NON_US_LOCATIONS = ["UK", "Australia", "Mexico", "argentina", "canada", "new zealand"]
US_LOCATIONS = ["Chicago, IL", "New York, NY"]

_FRED_MACRO = {
    "unemployment_rate": 3.7,
    "job_openings": 8000,
    "avg_hourly_earnings": 34.55,
    "fed_funds_rate": 5.25,
    "cpi_inflation": 3.1,
}


def _sheet_text(ws) -> str:
    parts = []
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if val is not None:
                parts.append(str(val))
    return "\n".join(parts)


def _non_us_uber_data(**overrides) -> dict:
    """Real-incident shape: 6 non-US markets, industry as selected in the
    wizard for the shipped Uber brief ("Hospitality & Travel" ->
    hospitality_travel, one of the 10 legacy keys missing from
    INDUSTRY_LABOUR_MARKET)."""
    data = {
        "client_name": "uber",
        "industry": "hospitality_travel",
        "locations": list(NON_US_LOCATIONS),
        "roles": ["Commercial Cab Driver"],
        "target_roles": ["Commercial Cab Driver"],
        "budget": "GBP 2,000,000",
        "competitors": [],
        "_synthesized": {},
        "_enriched": {},
    }
    data.update(overrides)
    return data


def _us_plan_data(industry: str = "tech_engineering", **overrides) -> dict:
    """False-positive guard fixture: a genuinely US-only plan. Default
    industry (tech_engineering) has real curated BLS/JOLTS data so the
    "must still show real data" assertions aren't vacuously true."""
    data = {
        "client_name": "Acme Corp",
        "industry": industry,
        "locations": list(US_LOCATIONS),
        "roles": ["Software Engineer"],
        "target_roles": ["Software Engineer"],
        "budget": "$500,000",
        "competitors": [],
        "_synthesized": {},
        "_enriched": {},
    }
    data.update(overrides)
    return data


def _build_market_intel_ws(data: dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    excel_v2._build_sheet_market_intelligence(ws, data, research_mod=research)
    return ws


def _build_channels_ws(data: dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    excel_v2._build_sheet_channels(ws, data, research_mod=research)
    return ws


# ===========================================================================
# FIX 1 -- Country column hardcode
# ===========================================================================
def test_fix1_non_us_plan_never_shows_united_states_country():
    """Real repro: build authentic loc_profiles via the ACTUAL synthesizer
    (data_synthesizer.fuse_location_profiles), not a hand-rolled stand-in,
    so this breaks if the nesting contract this bug relies on ever changes."""
    data = _non_us_uber_data()
    country_data = {
        "UK": {"name": "United Kingdom", "population": 67_000_000, "region": "Europe"},
        "Australia": {
            "name": "Australia",
            "population": 26_000_000,
            "region": "Oceania",
        },
        "Mexico": {
            "name": "Mexico",
            "population": 128_000_000,
            "region": "North America",
        },
        "argentina": {
            "name": "Argentina",
            "population": 45_000_000,
            "region": "South America",
        },
        "canada": {
            "name": "Canada",
            "population": 38_000_000,
            "region": "North America",
        },
        "new zealand": {
            "name": "New Zealand",
            "population": 5_000_000,
            "region": "Oceania",
        },
    }
    enriched = {"country_data": country_data}
    loc_profiles = data_synthesizer.fuse_location_profiles(enriched, {}, data)

    # Precondition sanity: confirm the synthesizer really nests country under
    # country_info.name and never sets a flat "country" key -- the exact
    # contract the bug depended on.
    for loc, profile in loc_profiles.items():
        assert "country" not in profile, f"{loc} unexpectedly has a flat country key"
        assert profile.get("country_info", {}).get("name"), f"{loc} missing country_info"

    data["_synthesized"] = {"location_profiles": loc_profiles}
    text = _sheet_text(_build_market_intel_ws(data))

    assert "United States" not in text
    for expected in [
        "United Kingdom",
        "Australia",
        "Mexico",
        "Argentina",
        "Canada",
        "New Zealand",
    ]:
        assert expected in text, f"expected country {expected!r} missing from sheet"


def test_fix1_us_plan_still_shows_united_states_country():
    """False-positive guard: a genuine US plan (with "City, ST" locations,
    the common case) must keep resolving to "United States" exactly as
    before."""
    data = _us_plan_data()
    text = _sheet_text(_build_market_intel_ws(data))
    assert "United States" in text
    for non_us in ["United Kingdom", "Australia", "Mexico", "Argentina", "Canada"]:
        assert non_us not in text


# ===========================================================================
# FIX 2 -- US macro block (FRED) gated on geography
# ===========================================================================
def test_fix2_macro_block_omitted_for_non_us_plan():
    data = _non_us_uber_data(
        _synthesized={"job_market_demand": {"r1": {"macro_economic": dict(_FRED_MACRO)}}}
    )
    text = _sheet_text(_build_market_intel_ws(data))
    assert "Macro Economic Context" not in text
    assert "Fed Funds Rate" not in text


def test_fix2_macro_block_present_for_us_plan():
    data = _us_plan_data(
        _synthesized={"job_market_demand": {"r1": {"macro_economic": dict(_FRED_MACRO)}}}
    )
    text = _sheet_text(_build_market_intel_ws(data))
    assert "Macro Economic Context" in text
    assert "Fed Funds Rate" in text


# ===========================================================================
# FIX 3 -- US sector metrics gated + label mangling + silent fallback
# ===========================================================================
def test_fix3_industry_metrics_omitted_for_non_us_plan():
    data = _non_us_uber_data()
    text = _sheet_text(_build_market_intel_ws(data))
    assert "Industry Metrics" not in text
    assert "BLS" not in text
    assert "JOLTS" not in text
    assert "Federal minimum wage" not in text  # US-only prose leak


def test_fix3_industry_metrics_present_and_labels_not_mangled_for_us_plan():
    """False-positive guard + label fix: tech_engineering has real curated
    BLS/JOLTS data -- must still render, with acronyms/country-suffix
    correctly cased."""
    data = _us_plan_data(industry="tech_engineering")
    text = _sheet_text(_build_market_intel_ws(data))
    assert "Industry Metrics" in text
    assert "Total Employment US" in text
    assert "Total Employment Us" not in text
    assert "BLS Sector Code" in text
    assert "Bls Sector Code" not in text
    assert "Job Openings Rate JOLTS" in text
    assert "Job Openings Rate Jolts" not in text


def test_fix3_generic_fallback_disclosed_not_silent_on_us_plan():
    """hospitality_travel has no curated entry in INDUSTRY_LABOUR_MARKET and
    silently substituted general_entry_level's retail/food-service stats.
    On a US plan the section must still render (gated on geo, not on
    coverage) but must disclose the substitution."""
    data = _us_plan_data(industry="hospitality_travel")
    text = _sheet_text(_build_market_intel_ws(data))
    assert "Industry Metrics" in text
    assert "generic cross-sector" in text.lower()


# ===========================================================================
# FIX 4 -- median_salary currency glyph must match the ISO code
# ===========================================================================
def test_fix4_median_salary_glyph_matches_currency_code():
    ctx = research.get_labour_market_intelligence(
        "hospitality_travel", ["UK", "Australia", "Mexico", "argentina"]
    )
    by_loc = {lc["location"]: lc["median_salary"] for lc in ctx["location_contexts"]}
    assert by_loc["UK"] == "£42,000 (GBP)"
    assert by_loc["Australia"] == "A$55,000 (AUD)"
    assert by_loc["Mexico"] == "MX$9,500 (MXN)"
    assert by_loc["argentina"] == "AR$8,000 (ARS)"
    # The exact pre-fix defect string must never appear again.
    for val in by_loc.values():
        assert val != "$42,000 (GBP)"


def test_fix4_us_median_salary_stays_dollar_and_usd():
    ctx = research.get_labour_market_intelligence("tech_engineering", ["Chicago"])
    lc = ctx["location_contexts"][0]
    assert lc["country"] == "United States"
    assert lc["median_salary"].startswith("$")
    assert "USD" in lc["median_salary"]


# ===========================================================================
# FIX 5 -- Competitor count must match the rendered table
# ===========================================================================
def test_fix5_competitor_count_matches_rendered_fallback_table():
    """Real bug precondition: an EMPTY competitors dict snapshotted a
    competitor_count of 0 at synthesis time, while the Competitor Analysis
    table renders 5 rows from a later industry-fallback list."""
    data = _non_us_uber_data(
        _synthesized={
            "competitive_intelligence": {
                "competitors": {},
                "market_positioning": {
                    "industry_sector": "hospitality_travel",
                    "is_public_company": False,
                    "competitor_count": 0,
                    "has_sec_filings": False,
                },
            }
        }
    )
    text = _sheet_text(_build_market_intel_ws(data))
    assert "Competitor Count: 0" not in text
    assert "Competitor Count: 5" in text


def test_fix5_competitor_count_matches_when_no_fallback_rows_either():
    """When there truly are zero rendered competitor rows (no explicit
    competitors AND no industry-fallback list for this industry), the count
    must still agree with the table -- i.e. also read 0, not some other
    stale number."""
    data = _non_us_uber_data(
        industry="_no_such_industry_key_",
        _synthesized={
            "competitive_intelligence": {
                "competitors": {},
                "market_positioning": {
                    "industry_sector": "_no_such_industry_key_",
                    "is_public_company": False,
                    "competitor_count": 7,  # deliberately wrong/stale
                    "has_sec_filings": False,
                },
            }
        },
    )
    text = _sheet_text(_build_market_intel_ws(data))
    assert "Competitor Count: 7" not in text
    assert "Competitor Count: 0" in text


# ===========================================================================
# FIX 6 -- Channels & Strategy false "not available" claim
# ===========================================================================
def test_fix6_niche_channels_note_no_false_absence_claim_and_surfaces_real_boards():
    data = _non_us_uber_data()
    text = _sheet_text(_build_channels_ws(data))
    assert "was not available for this campaign" not in text
    for platform in ["Indeed UK", "Seek", "OCC Mundial", "Bumeran", "Trade Me Jobs"]:
        assert platform in text, f"expected local platform {platform!r} missing"


def test_fix6_us_plan_niche_channels_unaffected():
    """False-positive guard: a US plan with a US-only niche-board industry
    (aerospace_defense) must keep rendering the real US niche board list,
    not the non-US note path."""
    data = _us_plan_data(industry="aerospace_defense")
    text = _sheet_text(_build_channels_ws(data))
    assert "was not available for this campaign" not in text
    assert "not shown because this plan targets a non-US market" not in text


if __name__ == "__main__":
    import pytest as _pytest

    raise SystemExit(_pytest.main([__file__, "-v"]))
