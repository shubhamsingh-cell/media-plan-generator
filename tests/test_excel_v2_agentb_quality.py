"""Regression tests for the Agent B bundle-quality pass on excel_v2.py.

Covers:
  1. Locale delegation: ``excel_v2._is_us_plan`` is a thin delegate to
     ``plan_geo.is_us_plan`` (single source of truth) instead of a
     re-implemented, drift-prone locale resolver.
  2. 90-Day Forecast footing: every printed monthly row (Applications,
     Hires) sums EXACTLY to its printed total (largest-remainder
     reconciliation via display_format.reconcile_monthly_to_total), and a
     campaign longer than ~13 weeks burns only the ramp-weighted
     first-90-days share of budget, never the full campaign budget.
  3. No raw snake_case leaks into any client-facing string cell of the
     generated workbook.
  4. Zero-hire channels show "—" for Cost Per Hire, never a fabricated
     budget-as-CPH number.
  5. Duration labels come from display_format.weeks_to_duration_label
     (round-trip safe with parse_duration_to_weeks).
  6. Quality Intelligence salary rows with confidence == "estimated" carry
     an "(est.)" badge.

Runs under pytest, or standalone:
``python3 tests/test_excel_v2_agentb_quality.py``.
"""

from __future__ import annotations

import io
import re
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl  # noqa: E402

import budget_engine  # noqa: E402
import display_format  # noqa: E402
import excel_v2  # noqa: E402
import gold_standard  # noqa: E402
import plan_geo  # noqa: E402

_SNAKE_CASE_TOKEN_RE = re.compile(r"[a-zA-Z][a-zA-Z0-9]*_[a-zA-Z0-9_]*")


def _build_alloc(total_budget, roles, locations, industry, channels, **kwargs):
    return budget_engine.calculate_budget_allocation(
        total_budget=total_budget,
        roles=roles,
        locations=locations,
        industry=industry,
        channel_percentages=channels,
        **kwargs,
    )


def _base_data(**overrides):
    roles = [
        {"title": "Propane CDL Driver", "count": 20, "tier": "mid"},
        {"title": "Registered Nurse", "count": 10, "tier": "senior"},
    ]
    locations = [
        {"city": "Denver", "state": "CO", "country": "United States"},
        {"city": "Boise", "state": "ID", "country": "United States"},
    ]
    channels = {
        "niche_boards": 15,
        "programmatic_dsp": 25,
        "global_boards": 20,
        "social_media": 10,
        "employer_branding": 10,
        "regional_boards": 20,
    }
    alloc = _build_alloc(
        250_000,
        roles,
        locations,
        "logistics_supply_chain",
        channels,
        collar_type="blue_collar",
        campaign_start_month=3,
    )
    data = {
        "client_name": "PROPANE UNLIMITED CO",
        "industry": "logistics_supply_chain",
        "budget": "$250,000",
        "campaign_duration": "18 months",
        "campaign_weeks": 78,
        "campaign_start_month": 3,
        "hire_volume": "500 hires",
        "work_environment": "onsite",
        "roles": [r["title"] for r in roles],
        "target_roles": roles,
        "locations": [f"{l['city']}, {l['state']}" for l in locations],
        "competitors": ["Suburban Propane", "AmeriGas"],
        "_budget_allocation": alloc,
    }
    data.update(overrides)
    return data


def _generate_wb(data: dict):
    raw = excel_v2.generate_excel_v2(data)
    assert isinstance(raw, (bytes, bytearray)) and len(raw) > 0
    return openpyxl.load_workbook(io.BytesIO(raw), data_only=False)


# ---------------------------------------------------------------------------
# 1. Locale delegation
# ---------------------------------------------------------------------------
def test_is_us_plan_delegates_to_plan_geo():
    cases = [
        {"locations": ["New York, NY"]},
        {"locations": ["Massachusetts"]},
        {"locations": ["London, UK"]},
        {"locations": ["Denver, CO", "Auckland, NZ"]},
        {"target_region": "us_only"},
        {"target_region": "global"},
        {},
    ]
    for data in cases:
        assert excel_v2._is_us_plan(data) == plan_geo.is_us_plan(data), data


def test_non_us_signals_delegates_to_plan_geo():
    data = {"locations": ["New York, NY", "London, UK"]}
    assert excel_v2._non_us_signals(data) == plan_geo.non_us_signals(data)
    assert excel_v2._non_us_signals(data) == ["London, UK"]


def test_us_plan_shows_named_niche_boards():
    """A correctly-detected US plan (previously misdetected as non-US by
    the standalone locale bug) must show NAMED US niche boards for a
    logistics/healthcare brief, not the non-US suppression note."""
    data = _base_data()
    assert plan_geo.is_us_plan(data) is True
    wb = _generate_wb(data)
    ws = wb["Niche Board Matching"]
    text = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    joined = " ".join(text)
    assert any(
        board in joined
        for board in ("CDLjobs.com", "TruckingJobs.com", "WarehouseJobs.com")
    ), "Expected named US logistics niche boards on a correctly-detected US plan"
    assert "targets a non-US market" not in joined


# ---------------------------------------------------------------------------
# 2. 90-Day Forecast footing + ramp-weighted first-90-days share
# ---------------------------------------------------------------------------
def test_90_day_forecast_monthly_rows_foot_to_total():
    data = _base_data()  # 78-week (18-month) campaign
    wb = _generate_wb(data)
    ws = wb["90-Day Forecast"]

    rows = list(ws.iter_rows(min_col=2, values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Metric")
    metric_rows = {r[0]: r for r in rows[header_idx + 1 : header_idx + 5] if r and r[0]}

    for metric in ("Applications", "Hires"):
        row = metric_rows[metric]
        month_vals = row[1:4]
        total = row[4]
        assert all(isinstance(v, int) for v in month_vals), (
            metric,
            month_vals,
        )
        assert sum(month_vals) == total, (
            f"{metric} monthly values {month_vals} must sum EXACTLY to "
            f"printed total {total}"
        )


def test_90_day_forecast_scales_to_first_90_days_for_long_campaign():
    """A campaign longer than ~13 weeks must burn only the ramp-weighted
    first-90-days share of budget in this sheet, never the full budget."""
    data = _base_data()  # 78 weeks
    total_budget = 250_000.0
    wb = _generate_wb(data)
    ws = wb["90-Day Forecast"]

    rows = list(ws.iter_rows(min_col=2, values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Metric")
    spend_row = next(r for r in rows[header_idx + 1 :] if r and r[0] == "Spend")
    ninety_day_total_spend = spend_row[4]

    expected_scale = 13.0 / 78
    expected_total = total_budget * expected_scale
    assert abs(ninety_day_total_spend - expected_total) < 1.0, (
        ninety_day_total_spend,
        expected_total,
    )
    # Must be materially less than the full budget -- the defect this
    # guards against is showing 100% of an 18-month budget spent in 90 days.
    assert ninety_day_total_spend < total_budget * 0.5

    # The duration footnote must state the actual remaining-budget math.
    all_text = " ".join(
        c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)
    )
    assert "remaining" in all_text.lower()


def test_90_day_forecast_short_campaign_uses_full_budget():
    """A campaign <=13 weeks is itself within the 90-day window, so the
    forecast should show the FULL budget (scale factor 1.0). Since the
    duration-framing fix (Jesse Ofner 2026-07-31) the sheet is titled by the
    campaign's own length instead of "90-Day Forecast"."""
    data = _base_data(campaign_duration="10 weeks", campaign_weeks=10)
    total_budget = 250_000.0
    wb = _generate_wb(data)
    ws = wb["10-Week Forecast"]

    rows = list(ws.iter_rows(min_col=2, values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Metric")
    spend_row = next(r for r in rows[header_idx + 1 :] if r and r[0] == "Spend")
    assert abs(spend_row[4] - total_budget) < 1.0


# ---------------------------------------------------------------------------
# 3. No raw snake_case leaks into client-facing string cells
# ---------------------------------------------------------------------------
_ALLOWED_UNDERSCORE_SUBSTRINGS = ("http",)  # URLs may legitimately contain underscores


def test_no_snake_case_in_client_facing_cells():
    data = _base_data()
    wb = _generate_wb(data)
    offenders = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or not v:
                    continue
                if v.startswith("="):  # formulas
                    continue
                if any(tok in v for tok in _ALLOWED_UNDERSCORE_SUBSTRINGS):
                    continue
                if _SNAKE_CASE_TOKEN_RE.search(v):
                    offenders.append((ws.title, c.coordinate, v))
    assert (
        not offenders
    ), f"Raw snake_case leaked into client-facing cells: {offenders[:20]}"


# ---------------------------------------------------------------------------
# 4. Zero-hire channel CPH == "—"
# ---------------------------------------------------------------------------
def test_zero_hire_channel_shows_dash_cph_not_budget_as_cph():
    data = _base_data()  # includes employer_branding (brand channel, 0 hires)
    wb = _generate_wb(data)
    ws = wb["ROI Projections"]

    rows = list(ws.iter_rows(min_col=2, values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Channel Name")
    header = rows[header_idx]
    hires_col = header.index("Proj. Hires")
    cph_col = header.index("Cost Per Hire")

    saw_zero_hire_row = False
    for r in rows[header_idx + 1 :]:
        if not r or r[0] is None:
            continue
        if not isinstance(r[hires_col], (int, float)):
            continue
        if r[hires_col] == 0:
            saw_zero_hire_row = True
            assert r[cph_col] == "—", (
                f"Zero-hire channel {r[0]!r} must show — for CPH, got "
                f"{r[cph_col]!r} (a budget-as-CPH fabrication)"
            )
    assert saw_zero_hire_row, "Expected at least one zero-hire channel in this fixture"


# ---------------------------------------------------------------------------
# 5. Duration labels via display_format.weeks_to_duration_label
# ---------------------------------------------------------------------------
def test_duration_label_matches_display_format():
    for weeks in (10, 26, 78, 130):
        data = _base_data(campaign_weeks=weeks, campaign_duration="")
        data.pop("campaign_duration_canonical", None)
        resolved = excel_v2._resolve_campaign_duration(data)
        assert resolved == display_format.weeks_to_duration_label(weeks)
        # Round-trip safety: parsing the label back gives the same weeks.
        assert display_format.parse_duration_to_weeks(resolved) == weeks


def test_duration_18_months_never_becomes_17():
    """Regression: '18 months' must never silently round-trip to '17
    months' through a lossy weeks-per-month conversion."""
    data = _base_data(campaign_weeks=78, campaign_duration="18 months")
    resolved = excel_v2._resolve_campaign_duration(data)
    assert "18 months" in resolved
    assert "17 months" not in resolved


# ---------------------------------------------------------------------------
# 6. ESTIMATED badges on Quality Intelligence salary rows
# ---------------------------------------------------------------------------
def test_estimated_salary_rows_carry_badge():
    # "Yard Coordinator" has no keyword match in gold_standard's
    # _ROLE_SALARY_RANGES, so it gets a tier-scaled generic estimate
    # (confidence="estimated") alongside the benchmark-matched roles --
    # guarantees this fixture actually exercises the (est.) badge path.
    extra_roles = [{"title": "Yard Coordinator", "count": 5, "tier": "mid"}]
    data = _base_data()
    data["target_roles"] = data["target_roles"] + extra_roles
    data["roles"] = data["roles"] + [r["title"] for r in extra_roles]
    gold = gold_standard.apply_all_quality_gates(data)
    data["_gold_standard"] = gold

    # Sanity: this fixture's gold-standard output actually contains a mix
    # of benchmark and estimated per-role salary rows (else the assertion
    # below would vacuously pass).
    city_data = gold.get("city_level_data") or {}
    has_estimated = any(
        sal.get("confidence") == "estimated"
        for info in city_data.values()
        for sal in (info.get("per_role_salary") or {}).values()
    )
    assert has_estimated, "Fixture must produce at least one estimated salary row"

    wb = _generate_wb(data)
    assert "Quality Intelligence" in wb.sheetnames
    ws = wb["Quality Intelligence"]
    text_cells = [
        c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)
    ]
    assert any("Salary Intelligence" in t for t in text_cells)
    assert any("(est.)" in t for t in text_cells)


def test_quality_intelligence_program_structure_no_double_claim():
    data = _base_data()
    gold = gold_standard.apply_all_quality_gates(data)
    data["_gold_standard"] = gold
    wb = _generate_wb(data)
    ws = wb["Quality Intelligence"]
    text_cells = [
        c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)
    ]
    assert any("Recommended Program Structure" in t for t in text_cells)
    assert any("Total Program" in t for t in text_cells)
    # Media must be presented at 100%, not the old 72%-of-total framing.
    assert any(
        "Media (this plan's full budget)" in t or t == "Media (this plan's full budget)"
        for t in text_cells
    )


# ---------------------------------------------------------------------------
# 7. Executive Summary "Seasonal Patterns" row honors gold_standard's
#    sub-vertical activation-calendar override instead of always quoting the
#    generic KB industry seasonal_patterns benchmark (W3C follow-up: the
#    Manpower-AmeriGas Executive Summary kept showing generic freight/
#    e-commerce "Peak Months: August-November" text even after
#    gold_standard.build_activation_calendar started applying a
#    fuel_heating_delivery seasonal override for propane clients).
# ---------------------------------------------------------------------------
def _generate_wb_with_kb(data: dict):
    """Like ``_generate_wb`` but wires a real ``load_kb_fn`` so the
    "Recruitment Benchmarks" section (which the Seasonal Patterns row lives
    in) actually renders -- that section is skipped entirely when
    ``load_kb_fn`` is None, which is fine for the other tests in this file
    but not for one that asserts on Seasonal Patterns content."""
    import kb_loader

    raw = excel_v2.generate_excel_v2(data, load_kb_fn=kb_loader.load_knowledge_base)
    assert isinstance(raw, (bytes, bytearray)) and len(raw) > 0
    return openpyxl.load_workbook(io.BytesIO(raw), data_only=False)


def _seasonal_patterns_cell(wb) -> str:
    ws = wb["Executive Summary"]
    rows = list(ws.iter_rows())
    for r in rows:
        for c in r:
            if c.value == "Seasonal Patterns":
                # Value lives a couple columns over in the same row (label
                # column, then a merged/value column) -- grab the first
                # non-empty string cell after the label.
                for c2 in r[c.column :]:
                    if isinstance(c2.value, str) and c2.value.strip():
                        return c2.value
    return ""


def test_seasonal_patterns_row_uses_subvertical_override_when_matched():
    # _base_data()'s client ("PROPANE UNLIMITED CO") matches the
    # fuel_heating_delivery keyword set in a logistics_supply_chain plan.
    data = _base_data()
    gold = gold_standard.apply_all_quality_gates(data)
    data["_gold_standard"] = gold
    assert (gold.get("activation_calendar") or {}).get("subvertical") == (
        "fuel_heating_delivery"
    )

    wb = _generate_wb_with_kb(data)
    seasonal_text = _seasonal_patterns_cell(wb)
    assert seasonal_text, "Seasonal Patterns row must be present"
    assert "Fuel & Heating Delivery" in seasonal_text
    assert "Peak Months" in seasonal_text
    # Must NOT show the generic KB logistics benchmark's freight/e-commerce
    # framing, which is backwards for a propane/heating-fuel client.
    assert "E-commerce" not in seasonal_text
    assert "holiday logistics surge" not in seasonal_text.lower()


def test_seasonal_patterns_row_falls_back_to_generic_when_no_subvertical_match():
    data = _base_data(
        client_name="Acme Warehousing Co",
        competitors=["XPO Logistics", "Ryder"],
    )
    data["target_roles"] = [
        {"title": "Warehouse Associate", "count": 20, "tier": "mid"}
    ]
    data["roles"] = ["Warehouse Associate"]
    gold = gold_standard.apply_all_quality_gates(data)
    data["_gold_standard"] = gold
    assert (gold.get("activation_calendar") or {}).get("subvertical") is None

    wb = _generate_wb_with_kb(data)
    seasonal_text = _seasonal_patterns_cell(wb)
    assert seasonal_text, "Seasonal Patterns row must be present"
    # Unmatched plans keep the pre-existing generic KB benchmark text.
    assert "Fuel & Heating Delivery" not in seasonal_text
    assert "Peak Months" in seasonal_text


if __name__ == "__main__":
    test_is_us_plan_delegates_to_plan_geo()
    test_non_us_signals_delegates_to_plan_geo()
    test_us_plan_shows_named_niche_boards()
    test_90_day_forecast_monthly_rows_foot_to_total()
    test_90_day_forecast_scales_to_first_90_days_for_long_campaign()
    test_90_day_forecast_short_campaign_uses_full_budget()
    test_no_snake_case_in_client_facing_cells()
    test_zero_hire_channel_shows_dash_cph_not_budget_as_cph()
    test_duration_label_matches_display_format()
    test_duration_18_months_never_becomes_17()
    test_estimated_salary_rows_carry_badge()
    test_quality_intelligence_program_structure_no_double_claim()
    test_seasonal_patterns_row_uses_subvertical_override_when_matched()
    test_seasonal_patterns_row_falls_back_to_generic_when_no_subvertical_match()
    print("OK")
