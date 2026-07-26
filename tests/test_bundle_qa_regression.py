"""Regression gate: regenerate the two reference bundles (manpower/logistics
and atria/healthcare -- the ones the July bundle audits used) and assert
``bundle_qa.run_bundle_qa`` reports ZERO critical findings.

This is the test that turns "a human reviewer eventually caught it" into
"CI catches it" -- the whole reason bundle_qa.py exists is that this exact
class of defect (snake_case leaking into a cell, a fabricated "beating
benchmark" badge, near-duplicate competitor prose, ...) shipped twice
before a human flagged it.

Also includes focused positive-control unit tests proving each check
actually fires on an injected defect (not just that the regexes exist) --
a linter with zero findings on clean input and zero findings on broken
input is worse than no linter, because it looks like coverage.

Runs under pytest, or standalone:
``python3 tests/test_bundle_qa_regression.py``.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import bundle_qa  # noqa: E402
import tools_regen_bundles as trb  # noqa: E402

_REFERENCE_BRIEFS = {
    "manpower": trb.MANPOWER_BRIEF,
    "atria": trb.ATRIA_BRIEF,
}


def _regen_and_scan(slug: str, tmp_path: Path) -> tuple[dict, list[dict]]:
    brief = _REFERENCE_BRIEFS[slug]
    result = trb.generate_bundle(brief, tmp_path, slug)
    assert not result["errors"], f"{slug} bundle failed to generate: {result['errors']}"
    findings = bundle_qa.run_bundle_qa(
        result["pptx_bytes"], result["xlsx_bytes"], result["data"]
    )
    return result, findings


def test_manpower_bundle_has_zero_critical_findings(tmp_path):
    _, findings = _regen_and_scan("manpower", tmp_path)
    critical = [f for f in findings if f["severity"] == "critical"]
    assert not critical, (
        "manpower/logistics reference bundle has critical bundle_qa "
        f"findings: {critical}"
    )


def test_atria_bundle_has_zero_critical_findings(tmp_path):
    _, findings = _regen_and_scan("atria", tmp_path)
    critical = [f for f in findings if f["severity"] == "critical"]
    assert not critical, (
        "atria/healthcare reference bundle has critical bundle_qa "
        f"findings: {critical}"
    )


def test_findings_have_the_required_shape(tmp_path):
    """Every finding -- on either bundle -- carries the documented API
    shape, regardless of severity (structural contract, not a content
    check)."""
    for slug in _REFERENCE_BRIEFS:
        _, findings = _regen_and_scan(slug, tmp_path)
        for f in findings:
            assert f["severity"] in ("critical", "warn")
            assert isinstance(f["code"], str) and f["code"]
            assert isinstance(f["message"], str) and f["message"]
            assert isinstance(f["location"], str)


# ---------------------------------------------------------------------------
# Positive controls: prove each check actually fires
# ---------------------------------------------------------------------------
def test_detects_snake_case_leak():
    units = [bundle_qa._TextUnit("Rates: warehouse_hourly $12-$18", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {}, findings)
    assert any(f["code"] == "snake_case_leak" for f in findings)


def test_detects_multi_underscore_snake_case_leak():
    """The literal spec regex r'\\b[a-z0-9]+_[a-z0-9]+\\b' cannot match a
    token with 2+ underscores (no internal \\b for a fixed two-group
    pattern to anchor on) -- bundle_qa uses a repeating-group regex instead
    so a real KB key like "cdl_drivers_hourly" is still caught."""
    units = [bundle_qa._TextUnit("cdl_drivers_hourly: $25-$50", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {}, findings)
    assert any(f["code"] == "snake_case_leak" for f in findings)


def test_ignores_snake_case_inside_urls_and_emails():
    units = [
        bundle_qa._TextUnit("See our_data at https://example.com/api_docs", "X!A1"),
        bundle_qa._TextUnit("Contact hiring_team@example.com for details", "X!A2"),
    ]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {}, findings)
    assert not any(f["code"] == "snake_case_leak" for f in findings)


def test_detects_pluralization_artifact():
    units = [bundle_qa._TextUnit("5 role(s) selected", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {}, findings)
    assert any(f["code"] == "pluralization_artifact" for f in findings)


def test_detects_raw_float_precision():
    units = [bundle_qa._TextUnit("CPA is $12.345678", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {}, findings)
    assert any(f["code"] == "raw_float_precision" for f in findings)


def test_detects_dollar_point_zero_k():
    units = [bundle_qa._TextUnit("Budget: $150.0K", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {}, findings)
    assert any(f["code"] == "dollar_point_zero_k" for f in findings)


def test_detects_over_ongoing_grammar():
    """prod-Atria defect: 'Finalize weekly budget (25000 over Ongoing) and
    success metrics' -- an unbounded duration spliced verbatim into the
    '<budget> over <duration>' template, plus the raw unformatted budget
    digits that shipped alongside it. Both artifact guards must fire."""
    units = [
        bundle_qa._TextUnit(
            "Finalize weekly budget (25000 over Ongoing) and success metrics",
            "X!A1",
        )
    ]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {}, findings)
    codes = {f["code"] for f in findings}
    assert "over_ongoing_grammar" in codes
    assert "raw_number_before_over" in codes


def test_allows_fixed_ongoing_phrasing_and_formatted_money():
    """The corrected phrasing this codebase now emits ('$25K, ongoing' /
    '$25K on an ongoing basis' / '$25K over 6 months') must NOT
    false-positive on either new check."""
    units = [
        bundle_qa._TextUnit(
            "Finalize weekly budget ($25K, ongoing) and success metrics",
            "X!A1",
        ),
        bundle_qa._TextUnit(
            "Launch within 2 business weeks of feed integration — $25K "
            "on an ongoing basis",
            "X!A2",
        ),
        bundle_qa._TextUnit(
            "Finalize weekly budget ($25K over 6 months) and success metrics",
            "X!A3",
        ),
    ]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {}, findings)
    codes = {f["code"] for f in findings}
    assert "over_ongoing_grammar" not in codes
    assert "raw_number_before_over" not in codes


def test_raw_number_before_over_ignores_out_of_range_digit_runs():
    """The 4-7 digit window targets an unformatted BUDGET specifically -- a
    short 2-digit rank or an 8+ digit reference number before 'over' is a
    different shape and must not be flagged."""
    units = [
        bundle_qa._TextUnit("Rank 12 over 500 applicants", "X!A1"),
        bundle_qa._TextUnit("Reference 123456789 over quota", "X!A2"),
    ]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {}, findings)
    assert not any(f["code"] == "raw_number_before_over" for f in findings)


def test_detects_location_advisory_leak_unconditionally():
    units = [bundle_qa._TextUnit("Location advisory: verify targeting", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {}, findings)
    assert any(f["code"] == "location_advisory_leak" for f in findings)


def test_detects_non_us_text_on_us_plan():
    units = [bundle_qa._TextUnit("This reflects a non-US market signal", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {"locations": ["Austin, TX"]}, findings)
    assert any(f["code"] == "non_us_text_on_us_plan" for f in findings)


def test_allows_non_us_text_on_non_us_plan():
    units = [bundle_qa._TextUnit("This reflects a non-US market signal", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {"locations": ["London, UK"]}, findings)
    assert not any(f["code"] == "non_us_text_on_us_plan" for f in findings)


def test_detects_ai_training_vocab_on_non_ai_training_plan():
    units = [bundle_qa._TextUnit("Our AI Trainer network is ready", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(
        units,
        {"industry": "healthcare_medical", "roles": ["Registered Nurse"]},
        findings,
    )
    assert any(f["code"] == "ai_training_vocab_leak" for f in findings)


def test_allows_ai_training_vocab_on_ai_training_plan():
    units = [bundle_qa._TextUnit("Our AI Trainer network is ready", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(
        units, {"industry": "ai_training", "roles": ["AI Trainer"]}, findings
    )
    assert not any(f["code"] == "ai_training_vocab_leak" for f in findings)


def test_detects_duration_label_drifted_from_campaign_weeks():
    units = [bundle_qa._TextUnit("Duration: 6 months (~30 weeks)", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {"campaign_weeks": 24}, findings)
    assert any(
        f["code"] == "duration_label_drifted_from_campaign_weeks" for f in findings
    )


def test_allows_bucketed_duration_label_matching_campaign_weeks():
    """A "6 months (~24 weeks)" label is a legitimate fixed marketing
    bucket (4 weeks/month), NOT a 52/12 conversion -- it must not be
    flagged as long as it agrees with the bundle's own campaign_weeks."""
    units = [bundle_qa._TextUnit("Duration: 6 months (~24 weeks)", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {"campaign_weeks": 24}, findings)
    assert not any(f["code"].startswith("duration_") for f in findings), findings


def test_allows_subset_market_count_within_total():
    """Prod-Manpower false positive: the Risk Register's "High Competition
    (N market(s))" title reports a SUBSET of the plan's locations (how many
    are classified high/very_high hiring_intensity in competitor_map), not
    the plan's total market count -- e.g. "High Competition (1 market)" on
    a 6-location plan is correct, not a mismatch. Any subset count
    (<= n_locations) must not be flagged."""
    units = [bundle_qa._TextUnit("High Competition (1 market)", "slide 9")]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(
        units,
        {
            "locations": [
                "Massachusetts",
                "Maine",
                "New Hampshire",
                "Rhode Island",
                "Connecticut",
                "Denver, CO",
            ]
        },
        findings,
    )
    assert not any(f["code"] == "markets_count_mismatch" for f in findings), findings


def test_detects_impossible_market_count_exceeding_total():
    """A parenthetical market count can never legitimately exceed the
    plan's total location count -- that's the real invariant worth
    flagging (e.g. stale/hardcoded copy claiming more markets than the
    plan actually has)."""
    units = [
        bundle_qa._TextUnit(
            "This campaign spans (8 markets) with strong regional demand",
            "slide 3",
        )
    ]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(
        units, {"locations": ["Austin, TX", "Dallas, TX", "Houston, TX"]}, findings
    )
    matches = [f for f in findings if f["code"] == "markets_count_mismatch"]
    assert matches, findings
    assert "8 markets" in matches[0]["message"]
    assert "3 location(s)" in matches[0]["message"]


def test_detects_client_name_wrong_casing():
    units = [bundle_qa._TextUnit("Plan for atria Senior living", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_client_name_casing(
        units, {"client_name": "atria Senior living"}, findings
    )
    assert any(f["code"] == "client_name_wrong_casing" for f in findings)


def test_allows_canonical_client_name_casing():
    units = [bundle_qa._TextUnit("Plan for Atria Senior Living", "X!A1")]
    findings: list[dict] = []
    bundle_qa._check_client_name_casing(
        units, {"client_name": "atria Senior living"}, findings
    )
    assert not any(f["code"] == "client_name_wrong_casing" for f in findings)


def test_detects_near_duplicate_counter_strategy():
    """Two DIFFERENT competitors, same table/view, near-identical text --
    must flag."""
    findings: list[dict] = []
    bundle_qa._check_counter_strategy_distinctness(
        [
            (
                "UPS's presence in this pool means CDL drivers have "
                "options -- lead with total-comp clarity and a same-week "
                "interview slot.",
                "a",
                "UPS",
                "Market Intelligence#10",
            ),
            (
                "FedEx's presence in this pool means CDL drivers have "
                "options -- lead with total-comp clarity and a same-week "
                "interview slot.",
                "b",
                "FedEx",
                "Market Intelligence#10",
            ),
        ],
        findings,
    )
    assert any(f["code"] == "counter_strategy_near_duplicate" for f in findings)


def test_allows_distinct_counter_strategy():
    findings: list[dict] = []
    bundle_qa._check_counter_strategy_distinctness(
        [
            (
                "UPS is actively staffing CDL drivers through third-party "
                "agencies -- compress time-to-offer.",
                "a",
                "UPS",
                "Market Intelligence#10",
            ),
            (
                "Expect FedEx to keep pressure on CDL drivers; a faster "
                "interview-to-offer cycle is the clearest lever.",
                "b",
                "FedEx",
                "Market Intelligence#10",
            ),
        ],
        findings,
    )
    assert not any(f["code"] == "counter_strategy_near_duplicate" for f in findings)


def test_allows_same_competitor_near_duplicate_across_different_artifacts():
    """S92 bug #1: the SAME competitor legitimately carries the SAME
    composed sentence across DIFFERENT artifacts (deck card vs Market
    Intelligence row vs Quality Intelligence row) -- that's intentional
    single-sourcing from insight_composer, not competitors reading as
    interchangeable. Cross-view pairs must never be flagged, regardless of
    identity."""
    findings: list[dict] = []
    bundle_qa._check_counter_strategy_distinctness(
        [
            (
                "ManpowerGroup is actively staffing CDL drivers through "
                "third-party agencies -- compress time-to-offer.",
                "deck loc",
                "ManpowerGroup",
                "deck",
            ),
            (
                "ManpowerGroup is actively staffing CDL drivers through "
                "third-party agencies -- compress time-to-offer.",
                "Market Intelligence!F5",
                "ManpowerGroup",
                "Market Intelligence#10",
            ),
        ],
        findings,
    )
    assert not any(f["code"] == "counter_strategy_near_duplicate" for f in findings)


def test_allows_same_competitor_near_duplicate_within_same_table():
    """Same view, SAME competitor identity repeating (e.g. the same top
    employer leads two different cities) -- expected, not a defect."""
    findings: list[dict] = []
    bundle_qa._check_counter_strategy_distinctness(
        [
            (
                "ManpowerGroup is actively staffing CDL drivers through "
                "third-party agencies -- compress time-to-offer.",
                "Quality Intelligence!F5",
                "ManpowerGroup",
                "Quality Intelligence#4",
            ),
            (
                "ManpowerGroup is actively staffing CDL drivers through "
                "third-party agencies -- compress time-to-offer.",
                "Quality Intelligence!F6",
                "ManpowerGroup",
                "Quality Intelligence#4",
            ),
        ],
        findings,
    )
    assert not any(f["code"] == "counter_strategy_near_duplicate" for f in findings)


def test_skips_pair_when_identity_not_collectable():
    """If identity can't be established for a pair, skip it rather than
    flag it -- absence of evidence isn't evidence of a defect."""
    findings: list[dict] = []
    bundle_qa._check_counter_strategy_distinctness(
        [
            (
                "UPS's presence in this pool means CDL drivers have "
                "options -- lead with total-comp clarity and a same-week "
                "interview slot.",
                "a",
                None,
                "deck",
            ),
            (
                "FedEx's presence in this pool means CDL drivers have "
                "options -- lead with total-comp clarity and a same-week "
                "interview slot.",
                "b",
                None,
                "deck",
            ),
        ],
        findings,
    )
    assert not any(f["code"] == "counter_strategy_near_duplicate" for f in findings)


def test_collects_xlsx_counter_strategies_bounded_to_table():
    """S92 bug #2: _collect_xlsx_counter_strategies must stop at the table
    boundary, not sweep into an unrelated table below that happens to
    reuse the same column letter -- e.g. Quality Intelligence's "Role
    Difficulty Classification" table puts "Location Modifier" in the same
    column F that "Competitive Landscape & Counter-Strategies" used for
    "Counter-Strategy". No blank row separates the two tables here on
    purpose, so this only passes if the header-style-based stop (not just
    the blank-row stop) is working."""
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(
        start_color="FF1F3A93", end_color="FF1F3A93", fill_type="solid"
    )
    body_font = Font(bold=False)
    body_fill = PatternFill(
        start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid"
    )

    def _write_header_row(ws, row_idx, headers):
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=row_idx, column=ci, value=h)
            c.font = header_font
            c.fill = header_fill

    def _write_data_row(ws, row_idx, values):
        for ci, v in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=ci, value=v)
            c.font = body_font
            c.fill = body_fill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quality Intelligence"

    # Table 1: Competitive Landscape & Counter-Strategies -- F = Counter-Strategy
    _write_header_row(
        ws,
        1,
        [
            "City",
            "Top Employers",
            "Hiring Intensity",
            "Est. Competing Postings",
            "Why They Matter",
            "Counter-Strategy",
        ],
    )
    _write_data_row(
        ws,
        2,
        [
            "Houston",
            "ManpowerGroup",
            "High",
            500,
            "High hiring volume",
            "ManpowerGroup is actively staffing CDL drivers -- compress time-to-offer.",
        ],
    )

    # Table 2 immediately follows, NO blank row -- F = Location Modifier
    # (same column letter, unrelated field).
    _write_header_row(
        ws,
        3,
        [
            "Role Title",
            "Seniority Level",
            "Difficulty (1-10)",
            "Supply Level",
            "Avg Time-to-Fill",
            "Location Modifier",
            "Budget Weight",
            "Channel Emphasis",
            "Description",
        ],
    )
    _write_data_row(
        ws,
        4,
        [
            "CDL Driver",
            "Mid",
            6,
            "Moderate",
            "45 days",
            "+2.0 (Houston)",
            "1.1x",
            "Job Boards",
            "Standard CDL role",
        ],
    )

    out = bundle_qa._collect_xlsx_counter_strategies(wb)
    texts = [t for t, *_rest in out]
    assert any("ManpowerGroup" in t for t in texts)
    assert not any("Location Modifier" in t or "+2.0" in t for t in texts)
    assert len(out) == 1
    # Identity for the one collected row should resolve to the "Top
    # Employers" column value, not None.
    assert out[0][2] == "ManpowerGroup"


def test_detects_zero_hire_nonzero_cph():
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    ws.append(["Channel", "Proj. Hires", "Cost Per Hire"])
    ws.append(["Niche Boards", 0, 625.0])
    findings: list[dict] = []
    bundle_qa._check_zero_hire_honesty(wb, findings)
    assert any(f["code"] == "zero_hire_nonzero_cph" for f in findings)


def test_allows_zero_hire_with_dash_cph():
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    ws.append(["Channel", "Proj. Hires", "Cost Per Hire"])
    ws.append(["Niche Boards", 0, "—"])
    findings: list[dict] = []
    bundle_qa._check_zero_hire_honesty(wb, findings)
    assert not any(f["code"] == "zero_hire_nonzero_cph" for f in findings)


def test_zero_hire_honesty_does_not_bleed_into_a_later_differently_shaped_table():
    """CHECK-ARTIFACT REGRESSION (2026-07-26, tiny-budget false positive):
    the real "ROI Projections" sheet stacks a "Per-Channel ROI Analysis"
    table (Proj. Hires / Cost Per Hire columns) directly above a
    "Recruitment Funnel" table (Qualified / App->Qualified-rate columns in
    those SAME column positions), separated by one blank row. The correct
    ROI Analysis rows below all show '-' for Cost Per Hire on their 0-hire
    rows -- this must produce ZERO findings. The scan must never run past
    the blank row into the funnel table below and misread its Qualified
    count as "hires" and its rate as a dollar "Cost Per Hire".
    """
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ROI Projections"
    # Column positions mirror the real sheet exactly (leading blank column
    # A; "Proj. Hires" at index 4, "Cost Per Hire" at index 7).
    ws.append(
        [None, "Channel Name", "Budget ($)", "Proj. Applications", "Proj. Hires",
         "Confidence", "Hire Range", "Cost Per Hire"]
    )
    ws.append([None, "Regional Job Boards", 3575.15, 602, 2, "MEDIUM", "1 - 2", 1787.58])
    ws.append([None, "Social Media", 136, 1, 0, "LOW", "0", "—"])
    ws.append([])  # table boundary
    ws.append(
        [None, "Channel Name", "Clicks", "Applications", "Qualified", "Interviews",
         "Hires", "App→Qualified"]
    )
    # This funnel row's "Qualified" count (misread as hires_col=4, i.e. the
    # position "Proj. Hires" held in the table above) is 0, and its
    # "App→Qualified" rate (misread as cph_col=7, the position "Cost Per
    # Hire" held above) is a nonzero fraction -- exactly the real
    # tiny-budget shape that used to misfire as "0 hires but a numeric
    # Cost Per Hire".
    ws.append([None, "Social Media", 52, 1, 0, 0, "—", 0.124])
    findings: list[dict] = []
    bundle_qa._check_zero_hire_honesty(wb, findings)
    assert not any(f["code"] == "zero_hire_nonzero_cph" for f in findings), findings


def test_detects_forecast_footing_mismatch():
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "90-Day Forecast"
    ws.append([None] * 8)
    ws.append(
        [
            None,
            "Metric",
            "July 2026",
            "August 2026",
            "September 2026",
            "90-Day Total",
            "Trend",
            None,
        ]
    )
    ws.append(
        [None, "Applications", 100, 200, 300, 999, "Increasing", None]
    )  # 600 != 999
    findings: list[dict] = []
    bundle_qa._check_90_day_forecast_footing(wb, findings)
    assert any(f["code"] == "forecast_footing_mismatch" for f in findings)


def test_allows_correct_forecast_footing():
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "90-Day Forecast"
    ws.append([None] * 8)
    ws.append(
        [
            None,
            "Metric",
            "July 2026",
            "August 2026",
            "September 2026",
            "90-Day Total",
            "Trend",
            None,
        ]
    )
    ws.append([None, "Applications", 100, 200, 300, 600, "Increasing", None])
    findings: list[dict] = []
    bundle_qa._check_90_day_forecast_footing(wb, findings)
    assert not any(f["code"] == "forecast_footing_mismatch" for f in findings)


_FUNNEL_HEADERS = [
    "Channel Name",
    "Clicks",
    "Applications",
    "Qualified",
    "Interviews",
    "Hires",
    "App→Qualified",
    "Qualified→Interview",
    "Interview→Hire",
]


def test_allows_large_scale_funnel_rate_reconstruction_residual():
    """TOLERANCE FIX (2026-07-26, India/tech false positive): stage rates
    in this table are stored rounded to 4 decimal places (see
    budget_engine.compute_funnel_stages). Reconstructing an upstream
    integer count via round(prev_stage * printed_rate) then carries an
    inherent residual of up to 0.00005 * prev_stage -- real signal, not a
    footing bug -- once prev_stage is large enough (hundreds of thousands
    of applications on a real non-US enterprise plan). This is the exact
    real row from the nonus_single_india_tech probe brief: Interviews
    (49257) x Interview→Hire (0.15) rounds to 7389, but the actual
    (CPH-anchored, independently correct) Hires is 7391 -- a 2-count gap
    that a flat 1-count tolerance flags but the scaled tolerance allows."""
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ROI Projections"
    ws.append(_FUNNEL_HEADERS)
    ws.append(
        ["Niche / Industry Boards", 12529761, 1754166, 246285, 49257, 7391, 0.1404, 0.2, 0.15]
    )
    ws.append(["TOTAL", 12529761, 1754166, 246285, 49257, 7391, 0.1404, 0.2, 0.15])
    findings: list[dict] = []
    bundle_qa._check_recruitment_funnel_footing(wb, findings)
    assert not any(
        f["code"] == "funnel_rate_footing_mismatch" for f in findings
    ), findings


def test_detects_genuine_large_scale_funnel_rate_mismatch():
    """The scaled tolerance must not blunt the check into uselessness --
    a Hires value far outside even the scaled residual (interviews x rate
    rounds to 7389; an actual Hires of 9000 is nowhere close) must still
    fire."""
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ROI Projections"
    ws.append(_FUNNEL_HEADERS)
    ws.append(
        ["Niche / Industry Boards", 12529761, 1754166, 246285, 49257, 9000, 0.1404, 0.2, 0.15]
    )
    ws.append(["TOTAL", 12529761, 1754166, 246285, 49257, 9000, 0.1404, 0.2, 0.15])
    findings: list[dict] = []
    bundle_qa._check_recruitment_funnel_footing(wb, findings)
    assert any(f["code"] == "funnel_rate_footing_mismatch" for f in findings)


def test_detects_beating_badge_next_to_non_comparable_benchmark():
    units = [
        bundle_qa._TextUnit(
            "1.3x  ▲", "slide 8 / client", slide_idx=8, top=1000, left=100
        ),
        bundle_qa._TextUnit(
            "Varies", "slide 8 / industry", slide_idx=8, top=1000, left=5000
        ),
    ]
    findings: list[dict] = []
    bundle_qa._check_comparison_badges(units, findings)
    assert any(f["code"] == "fabricated_beating_badge" for f in findings)


def test_allows_beating_badge_next_to_real_benchmark():
    units = [
        bundle_qa._TextUnit(
            "1.3x  ▲", "slide 8 / client", slide_idx=8, top=1000, left=100
        ),
        bundle_qa._TextUnit(
            "1.0x", "slide 8 / industry", slide_idx=8, top=1000, left=5000
        ),
    ]
    findings: list[dict] = []
    bundle_qa._check_comparison_badges(units, findings)
    assert not any(f["code"] == "fabricated_beating_badge" for f in findings)


# ---------------------------------------------------------------------------
# 2026-07-25 audit: five new generation-time QA rules for blind spots a real
# shipped bundle (client Uber, GBP 2,000,000, 6 non-US markets, role
# "commercial cab driver", shipped 2026-07-23 with 31 critical defects,
# only 11 caught by run_bundle_qa pre-fix) exposed. Each rule below has (a)
# a focused unit test proving it fires on an injected defect and does NOT
# fire on the clean equivalent, and (b) an integration test against the
# real shipped bundle bytes (tests/fixtures/uber_shipped_2026_07_23/).
# ---------------------------------------------------------------------------
_UBER_FIXTURE_DIR = (
    Path(__file__).resolve().parent / "fixtures" / "uber_shipped_2026_07_23"
)

# 2026-07-25: Uber_Media_Plan.xlsx was silently dropped from commit 1cdd80e
# by the repo's old blanket `*.xlsx` .gitignore rule (fixed alongside this
# guard -- see .gitignore's `!tests/fixtures/**/*.xlsx` negation). The bytes
# do not exist anywhere: not on disk, not in any branch's git history
# (`git log --all --diff-filter=A -- '**/Uber_Media_Plan.xlsx'` finds
# nothing to restore from), and they cannot be regenerated because these
# tests assert the bundle trips five SPECIFIC real-defect rules that only
# exist in that one shipped file. SKIP (not delete/fabricate) the tests
# that need it, scoped to exactly the consumers of
# `_scan_uber_shipped_bundle()` -- every other test in this file, including
# ones that only touch the .pptx half of the fixture, is unaffected and
# still runs. Whoever has the original Uber_Media_Plan.xlsx bytes should
# `git add -f` them back into tests/fixtures/uber_shipped_2026_07_23/ -- the
# moment the file exists on disk again, this guard flips false and all six
# tests below re-enable themselves automatically, no code change needed.
_UBER_XLSX_AVAILABLE = (_UBER_FIXTURE_DIR / "Uber_Media_Plan.xlsx").is_file()
_requires_uber_xlsx = pytest.mark.skipif(
    not _UBER_XLSX_AVAILABLE,
    reason=(
        "MISSING FIXTURE: tests/fixtures/uber_shipped_2026_07_23/"
        "Uber_Media_Plan.xlsx does not exist. It was silently dropped from "
        "commit 1cdd80e (2026-07-25) by the old blanket `*.xlsx` "
        ".gitignore rule and cannot be regenerated -- these 6 tests assert "
        "against real defects specific to that shipped file. Restore the "
        "original bytes to that path (now un-ignored) to re-enable this "
        "test automatically."
    ),
)

_UBER_DATA = {
    "client_name": "Uber",
    "industry": "Hospitality & Travel",
    "budget": 2000000,
    "locations": ["UK", "Australia", "Mexico", "argentina", "canada", "new zealand"],
    "roles": ["commercial cab driver"],
    "currency": "GBP",
    "duration": "4 weeks",
}


def _scan_uber_shipped_bundle() -> list[dict]:
    pptx_bytes = (_UBER_FIXTURE_DIR / "Uber_Strategy_Deck.pptx").read_bytes()
    xlsx_bytes = (_UBER_FIXTURE_DIR / "Uber_Media_Plan.xlsx").read_bytes()
    return bundle_qa.run_bundle_qa(pptx_bytes, xlsx_bytes, dict(_UBER_DATA))


@_requires_uber_xlsx
def test_uber_shipped_bundle_all_five_new_rules_fire_and_nothing_crashed():
    """The real shipped bundle must trip every one of the five new rule
    codes, and none of the new checks may crash (no *_check_crashed
    finding for any of them)."""
    findings = _scan_uber_shipped_bundle()
    codes = {f["code"] for f in findings}
    for expected in (
        "us_data_on_non_us_plan",
        "currency_symbol_mixing",
        "campaign_duration_incoherence",
        "industry_client_conflict",
        "competitor_count_contradiction",
    ):
        assert expected in codes, f"{expected} did not fire on the shipped bundle: {codes}"
    assert not any(
        f["code"].endswith("_check_crashed") for f in findings
    ), [f for f in findings if f["code"].endswith("_check_crashed")]


# --- RULE 1: us_data_on_non_us_plan ----------------------------------------
def test_detects_us_data_on_non_us_plan_macro_vocab():
    units = [bundle_qa._TextUnit("Fed Funds Rate: 3.64%", "Market Intelligence!B46")]
    findings: list[dict] = []
    bundle_qa._check_us_data_on_non_us_plan(
        units, None, {"locations": ["London, UK"]}, findings
    )
    assert any(f["code"] == "us_data_on_non_us_plan" for f in findings)


def test_allows_us_data_vocab_on_us_plan():
    """This new rule must never fire on a genuinely US-only plan -- that is
    the existing non_us_text_on_us_plan check's job, in the other
    direction (bundle_qa.py:398)."""
    units = [bundle_qa._TextUnit("Fed Funds Rate: 3.64%", "Market Intelligence!B46")]
    findings: list[dict] = []
    bundle_qa._check_us_data_on_non_us_plan(
        units, None, {"locations": ["Austin, TX"]}, findings
    )
    assert not any(f["code"] == "us_data_on_non_us_plan" for f in findings)


def test_detects_country_cell_says_united_states_with_no_us_location():
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Market Intelligence"
    ws.append(["Location", "Country", "Why This Market"])
    ws.append(["UK", "United States", "UK was selected by client footprint."])
    findings: list[dict] = []
    bundle_qa._check_us_data_on_non_us_plan(
        [], wb, {"locations": ["London, UK"]}, findings
    )
    assert any(f["code"] == "us_data_on_non_us_plan" for f in findings)


def test_allows_correct_country_cell_on_non_us_plan():
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Market Intelligence"
    ws.append(["Location", "Country", "Why This Market"])
    ws.append(["UK", "United Kingdom", "UK was selected by client footprint."])
    findings: list[dict] = []
    bundle_qa._check_us_data_on_non_us_plan(
        [], wb, {"locations": ["London, UK"]}, findings
    )
    assert not any(f["code"] == "us_data_on_non_us_plan" for f in findings)


@_requires_uber_xlsx
def test_uber_shipped_bundle_catches_us_data_on_non_us_plan():
    findings = _scan_uber_shipped_bundle()
    matches = [f for f in findings if f["code"] == "us_data_on_non_us_plan"]
    assert matches
    messages = " ".join(f["message"] for f in matches)
    assert "Fed Funds" in messages
    assert "Thanksgiving" in messages or "Memorial Day" in messages
    assert any("Country column" in f["message"] for f in matches)


# --- RULE 2: currency_symbol_mixing -----------------------------------------
def test_detects_bare_dollar_on_gbp_plan():
    units = [
        bundle_qa._TextUnit(
            "Budget engine projects £7 average CPA across all channels, "
            "with $1,626 average cost-per-hire.",
            "slide 6 / TextBox 85 para 0",
        )
    ]
    findings: list[dict] = []
    bundle_qa._check_currency_symbol_mixing(units, {"currency": "GBP"}, findings)
    assert any(f["code"] == "currency_symbol_mixing" for f in findings)


def test_allows_marked_us_dollar_benchmark_on_gbp_plan():
    """'US$1,500-US$3,500' is the sanctioned honest marker (ppt_generator.
    _mark_usd) and must NOT be flagged even on a GBP plan."""
    units = [bundle_qa._TextUnit("US$1,500-US$3,500", "slide 5 / shape")]
    findings: list[dict] = []
    bundle_qa._check_currency_symbol_mixing(units, {"currency": "GBP"}, findings)
    assert not any(f["code"] == "currency_symbol_mixing" for f in findings)


def test_allows_correct_plan_symbol():
    units = [bundle_qa._TextUnit("£561,772 programmatic investment", "slide 5")]
    findings: list[dict] = []
    bundle_qa._check_currency_symbol_mixing(units, {"currency": "GBP"}, findings)
    assert not any(f["code"] == "currency_symbol_mixing" for f in findings)


def test_detects_dollar_with_foreign_currency_code_in_parens():
    units = [bundle_qa._TextUnit("$42,000 (GBP)", "Market Intelligence!E25")]
    findings: list[dict] = []
    bundle_qa._check_currency_symbol_mixing(units, {"currency": "GBP"}, findings)
    matches = [f for f in findings if f["code"] == "currency_symbol_mixing"]
    assert matches
    assert "GBP" in matches[0]["message"]


def test_currency_mixing_respects_usd_header_marker_exemption():
    """A column whose header explicitly reads "... (USD)" (the Intl
    Benchmarks sheet's own convention, excel_v2.py ~line 95-98) must not be
    flagged even though its bare "$" figures don't match a non-USD plan --
    the header IS the honest marker."""
    units = [
        bundle_qa._TextUnit("CPA Range (USD)", "Intl Benchmarks!F5", top=5, left=6),
        bundle_qa._TextUnit("$16-$52; $10-$39", "Intl Benchmarks!F7", top=7, left=6),
    ]
    findings: list[dict] = []
    bundle_qa._check_currency_symbol_mixing(units, {"currency": "GBP"}, findings)
    assert not any(f["code"] == "currency_symbol_mixing" for f in findings)


@_requires_uber_xlsx
def test_uber_shipped_bundle_catches_currency_symbol_mixing():
    findings = _scan_uber_shipped_bundle()
    matches = [f for f in findings if f["code"] == "currency_symbol_mixing"]
    assert matches
    messages = " ".join(f["message"] for f in matches)
    assert "GBP" in messages  # the "$42,000 (GBP)" shape
    assert "AUD" in messages  # "$55,000 (AUD)"
    assert "MXN" in messages  # "$9,500 (MXN)"


# --- RULE 3: campaign_duration_incoherence ----------------------------------
def test_detects_campaign_duration_incoherence():
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    ws.append(["CAMPAIGN OVERVIEW", None, None])
    ws.append(["£2.0M", "4 weeks", "6"])
    ws.append(["Budget", "Duration", "Locations"])
    ws2 = wb.create_sheet("90-Day Forecast")
    ws2.append(["90-DAY ROLLING FORECAST"])
    ws2.append(["Forecast Period", "Jul 01, 2026 - Sep 30, 2026"])
    ws2.append(["Campaign Duration", "4 weeks"])

    units = [
        bundle_qa._TextUnit("Weeks 1-2", "slide 8"),
        bundle_qa._TextUnit("Weeks 7-12", "slide 8"),
        bundle_qa._TextUnit(
            "Finalize weekly budget ($2M over 1-3 months) and success metrics",
            "slide 11",
        ),
    ]
    findings: list[dict] = []
    bundle_qa._check_campaign_duration_incoherence(units, wb, findings)
    assert any(f["code"] == "campaign_duration_incoherence" for f in findings)


def test_allows_coherent_campaign_duration_despite_fixed_90_day_forecast_window():
    """Regression guard for the false positive this rule's first draft
    produced against the existing manpower/atria reference bundles above:
    the "90-Day Forecast" sheet's "Forecast Period" is a FIXED ~13-week
    rolling window regardless of the plan's total campaign length (both
    reference bundles show the identical "Jul 01 - Sep 30, 2026" window on
    a 24-week and a 78-week campaign respectively). It must never, by
    itself, trigger a false "duration incoherence" against an otherwise
    self-consistent 24-week campaign (a 24-vs-13 week gap that WOULD
    exceed the tolerance if it were treated as authoritative)."""
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    ws.append(["CAMPAIGN OVERVIEW", None, None])
    ws.append(["$150,000", "6 months (~24 weeks)", "1"])
    ws.append(["Budget", "Duration", "Locations"])
    ws2 = wb.create_sheet("90-Day Forecast")
    ws2.append(["90-DAY ROLLING FORECAST"])
    ws2.append(["Forecast Period", "Jul 01, 2026 - Sep 30, 2026"])
    ws2.append(["Campaign Duration", "6 months (~24 weeks)"])

    units = [
        bundle_qa._TextUnit("Weeks 1-12", "slide 8"),
        bundle_qa._TextUnit("Weeks 13-24", "slide 8"),
    ]
    findings: list[dict] = []
    bundle_qa._check_campaign_duration_incoherence(units, wb, findings)
    assert not any(
        f["code"] == "campaign_duration_incoherence" for f in findings
    ), findings


@_requires_uber_xlsx
def test_uber_shipped_bundle_catches_campaign_duration_incoherence():
    findings = _scan_uber_shipped_bundle()
    matches = [f for f in findings if f["code"] == "campaign_duration_incoherence"]
    assert matches
    msg = matches[0]["message"]
    assert "4 weeks" in msg
    assert "Week 12" in msg
    assert "1-3 months" in msg


# --- RULE 4: industry_client_conflict ---------------------------------------
def test_detects_industry_client_conflict():
    findings: list[dict] = []
    bundle_qa._check_industry_client_conflict(
        {
            "client_name": "Uber",
            "industry": "Hospitality & Travel",
            "roles": ["commercial cab driver"],
        },
        findings,
    )
    matches = [f for f in findings if f["code"] == "industry_client_conflict"]
    assert matches
    assert "Hospitality" in matches[0]["message"]
    assert (
        "Rideshare" in matches[0]["message"]
        or "logistics_supply_chain" in matches[0]["message"]
    )


def test_allows_industry_matching_client_and_roles():
    findings: list[dict] = []
    bundle_qa._check_industry_client_conflict(
        {
            "client_name": "Uber",
            "industry": "Rideshare & Gig Economy",
            "roles": ["commercial cab driver"],
        },
        findings,
    )
    assert not any(f["code"] == "industry_client_conflict" for f in findings)


def test_industry_client_conflict_absence_of_signal_is_not_a_conflict():
    """FALSE-POSITIVE FIX (2026-07-26): a client name/role set that carries
    NO usable industry signal must never be treated as "disagreeing" with
    an explicit, specific industry selection -- the old implementation
    independently called classify_industry("", client, roles), whose own
    last resort (when nothing matches) is the generic
    "general_entry_level" bucket, and then flagged the diff between that
    generic catch-all and the real selection as a conflict. Absence of
    evidence is not disagreement. These three real-brief shapes (from a
    10-brief false-positive matrix) must all come back clean."""
    cases = [
        ("Corner Cafe", "Hospitality & Travel", ["Barista"]),
        ("Pearson", "Education", ["Curriculum Designer"]),
        ("Nimbus", "Finance & Banking", ["Analyst"]),
    ]
    for client_name, industry, roles in cases:
        findings: list[dict] = []
        bundle_qa._check_industry_client_conflict(
            {"client_name": client_name, "industry": industry, "roles": roles},
            findings,
        )
        assert not any(
            f["code"] == "industry_client_conflict" for f in findings
        ), f"false positive for {client_name}/{industry}: {findings}"


def test_industry_client_conflict_still_fires_on_a_real_positive_signal():
    """The origin defect must still be caught after the false-positive
    fix: an explicit "Hospitality & Travel" pick for a company actually
    named Uber, with a driver-type role, is a genuine, specific,
    positive-signal disagreement (not a generic fallback) and must still
    fire."""
    findings: list[dict] = []
    bundle_qa._check_industry_client_conflict(
        {
            "client_name": "Uber",
            "industry": "Hospitality & Travel",
            "roles": ["commercial cab driver"],
        },
        findings,
    )
    matches = [f for f in findings if f["code"] == "industry_client_conflict"]
    assert matches
    assert "Hospitality" in matches[0]["message"]
    assert (
        "Rideshare" in matches[0]["message"]
        or "logistics_supply_chain" in matches[0]["message"]
    )


def test_industry_client_conflict_check_never_raises_on_malformed_roles():
    findings: list[dict] = []
    bundle_qa._check_industry_client_conflict(
        {"client_name": "Uber", "roles": [{"weird": "shape"}]}, findings
    )  # must not raise -- absence of a crash IS the assertion here


@_requires_uber_xlsx
def test_uber_shipped_bundle_catches_industry_client_conflict():
    findings = _scan_uber_shipped_bundle()
    matches = [f for f in findings if f["code"] == "industry_client_conflict"]
    assert matches
    assert "Hospitality" in matches[0]["message"]


# --- RULE 5: competitor_count_contradiction ---------------------------------
def test_detects_competitor_count_contradiction():
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Market Intelligence"
    ws.append(["Name", "Industry", "Hiring Activity", "Counter-Strategy"])
    ws.append(["Marriott", "Hospitality & Travel", "Active (est.)", "..."])
    ws.append(["Hilton", "Hospitality & Travel", "Active (est.)", "..."])
    ws.append([])
    ws.append(["Market Positioning"])
    ws.append(
        [
            "Industry Sector: Hospitality & Travel; Is Public Company: No; "
            "Competitor Count: 0; Has Sec Filings: No"
        ]
    )
    findings: list[dict] = []
    bundle_qa._check_competitor_count_contradiction(wb, findings)
    assert any(f["code"] == "competitor_count_contradiction" for f in findings)


def test_allows_matching_competitor_count():
    pytest = __import__("pytest")
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Market Intelligence"
    ws.append(["Name", "Industry", "Hiring Activity", "Counter-Strategy"])
    ws.append(["Marriott", "Hospitality & Travel", "Active (est.)", "..."])
    ws.append(["Hilton", "Hospitality & Travel", "Active (est.)", "..."])
    ws.append([])
    ws.append(["Market Positioning"])
    ws.append(
        [
            "Industry Sector: Hospitality & Travel; Is Public Company: No; "
            "Competitor Count: 2; Has Sec Filings: No"
        ]
    )
    findings: list[dict] = []
    bundle_qa._check_competitor_count_contradiction(wb, findings)
    assert not any(f["code"] == "competitor_count_contradiction" for f in findings)


@_requires_uber_xlsx
def test_uber_shipped_bundle_catches_competitor_count_contradiction():
    findings = _scan_uber_shipped_bundle()
    matches = [
        f for f in findings if f["code"] == "competitor_count_contradiction"
    ]
    assert matches
    assert "Competitor Count: 0" in matches[0]["message"]
    assert "5 named competitor" in matches[0]["message"]


# --- RULE 6: unsourced_competitor_claim -------------------------------------
# Real shipped defect (Uber): the deck and workbook asserted specific,
# never-observed hiring BEHAVIOUR about real named companies as fact --
# "Marriott is actively competing for commercial cab driver candidates in
# UK", "Expect Hilton to keep pressure on...", "Hyatt's hiring activity ...
# puts direct pressure on...", "Hilton is ... drawing from the same...".
# None of it was backed by enrichment: the names came from a static
# per-industry fallback (a hotel-chain roster, because the industry was
# misclassified for a rideshare client) and the workbook's own Sources &
# Confidence sheet grades this exact data at 20%/grade F.
def test_detects_unsourced_competitor_claim_actively_competing():
    units = [
        bundle_qa._TextUnit(
            "Counter: Marriott is actively competing for commercial cab "
            "driver candidates in UK -- sharpen offer cadence and "
            "speed-to-contact to stay ahead of it.",
            "Market Intelligence!F12",
        )
    ]
    findings: list[dict] = []
    bundle_qa._check_unsourced_competitor_claim(units, findings)
    matches = [f for f in findings if f["code"] == "unsourced_competitor_claim"]
    assert matches
    assert "Marriott" in matches[0]["message"]


def test_detects_unsourced_competitor_claim_keeps_pressure_on():
    units = [
        bundle_qa._TextUnit(
            "Counter: Expect Hilton to keep pressure on commercial cab "
            "driver candidates in UK; a faster interview-to-offer cycle is "
            "the clearest lever to counter it.",
            "Quality Intelligence!G20",
        )
    ]
    findings: list[dict] = []
    bundle_qa._check_unsourced_competitor_claim(units, findings)
    matches = [f for f in findings if f["code"] == "unsourced_competitor_claim"]
    assert matches
    assert "Hilton" in matches[0]["message"]


def test_detects_unsourced_competitor_claim_puts_direct_pressure_on():
    units = [
        bundle_qa._TextUnit(
            "Why: Hyatt's hiring activity for similar roles puts direct "
            "pressure on commercial cab driver candidates in UK, while "
            "also competing for the same customer base.",
            "slide 7 / TextBox 33",
        )
    ]
    findings: list[dict] = []
    bundle_qa._check_unsourced_competitor_claim(units, findings)
    assert any(f["code"] == "unsourced_competitor_claim" for f in findings)


def test_detects_unsourced_competitor_claim_drawing_from_the_same():
    units = [
        bundle_qa._TextUnit(
            "Why: Hilton is a same-vertical employer drawing from the same "
            "commercial cab driver candidates in UK this plan targets.",
            "slide 7 / TextBox 27",
        )
    ]
    findings: list[dict] = []
    bundle_qa._check_unsourced_competitor_claim(units, findings)
    assert any(f["code"] == "unsourced_competitor_claim" for f in findings)


def test_detects_unsourced_competitor_claim_is_slower_to_respond():
    units = [
        bundle_qa._TextUnit(
            "IHG is slower to respond to commercial cab driver candidates "
            "in UK, so first-contact speed alone can decide the outcome.",
            "Quality Intelligence!H21",
        )
    ]
    findings: list[dict] = []
    bundle_qa._check_unsourced_competitor_claim(units, findings)
    assert any(f["code"] == "unsourced_competitor_claim" for f in findings)


def test_detects_unsourced_competitor_claim_hiring_directly():
    units = [
        bundle_qa._TextUnit(
            "Airbnb is hiring commercial cab driver candidates directly in "
            "this market.",
            "Quality Intelligence!H22",
        )
    ]
    findings: list[dict] = []
    bundle_qa._check_unsourced_competitor_claim(units, findings)
    assert any(f["code"] == "unsourced_competitor_claim" for f in findings)


def test_detects_unsourced_competitor_claim_especially_aggressive():
    units = [
        bundle_qa._TextUnit(
            "Marriott has been especially aggressive here recently -- "
            "treat this as a priority lane.",
            "Quality Intelligence!H23",
        )
    ]
    findings: list[dict] = []
    bundle_qa._check_unsourced_competitor_claim(units, findings)
    assert any(f["code"] == "unsourced_competitor_claim" for f in findings)


def test_allows_presence_capability_framing():
    """The ALLOWED framing from the policy -- naming a real employer as a
    plausible/likely competitor, true regardless of anything actually
    observed -- must never trip this rule."""
    units = [
        bundle_qa._TextUnit(
            "Marriott is a major hospitality employer in the UK and a "
            "likely competitor for this talent pool.",
            "Market Intelligence!F12",
        ),
        bundle_qa._TextUnit(
            "Hilton is a plausible competitor for commercial cab driver "
            "candidates in UK -- lead with total-comp clarity and a "
            "same-week interview slot.",
            "Quality Intelligence!G13",
        ),
    ]
    findings: list[dict] = []
    bundle_qa._check_unsourced_competitor_claim(units, findings)
    assert not any(f["code"] == "unsourced_competitor_claim" for f in findings)


def test_allows_generic_risk_and_advice_language():
    """Generic why/positioning prose naming no specific asserted behaviour,
    the new "inferred data" disclosure, and advice-only mitigation text must
    not be flagged -- only the specific asserted-behaviour verb phrases
    are banned, not the mere presence of a competitor's name."""
    units = [
        bundle_qa._TextUnit(
            "High hiring volume in London — these employers compete for "
            "the same Hospitality & Travel talent pool",
            "Quality Intelligence!D14",
        ),
        bundle_qa._TextUnit(
            "Competitor set inferred from industry classification; not "
            "verified against live posting data.",
            "Market Intelligence!B40",
        ),
        bundle_qa._TextUnit(
            "Mitigation: Monitor Marriott's job posting volumes weekly; "
            "adjust messaging",
            "slide 9 / TextBox 30",
        ),
    ]
    findings: list[dict] = []
    bundle_qa._check_unsourced_competitor_claim(units, findings)
    assert not any(f["code"] == "unsourced_competitor_claim" for f in findings)


def test_allows_insight_composer_corrected_counter_strategy_output():
    """False-positive guard on the ACTUAL fixed generator (not a
    hand-picked example): every skeleton in every competitor_type bucket,
    at every ordinal position, at both a neutral and a "high" intensity
    (which appends the escalation sentence), must render text this rule
    does not flag. This is also the plan whose competitor data came from
    real enrichment (rather than the static industry fallback) --
    compose_counter_strategy is the same single code path either way, so a
    real-enrichment plan's counter-strategy prose is covered by this same
    proof."""
    import insight_composer as ic

    units: list[bundle_qa._TextUnit] = []
    for bucket, bank in ic._SKELETON_BANKS.items():
        for ordinal in range(len(bank)):
            for intensity in ("", "high", "moderate"):
                sentence = ic.compose_counter_strategy(
                    "Marriott",
                    {
                        "role": "commercial cab driver",
                        "city": "UK",
                        "industry": "Rideshare & Gig Economy",
                        "competitor_type": bucket,
                        "ordinal": ordinal,
                        "intensity": intensity,
                    },
                )
                units.append(
                    bundle_qa._TextUnit(
                        f"Counter: {sentence}", "Quality Intelligence!X1"
                    )
                )

    findings: list[dict] = []
    bundle_qa._check_unsourced_competitor_claim(units, findings)
    flagged = [
        f["message"] for f in findings if f["code"] == "unsourced_competitor_claim"
    ]
    assert not flagged, flagged


def test_unsourced_competitor_claim_check_never_raises_on_malformed_input():
    findings: list[dict] = []
    bundle_qa._check_unsourced_competitor_claim(
        [
            bundle_qa._TextUnit("", "empty"),
            bundle_qa._TextUnit(None, "none"),  # type: ignore[arg-type]
        ],
        findings,
    )  # must not raise -- absence of a crash IS the assertion here


def _scan_uber_shipped_deck_only() -> list[dict]:
    """Scan the frozen fixture DECK alone. The sibling Uber_Media_Plan.xlsx
    that ``_scan_uber_shipped_bundle()`` above also reads is *.xlsx-
    gitignored (see .gitignore:6) and was therefore never committed -- it
    is absent from a fresh checkout regardless of this change (pre-existing
    gap, unrelated to this rule). This rule's own "fires on the real
    shipped bundle" proof does not need the workbook: the exact banned
    sentences (see module docstring above) live in the deck's own
    Competitive Landscape slide."""
    pptx_bytes = (_UBER_FIXTURE_DIR / "Uber_Strategy_Deck.pptx").read_bytes()
    return bundle_qa.run_bundle_qa(pptx_bytes, None, dict(_UBER_DATA))


def test_uber_shipped_bundle_catches_unsourced_competitor_claim():
    findings = _scan_uber_shipped_deck_only()
    matches = [f for f in findings if f["code"] == "unsourced_competitor_claim"]
    assert matches
    named_messages = " ".join(m["message"] for m in matches)
    # the real shipped deck asserts unverified behaviour about at least
    # Marriott and Hilton by name
    assert "Marriott" in named_messages
    assert "Hilton" in named_messages
    assert not any(
        f["code"] == "unsourced_competitor_claim_check_crashed" for f in findings
    )


if __name__ == "__main__":
    import pytest

    raise SystemExit(pytest.main([__file__, "-v"]))
