"""Regression test for S2: Excel Total-row SUM formulas must reference their
OWN column.

Bug (fixed 2026-07): the Executive Summary channel-allocation table's "live
totals row" (excel_v2.py, S89 block) wrote each SUM formula's hardcoded column
letter one column to the LEFT of the cell it was written into -- e.g. the
Total row's Amount($) cell (column D) contained ``=SUM(C19:C29)`` (summing the
Budget % column) instead of ``=SUM(D19:D29)``. This test builds a real
workbook via the production generator (using the real budget engine, so the
channel-allocation shape matches production exactly) and asserts, generically
across every worksheet, that every ``=SUM(<col><n1>:<col><n2>)`` formula cell
references its own column letter -- mirroring scripts/qc_checks.py's
``total-row-formula-alignment`` check.

Runs under pytest, or standalone: ``python3 tests/test_excel_total_row_formulas.py``.
"""

import io
import re
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl  # noqa: E402

import budget_engine  # noqa: E402
import excel_v2  # noqa: E402

_SUM_RE = re.compile(r"=SUM\(([A-Z]+)\d+:([A-Z]+)\d+\)")


def _build_workbook():
    """Build a realistic plan dict (real budget engine output) and generate
    the Excel workbook via the production generator."""
    roles = [
        {"title": "Registered Nurse", "count": 40, "tier": "mid"},
        {"title": "ICU Nurse", "count": 15, "tier": "senior"},
    ]
    locations = [
        {"city": "Dallas", "state": "TX", "country": "United States"},
        {"city": "Houston", "state": "TX", "country": "United States"},
    ]
    channels = {
        "Indeed": 22,
        "Programmatic Job Boards": 18,
        "LinkedIn": 14,
        "Google Search Ads": 10,
        "Meta (Facebook/Instagram)": 8,
        "ZipRecruiter": 7,
        "Nurse.com": 6,
        "Health eCareers": 5,
        "Glassdoor": 4,
        "TikTok": 3,
        "Craigslist": 3,
    }
    alloc = budget_engine.calculate_budget_allocation(
        total_budget=150_000,
        roles=roles,
        locations=locations,
        industry="healthcare",
        channel_percentages=channels,
        collar_type="white",
        campaign_start_month=9,
    )
    data = {
        "client_name": "Mercy Health Partners",
        "requester_name": "Test Runner",
        "requester_email": "test@example.com",
        "industry": "healthcare",
        "budget": "$150,000",
        "budget_period": "campaign",
        "campaign_duration": "3 months",
        "campaign_start_month": 9,
        "hire_volume": "90 hires",
        "work_environment": "onsite",
        "experience_level": "mixed",
        "roles": [r["title"] for r in roles],
        "target_roles": roles,
        "locations": [f"{l['city']}, {l['state']}" for l in locations],
        "_budget_allocation": alloc,
    }
    raw = excel_v2.generate_excel_v2(data)
    assert isinstance(raw, (bytes, bytearray)) and len(raw) > 0
    return openpyxl.load_workbook(io.BytesIO(raw), data_only=False)


def _find_bad_sum_formulas(wb):
    """Generic scan (mirrors scripts/qc_checks.py's total-row-formula-alignment
    check): every =SUM(<col><n1>:<col><n2>) formula cell must reference its
    OWN column letter on both sides of the range."""
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("=SUM("):
                    m = _SUM_RE.match(v)
                    if m:
                        col_letter = cell.coordinate.rstrip("0123456789")
                        if m.group(1) != col_letter or m.group(2) != col_letter:
                            bad.append(f"{ws.title}!{cell.coordinate}={v}")
    return bad


def test_total_row_sum_formulas_reference_own_column():
    wb = _build_workbook()
    bad = _find_bad_sum_formulas(wb)
    assert not bad, (
        "Total-row SUM formula(s) reference a column other than their own "
        f"(off-by-one column shift): {bad}"
    )


def test_executive_summary_totals_row_has_sum_formulas():
    """Sanity check that the Executive Summary sheet's live totals row is
    actually present and non-trivial (guards against a future refactor
    silently dropping the totals row and making the alignment test above
    vacuously pass)."""
    wb = _build_workbook()
    ws = wb["Executive Summary"]
    sum_cells = [
        cell.coordinate
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=SUM(")
    ]
    assert sum_cells, "Expected at least one =SUM(...) formula on Executive Summary"


if __name__ == "__main__":
    test_total_row_sum_formulas_reference_own_column()
    test_executive_summary_totals_row_has_sum_formulas()
    print("OK")
