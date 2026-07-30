"""Pinning tests: the workbook forecast sheet is framed by the campaign's
OWN duration when the plan fits inside the 90-day window.

Real reported defect (Jesse Ofner, 2026-07-31): a 1-month plan's workbook
always presented itself as a "90-Day Forecast" -- sheet title, "90-Day
Total" column, and three monthly periods -- so every short campaign read
as a 90-day plan. The wizard/resolver side was fixed the same day
(fix-duration-granularity: true "2 weeks/1/2/3 months" options,
resolve_campaign_weeks no longer buckets "1 month" to 12 weeks); this
pins the workbook side:

* <=13-week plans: sheet titled "<N>-Week Forecast", weekly columns for
  <=4-week plans (monthly otherwise), a "Campaign Total" column carrying
  the FULL budget, and NO 90-day language anywhere on the sheet.
* >13-week plans: the classic "90-Day Forecast" framing and S89
  first-90-days scaling stay exactly as before (see also
  tests/test_excel_v2_agentb_quality.py, which pins the scaling math).

Runs under pytest, or standalone:
``python3 tests/test_forecast_duration_framing.py``.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl  # noqa: E402

import budget_engine  # noqa: E402
import bundle_qa  # noqa: E402
import excel_v2  # noqa: E402

_TOTAL_BUDGET = 250_000.0


def _base_data(**overrides):
    roles = [
        {"title": "Propane CDL Driver", "count": 20, "tier": "mid"},
        {"title": "Registered Nurse", "count": 10, "tier": "senior"},
    ]
    locations = [
        {"city": "Denver", "state": "CO", "country": "United States"},
        {"city": "Boise", "state": "ID", "country": "United States"},
    ]
    channels = {
        "niche_boards": 15,
        "programmatic_dsp": 25,
        "global_boards": 20,
        "social_media": 10,
        "employer_branding": 10,
        "regional_boards": 20,
    }
    alloc = budget_engine.calculate_budget_allocation(
        total_budget=_TOTAL_BUDGET,
        roles=roles,
        locations=locations,
        industry="logistics_supply_chain",
        channel_percentages=channels,
        collar_type="blue_collar",
        campaign_start_month=3,
    )
    data = {
        "client_name": "PROPANE UNLIMITED CO",
        "industry": "logistics_supply_chain",
        "budget": "$250,000",
        "campaign_duration": "18 months",
        "campaign_weeks": 78,
        "campaign_start_month": 3,
        "hire_volume": "500 hires",
        "work_environment": "onsite",
        "roles": [r["title"] for r in roles],
        "target_roles": roles,
        "locations": [f"{l['city']}, {l['state']}" for l in locations],
        "competitors": ["Suburban Propane", "AmeriGas"],
        "_budget_allocation": alloc,
    }
    data.update(overrides)
    return data


def _forecast_ws(data: dict):
    raw = excel_v2.generate_excel_v2(data)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    title = excel_v2._forecast_sheet_title(data)
    assert title in wb.sheetnames, (title, wb.sheetnames)
    return wb[title]


def _header_row(ws) -> tuple:
    rows = list(ws.iter_rows(min_col=2, values_only=True))
    hdr = next(r for r in rows if r and r[0] == "Metric")
    return tuple(v for v in hdr if v is not None)


def _spend_row(ws) -> list[float]:
    rows = list(ws.iter_rows(min_col=2, values_only=True))
    hi = next(i for i, r in enumerate(rows) if r and r[0] == "Metric")
    spend = next(r for r in rows[hi + 1 :] if r and r[0] == "Spend")
    return [v for v in spend[1:] if isinstance(v, (int, float))]


def _all_text(ws) -> str:
    return " ".join(
        c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)
    )


# ---------------------------------------------------------------------------
# 1. A 4-week plan gets a weekly "4-Week Forecast" of its full budget.
# ---------------------------------------------------------------------------
def test_4_week_plan_gets_weekly_duration_framed_forecast():
    data = _base_data(campaign_duration="4 weeks", campaign_weeks=4)
    ws = _forecast_ws(data)

    assert ws.title == "4-Week Forecast"
    assert _header_row(ws) == (
        "Metric",
        "Week 1",
        "Week 2",
        "Week 3",
        "Week 4",
        "Campaign Total",
        "Trend",
    )

    spend = _spend_row(ws)
    total = spend[-1]
    assert abs(total - _TOTAL_BUDGET) < 1.0, total  # full budget, scale 1.0
    assert abs(sum(spend[:-1]) - total) < 0.51, spend  # weekly cells foot

    text = _all_text(ws)
    assert "4-WEEK FORECAST" in text  # section header (writer uppercases)
    # No 90-day framing and no first-90-days scaling language anywhere.
    for banned in ("90-day", "90 day", "first-90-days", "remaining"):
        assert banned not in text.lower(), banned


# ---------------------------------------------------------------------------
# 2. A 12-week plan keeps monthly columns but is titled by its own length.
# ---------------------------------------------------------------------------
def test_12_week_plan_titled_by_its_own_length_with_monthly_columns():
    data = _base_data(campaign_duration="1-3 months", campaign_weeks=12)
    ws = _forecast_ws(data)

    assert ws.title == "12-Week Forecast"
    hdr = _header_row(ws)
    assert hdr[0] == "Metric" and hdr[-2:] == ("Campaign Total", "Trend")
    month_cols = hdr[1:-2]
    assert len(month_cols) == 3  # ~3 calendar months
    assert all(str(m).split()[-1].isdigit() for m in month_cols)  # "March 2026"

    spend = _spend_row(ws)
    assert abs(spend[-1] - _TOTAL_BUDGET) < 1.0
    assert "90-Day" not in _all_text(ws)


# ---------------------------------------------------------------------------
# 3. A >13-week plan keeps the classic 90-Day framing (S89 pinned elsewhere
#    too -- this pins title + headers so a refactor can't silently drop it).
# ---------------------------------------------------------------------------
def test_long_plan_keeps_90_day_framing_and_scaling_language():
    data = _base_data()  # 78 weeks
    ws = _forecast_ws(data)

    assert ws.title == "90-Day Forecast"
    hdr = _header_row(ws)
    assert hdr[-2:] == ("90-Day Total", "Trend")
    assert len(hdr[1:-2]) == 3

    text = _all_text(ws)
    assert "90-DAY ROLLING FORECAST" in text  # section header (writer uppercases)
    assert "remaining" in text.lower()  # S89 first-90-days footnote intact


# ---------------------------------------------------------------------------
# 4. bundle_qa's footing rule finds the sheet under BOTH framings.
# ---------------------------------------------------------------------------
def test_bundle_qa_finds_forecast_sheet_under_both_framings():
    for overrides in (
        {},
        {"campaign_duration": "4 weeks", "campaign_weeks": 4},
    ):
        data = _base_data(**overrides)
        raw = excel_v2.generate_excel_v2(data)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        name = bundle_qa._forecast_sheet_name(wb)
        assert name == excel_v2._forecast_sheet_title(data), (overrides, name)
        # And the footing rule itself stays green on the generator's output.
        findings: list = []
        bundle_qa._check_90_day_forecast_footing(wb, findings)
        assert findings == [], (overrides, findings)


if __name__ == "__main__":
    test_4_week_plan_gets_weekly_duration_framed_forecast()
    test_12_week_plan_titled_by_its_own_length_with_monthly_columns()
    test_long_plan_keeps_90_day_framing_and_scaling_language()
    test_bundle_qa_finds_forecast_sheet_under_both_framings()
    print("OK")
