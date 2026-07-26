"""Regression tests: collapsing four contradictory campaign-duration
derivations into ONE canonical value carried on the plan data, read by
every surface that states or implies a duration.

Real shipped defect (Uber, brief campaign_duration="1-3 months"): app.py's
regex ladder mapped this string to campaign_weeks=12 ("Weeks 1-12" on the
deck's Implementation Timeline and the workbook's Optimization
Milestones), but the workbook's Executive Summary + 90-Day Forecast
independently re-derived "4 weeks" from the SAME raw string via a
DIFFERENT parser (display_format.parse_duration_to_weeks's generic 52/12
rule reads "1-3 months" as just "1 month" = ~4 weeks), while the deck's
Next Steps slide echoed the raw "1-3 months" string verbatim, never
resolved at all -- four derivations of one input. Fixed by routing every
surface through display_format.resolve_campaign_weeks /
resolve_campaign_duration_label (see that module's docstring, and the
delegation in app.py / excel_v2.py / ppt_generator.py, for the full
architecture). bundle_qa.py's campaign_duration_incoherence rule (added by
a prior wave) already DETECTS this spread; these tests assert it now finds
NOTHING to flag, plus a few direct checks independent of that rule's own
regex heuristics.

VACUOUSNESS: every test below fails on the pre-fix code -- confirmed
against a throwaway `git worktree add --detach HEAD` at the parent commit
cc6b06c -- because (a) excel_v2._resolve_campaign_duration's raw-string
fallback called display_format.parse_duration_to_weeks directly instead of
a phrase-ladder-aware resolver, (b) excel_v2's Optimization Milestones
table was a FIXED "Week 1-2" .. "Week 11-12" schedule regardless of the
plan's actual campaign_weeks, and (c) ppt_generator's Next Steps slide read
data["campaign_duration"] (the raw brief string) instead of the resolved
canonical value.

Runs under pytest, or standalone:
``python3 tests/test_campaign_duration_single_source_of_truth.py``.
"""

from __future__ import annotations

import io
import re
import sys
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl  # noqa: E402

import bundle_qa  # noqa: E402
import display_format  # noqa: E402
import excel_v2  # noqa: E402
import ppt_generator  # noqa: E402
import tools_regen_bundles as trb  # noqa: E402

_REQUIRED_DURATIONS = ["1-3 months", "4 weeks", "18 months", "Ongoing"]

# Any "Week N" or "Week N-M" mention, single or ranged (the Optimization
# Milestones table renders single-week labels, e.g. "Week 4", once a
# campaign is shorter than the default 6-stage schedule -- see
# display_format.scale_week_phases).
_WEEK_MENTION_RE = re.compile(r"\bWeeks?\s+(\d+)(?:\s*[-–]\s*(\d+))?\b")


def _brief(duration: str) -> dict:
    """A single-role, single-location brief -- enough surface area to
    exercise the Executive Summary, 90-Day Forecast, and every deck slide
    duration touches, without the overhead of a multi-role/multi-location
    fixture."""
    return {
        "client_name": "Acme Driver Co",
        "requester_name": "Shubham Singh Chandel",
        "requester_email": "shubhamsingh@joveo.com",
        "industry": "Logistics & Supply Chain",
        "budget": "$150,000",
        "campaign_duration": duration,
        "hire_volume": "300 hires",
        "work_environment": "onsite",
        "locations": ["Dallas, TX"],
        "roles": ["CDL A Driver"],
        "target_roles": [{"title": "CDL A Driver", "count": 300, "tier": "Hourly"}],
        "notes": "Blue collar CDL driver hiring.",
        "competitors": ["FedEx", "UPS"],
    }


def _generate_bundle(duration: str):
    """Build plan data for `duration` and generate both artifacts exactly
    as tools_regen_bundles.py / app.py's real pipeline would (the SAME
    helper the standing production-discriminator checks use)."""
    data = trb.build_plan_data(_brief(duration))
    pptx_bytes = ppt_generator.generate_pptx(data)
    import research as _research

    try:
        xlsx_bytes = excel_v2.generate_excel_v2(data, research_mod=_research)
    except TypeError:
        xlsx_bytes = excel_v2.generate_excel_v2(data)
    return data, pptx_bytes, xlsx_bytes


def _kv_cell_value(ws, label: str) -> str | None:
    """Value of the cell one row ABOVE a `label` cell (the Executive
    Summary's KV hero-card layout: label cell sits below its value)."""
    for row in ws.iter_rows():
        for c in row:
            if c.value == label:
                return ws.cell(row=c.row - 1, column=c.column).value
    return None


def _row_kv_value(ws, label: str) -> str | None:
    """Value of the first non-empty string cell to the RIGHT of a `label`
    cell in the SAME row (the 90-Day Forecast's inline KV row layout)."""
    for row in ws.iter_rows():
        for ci, c in enumerate(row):
            if c.value == label:
                for other in row[ci + 1 :]:
                    if isinstance(other.value, str) and other.value.strip():
                        return other.value
    return None


def _max_week_mentioned(text: str) -> int:
    best = 0
    for m in _WEEK_MENTION_RE.finditer(text):
        best = max(best, int(m.group(1)), int(m.group(2) or m.group(1)))
    return best


# ---------------------------------------------------------------------------
# 1. bundle_qa's own detector must find nothing to flag.
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("duration", _REQUIRED_DURATIONS)
def test_zero_campaign_duration_incoherence_findings(duration):
    """The campaign_duration_incoherence rule a prior wave added
    specifically to DETECT this spread must find NOTHING to flag once the
    cause is fixed."""
    data, pptx_bytes, xlsx_bytes = _generate_bundle(duration)
    findings = bundle_qa.run_bundle_qa(pptx_bytes, xlsx_bytes, data)
    incoherence = [f for f in findings if f["code"] == "campaign_duration_incoherence"]
    assert incoherence == [], f"duration={duration!r} findings: {incoherence}"


# ---------------------------------------------------------------------------
# 2. Direct single-duration check, independent of bundle_qa's own regexes.
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("duration", _REQUIRED_DURATIONS)
def test_executive_summary_and_forecast_state_the_same_duration(duration):
    """The Executive Summary 'Duration' cell and the 90-Day Forecast
    'Campaign Duration' cell must be the EXACT same string -- and that
    string must equal what excel_v2._resolve_campaign_duration resolves
    for this plan's data (the single source of truth)."""
    data, _pptx_bytes, xlsx_bytes = _generate_bundle(duration)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    es_duration = _kv_cell_value(wb["Executive Summary"], "Duration")
    fc_duration = _row_kv_value(wb["90-Day Forecast"], "Campaign Duration")

    assert es_duration, f"Executive Summary Duration cell missing for {duration!r}"
    assert fc_duration, f"90-Day Forecast Campaign Duration cell missing for {duration!r}"
    assert es_duration == fc_duration, (
        f"duration={duration!r}: Executive Summary said {es_duration!r} but "
        f"90-Day Forecast said {fc_duration!r}"
    )
    assert es_duration == excel_v2._resolve_campaign_duration(data)


# ---------------------------------------------------------------------------
# 3. The deck's phased timeline must never draw a roadmap longer than the
#    campaign itself.
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("duration", _REQUIRED_DURATIONS)
def test_deck_timeline_never_exceeds_campaign_length(duration):
    """A client must never see a roadmap phase longer than the campaign
    they are buying (a 4-week campaign must not render "Weeks 7-12")."""
    data, pptx_bytes, _xlsx_bytes = _generate_bundle(duration)
    campaign_weeks = int(data.get("campaign_weeks") or 0)
    assert campaign_weeks > 0

    findings: list = []
    units = bundle_qa._iter_pptx_texts(pptx_bytes, findings)
    max_week = max((_max_week_mentioned(u.text) for u in units), default=0)
    assert max_week > 0, f"no 'Week N[-M]' phase text found on the deck for {duration!r}"
    assert max_week <= campaign_weeks, (
        f"duration={duration!r} (campaign_weeks={campaign_weeks}): deck "
        f"timeline phase text reaches Week {max_week}, past the campaign's "
        "own length"
    )


# ---------------------------------------------------------------------------
# 4. Same invariant for the workbook's own Optimization Milestones table
#    (previously a FIXED "Week 1-2" .. "Week 11-12" schedule regardless of
#    campaign_weeks).
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("duration", _REQUIRED_DURATIONS)
def test_workbook_milestones_never_exceed_campaign_length(duration):
    data, _pptx_bytes, xlsx_bytes = _generate_bundle(duration)
    campaign_weeks = int(data.get("campaign_weeks") or 0)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["90-Day Forecast"]
    rows = list(ws.iter_rows(values_only=False))
    header_row = next(
        (
            ri
            for ri, row in enumerate(rows)
            for c in row
            if c.value == "Optimization Milestones"
        ),
        None,
    )
    assert header_row is not None, f"Optimization Milestones header missing for {duration!r}"

    max_week = 0
    saw_a_milestone = False
    for row in rows[header_row + 1 :]:
        label = row[1].value if len(row) > 1 else None
        if not isinstance(label, str) or not label.strip():
            # Blank row (or the next subsection) ends the milestones block.
            if saw_a_milestone:
                break
            continue
        if not label.lower().startswith("week"):
            break
        saw_a_milestone = True
        max_week = max(max_week, _max_week_mentioned(label))

    assert saw_a_milestone, f"no milestone rows found for {duration!r}"
    # Exact equality, not just "<=": the milestones must SPAN the actual
    # campaign (display_format.scale_week_phases always pins the last phase
    # to end exactly at campaign_weeks), catching BOTH directions of the
    # pre-fix bug -- a short campaign whose fixed milestones overshot past
    # its own end (e.g. "4 weeks" reaching "Week 11-12"), and a long
    # campaign whose fixed milestones fell short of its own end (e.g. "18
    # months"/78 weeks stuck at "Week 11-12" instead of reaching 78).
    assert max_week == campaign_weeks, (
        f"duration={duration!r} (campaign_weeks={campaign_weeks}): workbook "
        f"Optimization Milestones reach Week {max_week}, not the campaign's "
        "own end"
    )


# ---------------------------------------------------------------------------
# 5. The deck's Next Steps slide must never echo the raw dropdown string.
# ---------------------------------------------------------------------------
def test_next_steps_slide_never_echoes_the_raw_dropdown_range_string():
    """The exact shipped defect: the deck's Next Steps slide stated
    "... over 1-3 months" (the raw brief string) while every other surface
    had already resolved to "12 weeks" for the SAME plan."""
    data, pptx_bytes, _xlsx_bytes = _generate_bundle("1-3 months")
    canonical = excel_v2._resolve_campaign_duration(data)
    assert canonical == "12 weeks"

    findings: list = []
    units = bundle_qa._iter_pptx_texts(pptx_bytes, findings)
    all_text = "\n".join(u.text for u in units)
    assert "1-3 months" not in all_text, (
        "deck still echoes the raw brief duration string instead of the "
        f"canonical {canonical!r}"
    )
    assert "12 weeks" in all_text, "deck never states the canonical duration at all"


# ---------------------------------------------------------------------------
# 6. "Ongoing" grammar handling must survive the consolidation.
# ---------------------------------------------------------------------------
def test_ongoing_duration_never_reads_as_a_fixed_length():
    """Regression: an open-ended "Ongoing" campaign must never silently
    resolve to a specific fixed-length label (e.g. a stale/duplicated
    formatter turning 52 weeks into "1 year (~12 months)") -- the
    established grammar phrase across this codebase is "Ongoing (no fixed
    end date)" (see excel_v2._build_narrative_facts_block /
    _is_unbounded_duration, ppt_generator._interpolate_next_steps)."""
    data, pptx_bytes, xlsx_bytes = _generate_bundle("Ongoing")
    canonical = excel_v2._resolve_campaign_duration(data)
    assert canonical == "Ongoing (no fixed end date)"
    assert display_format.is_unbounded_duration(canonical)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert _kv_cell_value(wb["Executive Summary"], "Duration") == canonical
    assert _row_kv_value(wb["90-Day Forecast"], "Campaign Duration") == canonical

    findings: list = []
    units = bundle_qa._iter_pptx_texts(pptx_bytes, findings)
    all_text = "\n".join(u.text for u in units)
    assert "over Ongoing" not in all_text
    assert "1 year (~12 months)" not in all_text


# ---------------------------------------------------------------------------
# 7. display_format.scale_week_phases -- the shared partition utility the
#    Optimization Milestones table now uses -- on its own.
# ---------------------------------------------------------------------------
@pytest.mark.parametrize(
    "total_weeks,num_phases,expected",
    [
        (4, 6, [(1, 1), (2, 2), (3, 3), (4, 4), (4, 4), (4, 4)]),
        (12, 6, [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10), (11, 12)]),
        (24, 6, [(1, 4), (5, 8), (9, 12), (13, 16), (17, 20), (21, 24)]),
        (78, 6, [(1, 13), (14, 26), (27, 39), (40, 52), (53, 65), (66, 78)]),
    ],
)
def test_scale_week_phases_examples(total_weeks, num_phases, expected):
    assert display_format.scale_week_phases(total_weeks, num_phases) == expected


def test_scale_week_phases_last_phase_always_ends_at_total():
    for total in (1, 3, 5, 7, 10, 13, 24, 48, 52, 78, 156):
        for n in (1, 3, 6):
            phases = display_format.scale_week_phases(total, n)
            assert phases[-1][1] == total
            assert len(phases) == n
            for start, end in phases:
                assert 1 <= start <= end <= total


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-v"]))
