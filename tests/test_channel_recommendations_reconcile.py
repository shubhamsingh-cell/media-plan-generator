"""Regression tests for O2 (2026-07-03): the workbook must present ONE plan.

Bugs fixed (BUNDLE_QC_FINDINGS_2026-07-03.json findings 35/48/58/63/70/77 and
42/49/64):

1. Two contradictory media plans in one bundle. The "Channel Recommendations"
   sheet was populated by an independent engine (channel_recommender) that
   recomputed its own allocation %, CPC/apply-rate/hire-rate and channel mix,
   so it projected a different application funnel and CPA than the Executive
   Summary / ROI Projections / 90-Day Forecast (which use
   budget_engine.calculate_budget_allocation). Its Spend column also did not
   sum to the stated budget. O2 makes the sheet render the SAME
   ``_budget_allocation`` object as the rest of the plan.

2. The client's stated hiring goal vs the plan's projected hires was never
   reconciled. O2 adds an explicit, quantified goal-gap callout on the
   Executive Summary when the plan is materially short of goal.

3. Campaign duration was worded independently per sheet ("1-2 years" on one
   sheet, a 90-day/12-week window on others). O2 resolves a single canonical
   duration string that every sheet references.

Runs under pytest, or standalone:
``python3 tests/test_channel_recommendations_reconcile.py``.
"""

import io
import re
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl  # noqa: E402

import budget_engine  # noqa: E402
import excel_v2  # noqa: E402


def _build_workbook(hire_volume: str, campaign_duration: str, campaign_weeks=None):
    roles = [{"title": "Business Sales Representative", "count": 5000, "tier": "entry"}]
    locations = [{"city": "Phoenix", "state": "AZ", "country": "United States"}]
    channels = {
        "Programmatic DSP": 40,
        "Global Job Boards": 22,
        "Social Media": 20,
        "Niche Boards": 8,
        "Regional Boards": 7,
        "Employer Branding": 3,
    }
    alloc = budget_engine.calculate_budget_allocation(
        total_budget=240_000,
        roles=roles,
        locations=locations,
        industry="retail_ecommerce",
        channel_percentages=channels,
        collar_type="white",
        campaign_start_month=6,
    )
    data = {
        "client_name": "Omada.ai",
        "industry": "retail_ecommerce",
        "budget": "$240,000",
        "budget_period": "campaign",
        "campaign_duration": campaign_duration,
        "campaign_start_month": 6,
        "hire_volume": hire_volume,
        "work_environment": "remote",
        "roles": [r["title"] for r in roles],
        "target_roles": roles,
        "locations": ["Phoenix, AZ, United States"],
        "country": "United States",
        "_budget_allocation": alloc,
    }
    if campaign_weeks is not None:
        data["campaign_weeks"] = campaign_weeks
        data["campaign_duration_canonical"] = excel_v2._resolve_campaign_duration(data)
    raw = excel_v2.generate_excel_v2(data)
    if isinstance(raw, tuple):
        raw = raw[0]
    assert isinstance(raw, (bytes, bytearray)) and len(raw) > 0
    return openpyxl.load_workbook(io.BytesIO(raw), data_only=False), data


def _exec_projected_hires(ws) -> int:
    """The Executive Summary 'Projected Hires' metric card (value cell sits one
    row above the 'Projected Hires' label cell)."""
    for row in ws.iter_rows():
        for c in row:
            if c.value == "Projected Hires":
                val = ws.cell(row=c.row - 1, column=c.column).value
                return int(val or 0)
    raise AssertionError("Executive Summary 'Projected Hires' card not found")


def _channel_rec_total_row(ws):
    """Return (spend, apps, hires) from the Channel Recommendations TOTAL row."""
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "TOTAL":
            return (
                float(ws.cell(row=r, column=5).value or 0),  # Spend
                int(ws.cell(row=r, column=9).value or 0),  # Apps
                int(ws.cell(row=r, column=10).value or 0),  # Hires
            )
    raise AssertionError("Channel Recommendations TOTAL row not found")


def test_channel_recommendations_is_the_same_plan():
    """The Channel Recommendations sheet must project the SAME total hires as the
    Executive Summary (single plan), not an independently-computed second plan."""
    wb, _ = _build_workbook(hire_volume="5000 hires", campaign_duration="1-2 years")
    exec_hires = _exec_projected_hires(wb["Executive Summary"])
    _, _, cr_hires = _channel_rec_total_row(wb["Channel Recommendations"])
    assert cr_hires == exec_hires, (
        f"Channel Recommendations hires ({cr_hires}) must equal Executive "
        f"Summary hires ({exec_hires}); a divergence means the workbook is "
        "presenting two contradictory plans again."
    )


def test_channel_recommendations_spend_foots_to_budget():
    """The Channel Recommendations Spend column must sum to the stated budget
    (it now uses the same per-channel dollar amounts as every other sheet)."""
    wb, _ = _build_workbook(hire_volume="5000 hires", campaign_duration="1-2 years")
    cr = wb["Channel Recommendations"]
    cr_spend_total, _, _ = _channel_rec_total_row(cr)
    # Sum the per-channel Spend cells (column E = 5), excluding the TOTAL row.
    # A data row is one whose Hires cell (column J = 10) holds a number.
    per_channel = 0.0
    for r in range(1, cr.max_row + 1):
        label = cr.cell(row=r, column=2).value
        spend = cr.cell(row=r, column=5).value
        hires = cr.cell(row=r, column=10).value
        if (
            label not in (None, "", "TOTAL")
            and isinstance(spend, (int, float))
            and isinstance(hires, (int, float))
        ):
            per_channel += spend
    # Budget for this plan is 240,000; allow a small rounding tolerance because
    # per-channel dollar_amounts are rounded to cents upstream.
    assert abs(cr_spend_total - 240_000) <= 1.0, (
        f"Channel Recommendations TOTAL spend {cr_spend_total} must foot to the "
        "240,000 budget"
    )
    assert abs(per_channel - cr_spend_total) <= 1.0, (
        "Sum of per-channel Spend cells must equal the TOTAL row spend"
    )


def test_goal_gap_callout_present_when_short():
    """A large goal-vs-projection shortfall must be stated head-on."""
    wb, _ = _build_workbook(hire_volume="5000 hires", campaign_duration="1-2 years")
    ws = wb["Executive Summary"]
    found = None
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and "Hiring-goal gap" in c.value:
                found = c.value
    assert found, "Expected an explicit hiring-goal-gap callout on Executive Summary"
    # It must name both the projection and the goal.
    assert "5,000" in found or "5000" in found, "Callout must state the stated goal"
    assert "% of goal" in found, "Callout must quantify the shortfall as % of goal"


def test_goal_gap_callout_absent_when_goal_met():
    """When the plan meets (or is within 10% of) the goal, no gap callout fires."""
    # Goal 1 is trivially met by any positive projection.
    wb, _ = _build_workbook(hire_volume="1 hires", campaign_duration="3 months")
    ws = wb["Executive Summary"]
    for row in ws.iter_rows():
        for c in row:
            assert not (
                isinstance(c.value, str) and "Hiring-goal gap" in c.value
            ), "Gap callout should not appear when the goal is met"


def test_duration_is_consistent_across_sheets():
    """The same canonical duration string must appear on the Executive Summary
    and the 90-Day Forecast (single source of truth)."""
    wb, data = _build_workbook(
        hire_volume="5000 hires", campaign_duration="1-2 years", campaign_weeks=80
    )
    canonical = excel_v2._resolve_campaign_duration(data)
    assert "year" in canonical  # 1-2 years -> "1.5 years (~18 months)"

    es = wb["Executive Summary"]
    es_duration = None
    for row in es.iter_rows():
        for c in row:
            if c.value == "Duration":
                es_duration = es.cell(row=c.row - 1, column=c.column).value
    assert es_duration == canonical, (
        f"Executive Summary duration {es_duration!r} must equal canonical "
        f"{canonical!r}"
    )

    fc = wb["90-Day Forecast"]
    fc_has_canonical = any(
        isinstance(c.value, str) and canonical in c.value
        for row in fc.iter_rows()
        for c in row
    )
    assert fc_has_canonical, (
        "90-Day Forecast must reference the same canonical duration string"
    )


if __name__ == "__main__":
    test_channel_recommendations_is_the_same_plan()
    test_channel_recommendations_spend_foots_to_budget()
    test_goal_gap_callout_present_when_short()
    test_goal_gap_callout_absent_when_goal_met()
    test_duration_is_consistent_across_sheets()
    print("OK")
