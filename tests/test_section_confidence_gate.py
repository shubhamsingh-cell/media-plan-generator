"""Tests for the GENERAL CONFIDENCE GATE (excel_v2.py).

The Sources & Confidence sheet already grades every synthesis section that
carries a computed score (data_synthesizer.compute_confidence_scores) A-F --
on the real Uber bundle: Overall 47% F, Salary Intelligence 40% F, Location
Profiles 20% F, Ad Platform Analysis 40% F, Competitive Intelligence 20% F,
Workforce Insights 60% D. Before this fix, only ONE section (Competitive
Intelligence's "static fallback" competitor roster) changed its own
presentation in response to a low score; every other section rendered in
the identical confident register as a fully-sourced (grade A) section.

This file tests the general mechanism (excel_v2._section_confidence /
_confidence_gated_title / _write_confidence_gate_note) both at the unit
level and end-to-end through generate_excel_v2, across:
  - multiple industries, including "insurance" (NOT covered by
    research.INDUSTRY_LABOUR_MARKET -- see test_silent_fallback_disclosure.py)
  - a low-confidence section (F/D grade) that must hedge
  - a high-confidence section (A grade) that must NOT hedge -- the
    false-positive guard the brief calls out as mattering most
  - a US plan and a non-US (UK) plan

VACUOUSNESS: every test in this file is new with this change and was run
against a throwaway pre-fix worktree (git worktree add --detach HEAD at the
parent commit) where it failed with AttributeError (helpers didn't exist)
/ AssertionError (no hedge text present) -- see the task report for the
observed pre-fix output.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl  # noqa: E402

import excel_v2  # noqa: E402


# ---------------------------------------------------------------------------
# Unit-level tests on the gate helpers themselves
# ---------------------------------------------------------------------------


def test_grade_bands_match_reported_uber_bundle():
    """Sanity-lock the letter-grade bands against the real bundle's own
    reported grades, so the gate's thresholds can never silently drift
    from what Sources & Confidence itself already shows."""
    assert excel_v2._grade_from_score(0.47) == "F"  # Overall
    assert excel_v2._grade_from_score(0.40) == "F"  # Salary Intelligence
    assert excel_v2._grade_from_score(0.20) == "F"  # Location Profiles / Comp Intel
    assert excel_v2._grade_from_score(0.60) == "D"  # Workforce Insights
    assert excel_v2._grade_from_score(0.90) == "A"


def test_section_confidence_reads_per_section_dict():
    data = {
        "_synthesized": {
            "confidence_scores": {"per_section": {"salary_intelligence": 0.4}}
        }
    }
    result = excel_v2._section_confidence(data, "salary_intelligence")
    assert result == (0.4, "F")


def test_section_confidence_reads_sections_alias_key():
    data = {
        "_synthesized": {"confidence_scores": {"sections": {"workforce_insights": 0.6}}}
    }
    assert excel_v2._section_confidence(data, "workforce_insights") == (0.6, "D")


def test_section_confidence_none_when_not_scored():
    """A section with no computed score at all -> None, not "low
    confidence" -- the gate must never invent a hedge out of absent data."""
    data = {"_synthesized": {"confidence_scores": {"per_section": {}}}}
    assert excel_v2._section_confidence(data, "salary_intelligence") is None
    assert excel_v2._section_confidence({}, "salary_intelligence") is None


def test_confidence_gated_title_hedge_and_estimate_tiers():
    # Well-sourced: unchanged.
    assert excel_v2._confidence_gated_title("Salary Intelligence", (0.9, "A")) == (
        "Salary Intelligence"
    )
    # D-tier hedge: grade suffix, no "Estimate" label yet.
    hedged = excel_v2._confidence_gated_title("Workforce Trends", (0.6, "D"))
    assert hedged == "Workforce Trends (D Confidence)"
    # F-tier: explicit "Estimate" label.
    estimated = excel_v2._confidence_gated_title("Location Intelligence", (0.2, "F"))
    assert estimated == "Location Intelligence — Estimate (F Confidence)"
    # No score at all: unchanged.
    assert excel_v2._confidence_gated_title("Ad Platform Analysis", None) == (
        "Ad Platform Analysis"
    )


def test_write_confidence_gate_note_false_positive_guard():
    """The false-positive guard at the unit level: a well-sourced section
    (score >= hedge threshold) must leave the row counter COMPLETELY
    unchanged -- proof that nothing is written at all, not just that the
    text is short."""
    wb = openpyxl.Workbook()
    ws = wb.active
    row_before = 10
    row_after = excel_v2._write_confidence_gate_note(
        ws, row_before, "salary_intelligence", (0.9, "A")
    )
    assert row_after == row_before
    # No score at all -> also untouched.
    row_after_none = excel_v2._write_confidence_gate_note(
        ws, row_before, "salary_intelligence", None
    )
    assert row_after_none == row_before


def test_write_confidence_gate_note_hedge_and_estimate_text():
    wb = openpyxl.Workbook()
    ws = wb.active
    row = excel_v2._write_confidence_gate_note(
        ws, 5, "competitive_intelligence", (0.2, "F")
    )
    assert row == 6
    text = ws.cell(row=5, column=excel_v2.COL_START).value
    assert "20%" in text and "F" in text
    assert "not verified against" in text
    assert "directional estimates, not verified benchmarks" in text

    # D-tier (hedge but not the bottom-end "estimate" sentence).
    ws2 = openpyxl.Workbook().active
    row2 = excel_v2._write_confidence_gate_note(
        ws2, 5, "workforce_insights", (0.6, "D")
    )
    text2 = ws2.cell(row=5, column=excel_v2.COL_START).value
    assert "60%" in text2 and "not verified against" in text2
    assert "directional estimates" not in text2


# ---------------------------------------------------------------------------
# End-to-end: generate_excel_v2 with controlled confidence scores
# ---------------------------------------------------------------------------


def _confidence_plan_data(
    industry: str,
    locations: list[str],
    per_section: dict,
    **overrides,
) -> dict:
    data = {
        "client_name": "Pearson",
        "company_name": "Pearson",
        "industry": industry,
        "budget": "$150,000",
        "locations": locations,
        "roles": ["Curriculum Designer"],
        "target_roles": ["Curriculum Designer"],
        "campaign_duration": "3 months",
        "hire_volume": "100",
        "work_environment": "remote",
        "_enriched": {},
        "_budget_allocation": {},
        "_synthesized": {
            "confidence_scores": {"overall": 0.5, "per_section": per_section},
            "salary_intelligence": {
                "Curriculum Designer": {
                    "min": 30000,
                    "p25": 35000,
                    "median": 45000,
                    "p75": 50000,
                    "max": 60000,
                    "confidence": 0.5,
                },
            },
            "ad_platform_analysis": {},
            "location_profiles": {},
            "competitive_intelligence": {},
            "workforce_insights": {
                "hiring_trends": {"note": "Demand is rising for this role."},
            },
        },
    }
    data.update(overrides)
    return data


def _sheet_text(ws) -> str:
    parts = []
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if val is not None:
                parts.append(str(val))
    return "\n".join(parts)


def test_low_confidence_sections_hedge_on_uncovered_industry_non_us_plan():
    """'insurance' is NOT one of the 12 keys in
    research.INDUSTRY_LABOUR_MARKET/INDUSTRY_COMPETITORS -- also exercises
    the "uncovered industry" + "non-US plan" legs of the generalisation
    requirement. Every scored section here is F/D grade and must hedge."""
    data = _confidence_plan_data(
        industry="insurance",
        locations=["London, United Kingdom"],
        per_section={
            "salary_intelligence": 0.4,  # F
            "ad_platform_analysis": 0.4,  # F
            "location_profiles": 0.2,  # F
            "competitive_intelligence": 0.2,  # F
            "workforce_insights": 0.6,  # D
        },
    )
    raw = excel_v2.generate_excel_v2(data)
    wb = openpyxl.load_workbook(io.BytesIO(raw))

    mi_text = _sheet_text(wb["Market Intelligence"])
    ch_text = _sheet_text(wb["Channels & Strategy"])

    # Hedged headers (grade suffix on every gated section).
    assert "SALARY INTELLIGENCE — ESTIMATE (F CONFIDENCE)" in mi_text
    assert "LOCATION INTELLIGENCE — ESTIMATE (F CONFIDENCE)" in mi_text
    assert "COMPETITIVE LANDSCAPE — ESTIMATE (F CONFIDENCE)" in mi_text
    assert "WORKFORCE TRENDS (D CONFIDENCE)" in mi_text
    assert "AD PLATFORM ANALYSIS — ESTIMATE (F CONFIDENCE)" in ch_text

    # Visible provenance lines with confidence % + grade, plus the bottom-
    # tier "estimate" sentence for the F-grade sections only.
    assert "40% (F)" in mi_text or "40% (F)" in ch_text
    assert "20% (F)" in mi_text
    assert "60% (D)" in mi_text
    assert mi_text.count("not verified against") >= 3
    assert "directional estimates, not verified benchmarks" in mi_text


def test_high_confidence_section_never_hedges_false_positive_guard():
    """The false-positive guard: a genuinely well-sourced section (grade A)
    must render IDENTICALLY to a plan with no confidence data at all --
    plain header, no grade suffix, no provenance footnote, no hedge
    language anywhere near it. Uses a covered US industry/plan."""
    data = _confidence_plan_data(
        industry="healthcare_medical",
        locations=["Dallas, TX"],
        per_section={
            "salary_intelligence": 0.9,  # A
            "workforce_insights": 0.9,  # A
        },
    )
    raw = excel_v2.generate_excel_v2(data)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    mi_text = _sheet_text(wb["Market Intelligence"])

    # Plain, un-suffixed headers.
    assert "SALARY INTELLIGENCE" in mi_text
    assert "SALARY INTELLIGENCE (" not in mi_text
    assert "SALARY INTELLIGENCE —" not in mi_text
    assert "WORKFORCE TRENDS" in mi_text
    assert "WORKFORCE TRENDS (" not in mi_text
    assert "WORKFORCE TRENDS —" not in mi_text

    # No hedge/provenance language from THIS mechanism leaked in. (Not
    # asserting "not verified against" is absent -- that phrase is shared
    # vocabulary with the pre-existing, unrelated Competitive Intelligence
    # "static fallback" disclosure, see _comp_from_static_fallback, which
    # can legitimately fire here too since this fixture supplies no
    # competitor data; "Data confidence for this section" and the
    # estimate sentence below are THIS gate's own unique markers.)
    assert "directional estimates, not verified benchmarks" not in mi_text
    assert "Data confidence for this section" not in mi_text


def test_unscored_plan_renders_exactly_as_before_the_mechanism_existed():
    """A plan generated with no confidence_scores at all (the vast
    majority of unit-test fixtures elsewhere in this suite) must be
    completely unaffected -- _section_confidence returns None for every
    section, and None is always a no-op for both title and footnote."""
    data = _confidence_plan_data(
        industry="tech_engineering",
        locations=["Austin, TX"],
        per_section={},  # nothing scored at all
    )
    raw = excel_v2.generate_excel_v2(data)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    mi_text = _sheet_text(wb["Market Intelligence"])
    assert "SALARY INTELLIGENCE" in mi_text
    assert "Data confidence for this section" not in mi_text
    assert "Confidence)" not in mi_text


if __name__ == "__main__":
    import pytest

    raise SystemExit(pytest.main([__file__, "-v"]))
