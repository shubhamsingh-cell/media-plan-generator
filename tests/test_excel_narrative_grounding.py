"""Grounding-validator + no-fabrication-guarantee tests for the Excel
deliverable's Executive Strategic Summary narrative (excel_v2.py).

House rule: never fabricate data. Root cause this guards against: the
narrative's system prompt used to command "Every sentence must contain a
number or specific insight" while only ~10 real numbers were ever passed to
the model -- so it reliably invented the rest. Verified against a real prod
bundle, none of these existed in the plan data: a "$5,000 industry average
cost-per-hire", a "1:2.4 cost-to-value ratio", "$360,000 in tangible value",
a "22% supply gap", a "38% reduction", "$7,500 lost revenue per unfilled
seat".

Fix (see excel_v2.py, the block above ``_build_sheet_executive_summary``):
  1. The prompt drops the "every sentence needs a number" mandate and is
     given a FACTS block built from every real number the plan already has
     (``_gather_narrative_grounding_context`` / ``_build_narrative_facts_block``).
  2. ``_narrative_is_grounded`` re-parses the model's own narrative after the
     call and rejects it if it cites any $/%%/ratio/large-integer figure that
     doesn't trace back (within a rounding tolerance) to a number in that
     SAME FACTS block.
  3. A rejected (or otherwise unavailable) narrative is replaced by
     ``_build_deterministic_executive_summary`` -- a template built ONLY
     from real plan fields, never a second LLM call.

This file covers:
  A. The grounding validator in isolation (extraction, tolerance rules,
     fabrication detection + naming).
  B. The end-to-end excel_v2.generate_excel_v2 flow with a mocked
     ``llm_router.call_llm``: a clean/grounded narrative is used verbatim; a
     fabricating narrative (the exact shapes from the bug report) is
     rejected and replaced by the deterministic fallback, which contains
     only real plan numbers.
  C. The deterministic fallback renders with all-real numbers (verified via
     the SAME grounding validator) for both reference briefs
     (tools_regen_bundles.MANPOWER_BRIEF / ATRIA_BRIEF).

See tests/test_excel_narrative_reliability.py for the LLM-call-plumbing
reliability tests (routing, timeout budget, exception handling) this file
does not re-cover.

Runs under pytest, or standalone: ``python3 tests/test_excel_narrative_grounding.py``.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path
from unittest import mock

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl  # noqa: E402
import pytest  # noqa: E402

import excel_v2  # noqa: E402


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
def _sheet_text(ws) -> str:
    parts = []
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if val is not None:
                parts.append(str(val))
    return "\n".join(parts)


def _all_text(wb) -> str:
    return "\n".join(_sheet_text(ws) for ws in wb.worksheets)


def _minimal_data_with_allocation(**overrides) -> dict:
    """A richer fixture than test_excel_narrative_reliability.py's
    _minimal_data(): includes a populated ``_budget_allocation`` so
    projected hires / blended CPH / top channels are all non-zero and the
    FACTS block (and therefore the grounding tests below) has real
    substance to check against."""
    data = {
        "client_name": "Acme Corp",
        "company_name": "Acme Corp",
        "industry": "logistics_supply_chain",
        "budget": "$150,000",
        "locations": ["Dallas, TX"],
        "roles": ["CDL Driver"],
        "target_roles": ["CDL Driver"],
        "campaign_duration": "6 months",
        "hire_volume": "400",
        "work_environment": "onsite",
        "_enriched": {},
        "_synthesized": {},
        "_budget_allocation": {
            "sufficiency": {"grade": "B"},
            "channel_allocations": {
                "programmatic_dsp": {
                    "dollar_amount": 90000,
                    "percentage": 60,
                    "projected_clicks": 45000,
                    "projected_applications": 3600,
                    "projected_hires": 210,
                    "cpc": 2.0,
                    "cpa": 25.0,
                    "roi_score": 8,
                },
                "niche_boards": {
                    "dollar_amount": 60000,
                    "percentage": 40,
                    "projected_clicks": 20000,
                    "projected_applications": 2000,
                    "projected_hires": 132,
                    "cpc": 3.0,
                    "cpa": 30.0,
                    "roi_score": 7,
                },
            },
        },
    }
    data.update(overrides)
    return data


# FACTS block equivalent to what `_minimal_data_with_allocation()` produces
# (captured once against real code + spot-checked; used by the validator
# unit tests in section A below, which test `_narrative_is_grounded` in
# isolation from the LLM plumbing).
_SAMPLE_FACTS = (
    "Client: Acme Corp\n"
    "Industry: Logistics & Supply Chain\n"
    "Budget: $150,000\n"
    "Duration: 6 months (~26 weeks)\n"
    "Stated Hiring Goal: 400\n"
    "Projected Hires: 342\n"
    "Blended Cost/Hire: $438.60\n"
    "Total Projected Applications: 5,600\n"
    "Blended Cost/Application: $26.79\n"
    "Budget Sufficiency Grade: B\n"
    "Hiring Goal Gap: 58 hires short of the 400 goal (86% of goal)\n"
    "Additional Budget To Close Gap: $25,439\n"
    "Top Channel — Programmatic (DSP): $90,000 (60% of budget) 210 hires CPA $25.00\n"
    "Top Channel — Niche / Industry Boards: $60,000 (40% of budget) 132 hires CPA $30.00\n"
)


# ---------------------------------------------------------------------------
# A. Grounding validator unit tests
# ---------------------------------------------------------------------------
def test_narrative_citing_only_facts_numbers_passes():
    allowed = excel_v2._allowed_numbers_from_facts_text(_SAMPLE_FACTS)
    text = (
        "This $150,000 plan is projected to deliver 342 hires against the "
        "stated goal of 400, putting it at roughly 86% of target. "
        "Programmatic (DSP) leads the mix at $90,000 (60% of budget) with "
        "210 projected hires at a $25.00 CPA, complemented by $60,000 in "
        "Niche / Industry Boards driving 132 hires at $30.00 CPA. Blended "
        "cost per hire lands at $438.60 across 5,600 projected applications "
        "($26.79 blended cost per application), a Budget Sufficiency Grade "
        "of B. Recommended next step: launch both channels within the "
        "first two weeks."
    )
    grounded, untraceable = excel_v2._narrative_is_grounded(text, allowed)
    assert grounded is True, f"unexpectedly rejected, untraceable={untraceable}"
    assert untraceable == []


def test_narrative_with_fabricated_figures_is_rejected_and_names_them():
    """The exact shapes from the bug report: a fake industry-average CPH, a
    fake cost-to-value ratio, a fake tangible-value dollar figure, fake
    supply-gap/reduction percentages, and a fake lost-revenue figure -- none
    present in _SAMPLE_FACTS."""
    allowed = excel_v2._allowed_numbers_from_facts_text(_SAMPLE_FACTS)
    text = (
        "The industry average cost-per-hire for logistics roles is $5,000, "
        "and against that benchmark this plan achieves a 1:2.4 "
        "cost-to-value ratio, generating $360,000 in tangible value. "
        "Deploying at scale closes a 22% supply gap in this market via a "
        "38% reduction in time-to-fill, and every unfilled seat otherwise "
        "costs roughly $7,500 in lost revenue."
    )
    grounded, untraceable = excel_v2._narrative_is_grounded(text, allowed)
    assert grounded is False
    for fig in ("$5,000", "1:2.4", "$360,000", "22%", "38%", "$7,500"):
        assert fig in untraceable, f"{fig!r} was not flagged; got {untraceable}"


def test_money_rounding_tolerance_allows_reasonable_rounding():
    """$3,125 in FACTS -> the model saying "$3,100" is a defensible round,
    not fabrication (matches the tolerance example in the design spec)."""
    allowed = excel_v2._allowed_numbers_from_facts_text("Blended Cost/Hire: $3,125\n")
    grounded, untraceable = excel_v2._narrative_is_grounded(
        "At roughly $3,100 per hire, this plan is efficient.", allowed
    )
    assert grounded is True, untraceable


def test_money_abbreviation_matches_base_value():
    """"$150K" citing a $150,000 FACTS value must be recognized as the SAME
    number, not a new one."""
    allowed = excel_v2._allowed_numbers_from_facts_text("Budget: $150,000\n")
    grounded, untraceable = excel_v2._narrative_is_grounded(
        "This $150K plan is aggressive for the timeline.", allowed
    )
    assert grounded is True, untraceable


def test_int_tolerance_is_tight_not_a_loophole():
    """Integers get only +/-1 tolerance -- a materially different hire count
    must still be caught."""
    allowed = excel_v2._allowed_numbers_from_facts_text("Projected Hires: 342\n")
    grounded_ok, _ = excel_v2._narrative_is_grounded(
        "The plan projects 342 hires.", allowed
    )
    assert grounded_ok is True

    grounded_bad, untraceable = excel_v2._narrative_is_grounded(
        "The plan projects 400 hires.", allowed
    )
    assert grounded_bad is False
    assert "400" in untraceable


def test_ordinary_prose_numbers_are_not_flagged():
    """Small integers used in ordinary sentence structure ("top 2 channels")
    must not be misread as invented statistics -- only large-integer,
    money, percent, and ratio tokens are checked."""
    allowed = excel_v2._allowed_numbers_from_facts_text("Budget: $150,000\n")
    grounded, untraceable = excel_v2._narrative_is_grounded(
        "The top 2 channels and 3 key risks are covered in the next "
        "5 sections, launching in the first 90 days.",
        allowed,
    )
    assert grounded is True, untraceable


def test_calendar_year_is_not_flagged_as_fabricated_statistic():
    allowed = excel_v2._allowed_numbers_from_facts_text("Budget: $150,000\n")
    grounded, untraceable = excel_v2._narrative_is_grounded(
        "This plan positions the client well for 2026 hiring demand.",
        allowed,
    )
    assert grounded is True, untraceable


def test_empty_narrative_is_trivially_grounded():
    allowed = excel_v2._allowed_numbers_from_facts_text(_SAMPLE_FACTS)
    grounded, untraceable = excel_v2._narrative_is_grounded("", allowed)
    assert grounded is True
    assert untraceable == []


# ---------------------------------------------------------------------------
# B. End-to-end: excel_v2.generate_excel_v2 with a mocked call_llm
# ---------------------------------------------------------------------------
def test_clean_grounded_narrative_is_used_verbatim():
    data = _minimal_data_with_allocation()
    narrative_text = (
        "This $150,000 plan is projected to deliver 342 hires against the "
        "stated goal of 400, roughly 86% of target, with a 58-hire gap that "
        "would need about $25,439 of additional budget to close. "
        "Programmatic (DSP) leads at $90,000 (60% of budget) driving 210 "
        "hires at a $25.00 CPA, alongside $60,000 in Niche / Industry "
        "Boards driving 132 hires at $30.00 CPA. Blended cost per hire is "
        "$438.60 across 5,600 projected applications. Recommended next "
        "step: launch both channels concurrently within the first two "
        "weeks."
    )

    def _fake_call_llm(**kwargs):
        return {
            "text": narrative_text,
            "provider": "deepseek",
            "model": "deepseek-v4-flash",
            "attempts": [],
        }

    with mock.patch("llm_router.call_llm", side_effect=_fake_call_llm):
        raw = excel_v2.generate_excel_v2(data)

    wb = openpyxl.load_workbook(io.BytesIO(raw))
    all_text = _all_text(wb)
    assert "EXECUTIVE STRATEGIC SUMMARY" in all_text.upper()
    assert narrative_text in all_text

    status = data.get("_narrative_status")
    assert status["generated"] is True
    assert status["status"] == "llm_grounded"


def test_fabricating_narrative_is_rejected_and_deterministic_fallback_renders():
    """The exact fabrication shapes from the bug report must be rejected,
    and the sheet must fall back to the deterministic, real-data-only
    template -- the fabricated text must not appear anywhere, and the
    rendered narrative sentence itself must not cite any of the invented
    figures (checked on the narrative text specifically, not a raw
    substring search across the whole workbook -- unrelated sections
    legitimately cite other numbers, e.g. a "6.22%" KB research stat, that
    would coincidentally substring-match a short fabricated figure like
    "22%")."""
    data = _minimal_data_with_allocation()
    fabricated_text = (
        "The industry average cost-per-hire for logistics roles is $5,000, "
        "and against that benchmark this plan achieves a 1:2.4 "
        "cost-to-value ratio, generating $360,000 in tangible value. "
        "Deploying at scale closes a 22% supply gap in this market via a "
        "38% reduction in time-to-fill, and every unfilled seat otherwise "
        "costs roughly $7,500 in lost revenue."
    )

    captured: dict = {}

    def _fake_call_llm(**kwargs):
        captured["prompt"] = kwargs["messages"][0]["content"]
        return {
            "text": fabricated_text,
            "provider": "deepseek",
            "model": "deepseek-v4-flash",
            "attempts": [],
        }

    with mock.patch("llm_router.call_llm", side_effect=_fake_call_llm):
        raw = excel_v2.generate_excel_v2(data)

    wb = openpyxl.load_workbook(io.BytesIO(raw))
    all_text = _all_text(wb)

    # The fabricated blob itself must be gone entirely.
    assert fabricated_text not in all_text

    # A real, deterministic summary must have rendered instead.
    assert "EXECUTIVE STRATEGIC SUMMARY" in all_text.upper()
    assert "This recruitment media plan for Acme Corp" in all_text

    _start = all_text.index("This recruitment media plan for Acme Corp")
    _narrative_only = all_text[_start : _start + 800].split("\n")[0]
    for fig in ("$5,000", "1:2.4", "$360,000", "22%", "38%", "$7,500"):
        assert fig not in _narrative_only, (
            f"fabricated figure {fig!r} leaked into the rendered narrative: "
            f"{_narrative_only!r}"
        )

    status = data.get("_narrative_status")
    assert status["generated"] is True
    assert status["status"] == "llm_rejected_fabrication"
    for fig in ("$5,000", "1:2.4", "$360,000", "22%", "38%", "$7,500"):
        assert fig in status["untraceable_figures"]

    # Every $/%%/ratio/large-int figure actually rendered must trace back to
    # a real number in the SAME FACTS block excel_v2 itself sent the (mocked)
    # LLM -- not a hand-maintained duplicate list.
    _prefix = "or invent any other number):\n"
    _suffix = "\n\nWrite as a senior recruitment strategist presenting to a VP of Talent Acquisition."
    prompt = captured["prompt"]
    facts_text = prompt[prompt.index(_prefix) + len(_prefix) : prompt.index(_suffix)]
    allowed = excel_v2._allowed_numbers_from_facts_text(facts_text)
    # Isolate just the rendered narrative sentence for a precise check.
    _start = all_text.index("This recruitment media plan for Acme Corp")
    _narrative_only = all_text[_start : _start + 800].split("\n")[0]
    grounded, untraceable = excel_v2._narrative_is_grounded(_narrative_only, allowed)
    assert grounded is True, f"deterministic fallback cited an untraceable figure: {untraceable}"


# ---------------------------------------------------------------------------
# C. Deterministic fallback renders with all-real numbers for both
#    reference briefs (tools_regen_bundles.MANPOWER_BRIEF / ATRIA_BRIEF).
# ---------------------------------------------------------------------------
def _fallback_narrative_and_allowed_numbers(brief: dict) -> tuple[str, dict]:
    """Build a reference brief through the SAME pipeline
    tools_regen_bundles.py / app.py use, force the deterministic fallback
    (via a mocked empty call_llm -- no network/API keys needed), and return
    the rendered narrative sentence plus the allowed-number set built from
    the EXACT FACTS block excel_v2 sent the (mocked) LLM -- so this checks
    production's own real numbers, not a hand-maintained duplicate list."""
    import tools_regen_bundles as trb
    from kb_loader import load_knowledge_base

    data = trb.build_plan_data(dict(brief))

    captured: dict = {}

    def _fake_call_llm(**kwargs):
        captured["prompt"] = kwargs["messages"][0]["content"]
        return {"text": "", "provider": "", "error": "no key (offline harness)", "attempts": []}

    with mock.patch("llm_router.call_llm", side_effect=_fake_call_llm):
        raw = excel_v2.generate_excel_v2(dict(data), load_kb_fn=load_knowledge_base)

    prompt = captured["prompt"]
    _prefix = "or invent any other number):\n"
    _suffix = "\n\nWrite as a senior recruitment strategist presenting to a VP of Talent Acquisition."
    facts_text = prompt[prompt.index(_prefix) + len(_prefix) : prompt.index(_suffix)]
    allowed = excel_v2._allowed_numbers_from_facts_text(facts_text)

    wb = openpyxl.load_workbook(io.BytesIO(raw))
    all_text = _all_text(wb)
    assert "EXECUTIVE STRATEGIC SUMMARY" in all_text.upper(), (
        "deterministic fallback did not render at all for this brief"
    )
    _marker = "This recruitment media plan for"
    _start = all_text.index(_marker)
    narrative = all_text[_start : _start + 800].split("\n")[0]
    return narrative, allowed


def test_deterministic_fallback_all_real_numbers_manpower_brief():
    import tools_regen_bundles as trb

    narrative, allowed = _fallback_narrative_and_allowed_numbers(trb.MANPOWER_BRIEF)
    grounded, untraceable = excel_v2._narrative_is_grounded(narrative, allowed)
    assert grounded is True, f"fallback cited an untraceable figure: {untraceable}\n{narrative}"


def test_deterministic_fallback_all_real_numbers_atria_brief():
    import tools_regen_bundles as trb

    narrative, allowed = _fallback_narrative_and_allowed_numbers(trb.ATRIA_BRIEF)
    grounded, untraceable = excel_v2._narrative_is_grounded(narrative, allowed)
    assert grounded is True, f"fallback cited an untraceable figure: {untraceable}\n{narrative}"


# ---------------------------------------------------------------------------
# D. Curated derived-number allowlist (hires/week, budget/month, per-channel
#    CPH/%%-share, ...) -- legitimate arithmetic on FACTS that a grounded LLM
#    narrative naturally states even though the derived figure itself is
#    never its own line in the FACTS block. `_curated_narrative_derivations`
#    is a fixed, named list computed from the SAME `ctx`
#    `_gather_narrative_grounding_context` builds -- NOT a general solver --
#    and `_build_narrative_allowed_numbers` folds it into the SAME
#    allowed-number set `_narrative_is_grounded` checks against.
# ---------------------------------------------------------------------------
def _narrative_ctx_for_brief(brief: dict):
    """Build the exact `ctx` `_build_sheet_executive_summary` would hand to
    `_build_narrative_facts_block` / `_curated_narrative_derivations` for a
    reference brief, via the SAME pipeline tools_regen_bundles.py / app.py
    use -- so this exercises production's own real numbers, not a
    hand-duplicated arithmetic model."""
    import tools_regen_bundles as trb
    from kb_loader import load_knowledge_base

    data = trb.build_plan_data(dict(brief))
    budget_alloc = data.get("_budget_allocation", {})
    channel_allocs = budget_alloc.get("channel_allocations", {})
    header_hires = sum(int(ch.get("projected_hires") or 0) for ch in channel_allocs.values())
    budget_num = excel_v2._get_budget_numeric(data)
    header_cph = round(budget_num / max(header_hires, 1), 2) if header_hires > 0 else 0
    duration = excel_v2._resolve_campaign_duration(data)
    return excel_v2._gather_narrative_grounding_context(
        data,
        client_name=data.get("client_name"),
        industry=data.get("industry"),
        industry_label=excel_v2._get_industry_label(data.get("industry")),
        budget_num=budget_num,
        duration=duration,
        locations=excel_v2._get_locations(data),
        roles=excel_v2._get_roles(data),
        hire_volume=data.get("hire_volume"),
        header_hires=header_hires,
        header_cph=header_cph,
        sufficiency=budget_alloc.get("sufficiency", {}),
        channel_allocs=channel_allocs,
        load_kb_fn=load_knowledge_base,
    )


def test_curated_derivations_include_hires_per_week_and_budget_per_month():
    """Direct check that the curated allowlist actually computes the two
    derivations the fix is named for, with sane values traceable to the
    MANPOWER_BRIEF's own real FACTS (budget $150,000, ~24-week duration, 48
    projected hires -- see the KB-driven `calculate_budget_allocation` log
    line for this exact brief)."""
    import tools_regen_bundles as trb

    ctx = _narrative_ctx_for_brief(trb.MANPOWER_BRIEF)
    derivations = excel_v2._curated_narrative_derivations(ctx)

    assert "hires_per_week" in derivations
    category, value = derivations["hires_per_week"]
    assert category == "int"
    assert value == pytest.approx(ctx["header_hires"] / ctx["duration_weeks"], rel=1e-6)

    assert "budget_per_month" in derivations
    category, value = derivations["budget_per_month"]
    assert category == "money"
    assert value == pytest.approx(ctx["budget_num"] / ctx["duration_months"], rel=1e-6)


def _narrative_ctx_for_data(data: dict, load_kb_fn=None):
    """Same as `_narrative_ctx_for_brief`, but for an already-built plan
    `data` dict (e.g. `_minimal_data_with_allocation()`) rather than a raw
    brief run through the full tools_regen_bundles/budget_engine pipeline.
    `load_kb_fn` defaults to None (no real KB) to match how the existing
    section B end-to-end tests call `excel_v2.generate_excel_v2(data)` --
    with a REAL KB loaded, a synthetic fixture's own numbers can
    coincidentally fall within a real KB benchmark's rounding tolerance of
    an unrelated fabricated figure, which is a property of that specific
    KB row, not of the grounding logic under test here."""
    budget_alloc = data.get("_budget_allocation", {})
    channel_allocs = budget_alloc.get("channel_allocations", {})
    header_hires = sum(int(ch.get("projected_hires") or 0) for ch in channel_allocs.values())
    budget_num = excel_v2._get_budget_numeric(data)
    header_cph = round(budget_num / max(header_hires, 1), 2) if header_hires > 0 else 0
    duration = excel_v2._resolve_campaign_duration(data)
    return excel_v2._gather_narrative_grounding_context(
        data,
        client_name=data.get("client_name", "Client"),
        industry=data.get("industry", "general_entry_level"),
        industry_label=excel_v2._get_industry_label(data.get("industry", "general_entry_level")),
        budget_num=budget_num,
        duration=duration,
        locations=excel_v2._get_locations(data),
        roles=excel_v2._get_roles(data),
        hire_volume=data.get("hire_volume"),
        header_hires=header_hires,
        header_cph=header_cph,
        sufficiency=budget_alloc.get("sufficiency", {}),
        channel_allocs=channel_allocs,
        load_kb_fn=load_kb_fn,
    )


def test_fabricated_prod_text_still_rejected_with_enriched_allowed_numbers():
    """The exact prod fabrication shapes must STILL be rejected once the
    allowed-number set includes both the enriched FACTS block AND the
    curated per-week/per-month/per-channel derivations -- the enrichment
    widens what's ALLOWED, it must never widen what's traceable to an
    unrelated invented figure."""
    data = _minimal_data_with_allocation()
    ctx = _narrative_ctx_for_data(data)
    facts_text = excel_v2._build_narrative_facts_block(ctx)
    allowed = excel_v2._build_narrative_allowed_numbers(ctx, facts_text)

    # Sanity: the curated derivations actually widened the allowed set
    # beyond FACTS-text-only, so this test is exercising the enrichment,
    # not just re-testing the pre-existing FACTS-only validator.
    facts_only_allowed = excel_v2._allowed_numbers_from_facts_text(facts_text)
    assert allowed["money"] > facts_only_allowed["money"] or allowed["int"] > facts_only_allowed["int"]

    fabricated_text = (
        "The industry average cost-per-hire for logistics roles is $5,000, "
        "and against that benchmark this plan achieves a 1:2.4 "
        "cost-to-value ratio, generating $360,000 in tangible value. "
        "Deploying at scale closes a 22% supply gap in this market via a "
        "38% reduction in time-to-fill, and every unfilled seat otherwise "
        "costs roughly $7,500 in lost revenue."
    )
    grounded, untraceable = excel_v2._narrative_is_grounded(fabricated_text, allowed)
    assert grounded is False
    for fig in ("$5,000", "1:2.4", "$360,000", "22%", "38%", "$7,500"):
        assert fig in untraceable, f"{fig!r} was not flagged; got {untraceable}"


def test_grounded_narrative_with_perweek_permonth_derivations_passes():
    """A realistic executive-summary narrative that cites FACTS numbers
    verbatim PLUS legitimate per-week/per-month rates derived from them
    (budget/month, hires/month, hires/week, applications/week -- none of
    which appear as their own FACTS line) must be accepted as grounded end
    to end, not forced into the deterministic fallback. Before the curated
    allowlist, every one of the derived figures below would have been
    rejected as an untraceable fabrication."""
    data = _minimal_data_with_allocation()
    # $150,000 / 6 months / 342 hires / 5,600 applications (see
    # _minimal_data_with_allocation) -> budget/month = $25,000 (exact),
    # hires/month = 57 (exact), hires/week (26 weeks) ~= 13.15, and
    # applications/week ~= 215.4 -- none of these four figures are their
    # own FACTS line, only derivable from FACTS budget/duration/hires/apps.
    narrative_text = (
        "This $150,000 plan runs over 6 months, pacing at roughly $25,000 "
        "a month and delivering about 57 hires a month -- close to 13 "
        "hires a week off a base of roughly 215 weekly applications. "
        "Programmatic (DSP) leads the channel mix at $90,000 (60% of "
        "budget), complemented by $60,000 in Niche / Industry Boards. "
        "Against the stated goal of 400, this pace would need to "
        "accelerate. Recommended next step: reassess channel pacing at "
        "the 8-week mark."
    )

    def _fake_call_llm(**kwargs):
        return {
            "text": narrative_text,
            "provider": "deepseek",
            "model": "deepseek-v4-flash",
            "attempts": [],
        }

    with mock.patch("llm_router.call_llm", side_effect=_fake_call_llm):
        raw = excel_v2.generate_excel_v2(data)

    wb = openpyxl.load_workbook(io.BytesIO(raw))
    all_text = _all_text(wb)
    assert "EXECUTIVE STRATEGIC SUMMARY" in all_text.upper()
    assert narrative_text in all_text, "grounded narrative with derivations was rejected"

    status = data.get("_narrative_status")
    assert status["status"] == "llm_grounded", (
        f"expected llm_grounded, got {status.get('status')}: "
        f"{status.get('untraceable_figures')}"
    )
    assert status["generated"] is True


if __name__ == "__main__":
    _failures = 0
    for _name, _fn in sorted(globals().items()):
        if _name.startswith("test_") and callable(_fn):
            try:
                _fn()
                print(f"PASS {_name}")
            except AssertionError as exc:
                _failures += 1
                print(f"FAIL {_name}: {exc}")
            except Exception as exc:  # noqa: BLE001
                _failures += 1
                print(f"ERROR {_name}: {exc}")
    if _failures:
        sys.exit(1)
