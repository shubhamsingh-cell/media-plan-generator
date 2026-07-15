"""S95 follow-up: `X or ""[:N]` operator-precedence no-ops + hard mid-word
character slices, verified by targeted repro before fixing.

Python precedence: slicing binds tighter than `or`, so
``lc.get("context_note") or ""[:80]`` slices the empty-string LITERAL and
the intended cap is silently a no-op. Audited sites fixed here:

  1. excel_v2 Market Intelligence "Location Economic Context" Context cell
     (`context_note or ""[:80]`) -- the 80-char cap never applied; a long
     note rendered in full. Now `_truncate_at_word_boundary(..., 80)`.
  2. excel_v2 Recommended Channels rationale (`notes[:60]`) -- a genuine
     hard cut that landed mid-word, contradicting the S5 doctrine comment
     directly below it ("never hard-truncate the rationale mid-word --
     write it in full"). Now appends the notes in full (the cell already
     wraps via `_ALIGN_WRAP`).
  3. excel_v2 Location Intelligence "Key Industries" cell
     (`industry_str[:80]`) -- correctly guarded but hard-cut mid-word.
     Now `_truncate_at_word_boundary(..., 80)`.
  4. data_synthesizer.fuse_competitive_intelligence company summary
     (`wiki_data.get("summary") or ""[:500]`) -- the 500 cap was a no-op,
     leaving the field unbounded. Fixed with a LOCAL word-boundary helper
     (not a hard slice: the excel renderer's own 500-char word-boundary cap
     no-ops on text already <= 500, so a mid-word data-layer cut at exactly
     500 would render verbatim -- the exact prod-defect class S95 fixed).

A source-level guard test keeps the precedence pattern from reappearing.
The repo-wide sweep is DONE: the guard now covers every audited module
(see the tuple in test_slice_precedence_pattern_absent_from_audited_files);
behavioral coverage for the sweep's sites lives in the sibling
tests/test_truncation_precedence_{nova,skill_target,market_intel,misc}.py.

Runs under pytest, or standalone:
``python3 tests/test_truncation_precedence_fixes.py``.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl  # noqa: E402

import data_synthesizer  # noqa: E402
import excel_v2  # noqa: E402


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
def _sheet_strings(ws):
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if isinstance(val, str):
                yield val


def _assert_word_boundary_cut(cell: str, source: str, cap: int):
    """The cell must be a clean word-boundary prefix of `source`, ending in a
    single ellipsis, never longer than cap + the ellipsis glyph."""
    assert len(cell) <= cap + 1
    assert cell.endswith("…")
    assert ".." not in cell
    body = cell[:-1].strip()
    assert source.startswith(body), f"not a prefix: {body!r}"
    # The next source character after the kept body must not be mid-word.
    assert not source[len(body)].isalnum(), f"mid-word cut: ...{body[-15:]!r}"


class _ResearchStub:
    """Feeds only the labour-market payload `_build_sheet_market_intelligence`
    asks for; everything else in that builder comes from `data`."""

    def __init__(self, context_note: str):
        self._note = context_note

    def get_labour_market_intelligence(self, industry, locations):
        return {
            "location_contexts": [
                {
                    "location": "Dallas, TX",
                    "country": "United States",
                    "unemployment_rate": "3.9%",
                    "median_salary": "$52,000",
                    "context_note": self._note,
                }
            ]
        }

    def get_location_info(self, loc):
        return {}


_LONG_NOTE = (
    "Dallas-Fort Worth remains one of the tightest logistics labour markets "
    "in the country, with warehouse and last-mile operators competing for "
    "the same CDL-licensed driver pool across the metroplex."
)  # 195 chars -- well past the 80-char cap

_LONG_INDUSTRIES = (
    "Transportation & Warehousing, Healthcare & Social Assistance, "
    "Professional & Business Services, Financial Activities, Manufacturing"
)  # 131 chars -- past the 80-char cap; the old hard cut landed inside
#     "Business" ("...Professional & Bus")

_LONG_DESC = (
    "Specialised job board for commercial drivers with verified CDL "
    "credential screening and regional route-preference matching built in."
)  # 133 chars -- the old notes[:60] cut this to "...with verified C"


def _market_intel_data(**overrides) -> dict:
    data = {
        "client_name": "Acme Corp",
        "company_name": "Acme Corp",
        "industry": "logistics_supply_chain",
        "budget": "$150,000",
        "locations": ["Dallas, TX"],
        "roles": ["CDL Driver"],
        "target_roles": ["CDL Driver"],
        "campaign_duration": "6 months",
        "_enriched": {},
        "_synthesized": {
            "location_profiles": {"Dallas, TX": {"key_industries": _LONG_INDUSTRIES}}
        },
        "_budget_allocation": {},
    }
    data.update(overrides)
    return data


def _build_market_intel_sheet(data, research_mod):
    wb = openpyxl.Workbook()
    ws = wb.active
    excel_v2._build_sheet_market_intelligence(ws, data, research_mod=research_mod)
    return ws


# ===========================================================================
# Guard: the precedence pattern must not reappear in the audited files
# ===========================================================================
def test_slice_precedence_pattern_absent_from_audited_files():
    """`or ""[...]` / `or ''[...]` / `or [][...]` always slice the empty
    LITERAL -- any occurrence is a bug by construction. The repo-wide sweep
    is done: guard every audited module."""
    pattern = re.compile(r"""or\s+(""|''|\[\])\s*\[""")
    for fname in (
        "excel_v2.py",
        "data_synthesizer.py",
        "nova.py",
        "skill_target.py",
        "competitive_intel.py",
        "market_pulse.py",
        "market_intel_reports.py",
        "research.py",
        "api_portal.py",
        "api_enrichment.py",
        "llm_router.py",
        "quick_plan.py",
        "social_plan.py",
        "ppt_generator.py",
        "archive/excel_legacy.py",
    ):
        src = (PROJECT_ROOT / fname).read_text(encoding="utf-8")
        hits = [
            f"{fname}:{i}"
            for i, line in enumerate(src.splitlines(), 1)
            if pattern.search(line)
        ]
        assert not hits, f"slice-binds-to-literal precedence bug at: {hits}"


# ===========================================================================
# Site 1: Location Economic Context "Context" cell (was a no-op cap)
# ===========================================================================
def test_location_context_note_capped_at_word_boundary():
    ws = _build_market_intel_sheet(_market_intel_data(), _ResearchStub(_LONG_NOTE))
    hits = [v for v in _sheet_strings(ws) if "tightest logistics" in v]
    assert hits, "Location Economic Context row did not render"
    assert hits[0] != _LONG_NOTE, "the 80-char cap is still a no-op"
    _assert_word_boundary_cut(hits[0], _LONG_NOTE, 80)


def test_location_context_note_short_note_unchanged():
    short = "Tight driver market."
    ws = _build_market_intel_sheet(_market_intel_data(), _ResearchStub(short))
    assert short in list(_sheet_strings(ws))


# ===========================================================================
# Site 2: Recommended Channels rationale (was a mid-word notes[:60] cut)
# ===========================================================================
def test_vetted_channel_rationale_carries_full_notes():
    data = {
        "client_name": "Acme Corp",
        "industry": "logistics_supply_chain",
        "budget": "$150,000",
        "locations": ["Dallas, TX"],
        "roles": ["CDL Driver"],
        "target_roles": ["CDL Driver"],
        "_enriched": {},
        "_synthesized": {},
        "_budget_allocation": {
            "channel_allocations": {
                "programmatic_dsp": {
                    "dollar_amount": 90000,
                    "percentage": 60,
                    "cpc": 2.0,
                }
            }
        },
        "_channels_db": {
            "niche": {"DriverBoard Pro": {"description": _LONG_DESC, "cpc": 2.5}}
        },
    }
    wb = openpyxl.Workbook()
    ws = wb.active
    excel_v2._build_sheet_channels(ws, data)

    hits = [v for v in _sheet_strings(ws) if "Specialised job board" in v]
    assert hits, "vetted-channel rationale cell did not render"
    rationale = hits[0]
    assert _LONG_DESC in rationale, (
        "rationale must carry the notes IN FULL (S5 doctrine: never "
        f"hard-truncate the rationale) -- got: {rationale!r}"
    )
    # The old defect: exactly the first 60 chars, cut mid-word ("...verified C").
    assert _LONG_DESC[:60] + ";" not in rationale
    assert not rationale.endswith(_LONG_DESC[:60])


# ===========================================================================
# Site 3: Location Intelligence "Key Industries" cell (was a hard 80 cut)
# ===========================================================================
def test_location_key_industries_capped_at_word_boundary():
    ws = _build_market_intel_sheet(
        _market_intel_data(), _ResearchStub("Tight driver market.")
    )
    hits = [v for v in _sheet_strings(ws) if "Transportation & Warehousing" in v]
    assert hits, "Key Industries cell did not render"
    assert hits[0] != _LONG_INDUSTRIES[:80], "still a hard mid-word 80-char slice"
    _assert_word_boundary_cut(hits[0], _LONG_INDUSTRIES, 80)


def test_location_key_industries_short_value_unchanged():
    short = "Logistics, Healthcare"
    data = _market_intel_data(
        _synthesized={"location_profiles": {"Dallas, TX": {"key_industries": short}}}
    )
    ws = _build_market_intel_sheet(data, _ResearchStub("Tight driver market."))
    assert short in list(_sheet_strings(ws))


# ===========================================================================
# Site 4: data_synthesizer company summary (was unbounded via the no-op)
# ===========================================================================
def _fuse(summary):
    wiki = {
        "description": "Atria Senior Living is a senior living operator.",
        "url": "https://en.wikipedia.org/wiki/Atria",
    }
    if summary is not None:
        wiki["summary"] = summary
    return data_synthesizer.fuse_competitive_intelligence(
        {"company_info": wiki},
        {},
        {"company_name": "Atria Senior Living", "industry": "healthcare_medical"},
    )


def test_company_summary_capped_at_word_boundary_in_synthesizer():
    long_summary = (
        "Atria Senior Living operates senior communities across the country. "
        * 20
    ).strip()  # ~1.4k chars
    prof = _fuse(long_summary).get("company_profile", {})
    out = prof.get("summary", "")
    assert out != long_summary, "the 500-char cap is still a no-op"
    _assert_word_boundary_cut(out, long_summary, 500)


def test_company_summary_short_and_missing_variants():
    short = "Atria Senior Living operates senior communities."
    assert _fuse(short).get("company_profile", {}).get("summary") == short
    assert _fuse(None).get("company_profile", {}).get("summary") == ""


def test_sec_recent_filings_capped_at_five():
    """`_sec.get("recent_filings") or [][:5]` sliced the empty-list literal --
    the filings list flowed through unbounded (repro: 12 in, 12 out)."""
    result = data_synthesizer.fuse_competitive_intelligence(
        {
            "company_info": {
                "description": "Atria Senior Living is a senior living operator."
            },
            "sec_data": {
                "cik": "0000001",
                "recent_filings": [f"10-K {i}" for i in range(12)],
            },
        },
        {},
        {"company_name": "Atria Senior Living", "industry": "healthcare_medical"},
    )
    filings = result["company_sec"]["filings"]
    assert len(filings) == 5
    assert filings == [f"10-K {i}" for i in range(5)]


def test_narrative_context_salary_sources_capped_at_three():
    """`", ".join(sal.get("sources") or [][:3])` joined ALL sources into the
    LLM-prompt context -- the 3-source token-budget cap was a no-op."""
    text = data_synthesizer._build_narrative_context(
        {
            "salary_intelligence": {
                "cdl_driver": {
                    "median": 62000,
                    "p25": 52000,
                    "p75": 74000,
                    "sources": ["BLS", "DataUSA", "Indeed", "Glassdoor", "Payscale"],
                }
            }
        },
        {"client_name": "Acme Corp"},
    )
    assert "BLS, DataUSA, Indeed" in text
    assert "Glassdoor" not in text
    assert "Payscale" not in text


def test_data_synthesizer_truncate_helper_edge_cases():
    t = data_synthesizer._truncate_at_word_boundary
    assert t("short text", 500) == "short text"
    assert t("", 80) == ""
    assert t(None, 80) == ""
    assert t("anything", 0) == ""
    # Single overlong word: hard-cut fallback rather than "".
    out = t("x" * 200, 50)
    assert out.endswith("…") and len(out) == 51


if __name__ == "__main__":
    _failures = 0
    for _name, _fn in sorted(globals().items()):
        if _name.startswith("test_") and callable(_fn):
            try:
                _fn()
                print(f"PASS {_name}")
            except AssertionError as exc:
                _failures += 1
                print(f"FAIL {_name}: {exc}")
            except Exception as exc:  # noqa: BLE001
                _failures += 1
                print(f"ERROR {_name}: {exc}")
    if _failures:
        sys.exit(1)
