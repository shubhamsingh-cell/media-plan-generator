"""S95: three excel_v2 narrative/text defects verified on a real prod Atria
bundle.

  1. "over Ongoing" grammar in the deterministic executive-summary narrative
     (`_build_deterministic_executive_summary`) -- an unbounded/"Ongoing"
     campaign duration produced "...with a $25,000 budget and over Ongoing
     targeting the Healthcare & Medical sector." (no verb, no comma). Fixed
     via `_is_unbounded_duration` + restructured sentence-building in both
     the deterministic template AND the FACTS block fed to the LLM
     (`_build_narrative_facts_block`), so the model is never told "Duration:
     Ongoing" bare either.

  2. Market Intelligence "Company Profile" description cell hard-truncated
     mid-word (e.g. "...manages ... communities in more than 200 l"). Fixed
     via a new `_truncate_at_word_boundary` helper applied to the
     "Description"/"Summary" company-profile fields.

  3. The narrative call was starved by llm_router's global LLM concurrency
     limiter under contention -- "Server busy -- too many concurrent LLM
     requests. Please retry shortly." with attempts=[] rejected the call
     before any provider ran, always losing to the deterministic fallback.
     Fixed via a bounded (<=2 retries, short backoff, <~8s total added
     wait) busy-retry loop scoped ONLY to that specific concurrency signal
     -- genuine provider errors and grounding rejections keep their
     existing single-shot paths.

Mocking patterns reused from tests/test_excel_narrative_grounding.py
(``_minimal_data_with_allocation``-style fixture, mocked ``llm_router.call_llm``)
and tests/test_excel_narrative_reliability.py (``_build``/``_sheet_text`` helpers).

Runs under pytest, or standalone: ``python3 tests/test_excel_v2_s95_bundle_quality.py``.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path
from unittest import mock

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl  # noqa: E402

import excel_v2  # noqa: E402


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
def _sheet_text(ws) -> str:
    parts = []
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if val is not None:
                parts.append(str(val))
    return "\n".join(parts)


def _all_text(wb) -> str:
    return "\n".join(_sheet_text(ws) for ws in wb.worksheets)


def _minimal_data_with_allocation(**overrides) -> dict:
    """Same shape as test_excel_narrative_grounding.py's fixture of the
    same name: a populated ``_budget_allocation`` so the FACTS block (and
    therefore grounding) has real substance."""
    data = {
        "client_name": "Acme Corp",
        "company_name": "Acme Corp",
        "industry": "logistics_supply_chain",
        "budget": "$150,000",
        "locations": ["Dallas, TX"],
        "roles": ["CDL Driver"],
        "target_roles": ["CDL Driver"],
        "campaign_duration": "6 months",
        "hire_volume": "400",
        "work_environment": "onsite",
        "_enriched": {},
        "_synthesized": {},
        "_budget_allocation": {
            "sufficiency": {"grade": "B"},
            "channel_allocations": {
                "programmatic_dsp": {
                    "dollar_amount": 90000,
                    "percentage": 60,
                    "projected_clicks": 45000,
                    "projected_applications": 3600,
                    "projected_hires": 210,
                    "cpc": 2.0,
                    "cpa": 25.0,
                    "roi_score": 8,
                },
                "niche_boards": {
                    "dollar_amount": 60000,
                    "percentage": 40,
                    "projected_clicks": 20000,
                    "projected_applications": 2000,
                    "projected_hires": 132,
                    "cpc": 3.0,
                    "cpa": 30.0,
                    "roi_score": 7,
                },
            },
        },
    }
    data.update(overrides)
    return data


# A grounded narrative for _minimal_data_with_allocation() (identical to the
# one proven-grounded in test_excel_narrative_grounding.py's
# test_clean_grounded_narrative_is_used_verbatim) -- reused here so tests
# that need a call_llm response that WILL pass `_narrative_is_grounded`
# don't have to re-derive one.
_CLEAN_GROUNDED_TEXT = (
    "This $150,000 plan is projected to deliver 342 hires against the "
    "stated goal of 400, roughly 86% of target, with a 58-hire gap that "
    "would need about $25,439 of additional budget to close. "
    "Programmatic (DSP) leads at $90,000 (60% of budget) driving 210 "
    "hires at a $25.00 CPA, alongside $60,000 in Niche / Industry "
    "Boards driving 132 hires at $30.00 CPA. Blended cost per hire is "
    "$438.60 across 5,600 projected applications. Recommended next "
    "step: launch both channels concurrently within the first two "
    "weeks."
)

_BUSY_ERROR = "Server busy -- too many concurrent LLM requests. Please retry shortly."


# ===========================================================================
# Defect 1: "over Ongoing" grammar
# ===========================================================================
def test_deterministic_narrative_ongoing_duration_reads_naturally():
    # `_fmt_currency`'s "$" default is the ACTIVE plan currency, tracked
    # per-thread and normally reset by `generate_excel_v2` -> `_set_active_
    # currency`. This test calls `_build_deterministic_executive_summary`
    # directly (bypassing that reset), so pin it to USD explicitly --
    # otherwise it inherits whatever currency an earlier test in the SAME
    # pytest process/thread last left active (e.g. GBP from a UK-plan test).
    excel_v2._set_active_currency(None)
    ctx = {
        "client_name": "Atria Senior Living",
        "budget_num": 25000,
        "duration": "Ongoing",
        "industry_label": "Healthcare & Medical",
    }
    text = excel_v2._build_deterministic_executive_summary(ctx)
    assert "on an ongoing basis" in text
    assert "over Ongoing" not in text
    # Grammatical: a comma precedes "targeting" so the sentence reads as one
    # flowing clause, not a dangling fragment.
    assert ", targeting the Healthcare & Medical sector." in text
    assert text == (
        "This recruitment media plan for Atria Senior Living with a "
        "$25,000 budget on an ongoing basis, targeting the Healthcare & "
        "Medical sector."
    )


def test_deterministic_narrative_ongoing_lowercase_and_empty_variants():
    """Case-insensitive + the empty/"not specified" family all route through
    the SAME unbounded-duration branch (`_is_unbounded_duration`)."""
    excel_v2._set_active_currency(None)
    base_ctx = {
        "client_name": "Atria Senior Living",
        "budget_num": 25000,
        "industry_label": "Healthcare & Medical",
    }
    for dur in ("ongoing", "Ongoing", "ONGOING", "  ongoing  "):
        ctx = dict(base_ctx, duration=dur)
        text = excel_v2._build_deterministic_executive_summary(ctx)
        assert "on an ongoing basis" in text, f"failed for duration={dur!r}"
        assert "over " + dur.strip() not in text


def test_deterministic_narrative_normal_duration_reads_naturally():
    excel_v2._set_active_currency(None)
    ctx = {
        "client_name": "Amerigas Test",
        "budget_num": 150000,
        "duration": "6 months",
        "industry_label": "Logistics & Supply Chain",
    }
    text = excel_v2._build_deterministic_executive_summary(ctx)
    assert "over 6 months," in text
    assert "on an ongoing basis" not in text
    assert text == (
        "This recruitment media plan for Amerigas Test with a $150,000 "
        "budget over 6 months, targeting the Logistics & Supply Chain "
        "sector."
    )


def test_facts_block_formats_ongoing_duration_sensibly_for_llm_prompt():
    """The FACTS block handed to the LLM must never read "Duration:
    Ongoing" bare -- give the model enough framing that it doesn't echo
    back the same "over Ongoing" phrasing the deterministic path used to."""
    ctx = {"duration": "Ongoing"}
    facts = excel_v2._build_narrative_facts_block(ctx)
    assert "Duration: Ongoing (no fixed end date)" in facts
    assert "over Ongoing" not in facts

    ctx_normal = {"duration": "6 months"}
    facts_normal = excel_v2._build_narrative_facts_block(ctx_normal)
    assert "Duration: 6 months" in facts_normal


def test_is_unbounded_duration_helper():
    assert excel_v2._is_unbounded_duration("Ongoing") is True
    assert excel_v2._is_unbounded_duration("ongoing") is True
    assert excel_v2._is_unbounded_duration("") is True
    assert excel_v2._is_unbounded_duration(None) is True
    assert excel_v2._is_unbounded_duration("6 months") is False
    assert excel_v2._is_unbounded_duration("13 weeks") is False


# ===========================================================================
# Defect 2: mid-word company-description truncation
# ===========================================================================
def test_truncate_at_word_boundary_never_cuts_mid_word():
    long_text = (
        "Atria Senior Living is a private company that manages senior "
        "living communities in more than 200 locations across 30 states "
        "and Canada, offering independent living, assisted living, and "
        "memory care services to thousands of residents nationwide."
    )
    out = excel_v2._truncate_at_word_boundary(long_text, 100)
    assert len(out) <= 101  # 100 chars + the single ellipsis glyph
    assert out.endswith("…")
    assert not out.endswith("..")
    assert ".." not in out
    # The cut must land on a real word boundary -- every word in the
    # truncated body (minus the ellipsis) must be a COMPLETE prefix word
    # from the original text, never a fragment like "20" from "200".
    body = out[:-1].strip()
    assert long_text.startswith(body)
    assert (
        not long_text[len(body) : len(body) + 1].isalnum()
        or long_text[len(body)] == " "
    )


def test_truncate_at_word_boundary_no_op_when_already_short():
    short = "Short company description."
    assert excel_v2._truncate_at_word_boundary(short, 500) == short


def test_truncate_at_word_boundary_single_overlong_word_falls_back_to_hard_cut():
    # No whitespace to back off to at all -- must not crash or return "".
    out = excel_v2._truncate_at_word_boundary("x" * 200, 50)
    assert out.endswith("…")
    assert len(out) == 51


def test_company_description_truncated_at_word_boundary_end_to_end():
    """The Market Intelligence "Company Profile" Description cell (the
    actual D-column cell the prod defect was found in) must never render a
    mid-word cut for a long Wikipedia-sourced description."""
    long_desc = (
        "Atria Senior Living is a private company that manages senior "
        "living communities in more than 200 locations across 30 states "
        "and Canada, offering independent living, assisted living, and "
        "memory care services to thousands of residents nationwide, with "
        "a mission centered on compassionate, high-quality senior care "
        "delivered by dedicated staff across every community it operates."
    ) * 2  # long enough to force truncation at the 500-char cap
    data = {
        "client_name": "Atria Senior Living",
        "company_name": "Atria Senior Living",
        "industry": "healthcare_medical",
        "budget": "$300,000",
        "locations": ["New York, NY"],
        "roles": ["Nurse"],
        "target_roles": ["Nurse"],
        "campaign_duration": "18 months",
        "hire_volume": "500",
        "work_environment": "onsite",
        "_enriched": {},
        "_synthesized": {
            "competitive_intelligence": {"company_profile": {"description": long_desc}}
        },
        "_budget_allocation": {},
    }

    def _fake_call_llm(**kwargs):
        return {
            "text": "",
            "provider": "",
            "error": "no key (test harness)",
            "attempts": [],
        }

    with mock.patch("llm_router.call_llm", side_effect=_fake_call_llm):
        raw = excel_v2.generate_excel_v2(data)

    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb["Market Intelligence"]

    found = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "manages senior" in cell.value:
                found = cell.value
    assert found is not None, "Company Profile description cell did not render"
    assert found != long_desc, "description was not truncated at all"
    assert len(found) <= 501
    assert found.endswith("…")
    assert not found.endswith("..")
    assert ".." not in found
    # No mid-word cut: the truncated body must be a clean prefix ending on
    # a word boundary of the original text.
    body = found[:-1].strip()
    assert long_desc.startswith(body)
    assert long_desc[len(body)] in (" ",)


# ===========================================================================
# Defect 3: LLM-concurrency-limiter busy retry
# ===========================================================================
def test_busy_retry_recovers_on_third_call():
    """Attempt 1 and 2 are rejected by the concurrency limiter (empty text,
    "Server busy..." error, attempts=[]); attempt 3 succeeds with a clean
    grounded narrative -- the narrative must land (llm_grounded),
    retried_on_busy must be recorded, and time.sleep must have been called
    (patched so the test doesn't actually wait)."""
    data = _minimal_data_with_allocation()
    calls: list = []

    def _fake_call_llm(**kwargs):
        calls.append(kwargs)
        if len(calls) < 3:
            return {
                "text": "",
                "provider": "",
                "attempts": [],
                "error": _BUSY_ERROR,
            }
        return {
            "text": _CLEAN_GROUNDED_TEXT,
            "provider": "deepseek",
            "model": "deepseek-v4-flash",
            "attempts": [{"provider": "deepseek", "status": "success"}],
        }

    with mock.patch("llm_router.call_llm", side_effect=_fake_call_llm), mock.patch(
        "excel_v2.time.sleep"
    ) as mock_sleep:
        raw = excel_v2.generate_excel_v2(data)

    assert len(calls) == 3, "expected the initial call plus exactly 2 busy retries"
    assert mock_sleep.called, "busy retry must back off via time.sleep"

    wb = openpyxl.load_workbook(io.BytesIO(raw))
    all_text = _all_text(wb)
    assert "EXECUTIVE STRATEGIC SUMMARY" in all_text.upper()
    assert _CLEAN_GROUNDED_TEXT in all_text

    status = data.get("_narrative_status")
    assert status["status"] == "llm_grounded"
    assert status["generated"] is True
    assert status.get("retried_on_busy") is True


def test_busy_retry_exhausted_falls_back_to_deterministic():
    """All 3 attempts (initial + 2 retries) hit the concurrency limiter --
    must fall back to the deterministic template, status fallback_template,
    reason mentions concurrency, and generation must not crash."""
    data = _minimal_data_with_allocation()
    calls: list = []

    def _fake_call_llm(**kwargs):
        calls.append(kwargs)
        return {"text": "", "provider": "", "attempts": [], "error": _BUSY_ERROR}

    with mock.patch("llm_router.call_llm", side_effect=_fake_call_llm), mock.patch(
        "excel_v2.time.sleep"
    ) as mock_sleep:
        raw = excel_v2.generate_excel_v2(data)  # must not raise

    assert len(calls) == 3, "expected the initial call plus exactly 2 busy retries"
    assert mock_sleep.called

    wb = openpyxl.load_workbook(io.BytesIO(raw))
    all_text = _all_text(wb)
    assert "EXECUTIVE STRATEGIC SUMMARY" in all_text.upper()
    assert "This recruitment media plan for Acme Corp" in all_text

    status = data.get("_narrative_status")
    assert status["generated"] is True
    assert status["status"] == "fallback_template"
    assert "concurrency" in status["reason"].lower()


def test_non_busy_empty_response_does_not_trigger_busy_retry():
    """Regression guard: a genuine provider failure (NOT the concurrency
    signal) must keep its existing single-shot behavior -- no extra calls,
    no sleep."""
    data = _minimal_data_with_allocation()
    calls: list = []

    def _fake_call_llm(**kwargs):
        calls.append(kwargs)
        return {
            "text": "",
            "provider": "",
            "attempts": [
                {"provider": "deepseek", "status": "failed", "error": "timeout"}
            ],
            "error": "All LLM providers unavailable or failed",
        }

    with mock.patch("llm_router.call_llm", side_effect=_fake_call_llm), mock.patch(
        "excel_v2.time.sleep"
    ) as mock_sleep:
        raw = excel_v2.generate_excel_v2(data)

    assert len(calls) == 1, "a genuine provider error must not trigger busy retries"
    assert not mock_sleep.called

    status = data.get("_narrative_status")
    assert status["status"] == "fallback_template"
    assert status.get("retried_on_busy") is not True
    assert status["reason"] == "All LLM providers unavailable or failed"

    wb = openpyxl.load_workbook(io.BytesIO(raw))
    assert "EXECUTIVE STRATEGIC SUMMARY" in _all_text(wb).upper()


def test_is_llm_concurrency_busy_error_matcher():
    assert excel_v2._is_llm_concurrency_busy_error(_BUSY_ERROR) is True
    assert (
        excel_v2._is_llm_concurrency_busy_error("too many concurrent requests") is True
    )
    assert excel_v2._is_llm_concurrency_busy_error("concurrency limit reached") is True
    assert (
        excel_v2._is_llm_concurrency_busy_error(
            "All LLM providers unavailable or failed"
        )
        is False
    )
    assert excel_v2._is_llm_concurrency_busy_error("") is False
    assert excel_v2._is_llm_concurrency_busy_error(None) is False


if __name__ == "__main__":
    _failures = 0
    for _name, _fn in sorted(globals().items()):
        if _name.startswith("test_") and callable(_fn):
            try:
                _fn()
                print(f"PASS {_name}")
            except AssertionError as exc:
                _failures += 1
                print(f"FAIL {_name}: {exc}")
            except Exception as exc:  # noqa: BLE001
                _failures += 1
                print(f"ERROR {_name}: {exc}")
    if _failures:
        sys.exit(1)
