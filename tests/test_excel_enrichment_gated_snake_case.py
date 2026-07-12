"""S94: regression tests for the 10 bundle_qa criticals found on a REAL
prod-generated bundle (live enrichment + full env) that never reproduce
locally, because the culprit sections only render when enrichment/KB data
is actually present:

  1. Market Intelligence!B64  -- "Industry Sector: logistics_supply_chain"
     (competitive_intelligence.market_positioning.industry_sector rendered
     as its raw internal key instead of the client-facing label).
  2. Market Intelligence!B94-B101 -- "- Key: appcast_benchmark_2023; ..."
     (workforce_insights.relevant_research list items leak the raw KB
     source key alongside Title/Publisher/Year, which already cover it).
  3. Market Intelligence!D107 -- "Occupation Key: warehousing_logistics"
     (workforce_insights.appcast_2026_benchmarks.occupation_benchmarks
     .occupation_key rendered as its raw Appcast occupation-bucket key).

The two local reference bundles (tests/test_bundle_qa_regression.py) never
exercise these code paths because `tools_regen_bundles` generates offline,
without live enrichment/KB data -- these dicts are simply absent/empty
locally, so bundle_qa's snake_case check has nothing to flag. This file
builds a minimal *enrichment-shaped* fixture (the exact dict shapes
data_synthesizer.fuse_competitive_intelligence /
fuse_workforce_insights produce, reproduced from the prod .xlsx and from
reading the producing code directly) to force those sections to render
in a fast, offline, network-free test, and asserts:
  (a) the rendered text is clean, human-readable text (no raw snake_case
      leak of the specific keys above), and
  (b) bundle_qa.run_bundle_qa's snake_case_leak check reports ZERO
      critical findings against the fixture-generated workbook -- the same
      check that caught these 10 findings on the real prod bundle.

Runs under pytest, or standalone:
``python3 tests/test_excel_enrichment_gated_snake_case.py``.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl  # noqa: E402

import bundle_qa  # noqa: E402
import excel_v2  # noqa: E402


# ---------------------------------------------------------------------------
# Fixture: the exact enrichment-gated shapes these 3 leaks came from
# ---------------------------------------------------------------------------
def _enrichment_shaped_data(**overrides) -> dict:
    """A plan data dict whose `_synthesized` carries the same
    competitive_intelligence / workforce_insights shapes
    data_synthesizer.fuse_competitive_intelligence /
    fuse_workforce_insights produce once live enrichment + KB data is
    present -- copied from those functions (data_synthesizer.py) and from
    the raw values observed in the real prod .xlsx."""
    data = {
        "client_name": "Amerigas",
        "company_name": "Amerigas",
        "industry": "logistics_supply_chain",
        "budget": "$150,000",
        "locations": ["Dallas, TX"],
        "roles": ["CDL Driver"],
        "target_roles": ["CDL Driver"],
        "campaign_duration": "3 months",
        "hire_volume": "50",
        "work_environment": "onsite",
        "_enriched": {},
        "_budget_allocation": {},
        "_synthesized": {
            # Shape produced by fuse_competitive_intelligence (data_synthesizer.py)
            "competitive_intelligence": {
                "company_profile": {"name": "Amerigas", "is_public": False},
                "competitors": {},
                "market_positioning": {
                    "industry_sector": "logistics_supply_chain",
                    "is_public_company": False,
                    "competitor_count": 0,
                    "has_sec_filings": False,
                },
            },
            # Shape produced by fuse_workforce_insights (data_synthesizer.py)
            "workforce_insights": {
                "relevant_research": [
                    {
                        "key": "appcast_benchmark_2023",
                        "title": (
                            "Recruitment Marketing Benchmark Report 2023 "
                            "(7th Annual)"
                        ),
                        "publisher": "Appcast",
                        "year": 2023,
                        "finding_count": 17,
                        "top_findings": ["Labor supply rebounded in 2022"],
                    },
                    {
                        "key": "talroo_frontline_hiring_2025",
                        "title": "Frontline & High-Volume Hiring Performance Data 2025",
                        "publisher": "Talroo",
                        "year": 2025,
                        "finding_count": 6,
                        "top_findings": [],
                    },
                ],
                "appcast_2026_benchmarks": {
                    "source": (
                        "Appcast 10th Annual Recruitment Marketing "
                        "Benchmark Report (2026)"
                    ),
                    "data_year": 2025,
                    "occupation_benchmarks": {
                        "occupation_key": "warehousing_logistics",
                        "cpa": 14.26,
                        "cph": 631,
                        "apply_rate": 0.055,
                        "cost_per_screen": 70,
                        "cost_per_interview": 155,
                        "cost_per_offer": 302,
                    },
                },
            },
        },
    }
    data.update(overrides)
    return data


def _build(data: dict):
    raw = excel_v2.generate_excel_v2(data)
    assert isinstance(raw, (bytes, bytearray)) and len(raw) > 0
    return raw


def _all_text_cells(wb) -> list[str]:
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value:
                    out.append(cell.value)
    return out


# ---------------------------------------------------------------------------
# 1. Industry Sector -- must resolve through INDUSTRY_LABEL_MAP
# ---------------------------------------------------------------------------
def test_industry_sector_value_is_humanized_not_raw_key():
    raw = _build(_enrichment_shaped_data())
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    texts = _all_text_cells(wb)
    joined = "\n".join(texts)

    assert "logistics_supply_chain" not in joined
    matches = [t for t in texts if "Industry Sector:" in t]
    assert matches, "Market Positioning line did not render"
    # Canonical shared_utils.INDUSTRY_LABEL_MAP label, not naive title-case.
    assert "Logistics & Supply Chain" in matches[0]


# ---------------------------------------------------------------------------
# 2. Data Provenance research list -- raw "Key:" component must be gone
# ---------------------------------------------------------------------------
def test_relevant_research_drops_raw_key_field():
    raw = _build(_enrichment_shaped_data())
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    texts = _all_text_cells(wb)
    joined = "\n".join(texts)

    assert "appcast_benchmark_2023" not in joined
    assert "talroo_frontline_hiring_2025" not in joined
    # Dropped entirely (not just humanized) -- the research-report list item
    # must never start with a "Key:" field. (Substring match on the whole
    # workbook would false-positive on the unrelated, legitimately-labeled
    # "Occupation Key:" line covered by test_occupation_key_value_is_humanized.)
    research_lines = [t for t in texts if "Recruitment Marketing Benchmark" in t]
    assert research_lines, "relevant_research list item did not render"
    assert "Key:" not in research_lines[0]

    # Title/Publisher/Year -- what a client should actually see -- survive.
    assert "Recruitment Marketing Benchmark Report 2023" in research_lines[0]
    assert "Publisher: Appcast" in research_lines[0]


# ---------------------------------------------------------------------------
# 3. Occupation Key -- must be humanized, not the raw Appcast bucket key
# ---------------------------------------------------------------------------
def test_occupation_key_value_is_humanized():
    raw = _build(_enrichment_shaped_data())
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    texts = _all_text_cells(wb)
    joined = "\n".join(texts)

    assert "warehousing_logistics" not in joined
    matches = [t for t in texts if "Occupation Key:" in t]
    assert matches, "Occupation Benchmarks line did not render"
    assert "Warehousing Logistics" in matches[0]
    # CPA/CPH acronyms must stay correctly cased (existing _TITLE_ACRONYMS
    # behavior for dict KEYS -- not touched by this fix, still must hold).
    assert "CPA:" in matches[0] and "CPH:" in matches[0]


# ---------------------------------------------------------------------------
# 4. bundle_qa itself: the actual linter that caught these on prod
# ---------------------------------------------------------------------------
def test_bundle_qa_reports_zero_snake_case_criticals_on_fixture_workbook():
    """Run the SAME check that flagged all 10 criticals on the real prod
    bundle against this fixture-generated workbook -- proves the fix closes
    the leak class the linter actually looks for, not just the 3 example
    strings asserted above."""
    data = _enrichment_shaped_data()
    xlsx_bytes = _build(data)

    findings = bundle_qa.run_bundle_qa(None, xlsx_bytes, data)
    snake_case_findings = [f for f in findings if f["code"] == "snake_case_leak"]
    assert not snake_case_findings, (
        f"bundle_qa still finds snake_case leaks in the enrichment-gated "
        f"sections: {snake_case_findings}"
    )


def test_bundle_qa_check_still_fires_on_a_genuine_new_leak():
    """Negative control, at the bundle_qa layer itself (not excel_v2): the
    fix must not have accidentally neutered the snake_case_leak detector --
    text that reaches the workbook via ANY path (not just _flatten_value,
    which now humanizes raw-identifier-shaped values) must still be
    flagged. Uses bundle_qa's own detection entry point directly, the same
    way tests/test_bundle_qa_regression.py's positive controls do."""
    units = [
        bundle_qa._TextUnit(
            "Some brand-new field: totally_unresolved_raw_key_value that "
            "did not go through excel_v2's dict flattening at all",
            "X!A1",
        )
    ]
    findings: list[dict] = []
    bundle_qa._check_text_patterns(units, {}, findings)
    assert any(f["code"] == "snake_case_leak" for f in findings), (
        "bundle_qa's snake_case_leak check no longer fires on a genuine "
        "leak -- the fix over-broadly suppressed the check."
    )


if __name__ == "__main__":
    _failures = 0
    for _name, _fn in sorted(globals().items()):
        if _name.startswith("test_") and callable(_fn):
            try:
                _fn()
                print(f"PASS {_name}")
            except AssertionError as exc:
                _failures += 1
                print(f"FAIL {_name}: {exc}")
            except Exception as exc:  # noqa: BLE001
                _failures += 1
                print(f"ERROR {_name}: {exc}")
    sys.exit(1 if _failures else 0)
